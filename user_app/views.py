import io
import json
import os
from datetime import datetime

import jdatetime
import openpyxl
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.contenttypes.models import ContentType
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum, ProtectedError
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import get_template
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.views.generic import TemplateView, CreateView, ListView, DetailView
from openpyxl.styles import Font, Alignment, PatternFill
from pypdf import PdfWriter
from weasyprint import CSS, HTML

from admin_panel.forms import UnifiedChargePaymentForm
from middleAdmin_panel.views import middle_admin_required
from notifications.models import Notification, SupportUser
from polls.templatetags.poll_extras import show_jalali
from user_app import helper
from admin_panel.models import Announcement, UnifiedCharge, MessageToUser, MessageReadStatus, Expense, Fund
from user_app.forms import LoginForm, MobileLoginForm, UserPayForm, UserPayMoneyForm
from user_app.models import User, Unit, Bank, MyHouse, CalendarNote, UserPayMoney, UserPayMoneyDocument

# def index(request):
#     form = LoginForm(request.POST or None)
#
#     if request.method == 'POST' and form.is_valid():
#         mobile = form.cleaned_data['mobile']
#         password = form.cleaned_data['password']
#         print("Mobile:", mobile)
#         print("Password entered:", password)
#
#         user = authenticate(request, username=mobile, password=password)
#         # print("Authenticated user:", user)
#
#         if user:
#             if user.is_superuser:
#                 messages.error(request, 'شما مجوز ورود از این صفحه را ندارید.')
#             elif not user.is_active:
#                 messages.error(request, 'حساب کاربری شما غیرفعال است.')
#             else:
#                 login(request, user)
#
#                 if user.is_middle_admin:
#                     # بررسی ثبت ساختمان
#                     has_house = MyHouse.objects.filter(user=user).exists()
#                     if has_house:
#                         return redirect('middle_admin_dashboard')
#                     else:
#                         return redirect('middle_manage_house')
#
#                 return redirect('user_panel')
#         else:
#             messages.error(request, 'ورود ناموفق: شماره موبایل یا کلمه عبور نادرست است.')
#
#     return render(request, 'index.html', {'form': form})
from django.db.models import Exists, OuterRef


@login_required
def switch_to_manager(request):
    """
    برگشت به محیط مدیر ساختمان
    """
    # فقط وقتی قبلا سوییچ شده به ساکن
    if request.session.get('active_context') == 'resident':
        # پاک کردن session ساکن
        request.session.pop('active_context', None)
        request.session.pop('active_unit_id', None)
        request.session.pop('active_building_id', None)

    # redirect به داشبورد مدیر
    return redirect('middle_admin_dashboard')


def index(request):
    form = LoginForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        mobile = form.cleaned_data['mobile']
        password = form.cleaned_data['password']

        user = authenticate(request, username=mobile, password=password)

        if user:
            if user.is_superuser:
                messages.error(request, 'شما مجوز ورود از این صفحه را ندارید.')

            elif not user.is_active:
                messages.error(request, 'حساب کاربری شما غیرفعال است.')

            # ⛔ مالک با مستاجر فعال
            elif Unit.objects.filter(
                    user=user,
                    renters__renter_is_active=True
            ).exists():
                messages.error(
                    request,
                    'برای واحد شما مستاجر فعال ثبت شده است و امکان ورود مالک وجود ندارد.'
                )

            else:
                login(request, user)

                if user.is_middle_admin:
                    has_house = MyHouse.objects.filter(user=user).exists()
                    if has_house:
                        return redirect('middle_admin_dashboard')
                    else:
                        return redirect('middle_manage_house')

                return redirect('user_panel')


        else:
            messages.error(request, 'ورود ناموفق: شماره موبایل یا کلمه عبور نادرست است.')

    return render(request, 'index.html', {'form': form})


def mobile_login(request):
    form = MobileLoginForm(request.POST or None)
    if request.method == 'POST':
        mobile = request.POST.get('mobile')
        if mobile:
            user = User.objects.filter(mobile=mobile).first()
            if user:
                otp = helper.get_random_otp()
                # helper.send_otp(mobile, otp)  # Uncomment to send OTP
                print("OTP1:", otp)
                user.otp = otp
                user.otp_create_time = timezone.now()
                user.save()

                request.session['user_mobile'] = user.mobile
                return HttpResponseRedirect(reverse('verify_otp'))
            else:
                messages.error(request, 'کاربر با این شماره همراه یافت نشد!')
                return redirect(reverse('mobile_login'))
    return render(request, 'mobile_Login.html', {'form': form})


def verify_otp(request):
    mobile = request.session.get('user_mobile')

    if not mobile:
        return redirect('mobile_login')

    try:
        user = User.objects.get(mobile=mobile)

        if request.method == "POST":
            # Combine OTP inputs from multiple fields
            otp_input = ''.join([request.POST.get(f'otp{i}', '') for i in range(1, 6)])

            # Check OTP expiration
            if not helper.check_otp_expiration(user.mobile):
                messages.error(request, "رمز یکبار مصرف منقضی شده است. تلاش مجدد!")
                return redirect('mobile_login')

            # Validate OTP
            if str(user.otp) != otp_input:
                messages.error(request, "رمز یکبار مصرف وارد شده اشتباه است!")
                return redirect('verify_otp')
            else:
                login(request, user)
                # messages.success(request, "ورود موفق")
                return redirect('user_panel')

        return render(
            request,
            'verify_login.html',
            {
                "mobile": mobile,
                "user": user,
            }
        )

    except User.DoesNotExist:
        messages.error(request, "کاربری با این شماره موبایل یافت نشد.")
        return redirect('mobile_login')


def resend_otp(request):
    try:
        mobile = request.session.get('user_mobile')
        user = User.objects.get(mobile=mobile)

        new_otp = helper.get_random_otp()
        # helper.send_otp(mobile, new_otp)
        print(f' new_otp: {new_otp}')
        user.otp = new_otp
        user.otp_create_time = timezone.now()
        user.save()
        messages.add_message(request, messages.SUCCESS, "رمز یکبار مصرف جدید ارسال شد!")

        # Redirect back to the verification page
        return HttpResponseRedirect(reverse('verify_otp'))

    except User.DoesNotExist:
        messages.error(request, 'User does not exist.')
        return HttpResponseRedirect(reverse('index'))


def logout_user(request):
    logout(request)
    return redirect('index')


def site_header_component(request):
    context = {
        'user': request.user,
    }
    return render(request, 'partials/notification_template.html', context)


# @login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def user_panel(request):
    user = request.user
    # اگر سوپریوزر در حال impersonate است، اجازه ورود بده
    if request.session.get('impersonator_id'):
        # impersonation فعال است، user همان target است
        pass
    # else:
    #     # کاربر معمولی فقط ساکن باشد
    #     if not user.is_unit:
    #         return redirect('/index/')  # یا مسیر login خودتان

    # -------------------------
    # Unified Charges
    # -------------------------
    units = Unit.objects.filter(
        is_active=True
    ).filter(
        Q(user=user) |  # مالک
        Q(renters__user=user, renters__renter_is_active=True)  # مستاجر فعال
    ).distinct()

    unified_qs = UnifiedCharge.objects.filter(
        unit__in=units,
        send_notification=True,
    ).select_related('unit')

    paid_unified_qs = unified_qs.filter(is_paid=True)
    print(paid_unified_qs.count())
    unpaid_unified_qs = unified_qs.filter(is_paid=False)

    # Update penalty ONLY for unpaid charges
    # for charge in unpaid_unified_qs:
    #     charge.update_penalty()

    # -------------------------
    # STATISTICS
    # -------------------------
    total_charge = unified_qs.count()
    total_charge_unpaid = unpaid_unified_qs.count()

    total_paid_amount = (
            paid_unified_qs.aggregate(
                total=Sum("total_charge_month")
            )["total"] or 0
    )

    # -------------------------
    # LAST CHARGES
    # -------------------------
    last_charges = unified_qs.order_by('-created_at')[:7]

    # -------------------------
    # TICKETS
    # -------------------------
    tickets = SupportUser.objects.filter(user=user).order_by('-created_at')[:5]
    ticket_count = tickets.count()

    # -------------------------
    # UNITS & ANNOUNCEMENTS
    # -------------------------
    if user.is_middle_admin:
        units = (
            Unit.objects
            .filter(user__manager=user, is_active=True)
            .prefetch_related('renters')
        )
        announcements = (
            Announcement.objects
            .filter(user=user, is_active=True)
            .order_by('-created_at')[:5]
        )
    else:
        units = (
            Unit.objects
            .filter(user=user, is_active=True)
            .prefetch_related('renters')
        )
        announcements = (
            Announcement.objects
            .filter(user=user.manager, is_active=True)
            .order_by('-created_at')[:5]
        )

    # -------------------------
    # ACTIVE RENTERS PER UNIT
    # -------------------------
    units_with_details = [
        {
            "unit": unit,
            "active_renter": unit.renters.filter(renter_is_active=True).first()
        }
        for unit in units
    ]

    # -------------------------
    # CONTEXT
    # -------------------------
    context = {
        "user": user,
        "units": units,
        "units_with_details": units_with_details,

        "tickets": tickets,
        "ticket": ticket_count,

        "announcements": announcements,

        "total_charge": total_charge,
        "total_charge_unpaid": total_charge_unpaid,
        "total_paid_amount": total_paid_amount,

        "last_charges": last_charges,
        "unpaid_charges": unpaid_unified_qs,
    }

    return render(request, 'partials/home_template.html', context)


# def core_announce(request):
#     user = request.user
#     announcements = Announcement.objects.filter(
#             user=user,
#             is_active=True
#         )
#     return render(request, 'partials/core_template.html', {'announcements': announcements})


# ==================================
# ======================== Charges ===================
def fetch_user_charges(request):
    user = request.user
    query = request.GET.get('q', '').strip()
    paginate = request.GET.get('paginate', '20')

    # واحدهایی که کاربر به‌عنوان مالک یا مستاجر فعال دارد
    units = Unit.objects.filter(
        is_active=True
    ).filter(
        Q(user=user) |  # مالک
        Q(renters__user=user, renters__renter_is_active=True)  # مستاجر فعال
    ).distinct()

    charges = UnifiedCharge.objects.filter(
        unit__in=units,
        send_notification=True,
    ).select_related('unit')

    for charge in charges:
        charge.update_penalty()

    if query:
        charges = charges.filter(
            Q(title__icontains=query) |
            Q(details__icontains=query)
        )

    charges = charges.order_by('-created_at')

    try:
        paginate = int(paginate)
    except ValueError:
        paginate = 20

    paginator = Paginator(charges, paginate)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'query': query,
        'paginate': paginate,
    }
    return render(request, 'manage_charges.html', context)


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_charge_pdf(request, pk):
    charge = get_object_or_404(UnifiedCharge, pk=pk)

    house = None
    bank = None

    # 1️⃣ پیدا کردن خانه از روی واحد
    if charge.unit:
        # اگر Unit به MyHouse وصل است
        # ⚠️ اگر Unit هنوز FK به MyHouse نداره، باید اضافه بشه
        house = charge.unit.myhouse  # ← اسم FK واقعی توی مدل Unit

    # 2️⃣ بانک پیش‌فرض مربوط به همان خانه
    if house:
        bank = Bank.objects.filter(house=house, is_default=True, is_active=True).first()

    template = get_template('middleCharge/single_charge_pdf.html')
    html_string = template.render({'charge': charge,
                                   'house': house,
                                   'bank': bank,
                                   'font_url': request.build_absolute_uri('/static/fonts/Vazir.ttf')
                                   })

    css = CSS(string=f"""
        @page {{ size: A5 portrait; margin: 1cm; }}
        body {{
            font-family: 'Vazir', sans-serif;
        }}
        @font-face {{
            font-family: 'Vazira', sans-serif;
    
        }}
    """)

    pdf_file = io.BytesIO()
    HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[css])
    pdf_file.seek(0)

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment;filename=charge_unit:{charge.unit.unit}.pdf'
    return response


# ==============================================

@method_decorator(login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN), name='dispatch')
class AnnouncementListView(ListView):
    model = Announcement
    template_name = 'manage_announcement.html'
    context_object_name = 'announcements'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        user = self.request.user

        if user.is_middle_admin:
            queryset = Announcement.objects.filter(
                user=user,
                is_active=True
            )
        elif hasattr(user, 'manager') and user.manager:
            queryset = Announcement.objects.filter(
                user=user.manager,
                is_active=True
            )
        else:
            return Announcement.objects.none()

        query = self.request.GET.get('q', '')
        if query:
            queryset = queryset.filter(title__icontains=query)

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@method_decorator(login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN), name='dispatch')
class MessageListView(ListView):
    template_name = 'user_message.html'
    context_object_name = 'user_messages'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        user = self.request.user
        query = self.request.GET.get('q', '')
        units = Unit.objects.filter(
            is_active=True
        ).filter(
            Q(user=user) |  # مالک
            Q(renters__user=user, renters__renter_is_active=True)  # مستاجر فعال
        ).distinct()

        # پیام‌های فعال کاربر
        queryset = MessageToUser.objects.filter(
            notified_units__in=units,
            is_active=True
        ).distinct()

        # آپدیت read_status برای هر پیام و هر واحد کاربر
        for msg in queryset:
            for unit in units:  # هر واحد کاربر
                read_status, created = MessageReadStatus.objects.get_or_create(
                    message=msg,
                    unit=unit
                )
                if not read_status.is_read:
                    read_status.is_read = True
                    read_status.read_at = timezone.now()
                    read_status.save()

        # فیلتر جستجو
        if query:
            queryset = queryset.filter(
                Q(user__full_name__icontains=query) |
                Q(title__icontains=query) |
                Q(message__icontains=query)
            ).distinct()

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def user_profile(request):
    user = request.user

    # اگر فرم ارسال شد
    if request.method == 'POST':
        password_form = PasswordChangeForm(user, request.POST)
        if password_form.is_valid():
            password_form.save()
            update_session_auth_hash(request, user)  # مهم: کاربر بعد از تغییر رمز لاگ‌اوت نشود
            messages.success(request, 'رمز عبور با موفقیت تغییر کرد!')
            return redirect('user_profile')  # redirect باعث می‌شود پیام ظاهر شود
        else:
            messages.warning(request, 'رمز عبور باید حداقل 8 رقم و شامل حروف و اعداد باشد. مجددا بررسی فرمایید.')
    else:
        password_form = PasswordChangeForm(user)
    # اطلاعات واحد
    unit = Unit.objects.filter(
        is_active=True
    ).filter(
        Q(user=user) |  # مالک
        Q(renters__user=user, renters__renter_is_active=True)  # مستاجر فعال
    ).distinct().first()

    context = {
        'user_obj': user,
        'unit': unit,
        'password_form': password_form,
    }
    return render(request, 'my_profile.html', context)


# ===================== User Pay Money ====================================
@method_decorator(login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN), name='dispatch')
class UserPayMoneyViewCreateView(CreateView):
    model = UserPayMoney
    template_name = 'user_pay.html'
    form_class = UserPayMoneyForm
    success_url = reverse_lazy('user_pay_money')

    def form_valid(self, form):
        form.instance.user = self.request.user

        unit = Unit.objects.filter(
            Q(user=self.request.user) |  # مالک
            Q(renters__user=self.request.user),  # مستأجر
            is_active=True
        ).select_related('myhouse').distinct().first()

        if not unit:
            messages.error(self.request, 'هیچ واحد فعالی برای شما ثبت نشده است')
            return redirect(self.success_url)

        form.instance.unit = unit
        form.instance.house = unit.myhouse if hasattr(unit, 'myhouse') else None

        try:
            with transaction.atomic():
                self.object = form.save(commit=False)
                self.object.is_paid = False
                self.object.save()

                # ذخیره فایل‌ها
                for f in self.request.FILES.getlist('document'):
                    UserPayMoneyDocument.objects.create(
                        user_pay=self.object,
                        document=f
                    )

            messages.success(self.request, 'پرداخت با موفقیت ثبت شد (در انتظار پرداخت)')
            return redirect(self.success_url)

        except Exception as e:
            messages.error(self.request, f'خطا در ثبت پرداخت: {e}')
            return self.form_invalid(form)

    def get_queryset(self):
        queryset = UserPayMoney.objects.filter(user=self.request.user).order_by('-created_at')

        # فیلتر بر اساس بانک
        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank__id=bank_id)

        # فیلتر بر اساس amount
        amount = self.request.GET.get('amount')
        if amount:
            queryset = queryset.filter(amount__icontains=amount)

        # فیلتر بر اساس description
        description = self.request.GET.get('description')
        if description:
            queryset = queryset.filter(description__icontains=description)

        # فیلتر بر اساس details
        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس تاریخ
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')
        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y-%m-%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(register_date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y-%m-%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(register_date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')

        is_paid = self.request.GET.get('is_paid')
        if is_paid == '1':
            queryset = queryset.filter(is_paid=True)
        elif is_paid == '0':
            queryset = queryset.filter(is_paid=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        incomes = self.get_queryset()  # از get_queryset برای دریافت داده‌های فیلتر شده استفاده می‌کنیم
        paginator = Paginator(incomes, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_user_pays'] = UserPayMoney.objects.filter(user=self.request.user).count()
        unit = Unit.objects.filter(
            Q(user=self.request.user) |
            Q(renters__user=self.request.user),
            is_active=True
        ).first()

        context['banks'] = Bank.objects.filter(
            house=unit.myhouse,
            is_active=True
        )
        managed_users = User.objects.filter(Q(manager=self.request.user) | Q(pk=self.request.user.pk))
        context['units'] = Unit.objects.filter(is_active=True, user__in=managed_users)

        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def pay_user_edit(request, pk):
    pay = get_object_or_404(
        UserPayMoney,
        pk=pk,
        user=request.user
    )

    if request.method != 'POST':
        return redirect('user_pay_money')

    form = UserPayMoneyForm(
        request.POST,
        request.FILES,
        instance=pay
    )

    if not form.is_valid():
        messages.error(request, 'خطا در ویرایش پرداخت.')
        return redirect('user_pay_money')

    try:
        with transaction.atomic():
            pay = form.save()
            pay.user = request.user
            pay.save()

            files = request.FILES.getlist('document')
            for f in files:
                UserPayMoneyDocument.objects.create(
                    user_pay=pay,
                    document=f
                )

        messages.success(request, 'پرداخت با موفقیت ویرایش شد.')
        return redirect('user_pay_money')

    except Exception:
        messages.error(request, 'خطا در ویرایش پرداخت.')
        return redirect('user_pay_money')


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def user_pay_delete(request, pk):
    pay = get_object_or_404(UserPayMoney, id=pk)
    try:
        with transaction.atomic():
            # حذف Fund مربوطه
            pay_ct = ContentType.objects.get_for_model(UserPayMoney)
            Fund.objects.filter(content_type=pay_ct, object_id=pay.id).delete()

            pay.delete()

        messages.success(request, 'پرداخت با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, "امکان حذف وجود ندارد!")
    except Exception as e:
        messages.error(request, f"خطا در حذف: {str(e)}")

    return redirect(reverse('user_pay_money'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def user_delete_pay_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        pay_id = request.POST.get('pay_id')

        if not image_url or not pay_id:
            return JsonResponse({'status': 'error', 'message': 'پرداخت یافت نشد'})

        try:
            pay = get_object_or_404(UserPayMoney, id=pay_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = UserPayMoneyDocument.objects.filter(user_pay=pay, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except UserPayMoney.DoesNotExist:

            return JsonResponse({'status': 'error', 'message': 'پرداخت یافت نشد'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_user_pay_money_pdf(request):
    # خانه مرتبط با کاربر (در صورت نیاز)
    house = None
    if request.user.is_authenticated:
        # اگر کاربر مستأجر است، خانه‌ای که در آن واحد دارد
        house = MyHouse.objects.filter(
            Q(residents=request.user) | Q(user=request.user)
        ).order_by('-created_at').first()

    # Queryset اصلی بر اساس واحد
    payments = UserPayMoney.objects.filter(user=request.user).order_by('-register_date')

    # فیلترهای GET
    filter_fields = {
        'payment_date': 'payment_date__icontains',
        'description': 'title__icontains',  # عنوان پرداخت
        'payment_gateway': 'payment_gateway__icontains',
        'amount': 'amount__icontains',
        'transaction_reference': 'transaction_reference__icontains',
        'bank': 'bank__id__icontains',
        'is_paid': 'is_paid',
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            payments = payments.filter(**{lookup: value})

    # تنظیمات PDF
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        body {{
            font-family: 'BYekan', sans-serif;
        }}
        @font-face {{
            font-family: 'BYekan';
            src: url('{font_url}');
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 5px;
            text-align: center;
        }}
        th {{
            background-color: #FFD700;
        }}
    """)

    # Render template
    template = get_template("user_pay_pdf.html")
    context = {
        'payments': payments,
        'font_path': font_url,
        'today': datetime.now(),
        'house': house,
    }

    html = template.render(context)
    pdf_file = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[css])
    pdf_file.seek(0)

    # Response
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="userpaymoney_report.pdf"'
    return response


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_user_pay_money_excel(request):
    # Queryset اصلی بر اساس واحد
    payments = UserPayMoney.objects.filter(user=request.user).order_by('-register_date')

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "units"
    ws.sheet_view.rightToLeft = True

    # Title
    title_cell = ws.cell(row=1, column=1, value=f"لیست پرداخت های من")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    # Headers
    headers = [' بانک', 'شرح', 'مبلغ', 'تاریخ ثبت', 'تاریخ پرداخت', 'شماره رهگیری', 'روش پرداخت']
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for row_num, pay in enumerate(payments, start=3):
        ws.cell(row=row_num, column=1, value=pay.bank.bank_name if pay.bank else '')
        ws.cell(row=row_num, column=2, value=pay.description)
        ws.cell(row=row_num, column=3, value=pay.amount)
        ws.cell(row=row_num, column=4, value=show_jalali(pay.register_date))
        ws.cell(row=row_num, column=5, value=show_jalali(pay.payment_date))
        ws.cell(row=row_num, column=6, value=pay.transaction_reference)
        ws.cell(row=row_num, column=7, value=pay.payment_gateway)

    # Return response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=userpaymoney_report.xlsx'
    wb.save(response)
    return response
