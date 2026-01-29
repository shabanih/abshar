import io
from collections import defaultdict
from datetime import datetime
from decimal import Decimal

import jdatetime
import openpyxl
from django.apps import apps
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.contenttypes.models import ContentType
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Q, Sum, F, Count
from django.db.models.functions import Coalesce
from django.http import HttpResponse, Http404
from django.shortcuts import render, get_object_or_404, redirect
from django.template.context_processors import static
from django.template.loader import get_template, render_to_string
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.generic import ListView
from django.views.generic.edit import FormMixin

from admin_panel.forms import UnifiedChargePaymentForm
from admin_panel.models import Fund, Expense, Income, Property, ExpenseCategory, IncomeCategory, Maintenance, \
    UnifiedCharge, PersonCharge, FixPersonCharge, FixAreaCharge, AreaCharge, \
    FixCharge, ChargeByPersonArea, ChargeFixVariable, ChargeByFixPersonArea, PayMoney, ReceiveMoney, AdminFund
from middleAdmin_panel.views import middle_admin_required
from polls.templatetags.poll_extras import show_jalali
from user_app.forms import UnitReportForm
from user_app.models import Unit, MyHouse, UnitResidenceHistory, Bank, User
from openpyxl.styles import PatternFill, Font, Alignment
from pypdf import PdfWriter
from weasyprint import HTML, CSS


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middleAdmin_turnover(request):
    manager = request.user
    query = request.GET.get('q', '').strip()
    paginate = int(request.GET.get('paginate', 20) or 20)
    paginate = paginate if paginate > 0 else 20

    funds = AdminFund.objects.filter(user=manager).order_by('-payment_date')

    if query:
        funds = funds.filter(
            Q(payment_description__icontains=query) |
            Q(transaction_no__icontains=query)
        )

    total_fund = (
            AdminFund.objects
            .filter(user=request.user, is_paid=True)
            .aggregate(total=Sum('amount'))['total']
            or Decimal('0')
    )

    paginator = Paginator(funds, paginate)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'middleAdmin_turnover.html', {
        'funds': page_obj,
        'query': query,
        'paginate': paginate,
        'page_obj': page_obj,
        'total_fund': total_fund

    })


def to_jalali(date_obj):
    if not date_obj:
        return ''
    jalali_date = jdatetime.date.fromgregorian(date=date_obj)
    return jalali_date.strftime('%Y/%m/%d')


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def fund_middle_turnover(request):
    manager = request.user
    query = request.GET.get('q', '').strip()
    paginate = int(request.GET.get('paginate', 20) or 20)
    paginate = paginate if paginate > 0 else 20

    funds = Fund.objects.select_related('bank', 'content_type', 'unit').filter(
        Q(unit__user=manager) | Q(unit__user__manager=manager) | Q(user=manager)
    ).order_by('-payment_date')


    if query:
        funds = funds.filter(
            Q(payment_description__icontains=query) |
            Q(payer_name__icontains=query) |
            Q(receiver_name__icontains=query) |
            Q(transaction_no__icontains=query) |
            Q(creditor_amount__icontains=query) |
            Q(debtor_amount__icontains=query) |
            Q(bank__bank_name__icontains=query)
        )

    totals = funds.aggregate(
        total_debtor=Sum('debtor_amount'),
        total_creditor=Sum('creditor_amount'),
    )

    balance = (totals['total_debtor'] or 0) - (totals['total_creditor'] or 0)

    paginator = Paginator(funds, paginate)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'fund_middle_turnover.html', {
        'funds': page_obj,
        'query': query,
        'paginate': paginate,
        'page_obj': page_obj,
        'totals': totals,
        'balance': balance,
    })


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_middle_report_pdf(request):
    manager = request.user

    # Queryset اصلی
    funds = Fund.objects.filter(Q(user=manager) | Q(user__manager=manager))
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    # فیلترها
    query = request.GET.get('q', '').strip()
    if query:
        funds = funds.filter(
            Q(payment_description__icontains=query) |
            Q(unit__unit__icontains=query) |
            Q(payer_name__icontains=query) |
            Q(receiver_name__icontains=query) |
            Q(payment_gateway__icontains=query) |
            Q(debtor_amount__icontains=query) |
            Q(creditor_amount__icontains=query)
        )
    # for field, lookup in filter_fields.items():
    #     value = request.GET.get(field)
    #     if value:
    #         funds = funds.filter(**{lookup: value})

    # محاسبه totals و balance
    totals = funds.aggregate(
        total_income=Sum('debtor_amount'),
        total_expense=Sum('creditor_amount'),
    )
    balance = (totals['total_income'] or 0) - (totals['total_expense'] or 0)

    # PDF settings
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        body {{
            font-family: 'BYekan', sans-serif;
        }}
        @font-face {{
            font-family: 'BYekan';
            src: url('{font_url}');
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 5px;
            text-align: center;
        }}
        th {{
            background-color: #FFD700;
        }}
    """)

    template = get_template("middle_report_pdf.html")
    context = {
        'funds': funds,
        'query': query,
        'totals': totals,
        'balance': balance,
        'font_path': font_url,
        'today': datetime.now(),
        'house': house,
    }
    html = template.render(context)
    pdf_file = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[css])
    pdf_file.seek(0)

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="fund_report.pdf"'
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_middle_report_excel(request):
    manager = request.user
    report = Fund.objects.filter(Q(user=manager) | Q(user__manager=manager)).order_by('-payment_date')

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "units"
    ws.sheet_view.rightToLeft = True

    # Title
    title_cell = ws.cell(row=1, column=1, value=f"گردش مالی صندوق ")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # Headers
    headers = [' بانک', 'تاریخ پرداخت', 'شرح', 'پرداخت کننده/واریز کننده', 'روش پرداخت', 'بدهکار', 'بستانکار']

    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for row_num, fund in enumerate(report, start=3):
        ws.cell(row=row_num, column=1, value=f"{fund.bank.bank_name} - {fund.bank.account_no}")
        ws.cell(row=row_num, column=2, value=show_jalali(fund.payment_date))
        ws.cell(row=row_num, column=3, value=fund.payment_description)
        ws.cell(row=row_num, column=4, value=f"{fund.payer_name} - {fund.receiver_name}")
        ws.cell(row=row_num, column=5, value=fund.payment_gateway)
        ws.cell(row=row_num, column=6, value=fund.debtor_amount)
        ws.cell(row=row_num, column=7, value=fund.creditor_amount)

    # Return response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=fund-report.xlsx'
    wb.save(response)
    return response


# -------------------- Admin ------------------------------
@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def admin_fund_turnover(request):
    query = request.GET.get('q', '').strip()
    paginate = int(request.GET.get('paginate', 20) or 20)
    paginate = paginate if paginate > 0 else 20

    funds = (
        Fund.objects
        .select_related('bank', 'content_type')
        .all()
        .order_by('-payment_date')
    )

    if query:
        funds = funds.filter(
            Q(payment_description__icontains=query) |
            Q(payer_name__icontains=query) |
            Q(receiver_name__icontains=query) |
            Q(transaction_no__icontains=query) |
            Q(creditor_amount__icontains=query) |
            Q(debtor_amount__icontains=query) |
            Q(bank__bank_name__icontains=query)
        )

    totals = funds.aggregate(
        total_debtor=Sum('debtor_amount'),
        total_creditor=Sum('creditor_amount'),
    )

    balance = (totals['total_debtor'] or 0) - (totals['total_creditor'] or 0)

    paginator = Paginator(funds, paginate)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'admin_reports/admin_fund_turnover.html', {
        'funds': page_obj,
        'query': query,
        'paginate': paginate,
        'page_obj': page_obj,
        'totals': totals,
        'balance': balance,
    })


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def admin_export_middle_report_pdf(request):

    # Queryset اصلی
    funds = Fund.objects.all()
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    # فیلترها
    query = request.GET.get('q', '').strip()
    if query:
        funds = funds.filter(
            Q(payment_description__icontains=query) |
            Q(unit__unit__icontains=query) |
            Q(payer_name__icontains=query) |
            Q(receiver_name__icontains=query) |
            Q(payment_gateway__icontains=query) |
            Q(debtor_amount__icontains=query) |
            Q(creditor_amount__icontains=query)
        )

    # محاسبه totals و balance
    totals = funds.aggregate(
        total_income=Sum('debtor_amount'),
        total_expense=Sum('creditor_amount'),
    )
    balance = (totals['total_income'] or 0) - (totals['total_expense'] or 0)

    # PDF settings
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        body {{
            font-family: 'BYekan', sans-serif;
        }}
        @font-face {{
            font-family: 'BYekan';
            src: url('{font_url}');
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 5px;
            text-align: center;
        }}
        th {{
            background-color: #FFD700;
        }}
    """)

    template = get_template("admin_report_pdf.html")
    context = {
        'funds': funds,
        'query': query,
        'totals': totals,
        'balance': balance,
        'font_path': font_url,
        'today': datetime.now(),
        'house': house,
    }
    html = template.render(context)
    pdf_file = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[css])
    pdf_file.seek(0)

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="admin_fund_report.pdf"'
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def admin_export_middle_report_excel(request):
    manager = request.user
    report = Fund.objects.filter(Q(user=manager) | Q(user__manager=manager)).order_by('-payment_date')

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "units"
    ws.sheet_view.rightToLeft = True

    # Title
    title_cell = ws.cell(row=1, column=1, value=f"گردش مالی صندوق ")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # Headers
    headers = [' بانک', 'تاریخ پرداخت', 'شرح', 'پرداخت کننده/واریز کننده', 'روش پرداخت', 'بدهکار', 'بستانکار']

    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for row_num, fund in enumerate(report, start=3):
        ws.cell(row=row_num, column=1, value=f"{fund.bank.bank_name} - {fund.bank.account_no}")
        ws.cell(row=row_num, column=2, value=show_jalali(fund.payment_date))
        ws.cell(row=row_num, column=3, value=fund.payment_description)
        ws.cell(row=row_num, column=4, value=f"{fund.payer_name} - {fund.receiver_name}")
        ws.cell(row=row_num, column=5, value=fund.payment_gateway)
        ws.cell(row=row_num, column=6, value=fund.debtor_amount)
        ws.cell(row=row_num, column=7, value=fund.creditor_amount)

    # Return response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=fund-report.xlsx'
    wb.save(response)
    return response


# ========================================================================
@method_decorator(middle_admin_required, name='dispatch')
class UnitReportsTurnOver(FormMixin, ListView):
    model = Fund
    form_class = UnitReportForm
    template_name = 'unit_reports.html'
    context_object_name = 'transactions'
    paginate_by = 50

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_queryset(self):
        return Fund.objects.none()

    def form_valid(self, form):
        unit = form.cleaned_data['unit']

        # همیشه مالک
        # unit_user = unit.user

        transactions = Fund.objects.filter(
            unit=unit
        ).order_by('-payment_date')

        # ⚡ حتما object_list ست شود
        self.object_list = transactions

        totals = Fund.objects.filter(unit=unit).aggregate(
            total_income=Sum('debtor_amount'),
            total_expense=Sum('creditor_amount'),
        )

        balance = (totals['total_income'] or 0) - (totals['total_expense'] or 0)

        context = self.get_context_data(
            form=form,
            transactions=transactions,
            selected_unit=unit,
            balance=balance,
            totals=totals,
        )
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            return self.form_valid(form)
        return self.form_invalid(form)


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_units_report_excel(request):
    unit_id = request.GET.get('unit')
    unit = get_object_or_404(Unit, pk=unit_id)
    unit_user = unit.user

    report = Fund.objects.filter(user=unit_user).order_by('doc_number')

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "units"
    ws.sheet_view.rightToLeft = True

    # Title
    title_cell = ws.cell(row=1, column=1, value=f"گردش مالی واحد {unit.unit}")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Headers
    headers = [' بانک', 'تاریخ پرداخت', 'شرح', 'روش پرداخت', 'بدهکار', 'بستانکار']
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for row_num, fund in enumerate(report, start=3):
        ws.cell(row=row_num, column=1, value=f"{fund.bank.bank_name} - {fund.bank.account_no}")
        ws.cell(row=row_num, column=2, value=show_jalali(fund.payment_date))
        ws.cell(row=row_num, column=3, value=fund.payment_description)
        ws.cell(row=row_num, column=4, value=fund.payment_gateway)
        ws.cell(row=row_num, column=5, value=fund.debtor_amount)
        ws.cell(row=row_num, column=6, value=fund.creditor_amount)

    # Return response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=fund_unit_{unit.unit}_report.xlsx'
    wb.save(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_units_report_pdf(request):
    unit_id = request.GET.get('unit')
    unit = get_object_or_404(Unit, pk=unit_id)

    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    # Queryset اصلی بر اساس واحد (کاربر مالک)
    funds = Fund.objects.filter(unit=unit).order_by('payment_date')

    # فیلترها از GET
    filter_fields = {
        'payment_date': 'payment_date__icontains',
        'payment_description': 'payment_description__icontains',
        'payment_gateway': 'payment_gateway__icontains',
        'debtor_amount': 'debtor_amount__icontains',
        'creditor_amount': 'creditor_amount__icontains',
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            funds = funds.filter(**{lookup: value})

    # PDF settings
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        body {{
            font-family: 'BYekan', sans-serif;
        }}
        @font-face {{
            font-family: 'BYekan';
            src: url('{font_url}');
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 5px;
            text-align: center;
        }}
        th {{
            background-color: #FFD700;
        }}
    """)

    # Render template
    template = get_template("unit_report_pdf.html")
    context = {
        'funds': funds,
        'unit': unit,
        'font_path': font_url,
        'today': datetime.now(),
        'house': house,
    }

    html = template.render(context)
    pdf_file = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[css])
    pdf_file.seek(0)

    # Response
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="fund_unit_{unit.unit}_report.pdf"'
    return response


# ======================================================================================
@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def fund_turnover_user(request):
    user = request.user
    query = request.GET.get('q', '').strip()
    paginate = request.GET.get('paginate', '20')

    # واحدهایی که کاربر به آن دسترسی دارد
    user_units = Unit.objects.filter(
        Q(user=user) |  # مالک
        Q(renters__user=user, renters__renter_is_active=True)  # مستاجر فعال
    ).distinct()

    # مدیر ساختمان → همه واحدهای ساختمان
    if user.is_middle_admin:
        user_units = Unit.objects.filter(
            myhouse__residents=user,
            is_active=True
        ).distinct()

    funds = Fund.objects.filter(
        unit__in=user_units,
        user=request.user,  # پرداخت‌کننده خود کاربر
        is_initial=False  # حذف افتتاحیه‌ها
    )

    # جستجو
    if query:
        funds = funds.filter(
            Q(payment_description__icontains=query) |
            Q(payment_gateway__icontains=query) |
            Q(transaction_no__icontains=query) |
            Q(payment_date__icontains=query) |
            Q(amount__icontains=query)
        )

    funds = funds.order_by('-created_at')

    # مجموع
    total_amount = funds.aggregate(total=Sum('amount'))['total'] or 0

    # پیجینیشن
    try:
        paginate = int(paginate)
        if paginate <= 0:
            paginate = 20
    except ValueError:
        paginate = 20

    paginator = Paginator(funds, paginate)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'fund_turnover_user.html', {
        'funds': page_obj,
        'query': query,
        'paginate': paginate,
        'page_obj': page_obj,
        'total_amount': total_amount,
    })



@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_user_report_excel(request):
    user = request.user

    report = Fund.objects.filter(user=user).order_by('doc_number')

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "units"
    ws.sheet_view.rightToLeft = True

    # Title
    title_cell = ws.cell(row=1, column=1, value=f"لیست تراکنش های من ")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # Headers
    headers = ['تاریخ پرداخت', 'شرح', 'روش پرداخت', 'شماره تراکنش', 'مبلغ']
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for row_num, fund in enumerate(report, start=3):
        ws.cell(row=row_num, column=1, value=show_jalali(fund.payment_date))
        ws.cell(row=row_num, column=2, value=fund.payment_description)
        ws.cell(row=row_num, column=3, value=fund.payment_gateway)
        ws.cell(row=row_num, column=4, value=fund.transaction_no)
        ws.cell(row=row_num, column=5, value=fund.amount)

    # Return response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=fund_user_report.xlsx'
    wb.save(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_user_report_pdf(request):
    user = request.user

    funds = Fund.objects.filter(user=user).order_by('doc_number')

    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    # # Queryset اصلی بر اساس واحد (کاربر مالک)
    # funds = Fund.objects.filter(unit=unit).order_by('payment_date')

    # فیلترها از GET
    query = request.GET.get('q', '').strip()
    if query:
        funds = funds.filter(
            Q(payment_description__icontains=query) |
            Q(payment_gateway__icontains=query) |
            Q(transaction_no__icontains=query) |
            Q(payment_date__icontains=query) |
            Q(amount__icontains=query)

        )

    # PDF settings
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        body {{
            font-family: 'BYekan', sans-serif;
        }}
        @font-face {{
            font-family: 'BYekan';
            src: url('{font_url}');
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 5px;
            text-align: center;
        }}
        th {{
            background-color: #FFD700;
        }}
    """)

    # Render template
    template = get_template("user_report_pdf.html")
    context = {
        'funds': funds,
        'query': query,
        'font_path': font_url,
        'today': datetime.now(),
        'house': house,
    }

    html = template.render(context)
    pdf_file = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[css])
    pdf_file.seek(0)

    # Response
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="fund_user_report.pdf"'
    return response


# ============================================================
@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def charge_notify_report_list(request):
    search = request.GET.get('q', '').strip()

    charges = (
        UnifiedCharge.objects
        .filter(user=request.user, send_notification=True).order_by('-created_at'))

    # 🔍 فیلتر جستجو
    if search:
        q_obj = (
                Q(title__icontains=search) |
                Q(details__icontains=search) |
                Q(unit__unit__icontains=search) |
                Q(unit__user__full_name__icontains=search)
        )

        if search.isdigit():
            q_obj |= Q(base_charge=search) | Q(total_charge_month=search)

        charges = charges.filter(q_obj)

    # 🧮 annotate
    charges = charges.filter(user=request.user).annotate(
        unit_number=F('unit__unit'),
        user_full_name=F('unit__user__full_name')
    )

    # 📄 pagination
    paginate = request.GET.get('paginate', '20')
    if str(paginate).lower() == 'all':
        paginate = charges.count() or 1
    else:
        try:
            paginate = int(paginate)
        except ValueError:
            paginate = 20

    paginator = Paginator(charges, paginate)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'query': search,
        'paginate': paginate,
    }

    return render(request, 'charge_notify_report.html', context)


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def charge_units_list_report_pdf(request):
    manager = request.user
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    unified_qs = UnifiedCharge.objects.filter(user=manager).order_by('-created_at')

    # 🔍 جستجو
    query = request.GET.get('q', '').strip()

    unified_charges = unified_qs.filter(
        send_notification_date__isnull=False
    ).select_related('unit', 'unit__user')

    if query:
        search_q = (
                Q(unit__unit__icontains=query) |
                Q(unit__user__full_name__icontains=query)
        )

        try:
            value = Decimal(query)
            search_q |= (
                    Q(penalty_amount=value) |
                    Q(total_charge_month=value) |
                    Q(base_charge=value)
            )
        except:
            pass

        unified_charges = unified_charges.filter(search_q)

    unified_charges = unified_charges.order_by('-created_at')

    html_string = render_to_string(
        'middleCharge/middle_charges_detail_pdf.html',
        {
            'unified_charges': unified_charges,
            'query': query,
            'today': datetime.now(),
            'house': house,
            'font_url': request.build_absolute_uri('/static/fonts/Vazir.ttf')

        }
    )

    font_url = request.build_absolute_uri(static('fonts/Vazir.ttf'))
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        @font-face {{
            font-family: 'Vazir';
            src: url('{font_url}');
        }}
        body {{
            font-family: 'Vazir', sans-serif;
        }}
    """)

    pdf = HTML(string=html_string).write_pdf(stylesheets=[css])

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="charge_units_report.pdf"'
    return response

@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def charge_units_list_report_excel(request):
    unified_qs = UnifiedCharge.objects.all()

    # 🔍 جستجو
    query = request.GET.get('q', '').strip()

    unified_charges = unified_qs.filter(
        send_notification_date__isnull=False
    ).select_related('unit', 'unit__user')

    if query:
        search_q = (
                Q(unit__unit__icontains=query) |
                Q(unit__user__full_name__icontains=query)
        )

        try:
            value = Decimal(query)
            search_q |= (
                    Q(penalty_amount=value) |
                    Q(total_charge_month=value) |
                    Q(base_charge=value)
            )
        except:
            pass

        unified_charges = unified_charges.filter(search_q)

    unified_charges = unified_charges.order_by('-created_at')

    # -------------------------
    # Excel
    # -------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Charge Units"
    ws.sheet_view.rightToLeft = True

    # عنوان اصلی
    title_cell = ws.cell(row=1, column=1, value="لیست تراکنش های من")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)  # 9 ستون

    # هدرها
    headers = [
        '#', 'واحد', 'مالک / مستاجر', 'مبلغ پایه', 'جریمه',
        'مبلغ نهایی', 'تاریخ اعلام', 'مهلت پرداخت', 'وضعیت پرداخت'
    ]
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # داده‌ها
    row = 3  # داده‌ها از ردیف بعد از هدر شروع می‌شوند
    for index, uc in enumerate(unified_charges, start=1):
        ws.cell(row=row, column=1, value=index)
        ws.cell(row=row, column=2, value=uc.title)
        ws.cell(row=row, column=3, value=uc.unit.get_label())
        ws.cell(row=row, column=4, value=uc.base_charge)
        ws.cell(row=row, column=5, value=uc.penalty_amount)
        ws.cell(row=row, column=6, value=uc.total_charge_month)
        ws.cell(row=row, column=7, value=show_jalali(uc.send_notification_date))
        ws.cell(row=row, column=8, value=show_jalali(uc.payment_deadline_date))
        ws.cell(row=row, column=9, value="پرداخت شده" if uc.is_paid else "پرداخت نشده")
        row += 1

    # پاسخ Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=charge_units_report.xlsx'
    wb.save(response)
    return response


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def charge_units_report_pdf(request, charge_id):
    charge = get_object_or_404(UnifiedCharge, id=charge_id)
    units = Unit.objects.filter(unified_charges=charge, is_active=True).order_by('unit')
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()
    bank = Bank.get_default(request.user, house)
    html_string = render_to_string('middleCharge/single_charge_pdf.html', {
        'charge': charge,
        'units': units,
        'house': house,
        'bank': bank,
        'font_url': request.build_absolute_uri('/static/fonts/Vazir.ttf')
    })

    pdf = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="charge_{charge.id}_units.pdf"'
    return response


# ===========================================================
@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def debtor_units_report(request):
    query = request.GET.get('q', '').strip()
    # -------------------------
    # Unpaid charges
    # -------------------------
    charges = (
        UnifiedCharge.objects
        .filter(is_paid=False, unit__isnull=False, user=request.user)
        .select_related('unit')
        .order_by('-created_at')
    )

    # -------------------------
    # Update penalties for unpaid charges
    # -------------------------
    if query:
        charges = charges.filter(
            Q(unit__unit__icontains=query) |
            Q(unit__owner_name__icontains=query) |
            Q(unit__renters__renter_name__icontains=query)
        ).distinct()

    charges = charges.order_by('-created_at')

    # -------------------------
    # Update penalties for unpaid charges
    # -------------------------
    for charge in charges:
        charge.update_penalty()
        charge.save(update_fields=['total_charge_month', 'penalty_amount'])

    # -------------------------
    # Organize charges per unit
    # -------------------------
    units = defaultdict(lambda: {
        'id': None,
        'label': '',
        'total_debt': 0,
        'charges': []
    })

    for charge in charges:
        unit = charge.unit
        renter = unit.get_active_renter()
        label = (
            f"واحد {unit.unit} - {renter.renter_name}" if renter
            else f"واحد {unit.unit} - {unit.owner_name}"
        )

        data = units[unit.id]
        data['id'] = unit.id
        data['label'] = label
        data['total_debt'] += charge.total_charge_month or 0
        data['charges'].append(charge)

    # -------------------------
    # Sort units by total debt
    # -------------------------

    units_with_debt = sorted(
        units.values(),
        key=lambda x: x['id'],
        reverse=True
    )
    total_debt_all_units = sum(unit['total_debt'] for unit in units_with_debt)

    # Pagination
    # -------------------------
    paginate_by = int(request.GET.get('paginate', 20))
    page_number = request.GET.get('page', 1)
    paginator = Paginator(units_with_debt, paginate_by)
    page_obj = paginator.get_page(page_number)

    return render(request, 'debtor_creditor_report.html', {
        'units_with_debt': page_obj.object_list,
        'total_debt_all_units': total_debt_all_units,
        'page_obj': page_obj,
        'query': query,
    })


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_debtor_report_pdf(request):
    query = request.GET.get('q', '').strip()
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()
    # -------------------------
    # Unpaid charges
    # -------------------------
    charges = UnifiedCharge.objects.filter(is_paid=False, unit__isnull=False).select_related('unit')

    # -------------------------
    # Apply search query
    # -------------------------
    if query:
        charges = charges.filter(
            Q(unit__unit__icontains=query) |
            Q(unit__owner_name__icontains=query) |
            Q(unit__renters__renter_name__icontains=query)
        ).distinct()

    # -------------------------
    # Update penalties
    # -------------------------
    for charge in charges:
        charge.update_penalty()
        charge.save(update_fields=['total_charge_month', 'penalty_amount'])

    # -------------------------
    # Organize charges per unit
    # -------------------------
    units = defaultdict(lambda: {'id': None, 'label': '', 'total_debt': 0, 'charges': []})
    for charge in charges:
        unit = charge.unit
        renter = unit.get_active_renter()
        label = f"واحد {unit.unit} - {renter.renter_name}" if renter else f"واحد {unit.unit} - {unit.owner_name}"

        data = units[unit.id]
        data['id'] = unit.id
        data['label'] = label
        data['total_debt'] += charge.total_charge_month or 0
        data['charges'].append(charge)

    # -------------------------
    # Sort units by unit number
    # -------------------------
    units_with_debt = sorted(units.values(), key=lambda x: x['id'])

    # -------------------------
    # Total debt all units
    # -------------------------
    total_debt_all_units = sum(unit['total_debt'] for unit in units_with_debt)

    # -------------------------
    # Render HTML template
    # -------------------------
    template = get_template("debtor_report_pdf.html")
    context = {
        'units_with_debt': units_with_debt,
        'total_debt_all_units': total_debt_all_units,
        'today': datetime.now(),
        'house': house,
    }
    html = template.render(context)

    # -------------------------
    # Generate PDF
    # -------------------------
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        body {{ font-family: 'BYekan', sans-serif; }}
        @font-face {{
            font-family: 'BYekan';
            src: url('{font_url}');
        }}
    """)
    pdf_file = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[css])
    pdf_file.seek(0)

    response = HttpResponse(pdf_file.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="debtor_report.pdf"'
    return response


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_debtor_report_excel(request):
    query = request.GET.get('q', '').strip()

    # -------------------------
    # Unpaid charges
    # -------------------------
    charges = UnifiedCharge.objects.filter(is_paid=False, unit__isnull=False).select_related('unit')

    # Apply search query
    if query:
        charges = charges.filter(
            Q(unit__unit__icontains=query) |
            Q(unit__owner_name__icontains=query) |
            Q(unit__renters__renter_name__icontains=query)
        ).distinct()

    # Update penalties
    for charge in charges:
        charge.update_penalty()
        charge.save(update_fields=['total_charge_month', 'penalty_amount'])

    # Organize charges per unit
    units = defaultdict(lambda: {'id': None, 'label': '', 'total_debt': 0, 'charges': []})
    for charge in charges:
        unit = charge.unit
        renter = unit.get_active_renter()
        label = f"واحد {unit.unit} - {renter.renter_name}" if renter else f"واحد {unit.unit} - {unit.owner_name}"

        data = units[unit.id]
        data['id'] = unit.id
        data['label'] = label
        data['total_debt'] += charge.total_charge_month or 0
        data['charges'].append(charge)

    # Sort units by unit number
    units_with_debt = sorted(units.values(), key=lambda x: x['id'])
    total_debt_all_units = sum(u['total_debt'] for u in units_with_debt)

    # -------------------------
    # Create Excel workbook
    # -------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Debtor Report"
    ws.sheet_view.rightToLeft = True

    # Title
    title_cell = ws.cell(row=1, column=1, value="گزارش بدهکاران")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    # Total debt summary
    ws.cell(row=2, column=1, value="جمع کل بدهی همه واحدها")
    ws.cell(row=2, column=2, value=total_debt_all_units)
    ws.cell(row=2, column=2).font = Font(bold=True, color="FF0000")  # Red

    current_row = 4

    # Headers for unit charges
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)

    for unit in units_with_debt:
        # Unit title
        ws.cell(row=current_row, column=1, value=f"{unit['label']} - جمع بدهی: {unit['total_debt']}")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        current_row += 1

        # Column headers
        headers = ['#', 'شرح', 'شارژ ماهیانه', 'جریمه', 'قابل پرداخت', 'مهلت پرداخت']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
        current_row += 1

        # Charges
        for idx, charge in enumerate(unit['charges'], start=1):
            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=2, value=charge.title)
            ws.cell(row=current_row, column=3, value=charge.amount)
            ws.cell(row=current_row, column=4, value=charge.penalty_amount or 0)
            ws.cell(row=current_row, column=5, value=charge.total_charge_month)
            ws.cell(row=current_row, column=6, value=show_jalali(charge.payment_deadline_date))
            current_row += 1

        current_row += 1  # Empty row between units

    # -------------------------
    # Return Excel response
    # -------------------------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=debtor_report.xlsx'
    wb.save(response)
    return response


# ======================================================
@method_decorator(middle_admin_required, name='dispatch')
class HistoryUnitReports(FormMixin, ListView):
    model = UnitResidenceHistory
    form_class = UnitReportForm
    template_name = 'unit_history_reports.html'
    context_object_name = 'unit_histories'
    paginate_by = 40

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_queryset(self):
        queryset = UnitResidenceHistory.objects.none()
        self.selected_unit = None  # ⚡ مهم: برای context

        if self.request.method == 'POST':
            form = self.get_form()
            if form.is_valid():
                unit = form.cleaned_data['unit']
                self.selected_unit = unit  # ⚡ ذخیره برای context و export
                queryset = UnitResidenceHistory.objects.filter(
                    unit=unit
                ).order_by('-created_at')  # جدیدترین‌ها بالای جدول
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.get_form()
        context['selected_unit'] = getattr(self, 'selected_unit', None)
        return context

    def post(self, request, *args, **kwargs):
        # ⚡ مهم برای paginate کردن
        self.object_list = self.get_queryset()
        return self.render_to_response(self.get_context_data())


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_unit_history_report_pdf(request):
    # دریافت واحد انتخاب شده از query string
    unit_id = request.GET.get('unit')
    if not unit_id:
        return HttpResponse("واحدی انتخاب نشده است.", status=400)

    try:
        unit = Unit.objects.get(pk=unit_id)
    except Unit.DoesNotExist:
        return HttpResponse("واحد انتخاب شده وجود ندارد.", status=404)

    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    # Queryset سوابق سکونت آن واحد
    unit_histories = UnitResidenceHistory.objects.filter(unit=unit).order_by('-created_at')

    # اگر سابقه‌ای نیست، پیام بده یا خالی بفرست
    if not unit_histories.exists():
        return HttpResponse("هیچ سابقه سکونتی برای این واحد وجود ندارد.", status=404)

    query = request.GET.get('q', '').strip()
    if query:
        unit_histories = unit_histories.filter(
            Q(resident_type__icontains=query) |
            Q(name__icontains=query) |
            Q(mobile__icontains=query)

        )

    # PDF settings
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        body {{
            font-family: 'BYekan', sans-serif;
        }}
        @font-face {{
            font-family: 'BYekan';
            src: url('{font_url}');
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 5px;
            text-align: center;
        }}
        th {{
            background-color: #FFD700;
        }}
    """)

    # Template PDF (باید جدولی مطابق unit_histories داشته باشد)
    template = get_template("unit_history_report_pdf.html")
    context = {
        'unit': unit,
        'unit_histories': unit_histories,
        'today': datetime.now(),
        'font_path': font_url,
        'house': house
    }
    html = template.render(context)

    pdf_file = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[css])
    pdf_file.seek(0)

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="unit_{unit.unit}_history.pdf"'
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_unit_history_report_excel(request):
    unit_id = request.GET.get('unit')
    if not unit_id:
        return HttpResponse("واحدی انتخاب نشده است.", status=400)

    try:
        unit = Unit.objects.get(pk=unit_id)
    except Unit.DoesNotExist:
        return HttpResponse("واحد انتخاب شده وجود ندارد.", status=404)

    # Queryset سوابق سکونت آن واحد
    histories = UnitResidenceHistory.objects.filter(unit=unit).order_by('-created_at')
    if not histories.exists():
        return HttpResponse("هیچ سابقه سکونتی برای این واحد وجود ندارد.", status=404)

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"واحد {unit.unit}"
    ws.sheet_view.rightToLeft = True

    # Title
    title_cell = ws.cell(row=1, column=1, value=f"سوابق سکونت واحد {unit.unit}")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    # Headers
    headers = ['ردیف', 'نوع سکونت', 'نام', 'موبایل', 'تعداد نفرات', 'از تاریخ', 'تا تاریخ']

    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_num, h in enumerate(histories, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # ردیف
        ws.cell(row=row_num, column=2, value='مالک' if h.resident_type == 'owner' else 'مستاجر')
        ws.cell(row=row_num, column=3, value=h.name)
        ws.cell(row=row_num, column=4, value=h.mobile)
        ws.cell(row=row_num, column=5, value=h.people_count)
        ws.cell(row=row_num, column=6, value=show_jalali(h.from_date))
        ws.cell(row=row_num, column=7, value=show_jalali(h.to_date) if h.to_date else 'اکنون')

    # Return response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=unit_{unit.unit}_history.xlsx'
    wb.save(response)
    return response


# ====================================================================
@method_decorator(middle_admin_required, name='dispatch')
class ReportExpenseView(ListView):
    model = Expense
    template_name = 'expense_reports.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_queryset(self):
        queryset = Expense.objects.filter(user=self.request.user).order_by('-created_at')

        # فیلتر بر اساس category__title
        category_id = self.request.GET.get('category')
        if category_id:
            queryset = queryset.filter(category__id=category_id)

        # فیلتر بر اساس amount
        amount = self.request.GET.get('amount')
        if amount:
            queryset = queryset.filter(amount__icontains=amount)

        description = self.request.GET.get('description')
        if description:
            queryset = queryset.filter(description__icontains=description)

        doc_no = self.request.GET.get('doc_no')
        if doc_no:
            queryset = queryset.filter(doc_no__icontains=doc_no)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس date
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')

        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')
        is_paid = self.request.GET.get('is_paid')
        if is_paid == '1':
            queryset = queryset.filter(is_paid=True)
        elif is_paid == '0':
            queryset = queryset.filter(is_paid=False)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        expenses = self.get_queryset()  # از get_queryset برای دریافت داده‌های فیلتر شده استفاده می‌کنیم
        paginator = Paginator(expenses, 40)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_expense'] = Expense.objects.filter(user=self.request.user).count()
        context['categories'] = ExpenseCategory.objects.filter(user=self.request.user)
        context['total_amount'] = Expense.objects.filter(user=self.request.user).aggregate(total=Sum('amount'))[
                                      'total'] or 0
        context['banks'] = Bank.objects.filter(user=self.request.user)
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_expense_report_pdf(request):
    user = request.user
    expenses = Expense.objects.filter(user=user)
    total_amount = Expense.objects.filter(user=user).aggregate(total=Sum('amount'))['total'] or 0
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    filter_fields = {
        'category': 'category__id',
        'amount': 'amount__icontains',
        'doc_no': 'doc_no__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
        'is_paid': 'is_paid'
    }

    # Apply filters based on GET parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            expenses = expenses.filter(**filter_expression)

    # Handle date filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__lte=to_date)
    except ValueError:
        expenses = Expense.objects.none()

    # Log the filtered expenses for debugging
    print(expenses)

    # Font setup
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # Render HTML template
    template = get_template("expense_report_pdf.html")
    context = {
        'expenses': expenses,
        'font_path': font_url,
        'total_amount': total_amount,
        'house': house,
        'today': datetime.now(),
    }
    html = template.render(context)

    # Generate PDF
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # Generate the final PDF response
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="expenses_report.pdf"'
    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_expense_report_excel(request):
    expenses = Expense.objects.filter(user=request.user)

    # Filter fields
    filter_fields = {
        'category': 'category__id',
        'bank': 'bank__id',
        'amount': 'amount__icontains',
        'doc_no': 'doc_no__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
        'is_paid': 'is_paid'
    }

    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            expenses = expenses.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__lte=to_date)
    except ValueError:
        expenses = Expense.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "expenses"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست هزینه‌ها")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'موضوع هزینه', 'شرح سند', ' شماره سند', 'مبلغ', 'تاریخ سند', 'پرداخت به', 'شماره حساب',
               'تاریخ پرداخت', 'توضیحات']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, expense in enumerate(expenses, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        ws.cell(row=row_num, column=2, value=expense.category.title)
        ws.cell(row=row_num, column=3, value=expense.description)
        ws.cell(row=row_num, column=4, value=expense.doc_no)
        ws.cell(row=row_num, column=5, value=expense.amount)
        ws.cell(row=row_num, column=6, value=show_jalali(expense.date))
        ws.cell(row=row_num, column=7, value=expense.receiver_name)
        ws.cell(row=row_num, column=8, value=expense.bank.bank_name)
        ws.cell(row=row_num, column=9, value=show_jalali(expense.payment_date))
        ws.cell(row=row_num, column=10, value=expense.details)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=expenses_report.xlsx'
    wb.save(response)
    return response


# =======================================================
@method_decorator(middle_admin_required, name='dispatch')
class ReportIncomeView(ListView):
    model = Income
    template_name = 'income_reports.html'

    def get_queryset(self):
        queryset = Income.objects.filter(user=self.request.user).order_by('-created_at')

        # فیلتر بر اساس category
        category_id = self.request.GET.get('category')
        if category_id:
            queryset = queryset.filter(category__id=category_id)

        # فیلتر بر اساس بانک
        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank__id=bank_id)

        # فیلتر بر اساس واحد
        unit_id = self.request.GET.get('unit')
        if unit_id:
            queryset = queryset.filter(unit__id=unit_id)

        # فیلتر بر اساس amount
        amount = self.request.GET.get('amount')
        if amount:
            queryset = queryset.filter(amount__icontains=amount)

        # فیلتر بر اساس description
        description = self.request.GET.get('description')
        if description:
            queryset = queryset.filter(description__icontains=description)

        # فیلتر بر اساس doc_number
        doc_number = self.request.GET.get('doc_number')
        if doc_number:
            queryset = queryset.filter(doc_number__icontains=doc_number)

        # فیلتر بر اساس details
        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس تاریخ
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')
        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(doc_date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(doc_date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')

        is_paid = self.request.GET.get('is_paid')
        if is_paid == '1':
            queryset = queryset.filter(is_paid=True)
        elif is_paid == '0':
            queryset = queryset.filter(is_paid=False)

        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        incomes = self.get_queryset()  # از get_queryset برای دریافت داده‌های فیلتر شده استفاده می‌کنیم
        paginator = Paginator(incomes, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_incomes'] = Income.objects.filter(user=self.request.user).count()
        context['categories'] = IncomeCategory.objects.filter(user=self.request.user)
        context['banks'] = Bank.objects.filter(user=self.request.user)
        managed_users = User.objects.filter(Q(manager=self.request.user) | Q(pk=self.request.user.pk))
        context['units'] = Unit.objects.filter(is_active=True, user__in=managed_users)
        context['total_amount'] = Income.objects.filter(user=self.request.user).aggregate(total=Sum('amount'))[
                                      'total'] or 0

        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_income_report_pdf(request):
    incomes = Income.objects.filter(user=request.user)
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    filter_fields = {
        'category': 'category__id',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
        'is_paid': 'is_paid'
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            incomes = incomes.filter(**filter_expression)

    # فیلتر تاریخ
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__lte=to_date)
    except ValueError:
        incomes = Income.objects.none()

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # رندر قالب HTML
    template = get_template("income_report_pdf.html")
    context = {
        'incomes': incomes,
        'font_path': font_url,
        'house': house,
        'today': datetime.now()
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="incomes_report.pdf"'

    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_income_report_excel(request):
    incomes = Income.objects.filter(user=request.user)

    # Filter fields
    filter_fields = {
        'category': 'category__id',
        'bank': 'bank__id',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
        'is_paid': 'is_paid'
    }
    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            incomes = incomes.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__lte=to_date)
    except ValueError:
        expenses = Expense.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "incomes"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست درآمدها")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'موضوع درآمد', 'شرح سند', ' شماره سند', 'مبلغ', 'تاریخ سند', 'توضیحات', 'پرداخت کننده',
               'تاریخ پرداخت']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, income in enumerate(incomes, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        ws.cell(row=row_num, column=2, value=income.category.subject)
        ws.cell(row=row_num, column=3, value=income.description)
        ws.cell(row=row_num, column=4, value=income.doc_number)
        ws.cell(row=row_num, column=5, value=income.amount)
        ws.cell(row=row_num, column=6, value=show_jalali(income.doc_date))
        ws.cell(row=row_num, column=7, value=income.details)
        ws.cell(row=row_num, column=8, value=income.payer_name)
        ws.cell(row=row_num, column=9, value=show_jalali(income.payment_date))

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=incomes.xlsx'
    wb.save(response)
    return response


# ========================================================
@method_decorator(middle_admin_required, name='dispatch')
class ReportPropertyView(ListView):
    model = Property
    template_name = 'report_property.html'

    def get_queryset(self):
        queryset = Property.objects.filter(user=self.request.user)

        # فیلتر بر اساس amount
        property_name = self.request.GET.get('property_name')
        if property_name:
            queryset = queryset.filter(property_name__icontains=property_name)

        property_unit = self.request.GET.get('property_unit')
        if property_unit:
            queryset = queryset.filter(property_unit__icontains=property_unit)

        property_location = self.request.GET.get('property_location')
        if property_location:
            queryset = queryset.filter(property_location__icontains=property_location)

        property_code = self.request.GET.get('property_code')
        if property_code:
            queryset = queryset.filter(property_code__icontains=property_code)

        property_price = self.request.GET.get('property_price')
        if property_price:
            queryset = queryset.filter(property_price__icontains=property_price)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس date
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')

        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(property_purchase_date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(property_purchase_date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receives = self.get_queryset()
        paginator = Paginator(receives, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_properties'] = Property.objects.filter(user=self.request.user).count()
        context['properties'] = Property.objects.filter(user=self.request.user)
        context['total_amount'] = Property.objects.filter(user=self.request.user).aggregate(total=Sum('property_price'))[
                                      'total'] or 0
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_property_report_pdf(request):
    properties = Property.objects.filter(user=request.user)
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    filter_fields = {
        'property_name': 'property_name__icontains',
        'property_unit': 'property_unit__icontains',
        'property_location': 'property_location__icontains',
        'property_code': 'property_code__icontains',
        'property_price': 'property_price__icontains',
        'details': 'details__icontains',

    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            properties = properties.filter(**filter_expression)

    # فیلتر تاریخ
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            properties = properties.filter(property_purchase_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            properties = properties.filter(property_purchase_date__lte=to_date)
    except ValueError:
        properties = Property.objects.none()

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # رندر قالب HTML
    template = get_template("report_property_pdf.html")
    context = {
        'properties': properties,
        'font_path': font_url,
        'house': house,
        'today': datetime.now()

    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="properties_report.pdf"'

    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_property_report_excel(request):
    properties = Property.objects.filter(user=request.user)

    filter_fields = {
        'property_name': 'property_name__icontains',
        'property_unit': 'property_unit__icontains',
        'property_location': 'property_location__icontains',
        'property_code': 'property_code__icontains',
        'property_price': 'property_price__icontains',
        'details': 'details__icontains',

    }

    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            properties = properties.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            properties = properties.filter(property_purchase_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            properties = properties.filter(property_purchase_date__lte=to_date)
    except ValueError:
        properties = Property.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "properties"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست اموال ساختمان")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'نام اموال', 'واحد', ' شماره اموال', ' موقعیت ', 'ارزش', 'تاریخ خرید', 'توضیحات']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, property in enumerate(properties, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        ws.cell(row=row_num, column=2, value=property.property_name)
        ws.cell(row=row_num, column=3, value=property.property_unit)
        ws.cell(row=row_num, column=4, value=property.property_code)
        ws.cell(row=row_num, column=5, value=property.property_location)
        ws.cell(row=row_num, column=6, value=property.property_price)
        jalali_date = jdatetime.date.fromgregorian(date=property.property_purchase_date).strftime('%Y/%m/%d')
        ws.cell(row=row_num, column=7, value=jalali_date)
        ws.cell(row=row_num, column=8, value=property.details)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=properties_report.xlsx'
    wb.save(response)
    return response


# ==========================================================
@method_decorator(middle_admin_required, name='dispatch')
class ReportMaintenanceView(ListView):
    model = Maintenance
    template_name = 'report_maintenance.html'

    def get_queryset(self):
        queryset = Maintenance.objects.filter(user=self.request.user)

        # فیلتر بر اساس amount
        maintenance_description = self.request.GET.get('maintenance_description')
        if maintenance_description:
            queryset = queryset.filter(maintenance_description__icontains=maintenance_description)

        maintenance_price = self.request.GET.get('maintenance_price')
        if maintenance_price:
            queryset = queryset.filter(maintenance_price__icontains=maintenance_price)

        maintenance_status = self.request.GET.get('maintenance_status')
        if maintenance_status:
            queryset = queryset.filter(maintenance_status__icontains=maintenance_status)

        service_company = self.request.GET.get('service_company')
        if service_company:
            queryset = queryset.filter(service_company__icontains=service_company)

        maintenance_document_no = self.request.GET.get('maintenance_document_no')
        if maintenance_document_no:
            queryset = queryset.filter(maintenance_document_no__icontains=maintenance_document_no)

        maintenance_start_date = self.request.GET.get('maintenance_start_date')
        if maintenance_start_date and isinstance(maintenance_start_date, str):
            try:
                j_start = jdatetime.date.fromisoformat(maintenance_start_date)
                g_start = j_start.togregorian()
                queryset = queryset.filter(maintenance_start_date__gte=g_start)
            except (ValueError, TypeError) as e:
                print("Invalid date format:", maintenance_start_date, e)

        maintenance_end_date = self.request.GET.get('maintenance_end_date')
        if maintenance_end_date and isinstance(maintenance_end_date, str):
            try:
                j_start = jdatetime.date.fromisoformat(maintenance_end_date)
                g_start = j_start.togregorian()
                queryset = queryset.filter(maintenance_end_date__lte=g_start)
            except (ValueError, TypeError) as e:
                print("Invalid date format:", maintenance_end_date, e)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        maintenances = self.get_queryset()
        paginator = Paginator(maintenances, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['total_maintenances'] = maintenances.filter(user=self.request.user).count()
        context['maintenances'] = page_obj.object_list
        context['total_amount'] = maintenances.filter(user=self.request.user).aggregate(total=Sum('maintenance_price'))[
                                      'total'] or 0

        return context


def parse_jalali_to_gregorian(date_str):
    try:
        return jdatetime.date.fromisoformat(date_str.strip()).togregorian()
    except Exception:
        return None

@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_maintenance_report_pdf(request):
    maintenances = Maintenance.objects.filter(user=request.user).order_by('-created_at')
    total_amount = maintenances.filter(user=request.user).aggregate(total=Sum('maintenance_price'))[
                       'total'] or 0
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    filter_fields = {
        'maintenance_description': 'maintenance_description__icontains',
        'maintenance_start_date': 'maintenance_start_date__gte',
        'maintenance_end_date': 'maintenance_end_date__lte',
        'maintenance_price': 'maintenance_price__icontains',
        'maintenance_status': 'maintenance_status__icontains',
        'service_company': 'service_company__icontains',
        'maintenance_document_no': 'maintenance_document_no__icontains',
        'details': 'details__icontains',
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            if field in ['maintenance_start_date', 'maintenance_end_date']:
                gregorian_date = parse_jalali_to_gregorian(value)
                if gregorian_date:
                    maintenances = maintenances.filter(**{lookup: gregorian_date})
            else:
                maintenances = maintenances.filter(**{lookup: value.strip()})

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # رندر قالب HTML
    template = get_template("report_maintenance_pdf.html")
    context = {
        'maintenances': maintenances,
        'font_path': font_url,
        'today': datetime.now(),
        'house': house,
        'total_amount': total_amount
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="maintenances_report.pdf"'

    pdf_merger.write(response)
    return response

@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_maintenance_report_excel(request):
    maintenances = Maintenance.objects.filter(user=request.user)

    filter_fields = {
        'maintenance_description': 'maintenance_description__icontains',
        'maintenance_start_date': 'maintenance_start_date__gte',
        'maintenance_end_date': 'maintenance_end_date__lte',
        'maintenance_price': 'maintenance_price__icontains',
        'maintenance_status': 'maintenance_status__icontains',
        'service_company': 'service_company__icontains',
        'maintenance_document_no': 'maintenance_document_no__icontains',
        'details': 'details__icontains',
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            if field in ['maintenance_start_date', 'maintenance_end_date']:
                gregorian_date = parse_jalali_to_gregorian(value)
                if gregorian_date:
                    maintenances = maintenances.filter(**{lookup: gregorian_date})
            else:
                maintenances = maintenances.filter(**{lookup: value.strip()})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "maintenances"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست هزینه های تعمیرات و نگهداری")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'شرح کار', 'تاریخ شروع', ' تاریخ پایان', ' اجرت/دستمزد ', 'شرکت خدماتی', 'شماره فاکتور',
               'توضیحات', 'آخرین وضعیت']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, maintenance in enumerate(maintenances, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        ws.cell(row=row_num, column=2, value=maintenance.maintenance_description)
        jalali_date = jdatetime.date.fromgregorian(date=maintenance.maintenance_start_date).strftime('%Y/%m/%d')
        ws.cell(row=row_num, column=3, value=jalali_date)
        jalali_date = jdatetime.date.fromgregorian(date=maintenance.maintenance_end_date).strftime('%Y/%m/%d')
        ws.cell(row=row_num, column=4, value=jalali_date)
        ws.cell(row=row_num, column=5, value=maintenance.maintenance_price)
        ws.cell(row=row_num, column=6, value=maintenance.service_company)
        ws.cell(row=row_num, column=7, value=maintenance.maintenance_document_no)
        ws.cell(row=row_num, column=8, value=maintenance.details)
        ws.cell(row=row_num, column=9, value=maintenance.maintenance_status)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=maintenances_report.xlsx'
    wb.save(response)
    return response


# =================================================
@method_decorator(middle_admin_required, name='dispatch')
class PayReceiveReportView(ListView):
    model = Fund
    template_name = 'pay_receive_report.html'
    paginate_by = 50
    context_object_name = 'funds'

    def get_queryset(self):
        qs = Fund.objects.filter(
            user=self.request.user
        ).filter(
            Q(is_received_money=True) | Q(is_paid_money=True)
        ).order_by('-payment_date')

        if bank := self.request.GET.get('bank'):
            qs = qs.filter(bank_id=bank)

        if unit := self.request.GET.get('unit'):
            qs = qs.filter(unit_id=unit)

        if amount := self.request.GET.get('amount'):
            qs = qs.filter(amount=amount)

        q = self.request.GET.get('q')
        if q:
            search_filter = (
                    Q(receiver_name__icontains=q) |
                    Q(payer_name__icontains=q)
            )
            qs = qs.filter(search_filter)

        if payment_description := self.request.GET.get('payment_description'):
            qs = qs.filter(payment_description__icontains=payment_description)

        if doc_number := self.request.GET.get('doc_number'):
            qs = qs.filter(doc_number__exact=doc_number)

        if details := self.request.GET.get('details'):
            qs = qs.filter(details__icontains=details)

        transaction_type = self.request.GET.get('transaction_type')

        if transaction_type == 'received':
            qs = qs.filter(is_received_money=True)

        elif transaction_type == 'paid':
            qs = qs.filter(is_paid_money=True)

        try:
            if from_date := self.request.GET.get('from_date'):
                qs = qs.filter(
                    payment_date__gte=jdatetime.datetime.strptime(
                        from_date, '%Y/%m/%d'
                    ).togregorian().date()
                )

            if to_date := self.request.GET.get('to_date'):
                qs = qs.filter(
                    payment_date__lte=jdatetime.datetime.strptime(
                        to_date, '%Y/%m/%d'
                    ).togregorian().date()
                )
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست')

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        totals = self.object_list.aggregate(
            total_income=Sum('debtor_amount'),
            total_expense=Sum('creditor_amount'),
        )

        context['total_income'] = totals['total_income'] or 0
        context['total_expense'] = totals['total_expense'] or 0
        context['balance'] = context['total_income'] - context['total_expense']

        managed_users = User.objects.filter(
            Q(manager=self.request.user) | Q(pk=self.request.user.pk)
        )

        context.update({
            'banks': Bank.objects.filter(user=self.request.user),
            'units': Unit.objects.filter(is_active=True, user__in=managed_users),
        })

        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_pay_receive_report_pdf(request):
    manager = request.user

    # Queryset اصلی
    funds = Fund.objects.filter(
        Q(user=manager) | Q(user__manager=manager)
    ).filter(
        Q(is_received_money=True) | Q(is_paid_money=True)
    ).order_by('-payment_date')
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    transaction_type = request.GET.get('transaction_type')
    if transaction_type == 'received':
        funds = funds.filter(is_received_money=True)
    elif transaction_type == 'paid':
        funds = funds.filter(is_paid_money=True)
    else:
        funds = funds.filter(Q(is_received_money=True) | Q(is_paid_money=True))

    # فیلتر جستجو
    q = request.GET.get('q')
    if q:
        search_filter = Q(receiver_name__icontains=q) | Q(payer_name__icontains=q)
        funds = funds.filter(search_filter)

    filter_fields = {
        'bank': 'bank__id',
        'unit': 'unit__id',
        'receiver_name': 'payer_name',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__exact',
        'payment_description': 'payment_description__icontains',
        'details': 'details__icontains',
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            funds = funds.filter(**filter_expression)

    # فیلتر تاریخ
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            funds = funds.filter(payment_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            funds = funds.filter(payment_date__lte=to_date)
    except ValueError:
        funds = Fund.objects.none()

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
                @page {{ size: A4 landscape; margin: 1cm; }}
                body {{
                    font-family: 'BYekan', sans-serif;
                }}
                @font-face {{
                    font-family: 'BYekan';
                    src: url('{font_url}');
                }}
            """)

    # رندر قالب HTML
    template = get_template("pay_receive_report_pdf.html")
    context = {
        'funds': funds,
        'font_path': font_url,
        'house': house,
        'today': datetime.now()
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="pay_receive_reports.pdf"'

    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_pay_receive_report_excel(request):
    manager = request.user
    funds = Fund.objects.filter(
        Q(user=manager) | Q(user__manager=manager)
    ).filter(
        Q(is_received=True) | Q(is_paid=True)
    ).order_by('-payment_date')

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "units"
    ws.sheet_view.rightToLeft = True

    # Headers
    headers = [' بانک', 'تاریخ پرداخت', 'شرح', 'پرداخت کننده/واریز کننده', 'شماره سند', 'بدهکار', 'بستانکار']

    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for row_num, fund in enumerate(funds, start=3):
        ws.cell(row=row_num, column=1, value=f"{fund.bank.bank_name} - {fund.bank.account_no}")
        ws.cell(row=row_num, column=2, value=show_jalali(fund.payment_date))
        ws.cell(row=row_num, column=3, value=fund.payment_description)
        ws.cell(row=row_num, column=4, value=f"{fund.payer_name} - {fund.receiver_name}")
        ws.cell(row=row_num, column=5, value=fund.doc_number)
        ws.cell(row=row_num, column=6, value=fund.debtor_amount)
        ws.cell(row=row_num, column=7, value=fund.creditor_amount)

    # Return response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=report_pay_receive.xlsx'
    wb.save(response)
    return response


# ==============================================================
@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def house_balance_view(request):
    total_incomes_exclude_unpaid = Income.objects.filter(is_paid=False, user=request.user).aggregate(Sum('amount'))[
        'amount__sum']

    total_expenses_exclude_unpaid = Expense.objects.filter(is_paid=False, user=request.user).aggregate(Sum('amount'))[
        'amount__sum']

    total_incomes = Income.objects.filter(user=request.user).aggregate(Sum('amount'))[
        'amount__sum']
    total_pay_money = PayMoney.objects.filter(user=request.user).aggregate(Sum('amount'))[
        'amount__sum']
    total_receive_money = ReceiveMoney.objects.filter(is_paid=True, user=request.user).aggregate(Sum('amount'))[
        'amount__sum']
    total_expenses = Expense.objects.filter(is_paid=True, user=request.user).aggregate(Sum('amount'))[
        'amount__sum']

    total_assets = (total_incomes or 0) + (total_receive_money or 0)
    total_debts = (total_pay_money or 0) + (total_expenses or 0)

    total_amount_assets_debts = (total_assets or 0) - (total_debts or 0)

    funds = (
        Fund.objects
        .select_related('bank', 'content_type')
        .filter(Q(user=request.user) | Q(user__manager=request.user))
        .order_by('-payment_date')
    )

    totals = funds.aggregate(
        total_income=Sum('debtor_amount'),
        total_expense=Sum('creditor_amount'),
    )

    balance = (totals['total_income'] or 0) - (totals['total_expense'] or 0)

    total_charge_unpaid = \
    UnifiedCharge.objects.filter(is_paid=False, user=request.user).aggregate(Sum('total_charge_month'))[
        'total_charge_month__sum']

    context = {
        'total_incomes_exclude_unpaid': total_incomes_exclude_unpaid,
        'total_expenses_exclude_unpaid': total_expenses_exclude_unpaid,
        'total_incomes': total_incomes,
        'total_expenses': total_expenses,
        'total_pay_money': total_pay_money,
        'total_receive_money': total_receive_money,
        'total_assets': total_assets,
        'total_debts': total_debts,
        'balance': balance,
        'total_charge_unpaid': total_charge_unpaid,
        'total_amount_assets_debts': total_amount_assets_debts
    }
    return render(request, 'house_balance.html', context)


# ===========================================================

@login_required(login_url=settings.LOGIN_URL_ADMIN)
def middleFund_report_excel(request):
    user = request.user

    report = AdminFund.objects.filter(user=user).order_by('-payment_date')

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "units"
    ws.sheet_view.rightToLeft = True

    # Title
    title_cell = ws.cell(row=1, column=1, value=f"لیست تراکنش های من ")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # Headers
    headers = ['تاریخ پرداخت', 'شرح', 'روش پرداخت', 'شماره تراکنش', 'مبلغ']
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for row_num, fund in enumerate(report, start=3):
        ws.cell(row=row_num, column=1, value=show_jalali(fund.payment_date))
        ws.cell(row=row_num, column=2, value=fund.payment_description)
        ws.cell(row=row_num, column=3, value=fund.payment_gateway)
        ws.cell(row=row_num, column=4, value=fund.transaction_no)
        ws.cell(row=row_num, column=5, value=fund.amount)

    # Return response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=middle_fund_report.xlsx'
    wb.save(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def middleFund_report_pdf(request):
    user = request.user

    funds = AdminFund.objects.filter(user=user).order_by('-payment_date')

    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(user=user).order_by('-created_at').first()
        print(house)

    # # Queryset اصلی بر اساس واحد (کاربر مالک)
    # funds = Fund.objects.filter(unit=unit).order_by('payment_date')

    # فیلترها از GET
    query = request.GET.get('q', '').strip()
    if query:
        funds = funds.filter(
            Q(payment_description__icontains=query) |
            Q(payment_gateway__icontains=query) |
            Q(transaction_no__icontains=query) |
            Q(payment_date__icontains=query) |
            Q(amount__icontains=query)

        )

    # PDF settings
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        body {{
            font-family: 'BYekan', sans-serif;
        }}
        @font-face {{
            font-family: 'BYekan';
            src: url('{font_url}');
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 5px;
            text-align: center;
        }}
        th {{
            background-color: #FFD700;
        }}
    """)

    # Render template
    template = get_template("user_report_pdf.html")
    context = {
        'funds': funds,
        'query': query,
        'font_path': font_url,
        'today': datetime.now(),
        'house': house,
    }

    html = template.render(context)
    pdf_file = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[css])
    pdf_file.seek(0)

    # Response
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="fund_middle_report.pdf"'
    return response