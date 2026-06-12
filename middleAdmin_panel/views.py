import io
import json
import logging
import os
import time
from collections import defaultdict
from datetime import timezone, datetime, timedelta
from decimal import Decimal, InvalidOperation
from itertools import chain
from dateutil.relativedelta import relativedelta
from django.apps import apps
from django.conf.urls.static import static
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.contenttypes.models import ContentType
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db.models.functions import ExtractMonth
from django.utils import timezone
import jdatetime
import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import user_passes_test, login_required
from django.db import transaction, IntegrityError, models
from django.db.models import ProtectedError, Count, Q, Sum, F, Prefetch, Case, When, IntegerField, FilteredRelation
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import get_template, render_to_string
from django.urls import reverse, reverse_lazy
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import CreateView, UpdateView, DetailView, ListView
from openpyxl.styles import PatternFill, Alignment, Font
from pypdf import PdfWriter
from sweetify import sweetify
from weasyprint import CSS, HTML

from admin_panel import helper
from admin_panel.forms import announcementForm, BankForm, UnitForm, ExpenseCategoryForm, ExpenseForm, \
    IncomeCategoryForm, IncomeForm, ReceiveMoneyForm, PayerMoneyForm, PropertyForm, MaintenanceForm, FixChargeForm, \
    FixAreaChargeForm, AreaChargeForm, PersonChargeForm, FixPersonChargeForm, PersonAreaChargeForm, \
    PersonAreaFixChargeForm, VariableFixChargeForm, MyHouseForm, SmsForm, RenterAddForm, ExpensePayForm, IncomePayForm, \
    SmsCreditForm, SubscriptionPlanForm, CivilForm, SewageForm, ReceivePayForm, PayMoneyForm, PropertyPayForm, \
    MaintenancePayForm, TransferMoneyForm
from admin_panel.helper import send_notify_user_by_sms
from admin_panel.models import Announcement, ExpenseCategory, Expense, Fund, ExpenseDocument, IncomeCategory, Income, \
    IncomeDocument, ReceiveMoney, ReceiveDocument, PayMoney, PayDocument, Property, PropertyDocument, Maintenance, \
    MaintenanceDocument, FixCharge, AreaCharge, PersonCharge, \
    FixAreaCharge, FixPersonCharge, ChargeByPersonArea, \
    ChargeByFixPersonArea, ChargeFixVariable, SmsManagement, \
    UnifiedCharge, SmsCredit, SubscriptionPlan, Subscription, CivilManage, CivilDocument, CivilInstallment, \
    SewageManage, SewageDocument, SewageInstallment, BankFund, Coupon, CouponUsage
from admin_panel.services.calculators import CALCULATORS
from middleAdmin_panel.services.bank_services import BankTransactionService
from middleAdmin_panel.services.unit_services import UnitUpdateService
from notifications.models import Notification, SupportUser
from notifications.services.sms_service import SmsService
from polls.templatetags.poll_extras import show_jalali

from user_app.models import Bank, Unit, User, Renter, MyHouse, UnitResidenceHistory


def middle_admin_required(view_func):
    return user_passes_test(
        lambda u: u.is_authenticated and getattr(u, 'is_middle_admin', False),
        login_url=settings.LOGIN_URL_MIDDLE_ADMIN
    )(view_func)


# ========================== Subscription =====================

def check_coupon(request):
    code = request.GET.get("code", "").strip()

    if not code:
        return JsonResponse({"valid": False, "discount": 0, "message": "کد وارد نشده است"})

    try:
        coupon = Coupon.objects.get(code__iexact=code)

        if not coupon.is_valid():
            return JsonResponse({"valid": False, "discount": 0, "message": "کد منقضی یا غیرفعال است"})

        return JsonResponse({
            "valid": True,
            "discount": coupon.discount,  # درصد
            "message": "کد معتبر است"
        })

    except Coupon.DoesNotExist:
        return JsonResponse({"valid": False, "discount": 0, "message": "کد نامعتبر است"})


def buy_subscription(request):
    plans = SubscriptionPlan.objects.filter(is_active=True).order_by('duration')
    house = MyHouse.objects.filter(user=request.user).first()

    managed_users = request.user.managed_users.all()
    unit_count = Unit.objects.filter(
        Q(user=request.user) | Q(user__in=managed_users),
        is_active=True,
    ).count()

    # اشتراک فعال فعلی (اگر وجود داشته باشد)
    active_sub = Subscription.objects.filter(
        user=request.user,
        status='active'
    ).order_by("-end_date").first()

    active_subscription = False

    if active_sub and active_sub.end_date:
        remaining_days = (active_sub.end_date.date() - timezone.now().date()).days

        # اگر بیشتر از ۳ روز مانده باشد → اجازه خرید اشتراک جدید ندارد
        if remaining_days > 3:
            active_subscription = True

    if active_subscription:
        messages.warning(request, "هنوز بیش از ۳ روز به پایان اشتراک فعلی شما مانده است.")
        return redirect("middle_admin_dashboard")

    if request.method == "POST":
        try:
            units = int(request.POST.get("units_count"))
        except (ValueError, TypeError):
            messages.error(request, "تعداد واحد نامعتبر است.")
            return redirect("buy_subscription")

        if units < unit_count:
            messages.error(
                request,
                f"تعداد واحد وارد شده نمی‌تواند کمتر از تعداد ثبت‌شده ({unit_count}) باشد."
            )
            return redirect("buy_subscription")

        try:
            plan_id = int(request.POST.get("plan"))
            plan = SubscriptionPlan.objects.get(id=plan_id, is_active=True)
        except (ValueError, SubscriptionPlan.DoesNotExist):
            messages.error(request, "پلن انتخابی نامعتبر است.")
            return redirect("buy_subscription")

        # total = units * plan.price_per_unit
        total_amount = units * plan.price_per_unit
        discount_amount = 0
        coupon = None

        coupon_code = request.POST.get("coupon")

        discount_amount = 0
        coupon = None

        if coupon_code:
            try:
                coupon = Coupon.objects.get(code__iexact=coupon_code)

                # بررسی استفاده قبلی
                already_used = CouponUsage.objects.filter(
                    user=request.user,
                    coupon=coupon
                ).exists()

                if already_used:
                    messages.error(
                        request,
                        "شما قبلاً از این کد تخفیف استفاده کرده‌اید."
                    )
                    return redirect("buy_subscription")

                if not coupon.is_valid():
                    messages.error(
                        request,
                        "کد تخفیف منقضی یا غیرفعال است."
                    )
                    return redirect("buy_subscription")

                if coupon.discount > total_amount:
                    messages.error(
                        request,
                        "مبلغ کد تخفیف بیشتر از مبلغ کل سفارش است و قابل استفاده نیست."
                    )
                    return redirect("buy_subscription")

                discount_amount = coupon.discount

            except Coupon.DoesNotExist:
                messages.error(
                    request,
                    "کد تخفیف نامعتبر است."
                )
                return redirect("buy_subscription")

        final_amount = total_amount - discount_amount

        # تعیین تاریخ شروع
        if active_sub and active_sub.end_date:
            start_date = active_sub.end_date + timedelta(days=1)
        else:
            start_date = timezone.now()

        end_date = start_date + timedelta(days=plan.duration)

        Subscription.objects.create(
            user=request.user,
            house=house,  # اگر داری
            units_count=units,
            plan=plan,

            coupon=coupon,
            total_amount=total_amount,
            discount_amount=discount_amount,
            final_amount=final_amount,

            start_date=start_date,
            end_date=end_date,

            is_paid=False,
            status='active'
        )

        return redirect("subscription_success")

    return render(request, "middle_admin/middle_add_subscription.html", {
        "plans": plans,
        "unit_count": unit_count,
    })


def get_single_resident_building(user):
    units = Unit.objects.filter(
        Q(user=user) |
        Q(renters__user=user, renters__renter_is_active=True),
        is_active=True
    ).select_related('myhouse').distinct()

    buildings = set(
        unit.myhouse_id for unit in units if unit.myhouse_id
    )

    if len(buildings) == 1:
        return units.first()  # یک واحد از همان ساختمان
    return None


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def switch_to_resident(request):
    """
    سوییچ به محیط ساکن برای مدیر یا middleAdmin
    """
    user = request.user

    # فقط admin یا middleAdmin می‌توانند سوییچ کنند
    if not (user.is_staff or user.is_middle_admin):
        messages.error(request, 'دسترسی به پنل ساکن ندارید.')
        return redirect('middle_admin_dashboard')

    # فقط مدیرانی که ساکن هستند می‌توانند
    if not getattr(user, 'is_resident', False):
        messages.error(request, 'شما ساکن ساختمان نیستید.')
        return redirect('middle_admin_dashboard')

    unit = get_single_resident_building(user)
    if not unit:
        messages.error(request, 'جهت ورود به پنل ابتدا باید واحد خود را ثبت کنید.')
        return redirect('middle_admin_dashboard')

    # ✅ تنظیم session
    request.session['active_context'] = 'resident'
    request.session['active_unit_id'] = unit.id
    request.session['active_building_id'] = unit.myhouse.id

    return redirect('user_panel')


# ================================================================
@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_admin_dashboard(request):
    cache_key = f"middle_dashboard_{request.user.id}"
    cached_context = cache.get(cache_key)

    if cached_context:
        return render(
            request,
            'middleShared/home_template.html',
            cached_context
        )

    now = timezone.now().date()

    # =========================
    # Subscription
    # =========================
    subscription = (
        Subscription.objects
        .filter(user=request.user)
        .only(
            'id',
            'status',
            'end_date',
            'created_at'
        )
        .order_by('-created_at')
        .first()
    )

    if not subscription:
        return redirect('buy_subscription')

    subscription.expire_if_needed()
    subscription.refresh_from_db(fields=['status', 'end_date'])

    if (
            subscription.status == "expired"
            or not subscription.end_date
            or subscription.end_date.date() < now
    ):
        return redirect('buy_subscription')

    # =========================
    # Managed Users
    # =========================
    managed_users = request.user.managed_users.values_list(
        'id',
        flat=True
    )

    # =========================
    # Resident Unit
    # =========================
    resident_unit = get_single_resident_building(request.user)

    # =========================
    # Announcements
    # =========================
    announcements = (
        Announcement.objects
        .filter(
            is_active=True,
            user=request.user
        )
        .only(
            'id',
            'title',
            'created_at'
        )
        .order_by('-created_at')[:3]
    )

    # =========================
    # Units
    # =========================
    units = (
        Unit.objects
        .filter(
            myhouse__user=request.user,
            is_active=True
        )
        .select_related('myhouse')
        .distinct()
    )

    unit_count = units.count()

    # =========================
    # Unit Status Statistics
    # =========================
    empty_units_count = units.filter(
        status_residence='empty'
    ).count()

    renter_units_count = (
        units.filter(
            renters__renter_is_active=True
        )
        .distinct()
        .count()
    )

    owner_units_count = (
        units.exclude(
            status_residence='empty'
        )
        .exclude(
            renters__renter_is_active=True
        )
        .distinct()
        .count()
    )

    unit_status_stats = {
        'owner': owner_units_count,
        'renter': renter_units_count,
        'empty': empty_units_count,
    }

    has_unit_chart_data = any(unit_status_stats.values())

    # =========================
    # Expense Chart
    # =========================
    category_expenses = (
        Expense.objects
        .filter(
            house__user=request.user,
            is_active=True,
            is_paid=True
        )
        .values('category__title')
        .annotate(
            total_amount=Sum('amount')
        )
        .order_by('-total_amount')
    )

    expense_chart_data = {
        "labels": [
            item['category__title']
            for item in category_expenses
        ],
        "data": [
            item['total_amount'] or 0
            for item in category_expenses
        ]
    }

    # =========================
    # Income Chart
    # =========================
    income_by_category = (
        IncomeCategory.objects
        .filter(
            user=request.user,
            is_active=True
        )
        .annotate(
            total_amount=Sum(
                'incomes__amount',
                filter=Q(
                    incomes__is_paid=True,
                    incomes__is_active=True
                )
            ),
            incomes_count=Count(
                'incomes',
                filter=Q(
                    incomes__is_paid=True,
                    incomes__is_active=True
                )
            )
        )
        .filter(
            incomes_count__gt=0
        )
    )

    income_chart_data = {
        "labels": [
            cat.subject
            for cat in income_by_category
        ],
        "data": [
            cat.total_amount or 0
            for cat in income_by_category
        ]
    }

    has_income_chart_data = any(
        amount > 0
        for amount in income_chart_data["data"]
    )

    # =========================
    # Tickets
    # =========================
    tickets = (
        SupportUser.objects
        .select_related('user')
        .filter(
            Q(user=request.user)
            | Q(user__in=managed_users)
        )
        .order_by('-created_at')[:5]
    )

    # =========================
    # Fund Balance
    # =========================
    totals = (
        Fund.objects
        .filter(
            Q(user=request.user)
            | Q(user__in=managed_users)
        )
        .aggregate(
            total_income=Sum('debtor_amount'),
            total_expense=Sum('creditor_amount')
        )
    )

    balance = (
            (totals['total_income'] or 0)
            - (totals['total_expense'] or 0)
    )

    # =========================
    # Unpaid Charges Count
    # =========================
    unit_count_unpaid_charges = (
        UnifiedCharge.objects
        .filter(
            house__user=request.user,
            send_notification=True,
            is_paid=False,
            unit__isnull=False
        )
        .count()
    )

    # =========================
    # Monthly Paid Charges
    # =========================
    paid_charges = (
        UnifiedCharge.objects
        .filter(
            house__user=request.user,
            send_notification=True,
            unit__in=units,
            is_paid=True,
            payment_date__isnull=False
        )
        .annotate(
            month=ExtractMonth('payment_date')
        )
        .values('month')
        .annotate(
            count=Count('id')
        )
    )

    # =========================
    # Monthly Unpaid Charges
    # =========================
    unpaid_charges = (
        UnifiedCharge.objects
        .filter(
            house__user=request.user,
            send_notification=True,
            unit__in=units,
            is_paid=False,
            send_notification_date__isnull=False
        )
        .annotate(
            month=ExtractMonth('send_notification_date')
        )
        .values('month')
        .annotate(
            count=Count('id')
        )
    )

    paid_counts = {i: 0 for i in range(1, 13)}
    unpaid_counts = {i: 0 for i in range(1, 13)}

    for item in paid_charges:
        if item['month']:
            paid_counts[item['month']] = item['count']

    for item in unpaid_charges:
        if item['month']:
            unpaid_counts[item['month']] = item['count']

    months = list(range(1, 13))

    paid_data = [
        paid_counts[m]
        for m in months
    ]

    unpaid_data = [
        unpaid_counts[m]
        for m in months
    ]

    has_charge_data = any(paid_data) or any(unpaid_data)

    # =========================
    # Context
    # =========================
    context = {
        'announcements': announcements,
        'unit_count': unit_count,
        'fund_amount': balance,
        'tickets': tickets,
        'unit_count_unpaid_charges': unit_count_unpaid_charges,
        'resident_unit': resident_unit,
        'ownerRenterStats': unit_status_stats,
        'has_unit_chart_data': has_unit_chart_data,
        'expense_chart_data': expense_chart_data,
        'income_chart_data': income_chart_data,
        'months': months,
        'paid_data': paid_data,
        'unpaid_data': unpaid_data,
        'has_charge_data': has_charge_data,
        'has_income_chart_data': has_income_chart_data,
    }

    # =========================
    # Cache
    # =========================
    cache.set(
        cache_key,
        context,
        timeout=300
    )

    return render(
        request,
        'middleShared/home_template.html',
        context
    )


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_admin_login_view(request):
    if request.method == 'POST':
        mobile = request.POST.get('mobile')
        password = request.POST.get('password1')

        user = authenticate(request, mobile=mobile, password=password)
        if user is not None:
            if user.is_middle_admin:
                login(request, user)
                sweetify.success(request, f"{user.full_name} عزیز، با موفقیت وارد بخش مدیر ساختمان شدید!")
                return redirect(reverse('middle_manage_house'))
            else:
                logout(request)  # Log out any non-superuser who authenticated successfully
                messages.error(request, 'شما مجوز دسترسی به بخش مدیر ساختمان را ندارید!')
                return redirect(reverse('login_middle_admin'))
        else:
            messages.error(request, 'نام کاربری و یا رمز عبور اشتباه است!')
            return redirect(reverse('login_middle_admin'))

    return render(request, 'middleShared/middle_login.html')


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def logout__middle_admin(request):
    logout(request)
    return redirect('login_middle')


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def site_header_component(request):
    context = {
        'user': request.user,
    }
    return render(request, 'middleShared/notification_template.html', context)


# =====================================Middle Profile ============================
@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_profile(request):
    user = request.user

    # اگر فرم ارسال شد
    if request.method == 'POST':
        password_form = PasswordChangeForm(user, request.POST)
        if password_form.is_valid():
            password_form.save()
            update_session_auth_hash(request, user)  # مهم: کاربر بعد از تغییر رمز لاگ‌اوت نشود
            messages.success(request, 'رمز عبور با موفقیت تغییر کرد!')
            return redirect('user_profile')  # redirect باعث می‌شود پیام ظاهر شود
        else:
            messages.warning(request, 'رمز عبور باید حداقل 8 رقم و شامل حروف و اعداد باشد. مجددا بررسی فرمایید.')
    else:
        password_form = PasswordChangeForm(user)

    # اطلاعات واحد
    # unit = Unit.objects.filter(user=user).first()

    context = {
        'user_obj': user,
        # 'unit': unit,
        'password_form': password_form,
    }
    return render(request, 'middle_admin/middle_my_profile.html', context)


# ============================= Announcement ====================
@method_decorator(middle_admin_required, name='dispatch')
class MiddleAnnouncementView(CreateView):
    model = Announcement
    template_name = 'middle_admin/middle_send_announcement.html'
    form_class = announcementForm
    success_url = reverse_lazy('middle_announcement')

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.user = self.request.user
        self.object.house = MyHouse.objects.filter(user=self.request.user).first()  # یا .houses.first()

        self.object.save()

        messages.success(self.request, 'اطلاعیه با موفقیت ثبت گردید!')
        return super(MiddleAnnouncementView, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['announcements'] = Announcement.objects.filter(user=self.request.user).order_by('-created_at')
        return context


@method_decorator(middle_admin_required, name='dispatch')
class MiddleAnnouncementListView(ListView):
    model = Announcement
    template_name = 'middle_admin/middle_announcement.html'
    context_object_name = 'announcements'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        queryset = Announcement.objects.filter(
            user=self.request.user,
            is_active=True
        )

        if query:
            queryset = queryset.filter(title__icontains=query)

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@method_decorator(middle_admin_required, name='dispatch')
class MiddleAnnouncementUpdateView(UpdateView):
    model = Announcement
    template_name = 'middle_admin/middle_send_announcement.html'
    form_class = announcementForm
    success_url = reverse_lazy('middle_announcement')

    def form_valid(self, form):
        edit_instance = form.instance
        self.object = form.save(commit=False)
        self.object.user = self.request.user
        self.object.house = MyHouse.objects.filter(user=self.request.user).first()  # یا .houses.first()

        self.object.save()
        messages.success(self.request, 'اطلاعیه با موفقیت ویرایش گردید!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['announcements'] = Announcement.objects.filter(user=self.request.user).order_by('-created_at')
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_announcement_delete(request, pk):
    announce = get_object_or_404(Announcement, id=pk)
    print(announce.id)

    try:
        announce.delete()
        messages.success(request, 'اظلاعیه با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('middle_announcement'))


# ========================== Bank Views ========================
@method_decorator(middle_admin_required, name='dispatch')
class middleAddBankView(CreateView):
    model = Bank
    template_name = 'middle_admin/middle_add_my_bank.html'
    form_class = BankForm
    success_url = reverse_lazy('middle_manage_bank')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        form.instance.user = self.request.user
        response = super().form_valid(form)

        bank = self.object
        house = bank.house
        initial_fund = bank.initial_fund or Decimal('0')
        content_type = ContentType.objects.get_for_model(Bank)

        # اگر موجودی اولیه دارد
        if initial_fund > 0:

            # --- ثبت Fund افتتاحیه (در صورتی که وجود نداشته باشد)
            fund_exists = Fund.objects.filter(
                content_type=content_type,
                object_id=bank.id,
                is_initial=True
            ).exists()

            if not fund_exists:
                Fund.objects.create(
                    user=self.request.user,
                    bank=bank,
                    house=house,
                    payer_name=bank.account_holder_name,
                    receiver_name='صندوق',
                    payment_gateway='کارت به کارت',
                    content_type=content_type,
                    object_id=bank.id,
                    is_initial=True,
                    is_paid=True,
                    amount=initial_fund,
                    debtor_amount=initial_fund,
                    creditor_amount=Decimal('0'),
                    payment_date=bank.create_at.date(),
                    payment_description=f'افتتاحیه حساب بانک {bank.bank_name}'
                )

            # --- ثبت BankFund افتتاحیه (با استفاده از سرویس)
            bankfund_exists = BankFund.objects.filter(
                content_type=content_type,
                object_id=bank.id,
                transaction_type='deposit',
                payment_description__icontains='افتتاحیه'
            ).exists()

            if not bankfund_exists:
                BankTransactionService.deposit(
                    user=self.request.user,
                    bank=bank,
                    unit=None,
                    amount=initial_fund,
                    description=f'افتتاحیه حساب بانک {bank.bank_name}',
                    content_object=bank,
                    payment_date=bank.create_at.date(),
                    gateway="افتتاح حساب",
                    house=bank.house
                )

        else:
            # اگر موجودی اولیه ندارد، موجودی جاری باید صفر بماند
            if bank.current_balance != Decimal('0'):
                bank.current_balance = Decimal('0')
                bank.save(update_fields=['current_balance'])

        messages.success(self.request, 'حساب بانکی با موفقیت ثبت گردید!')
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banks'] = Bank.objects.filter(user=self.request.user)
        return context


@method_decorator(middle_admin_required, name='dispatch')
class middleBankUpdateView(UpdateView):
    model = Bank
    template_name = 'middle_admin/middle_add_my_bank.html'
    form_class = BankForm
    success_url = reverse_lazy('middle_manage_bank')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        bank = self.get_object()
        old_initial_fund = bank.initial_fund or Decimal('0')

        # ست کردن کاربر
        form.instance.user = self.request.user

        # ذخیره بانک
        response = super().form_valid(form)

        # بانک جدید
        bank.refresh_from_db()
        new_initial_fund = bank.initial_fund or Decimal('0')
        house = bank.house

        # اگر تغییری نکرده
        if old_initial_fund == new_initial_fund:
            messages.success(self.request, 'حساب بانکی بروزرسانی شد.')
            return response

        # محاسبه تغییر موجودی
        delta = new_initial_fund - old_initial_fund

        content_type = ContentType.objects.get_for_model(Bank)

        # Fund افتتاحیه
        initial_fund_obj = Fund.objects.filter(
            content_type=content_type,
            object_id=bank.id,
            is_initial=True
        ).first()

        # --------- 1) اگر افتتاحیه جدید مثبت است → Fund را بروزرسانی یا ایجاد کن ---------
        if new_initial_fund > 0:

            if initial_fund_obj:
                # بروزرسانی Fund
                initial_fund_obj.amount = new_initial_fund
                initial_fund_obj.debtor_amount = new_initial_fund
                initial_fund_obj.creditor_amount = 0
                initial_fund_obj.house = house
                initial_fund_obj.payment_description = f"افتتاحیه حساب بانک {bank.bank_name}"
                initial_fund_obj.save()
                Fund.recalc_final_amounts_from(initial_fund_obj)

            else:
                # ایجاد Fund جدید
                fund = Fund.objects.create(
                    user=self.request.user,
                    bank=bank,
                    house=house,
                    payer_name=bank.account_holder_name,
                    receiver_name='صندوق',
                    payment_gateway="کارت به کارت",
                    content_type=content_type,
                    object_id=bank.id,
                    is_initial=True,
                    is_paid=True,
                    amount=new_initial_fund,
                    debtor_amount=new_initial_fund,
                    creditor_amount=0,
                    payment_date=bank.create_at.date(),
                    payment_description=f"افتتاحیه حساب بانک {bank.bank_name}",
                )
                Fund.recalc_final_amounts_from(fund)

        # --------- 2) اگر افتتاحیه صفر شد → Fund افتتاحیه را حذف کن ---------
        else:
            if initial_fund_obj:
                next_fund = Fund.objects.filter(
                    bank=bank,
                    id__gt=initial_fund_obj.id
                ).order_by("id").first()

                initial_fund_obj.delete()

                if next_fund:
                    Fund.recalc_final_amounts_from(next_fund)

        # --------- 3) اعمال تغییر واقعی بر حساب بانکی ---------
        if delta > 0:
            # افزایش افتتاحیه = واریز به بانک
            BankTransactionService.deposit(
                user=self.request.user,
                bank=bank,
                unit=None,
                amount=delta,
                description=f"افزایش موجودی افتتاحیه بانک {bank.bank_name}",
                content_object=bank,
                payment_date=bank.create_at.date(),
                gateway="اصلاح افتتاحیه",
                house=bank.house
            )

        elif delta < 0:
            # کاهش افتتاحیه = برداشت از بانک
            BankTransactionService.withdraw(
                user=self.request.user,
                bank=bank,
                amount=abs(delta),
                description=f"کاهش موجودی افتتاحیه بانک {bank.bank_name}",
                content_object=bank,
                payment_date=bank.created_at.date(),
                gateway="اصلاح افتتاحیه",
                house=bank.house
            )

        messages.success(self.request, 'افتتاحیه حساب بانکی با موفقیت اصلاح شد!')
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['banks'] = Bank.objects.filter(user=self.request.user)
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def middle_bank_delete(request, pk):
    bank = get_object_or_404(Bank, id=pk)
    try:
        has_fund = Fund.objects.filter(bank=bank).exists()

        if has_fund:
            messages.error(
                request,
                'به دلیل وجود گردش مالی، امکان حذف این حساب بانکی وجود ندارد.'
            )
            return redirect('middle_manage_bank')
        bank.delete()
        messages.success(request, 'حساب بانکی با موفقیت حذف گردید!')
        return redirect(reverse('middle_manage_bank'))
    except Bank.DoesNotExist:
        messages.info(request, 'خطا در حذف')
        return redirect(reverse('middle_manage_bank'))


@method_decorator(middle_admin_required, name='dispatch')
class TransferMoneyView(CreateView):
    model = BankFund  # مدلی که قرار است در نهایت یک رکورد در آن ثبت شود
    form_class = TransferMoneyForm
    template_name = 'middle_admin/transfer_money_banks.html'
    success_url = reverse_lazy('middle_manage_bank')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        try:
            BankTransactionService.transfer(
                user=self.request.user,
                from_bank=form.cleaned_data['from_bank'],
                to_bank=form.cleaned_data['to_bank'],
                amount=form.cleaned_data['amount'],
                unit=None,
                description="انتقال وجه داخلی",
                transaction_no=form.cleaned_data['transaction_reference'],
                payment_date=form.cleaned_data['payment_date'],
                # content_object=...  (در صورت نیاز)
            )
            messages.success(self.request, "انتقال وجه با موفقیت انجام شد.")
            return redirect(self.success_url)
        except Exception as e:
            form.add_error(None, str(e))
            return self.form_invalid(form)


def transfer_list_view(request):
    # گرفتن تمام رکوردهای انتقال مربوط به این بانک
    all_transfers = BankFund.objects.filter(
        transfer_group_id__isnull=False,
        user=request.user
    ).order_by('-created_at')

    # دسته‌بندی بر اساس group_id
    grouped_transfers = {}
    for t in all_transfers:
        if t.transfer_group_id not in grouped_transfers:
            grouped_transfers[t.transfer_group_id] = {'withdraw': None, 'deposit': None}

        if t.transaction_type == 'withdraw':
            grouped_transfers[t.transfer_group_id]['withdraw'] = t
        else:
            grouped_transfers[t.transfer_group_id]['deposit'] = t

    context = {
        'grouped_transfers': grouped_transfers.items()  # لیست جفت‌ها
    }
    return render(request, 'middle_admin/transfer_list_view.html', context)


@transaction.atomic
def delete_transfer(request, pk):
    record = get_object_or_404(BankFund, pk=pk)

    if not record.transfer_group_id:
        # اگر این رکورد انتقال نیست، فقط همان یکی حذف شود
        record.delete()
        messages.success(request, "لغو انتقال با موفقیت انجام شد.")
        return redirect('middle_manage_bank')

    BankFund.objects.filter(
        transfer_group_id=record.transfer_group_id
    ).delete()
    messages.success(request, "لغو انتقال با موفقیت انجام شد.")
    return redirect('middle_manage_bank')


# =========================== unit Views ================================

@method_decorator(middle_admin_required, name='dispatch')
class MiddleUnitRegisterView(CreateView):
    model = Unit
    form_class = UnitForm
    template_name = 'middle_unit_templates/unit_register.html'
    success_url = reverse_lazy('middle_manage_unit')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        try:
            with transaction.atomic():
                # ---------- محدودیت بر اساس Subscription ----------
                subscription = Subscription.objects.filter(
                    user=self.request.user,
                ).order_by('-created_at').first()

                if not subscription:
                    # هیچ اشتراکی وجود ندارد
                    form.add_error(None, "هیچ اشتراک فعالی برای شما وجود ندارد. ابتدا اشتراک خریداری کنید.")
                    return self.form_invalid(form)

                # اگر اشتراک موجود است ولی trial و پرداخت نشده
                # if subscription.is_trial and not subscription.is_paid:
                #     form.add_error(None, "اشتراک رایگان شما منقضی شده یا پرداخت نشده است. ابتدا اشتراک خریداری کنید.")
                #     return self.form_invalid(form)

                current_units_count = Unit.objects.filter(myhouse__user=self.request.user).count()
                if current_units_count >= subscription.units_count:
                    messages.warning(self.request, f"شما مجاز به ثبت بیش از {subscription.units_count} واحد نیستید.")
                    # form.add_error(None, f"شما مجاز به ثبت بیش از {subscription.units_count} واحد نیستید.")
                    # return self.form_invalid(form)
                    return redirect('middle_unit_register')

                is_renter = str(form.cleaned_data.get('is_renter')).lower() == 'true'

                # -----------------------
                # 1️⃣ ساخت User مالک
                # -----------------------
                house = MyHouse.objects.filter(
                    user=self.request.user,
                    is_active=True
                ).first()
                owner_mobile = form.cleaned_data.get('owner_mobile')
                owner_name = form.cleaned_data.get('owner_name')

                owner_user, owner_created = User.objects.get_or_create(
                    mobile=owner_mobile,
                    defaults={
                        'username': owner_mobile,
                        'full_name': owner_name,
                        'house': house,
                        'is_active': True,
                        'manager': self.request.user,
                        'is_unit': True,
                    }
                )

                # اگر قبلاً وجود داشته → نقش‌ها آپدیت شود
                if not owner_created:
                    owner_user.is_unit = True
                    owner_user.manager = self.request.user
                    owner_user.house = house  # ✅ درست شد
                    owner_user.full_name = owner_name
                    owner_user.save(update_fields=['house', 'full_name', 'manager', 'is_unit'])

                # اگر تازه ساخته شده → پسورد ست شود
                owner_password = form.cleaned_data.get('owner_password')

                if owner_password:
                    owner_user.set_password(owner_password)
                    owner_user.save()
                # -----------------------
                # 2️⃣ ساخت Unit با مالک
                # -----------------------
                unit = form.save(commit=False)
                unit.user = owner_user
                unit.myhouse = MyHouse.objects.filter(
                    user=self.request.user,
                    is_active=True
                ).first()

                unit.save()

                # -----------------------
                # 3️⃣ اگر مستاجر دارد
                # -----------------------
                if is_renter:

                    renter_mobile = form.cleaned_data.get('renter_mobile')
                    renter_name = form.cleaned_data.get('renter_name')

                    renter_user, renter_created = User.objects.get_or_create(
                        mobile=renter_mobile,

                        defaults={
                            'username': renter_mobile,
                            'full_name': renter_name,
                            'is_active': True,
                            'manager': self.request.user,
                            'is_unit': True,
                            'house': house
                        }
                    )
                    renter_password = form.cleaned_data.get('renter_password')

                    if renter_password:
                        renter_user.set_password(renter_password)
                        renter_user.save()

                    # 👇 غیرفعال کردن مالک
                    owner_user.is_active = True
                    owner_user.save(update_fields=['is_active'])

                    # ساخت مستاجر
                    Renter.objects.create(
                        unit=unit,
                        user=renter_user,
                        myhouse=unit.myhouse,
                        renter_name=renter_name,
                        renter_mobile=renter_mobile,
                        renter_national_code=form.cleaned_data.get('renter_national_code'),
                        renter_people_count=form.cleaned_data.get('renter_people_count'),
                        contract_number=form.cleaned_data.get('contract_number'),
                        estate_name=form.cleaned_data.get('estate_name'),
                        renter_is_active=True,
                        start_date=form.cleaned_data.get('start_date'),
                        end_date=form.cleaned_data.get('end_date'),
                        first_charge_renter=form.cleaned_data.get('first_charge_renter'),
                        renter_payment_date=form.cleaned_data.get('renter_payment_date'),
                        renter_transaction_no=form.cleaned_data.get('renter_transaction_no'),
                        renter_bank=form.cleaned_data.get('renter_bank'),
                    )

                    unit.people_count = form.cleaned_data.get('renter_people_count') or 0

                else:
                    unit.people_count = form.cleaned_data.get('owner_people_count') or 0

                unit.save(update_fields=['people_count'])

                owner_bank = form.cleaned_data.get('owner_bank') or Bank.objects.first()
                first_charge_owner = int(form.cleaned_data.get('first_charge_owner') or 0)
                if first_charge_owner > 0:
                    Fund.objects.create(
                        user=owner_user,
                        unit=unit,
                        bank=owner_bank,
                        house=unit.myhouse,
                        debtor_amount=Decimal(first_charge_owner),
                        creditor_amount=0,
                        amount=Decimal(first_charge_owner),
                        is_initial=True,
                        payment_date=form.cleaned_data.get('owner_payment_date'),
                        payer_name=unit.get_label,
                        payment_description="شارژ اولیه مالک",
                        payment_gateway='کارت به کارت',
                        content_object=unit,
                        transaction_no=form.cleaned_data.get('owner_transaction_no'),
                    )
                    BankTransactionService.deposit(
                        user=self.request.user,
                        bank=owner_bank,
                        unit=unit,
                        amount=Decimal(first_charge_owner),
                        description=f"شارژ اولیه مالک {unit.get_label}",
                        content_object=unit,
                        payment_date=form.cleaned_data.get('owner_payment_date'),
                        transaction_no=form.cleaned_data.get('owner_transaction_no'),
                        gateway="شارژ واحد",
                        house=unit.myhouse
                    )

                # -------------------------
                # ایجاد شارژ اولیه مستاجر
                # -------------------------
                renter_bank = form.cleaned_data.get('renter_bank') or Bank.objects.first()
                first_charge_renter = int(form.cleaned_data.get('first_charge_renter') or 0)
                if first_charge_renter > 0:
                    Fund.objects.create(
                        user=renter_user,
                        unit=unit,
                        bank=renter_bank,
                        house=unit.myhouse,
                        debtor_amount=Decimal(first_charge_renter),
                        creditor_amount=0,
                        amount=Decimal(first_charge_renter),
                        is_initial=True,
                        payment_date=form.cleaned_data.get('renter_payment_date'),
                        payer_name=unit.get_label,
                        payment_description="شارژ اولیه مستاجر",
                        payment_gateway='کارت به کارت',
                        content_object=unit,
                        transaction_no=form.cleaned_data.get('renter_transaction_no'),
                    )
                    BankTransactionService.deposit(
                        user=self.request.user,
                        bank=renter_bank,
                        unit=unit,
                        amount=Decimal(first_charge_renter),
                        description=f"شارژ اولیه مستاجر {unit.get_label}",
                        content_object=unit,
                        payment_date=form.cleaned_data.get('renter_payment_date'),
                        transaction_no=form.cleaned_data.get('renter_transaction_no'),
                        gateway="شارژ واحد",
                        house=unit.myhouse
                    )

                messages.success(self.request, 'واحد با موفقیت ثبت شد.')
                return super().form_valid(form)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            form.add_error(None, f'خطا در ثبت اطلاعات: {e}')
            return self.form_invalid(form)
            # form.add_error(None, f'خطا در ثبت اطلاعات: {e}')
            # return self.form_invalid(form)


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def add_renter_to_unit(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)

    if request.method == 'POST':
        form = RenterAddForm(request.POST, user=request.user)
        if form.is_valid():
            renter_mobile = form.cleaned_data['renter_mobile']

            # ❌ جلوگیری از مستاجر فعال در واحد دیگر
            if Renter.objects.filter(
                    user__mobile=renter_mobile,
                    renter_is_active=True
            ).exclude(unit=unit).exists():
                form.add_error(
                    'renter_mobile',
                    'این شماره موبایل در واحد دیگری به عنوان مستاجر فعال ثبت شده است.'
                )
                return render(
                    request,
                    'middle_unit_templates/new_renter_register.html',
                    {'form': form, 'unit': unit}
                )

            with transaction.atomic():
                # -------------------------
                # غیرفعال کردن مستاجر قبلی
                # -------------------------
                unit.renters.filter(renter_is_active=True).update(renter_is_active=False)

                # -------------------------
                # گرفتن یا ساخت یوزر مستاجر
                # -------------------------
                renter_user, created = User.objects.get_or_create(
                    mobile=renter_mobile,
                    defaults={
                        'username': renter_mobile,
                        'full_name': form.cleaned_data['renter_name'],
                        'is_active': True,
                        'manager': request.user,
                        'is_unit': True,  # ← اضافه شد
                    }
                )

                # بروزرسانی اطلاعات یوزر
                renter_user.full_name = form.cleaned_data['renter_name']
                renter_user.is_active = True
                password = form.cleaned_data.get('password')
                if password:
                    renter_user.set_password(password)
                renter_user.save()

                # -------------------------
                # انتخاب بانک
                # -------------------------
                renter_bank = form.cleaned_data.get('renter_bank') or Bank.objects.first()

                # -------------------------
                # ایجاد مستاجر جدید
                # -------------------------
                Renter.objects.create(
                    unit=unit,
                    user=renter_user,
                    renter_bank=renter_bank,
                    myhouse=unit.myhouse,
                    renter_name=form.cleaned_data['renter_name'],
                    renter_mobile=renter_mobile,
                    renter_national_code=form.cleaned_data['renter_national_code'],
                    renter_people_count=form.cleaned_data['renter_people_count'],
                    start_date=form.cleaned_data['start_date'],
                    end_date=form.cleaned_data['end_date'],
                    contract_number=form.cleaned_data['contract_number'],
                    estate_name=form.cleaned_data['estate_name'],
                    first_charge_renter=form.cleaned_data.get('first_charge_renter') or 0,
                    renter_details=form.cleaned_data.get('renter_details'),
                    renter_is_active=True,
                    renter_payment_date=form.cleaned_data.get('renter_payment_date'),
                    renter_transaction_no=form.cleaned_data.get('renter_transaction_no'),
                )
                owner_user = unit.user
                # owner_user.is_active = False
                owner_user.save()
                # owner_user.save(update_fields=['is_active'])

                # -------------------------
                # بروزرسانی واحد
                # -------------------------
                unit.is_renter = True

                # -------------------------
                # بروزرسانی people_count
                # -------------------------
                active_renter = unit.get_active_renter()
                if active_renter:
                    unit.people_count = int(active_renter.renter_people_count or 0)
                else:
                    unit.people_count = int(unit.owner_people_count or 0)
                unit.save(update_fields=['is_renter', 'people_count'])

                # -------------------------
                # ایجاد شارژ اولیه مستاجر
                # -------------------------
                first_charge_renter = int(form.cleaned_data.get('first_charge_renter') or 0)
                if first_charge_renter > 0:
                    Fund.objects.create(
                        user=renter_user,
                        unit=unit,
                        house=unit.myhouse,
                        bank=renter_bank,
                        debtor_amount=Decimal(first_charge_renter),
                        creditor_amount=0,
                        amount=Decimal(first_charge_renter),
                        is_initial=True,
                        payment_date=form.cleaned_data.get('renter_payment_date'),
                        payer_name=unit.get_label,
                        payment_description="شارژ اولیه مستاجر",
                        payment_gateway='کارت به کارت',
                        content_object=unit,
                        transaction_no=form.cleaned_data.get('renter_transaction_no'),
                    )
                    BankTransactionService.deposit(
                        user=request.user,
                        bank=renter_bank,
                        unit=unit,
                        amount=Decimal(first_charge_renter),
                        description=f"شارژ اولیه مستاجر {unit.get_label}",
                        content_object=unit,
                        payment_date=form.cleaned_data.get('renter_payment_date'),
                        transaction_no=form.cleaned_data.get('renter_transaction_no'),
                        gateway="شارژ واحد",
                        house=unit.myhouse
                    )

                messages.success(request, 'مستاجر با موفقیت ثبت شد.')
                return redirect('middle_manage_unit')

    else:
        form = RenterAddForm(user=request.user)

    return render(
        request,
        'middle_unit_templates/new_renter_register.html',
        {'form': form, 'unit': unit}
    )


@method_decorator(middle_admin_required, name='dispatch')
class MiddleUnitUpdateView(UpdateView):
    model = Unit
    form_class = UnitForm
    template_name = 'middle_unit_templates/edit_unit.html'
    success_url = reverse_lazy('middle_manage_unit')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        try:
            UnitUpdateService(
                unit=self.object,
                form=form,
                request=self.request
            ).execute()

            messages.success(self.request, 'اطلاعات واحد با موفقیت ویرایش شد')
            return super().form_valid(form)

        except ValueError as e:
            # نمایش پیام واقعی
            if str(e) == 'duplicate_mobile':
                messages.error(self.request, 'شماره موبایل وارد شده قبلاً ثبت شده است.')
            else:
                messages.error(self.request, f'خطا در ثبت! ({e})')

            return self.form_invalid(form)


@method_decorator(middle_admin_required, name='dispatch')
class MiddleUnitInfoView(DetailView):
    model = Unit
    template_name = 'middle_unit_templates/unit_info.html'
    context_object_name = 'unit'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unit = self.object

        # --- جستجو ---
        q = self.request.GET.get('q', '').strip()
        page_number = self.request.GET.get('page')

        renters_qs = unit.renters.all()

        if q:
            renters_qs = renters_qs.filter(
                Q(renter_name__icontains=q) |
                Q(renter_mobile__icontains=q) |
                Q(renter_national_code__icontains=q) |
                Q(contract_number__icontains=q)
            )

        renters_qs = renters_qs.order_by(
            '-renter_is_active',
            '-start_date'
        )

        # --- صفحه‌بندی ---
        paginator = Paginator(renters_qs, 20)
        page_obj = paginator.get_page(page_number)

        context['renters'] = page_obj.object_list
        context['page_obj'] = page_obj
        context['q'] = q

        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_unit_delete(request, pk):
    unit = get_object_or_404(Unit, id=pk)

    # اگر رکورد مرتبط در Fund یا UnifiedCharge وجود داشته باشد، حذف امکان‌پذیر نیست
    if unit.funds.exists():
        messages.error(
            request,
            "امکان حذف وجود ندارد! برای این واحد در گردش صندوق رکورد ثبت شده است."
        )
        return redirect(reverse('middle_manage_unit'))

    try:
        # فقط زمانی مستاجر حذف می‌شود که Unit قابل حذف باشد
        unit.renters.all().delete()

        # حذف کاربر فقط اگر هیچ واحد دیگری نداشته باشد
        if unit.user and not Unit.objects.filter(user=unit.user).exclude(pk=unit.pk).exists():
            unit.user.delete()

        # حذف خود واحد
        unit.delete()
        messages.success(
            request,
            "واحد با موفقیت حذف گردید! مستاجرها و سوابق سکونت به درستی مدیریت شدند."
        )

    except ProtectedError:
        messages.error(request, "امکان حذف وجود ندارد!")

    return redirect(reverse('middle_manage_unit'))


@method_decorator(middle_admin_required, name='dispatch')
class MiddleUnitListView(ListView):
    model = Unit
    template_name = 'middle_unit_templates/unit_management.html'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None
        return int(paginate) if paginate and paginate.isdigit() else 20

    def get_queryset(self):
        user = self.request.user

        queryset = (
            Unit.objects
            .filter(myhouse__user=user)
            .prefetch_related('renters')
        )

        filters = Q()
        params = self.request.GET

        if params.get('unit', '').isdigit():
            filters &= Q(unit=int(params['unit']))

        if params.get('owner_name'):
            filters &= Q(owner_name__icontains=params['owner_name'])

        if params.get('owner_mobile'):
            filters &= Q(owner_mobile__icontains=params['owner_mobile'])

        if params.get('area', '').isdigit():
            filters &= Q(area=int(params['area']))

        if params.get('bedrooms_count', '').isdigit():
            filters &= Q(bedrooms_count=int(params['bedrooms_count']))

        if params.get('renter_name'):
            filters &= Q(renters__renter_name__icontains=params['renter_name'])

        if params.get('renter_mobile'):
            filters &= Q(renters__renter_mobile__icontains=params['renter_mobile'])

        if params.get('people_count', '').isdigit():
            filters &= Q(owner_people_count=int(params['people_count']))

        if params.get('status_residence'):
            filters &= Q(status_residence=params['status_residence'])

        qs = queryset.filter(filters).distinct().order_by('unit')

        for unit in qs:
            unit.active_renters = unit.renters.filter(renter_is_active=True)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context.update({
            'total_units': Unit.objects.filter(
                myhouse__user=self.request.user
            ).count(),
            'units': context['object_list'],
            'paginate': self.request.GET.get('paginate', '20'),
        })

        return context


def to_jalali(date_obj):
    if not date_obj:
        return ''
    jalali_date = jdatetime.date.fromgregorian(date=date_obj)
    return jalali_date.strftime('%Y/%m/%d')


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_units_excel(request):
    units = Unit.objects.filter(user__manager=request.user).order_by('unit')

    filter_fields = {
        'unit': 'unit__icontains',
        'owner_name': 'owner_name__icontains',
        'owner_mobile': 'owner_mobile__icontains',
        'renter_name': 'renter_name__icontains',
        'renter_mobile': 'renter_mobile__icontains',
        'status_residence': 'status_residence__icontains',
        'area': 'area__icontains',
        'bedrooms_count': 'bedrooms_count__icontains',
        'people_count': 'people_count__icontains',
    }

    # Apply filters based on GET parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            units = units.filter(**filter_expression)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "units"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست واحدها")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=22)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = [
        'واحد', 'طبقه', 'متراژ', 'تعداد خواب', 'شماره تلفن',
        'تعداد پارکینگ', 'شماره پارکینگ', 'موقعیت پارکینک', 'وضعیت سکونت',
        'نام مالک', 'تلفن مالک', 'کد ملی مالک', 'تاریخ خرید', 'تعداد نفرات',
        'نام مستاجر', 'تلفن مستاجر', 'کد ملی مستاجر',
        'تاریخ اجاره', 'تاریخ پایان', 'شماره قرارداد', 'اجاره دهنده', 'شارژ اولیه',
    ]

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, unit in enumerate(units, start=3):
        ws.cell(row=row_num, column=1, value=unit.unit)
        ws.cell(row=row_num, column=2, value=unit.floor_number)
        ws.cell(row=row_num, column=3, value=unit.area)
        ws.cell(row=row_num, column=4, value=unit.bedrooms_count)
        ws.cell(row=row_num, column=5, value=unit.unit_phone)
        ws.cell(row=row_num, column=6, value=unit.parking_count)
        ws.cell(row=row_num, column=7, value=unit.parking_number)
        ws.cell(row=row_num, column=8, value=unit.parking_place)
        ws.cell(row=row_num, column=9, value=unit.status_residence)
        ws.cell(row=row_num, column=10, value=unit.owner_name)
        ws.cell(row=row_num, column=11, value=unit.owner_mobile)
        ws.cell(row=row_num, column=12, value=unit.owner_national_code)
        ws.cell(row=row_num, column=13, value=to_jalali(unit.purchase_date))
        ws.cell(row=row_num, column=14, value=unit.people_count)
        renter = unit.renters.first()  # or .last(), or use filtering for current renter
        if renter:
            ws.cell(row=row_num, column=15, value=renter.renter_name)
            ws.cell(row=row_num, column=16, value=renter.renter_mobile)
            ws.cell(row=row_num, column=17, value=renter.renter_national_code)
            ws.cell(row=row_num, column=18, value=to_jalali(renter.start_date))
            ws.cell(row=row_num, column=19, value=to_jalali(renter.end_date))
            ws.cell(row=row_num, column=20, value=renter.contract_number)
            ws.cell(row=row_num, column=21, value=renter.estate_name)
            ws.cell(row=row_num, column=22, value=renter.first_charge_renter)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=units.xlsx'
    wb.save(response)
    return response


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_units_pdf(request):
    units = Unit.objects.filter(user__manager=request.user).order_by('unit')

    filter_fields = {
        'unit': 'unit__icontains',
        'owner_name': 'owner_name__icontains',
        'owner_mobile': 'owner_mobile__icontains',
        'renter_name': 'renter_name__icontains',
        'renter_mobile': 'renter_mobile__icontains',
        'status_residence': 'status_residence__icontains',
        'area': 'area__icontains',
        'bedrooms_count': 'bedrooms_count__icontains',
        'people_count': 'people_count__icontains',
    }

    # Apply filters based on GET parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            units = units.filter(**filter_expression)

    # PDF settings
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        body {{
            font-family: 'BYekan', sans-serif;
        }}
        @font-face {{
            font-family: 'BYekan';
            src: url('{font_url}');
        }}
    """)

    # Render template
    template = get_template("unit_templates/unit_pdf.html")
    context = {
        'units': units,
        'font_path': font_url,
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])
    page_pdf.seek(0)

    # Create response
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="filtered_units.pdf"'
    pdf_merger.write(response)

    return response


# ================================= Expense Views ==============================
@method_decorator(middle_admin_required, name='dispatch')
class MiddleExpenseCategoryView(CreateView):
    model = ExpenseCategory
    template_name = 'middle_expense_templates/add_category_expense.html'
    form_class = ExpenseCategoryForm
    success_url = reverse_lazy('middle_register_category_expense')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            self.object = form.save()
            messages.success(self.request, 'موضوع هزینه با موفقیت ثبت گردید!')
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'خطا در ثبت !')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = ExpenseCategory.objects.filter(user=self.request.user)
        return context


@method_decorator(middle_admin_required, name='dispatch')
class MiddleExpenseCategoryUpdate(UpdateView):
    model = ExpenseCategory
    template_name = 'middle_expense_templates/add_category_expense.html'
    form_class = ExpenseCategoryForm
    success_url = reverse_lazy('middle_register_category_expense')

    def form_valid(self, form):
        form.instance.user = self.request.user

        try:
            edit_instance = form.instance
            self.object = form.save()
            if self.object.is_default:
                messages.error(self.request, f' موضوع هزینه پیش فرض قابل ویرایش نیست!')

            messages.success(self.request, f' موضوع هزینه با موفقیت ویرایش گردید!')
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'خطا در ثبت !')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = ExpenseCategory.objects.filter(user=self.request.user)
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_expense_category_delete(request, pk):
    category = get_object_or_404(ExpenseCategory, id=pk)
    if category.is_default:
        messages.error(request, 'موضوع هزینه قابل حذف نیست!')
        return redirect(reverse('middle_register_category_expense'))

    try:
        category.delete()
        messages.success(request, 'موضوع هزینه با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('middle_register_category_expense'))


@method_decorator(middle_admin_required, name='dispatch')
class MiddleExpenseView(CreateView):
    model = Expense
    template_name = 'middle_expense_templates/expense_register.html'
    form_class = ExpenseForm
    success_url = reverse_lazy('middle_register_expense')

    def form_valid(self, form):
        files = self.request.FILES.getlist('document')
        if len(files) > 2:
            messages.error(self.request, "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
            return redirect('middle_register_expense')

        form.instance.user = self.request.user
        house = MyHouse.objects.filter(user=self.request.user, is_active=True).first()
        form.instance.house = house  # 🔹 این مهم است

        try:
            with transaction.atomic():
                self.object = form.save(commit=False)

                # هزینه هنوز پرداخت نشده
                self.object.is_paid = False
                self.object.save()

                # ذخیره فایل‌ها
                # files = self.request.FILES.getlist('document')
                for f in files:
                    ExpenseDocument.objects.create(
                        expense=self.object,
                        document=f
                    )

            messages.success(
                self.request,
                'هزینه با موفقیت ثبت شد (در انتظار پرداخت)'
            )
            return redirect(self.success_url)

        except Exception:
            messages.error(self.request, 'خطا در ثبت هزینه')
            return self.form_invalid(form)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_queryset(self):
        queryset = Expense.objects.filter(user=self.request.user).order_by('-created_at')

        # فیلتر بر اساس category__title
        category_id = self.request.GET.get('category')
        if category_id:
            queryset = queryset.filter(category__id=category_id)

        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank__id=bank_id)

        # فیلتر بر اساس amount
        amount = self.request.GET.get('amount')
        if amount:
            queryset = queryset.filter(amount__icontains=amount)

        description = self.request.GET.get('description')
        if description:
            queryset = queryset.filter(description__icontains=description)

        doc_no = self.request.GET.get('doc_no')
        if doc_no:
            queryset = queryset.filter(doc_no__icontains=doc_no)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        receiver_name = self.request.GET.get('receiver_name')
        if receiver_name:
            queryset = queryset.filter(receiver_name__icontains=receiver_name)

        # فیلتر بر اساس date
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')

        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y-%m-%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y-%m-%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')

        is_paid = self.request.GET.get('is_paid')
        if is_paid == '1':
            queryset = queryset.filter(is_paid=True)
        elif is_paid == '0':
            queryset = queryset.filter(is_paid=False)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        expenses = self.get_queryset()  # از get_queryset برای دریافت داده‌های فیلتر شده استفاده می‌کنیم
        paginator = Paginator(expenses, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_expense'] = Expense.objects.filter(user=self.request.user).count()
        context['categories'] = ExpenseCategory.objects.filter(user=self.request.user)
        context['banks'] = Bank.objects.filter(user=self.request.user)
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def expense_pay_view(request, expense_id):
    expense = get_object_or_404(
        Expense,
        id=expense_id,
        is_paid=False,
        is_active=True
    )

    if request.method == 'POST':
        form = ExpensePayForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                with transaction.atomic():
                    bank = form.cleaned_data['bank']
                    reference = form.cleaned_data.get('transaction_reference')
                    payment_date = form.cleaned_data.get('payment_date')
                    receiver_name = form.cleaned_data.get('receiver_name')
                    unit = form.cleaned_data['unit']

                    # 🔹 موجودی فعلی صندوق بانک انتخاب شده

                    current_balance = bank.current_balance

                    if current_balance < expense.amount:
                        messages.error(
                            request,
                            f'موجودی بانک کافی نیست. موجودی فعلی: {current_balance:,} تومان'
                        )
                        return redirect(request.META.get('HTTP_REFERER'))

                    # funds = Fund.objects.filter(user=request.user, bank=bank)
                    # print(f"Funds count for bank {bank.id}: {funds.count()}")
                    # for f in funds:
                    #     print(f"Fund: {f.id}, bank: {f.bank}, final: {f.final_amount}")
                    #
                    # bank_funds = Fund.objects.filter(user=request.user, bank=bank)
                    # total_debit = bank_funds.aggregate(Sum('debtor_amount'))['debtor_amount__sum'] or 0
                    # total_credit = bank_funds.aggregate(Sum('creditor_amount'))['creditor_amount__sum'] or 0
                    # current_final = Decimal(total_debit) - Decimal(total_credit)
                    #
                    # print(f'bank-fund:{current_final}')

                    # 🔴 بررسی موجودی
                    # if current_final < expense.amount:
                    #     messages.error(
                    #         request,
                    #         'موجودی صندوق کافی نیست'
                    #     )
                    #     return redirect(request.META.get('HTTP_REFERER'))

                    # 🔹 ثبت Fund (هزینه → بستانکار)
                    Fund.objects.create(
                        unit=unit if unit else None,
                        receiver_name=receiver_name if not unit else f' {unit.get_label()}',
                        user=request.user,
                        bank=bank,
                        house=expense.house,
                        content_object=expense,
                        amount=expense.amount,
                        debtor_amount=0,
                        creditor_amount=expense.amount,
                        payment_date=payment_date,
                        transaction_no=reference,
                        payment_gateway='کارت به کارت',
                        payment_description=f' هزینه:  {expense.category.title}',
                        is_paid=True,

                    )
                    BankTransactionService.withdraw(
                        user=request.user,
                        bank=bank,
                        unit=unit if unit else None,
                        amount=Decimal(expense.amount),
                        description=f' هزینه: {expense.category.title}',
                        content_object=expense,
                        payment_date=payment_date,
                        transaction_no=reference,
                        gateway='کارت به کارت',
                        receiver_name=receiver_name if not unit else f' {unit.get_label}',
                        house=expense.house
                    )

                    # 🔹 بروزرسانی Expense
                    expense.is_paid = True
                    expense.bank = bank
                    expense.transaction_reference = reference
                    expense.payment_date = payment_date
                    expense.unit = unit
                    expense.receiver_name = unit.get_label if unit else receiver_name

                    expense.save(update_fields=[
                        'is_paid',
                        'bank',
                        'transaction_reference',
                        'payment_date',
                        'unit',
                        'receiver_name'
                    ])

                messages.success(request, 'پرداخت با موفقیت انجام شد')
                return redirect('middle_register_expense')

            except ValidationError as e:
                messages.error(request, e.message)
            except Exception as e:
                messages.error(request, f'خطا در پرداخت: {e}')

    else:
        form = ExpensePayForm(user=request.user)

    return render(
        request,
        'middle_expense_templates/expense_pay.html',
        {
            'expense': expense,
            'form': form
        }
    )


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def expense_cancel_pay_view(request, expense_id):
    # Expense فقط اگر پرداخت شده باشد قابل لغو است
    expense = get_object_or_404(
        Expense,
        id=expense_id,
        is_paid=True,
        is_active=True
    )

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # پیدا کردن Fund مربوطه
                fund = Fund.objects.filter(
                    content_type__model='expense',
                    object_id=expense.id,
                    user=request.user,
                    is_paid=True
                ).first()

                if not fund:
                    messages.error(request, 'Fund مرتبط با این پرداخت پیدا نشد!')
                    return redirect(request.META.get('HTTP_REFERER'))

                # حذف Fund
                fund.delete()
                bank_fund = BankFund.objects.filter(
                    content_type__model='expense',
                    object_id=expense.id
                ).first()
                print(f'bank_fund: {bank_fund}')

                if bank_fund:
                    bank = bank_fund.bank

                    # اصلاح موجودی بانک
                    if bank_fund.transaction_type == 'withdraw':
                        bank.current_balance -= bank_fund.amount
                    else:
                        bank.current_balance += bank_fund.amount

                    bank.save(update_fields=['current_balance'])

                    bank_fund.delete()

                # بازمحاسبه موجودی صندوق از این Fund به بعد
                Fund.recalc_final_amounts_from(fund)

                # بازگرداندن Expense به حالت پرداخت‌نشده
                expense.is_paid = False
                expense.bank = None
                expense.transaction_reference = None
                expense.payment_date = None
                expense.save(update_fields=[
                    'is_paid',
                    'bank',
                    'transaction_reference',
                    'payment_date'
                ])

                messages.success(request, 'پرداخت با موفقیت لغو شد و صندوق و بانک اصلاح شد.')
                return redirect(request.META.get('HTTP_REFERER'))

        except Exception as e:
            messages.error(request, f'خطا در لغو پرداخت: {e}')
            return redirect(request.META.get('HTTP_REFERER'))

    # اگر GET باشد، فقط برگرد به صفحه قبل
    return redirect(request.META.get('HTTP_REFERER'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_expense_edit(request, pk):
    expense = get_object_or_404(Expense, pk=pk)

    # اگر پرداخت شده → ویرایش مجاز نیست
    if expense.is_paid:
        messages.warning(request, 'این هزینه پرداخت شده است و قابل ویرایش نیست.')
        return redirect('middle_register_expense')

    if request.method != 'POST':
        return redirect('middle_register_expense')

    form = ExpenseForm(request.POST, request.FILES, instance=expense, user=request.user)

    if not form.is_valid():
        messages.error(request, f'خطا در ویرایش فرم هزینه: {form.errors}')
        return redirect('middle_register_expense')

    try:
        new_amount = Decimal(form.cleaned_data['amount'] or 0)
    except Exception:
        messages.error(request, "مقدار مبلغ وارد شده معتبر نیست.")
        return redirect('middle_register_expense')

    with transaction.atomic():
        files = request.FILES.getlist('document')
        if len(files) > 2:
            messages.error(request, "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
            return redirect('middle_register_expense')
        # 🔹 ست کردن خانه مرتبط با کاربر
        house = MyHouse.objects.filter(user=request.user, is_active=True).first()
        form.instance.house = house

        # ذخیره Expense
        expense = form.save()

        # ذخیره فایل‌های جدید بدون حذف قبلی
        for f in files:
            ExpenseDocument.objects.create(expense=expense, document=f)

    messages.success(request, 'هزینه با موفقیت ویرایش شد.')
    return redirect('middle_register_expense')


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_expense_delete(request, pk):
    expense = get_object_or_404(Expense, id=pk)

    try:
        with transaction.atomic():
            # حذف Fund مربوطه
            expense_ct = ContentType.objects.get_for_model(Expense)
            Fund.objects.filter(content_type=expense_ct, object_id=expense.id).delete()

            # حذف خود Expense
            expense.delete()

        messages.success(request, 'هزینه با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, "امکان حذف وجود ندارد!")
    except Exception as e:
        messages.error(request, f"خطا در حذف: {str(e)}")

    return redirect(reverse('middle_register_expense'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
@csrf_exempt
def middle_delete_expense_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        expense_id = request.POST.get('expense_id')

        if not image_url or not expense_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            expense = get_object_or_404(Expense, id=expense_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = ExpenseDocument.objects.filter(expense=expense, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_export_expense_pdf(request):
    expenses = Expense.objects.filter(user=request.user)
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    filter_fields = {
        'category': 'category__id',
        'bank': 'bank__id',
        'amount': 'amount__icontains',
        'doc_no': 'doc_no__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
        'is_paid': 'is_paid',
    }

    # Apply filters based on GET parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            expenses = expenses.filter(**filter_expression)

    # Handle date filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__lte=to_date)
    except ValueError:
        expenses = Expense.objects.none()

    # Log the filtered expenses for debugging
    print(expenses)

    # Font setup
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # Render HTML template
    template = get_template("middle_expense_templates/expense_pdf.html")
    context = {
        'expenses': expenses,
        'font_path': font_url,
        'house': house,
        'today': timezone.now()
    }
    html = template.render(context)

    # Generate PDF
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # Generate the final PDF response
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="expenses.pdf"'
    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_export_expense_excel(request):
    expenses = Expense.objects.filter(user=request.user)

    # Filter fields
    filter_fields = {
        'category': 'category__id',
        'bank': 'bank__id',
        'amount': 'amount__icontains',
        'doc_no': 'doc_no__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
        'is_paid': 'is_paid'
    }

    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            expenses = expenses.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__lte=to_date)
    except ValueError:
        expenses = Expense.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "expenses"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست هزینه‌ها")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'موضوع هزینه', 'شرح سند', ' شماره سند', 'مبلغ', 'تاریخ سند', 'پرداخت به', 'شماره حساب',
               'تاریخ پرداخت', 'توضیحات']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, expense in enumerate(expenses, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        ws.cell(row=row_num, column=2, value=expense.category.title)
        ws.cell(row=row_num, column=3, value=expense.description)
        ws.cell(row=row_num, column=4, value=expense.doc_no)
        ws.cell(row=row_num, column=5, value=expense.amount)
        ws.cell(row=row_num, column=6, value=show_jalali(expense.date))
        ws.cell(row=row_num, column=7, value=expense.receiver_name)
        ws.cell(row=row_num, column=8, value=expense.bank.bank_name)
        ws.cell(row=row_num, column=9, value=show_jalali(expense.payment_date))
        ws.cell(row=row_num, column=10, value=expense.details)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=expenses.xlsx'
    wb.save(response)
    return response


# =========================== Income Views =========================
@method_decorator(middle_admin_required, name='dispatch')
class MiddleIncomeCategoryView(CreateView):
    model = IncomeCategory
    template_name = 'middle_income_templates/add_category_income.html'
    form_class = IncomeCategoryForm
    success_url = reverse_lazy('middle_register_category_income')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            self.object = form.save()

            messages.success(self.request, 'موضوع درآمد با موفقیت ثبت گردید!')
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'خطا در ثبت !')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['income_categories'] = IncomeCategory.objects.filter(user=self.request.user)
        return context


@method_decorator(middle_admin_required, name='dispatch')
class MiddleIncomeCategoryUpdate(UpdateView):
    model = IncomeCategory
    template_name = 'middle_income_templates/add_category_income.html'
    form_class = IncomeCategoryForm
    success_url = reverse_lazy('middle_register_category_income')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            edit_instance = form.instance
            self.object = form.save()
            messages.success(self.request, f' موضوع درآمد با موفقیت ویرایش گردید!')
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'خطا در ثبت !')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['income_categories'] = IncomeCategory.objects.filter(user=self.request.user)
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_income_category_delete(request, pk):
    income_category = get_object_or_404(IncomeCategory, id=pk)
    try:
        income_category.delete()
        messages.success(request, 'موضوع درآمد با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('middle_register_category_income'))


@method_decorator(middle_admin_required, name='dispatch')
class MiddleIncomeView(CreateView):
    model = Income
    template_name = 'middle_income_templates/income_register.html'
    form_class = IncomeForm
    success_url = reverse_lazy('middle_register_income')

    def form_valid(self, form):
        files = self.request.FILES.getlist('document')
        if len(files) > 2:
            messages.error(self.request, "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
            return redirect('middle_register_income')
        form.instance.user = self.request.user
        house = MyHouse.objects.filter(user=self.request.user, is_active=True).first()
        form.instance.house = house  # 🔹 این مهم است

        try:
            with transaction.atomic():
                self.object = form.save(commit=False)

                # هزینه هنوز پرداخت نشده
                self.object.is_paid = False
                self.object.save()

                # ذخیره فایل‌ها
                # files = self.request.FILES.getlist('document')
                for f in files:
                    IncomeDocument.objects.create(
                        income=self.object,
                        document=f
                    )

            messages.success(
                self.request,
                'درآمد با موفقیت ثبت شد (در انتظار دریافت)'
            )
            return redirect(self.success_url)

        except Exception:
            messages.error(self.request, 'خطا در ثبت درآمد')
            return self.form_invalid(form)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_queryset(self):
        queryset = Income.objects.filter(user=self.request.user).order_by('-created_at')

        # فیلتر بر اساس category
        category_id = self.request.GET.get('category')
        if category_id:
            queryset = queryset.filter(category__id=category_id)

        # فیلتر بر اساس بانک
        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank__id=bank_id)

        # فیلتر بر اساس واحد
        unit_id = self.request.GET.get('unit')
        if unit_id:
            queryset = queryset.filter(unit__id=unit_id)

        # فیلتر بر اساس amount
        amount = self.request.GET.get('amount')
        if amount:
            queryset = queryset.filter(amount__icontains=amount)

        # فیلتر بر اساس description
        description = self.request.GET.get('description')
        if description:
            queryset = queryset.filter(description__icontains=description)

        # فیلتر بر اساس doc_number
        doc_number = self.request.GET.get('doc_number')
        if doc_number:
            queryset = queryset.filter(doc_number__icontains=doc_number)

        # فیلتر بر اساس details
        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        payer_name = self.request.GET.get('payer_name')
        if payer_name:
            queryset = queryset.filter(payer_name__icontains=payer_name)

        # فیلتر بر اساس تاریخ
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')
        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y-%m-%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(doc_date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y-%m-%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(doc_date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')

        is_paid = self.request.GET.get('is_paid')
        if is_paid == '1':
            queryset = queryset.filter(is_paid=True)
        elif is_paid == '0':
            queryset = queryset.filter(is_paid=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        incomes = self.get_queryset()  # از get_queryset برای دریافت داده‌های فیلتر شده استفاده می‌کنیم
        paginator = Paginator(incomes, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_incomes'] = Income.objects.filter(user=self.request.user).count()
        context['categories'] = IncomeCategory.objects.filter(user=self.request.user)
        context['banks'] = Bank.objects.filter(user=self.request.user)
        managed_users = User.objects.filter(Q(manager=self.request.user) | Q(pk=self.request.user.pk))
        context['units'] = Unit.objects.filter(is_active=True, user__in=managed_users)

        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def income_pay_view(request, income_id):
    income = get_object_or_404(
        Income,
        id=income_id,
        is_paid=False,
        is_active=True
    )

    if request.method == 'POST':
        form = IncomePayForm(request.POST, user=request.user)

        if form.is_valid():
            try:
                with transaction.atomic():

                    bank = form.cleaned_data['bank']
                    reference = form.cleaned_data.get('transaction_reference')
                    payment_date = form.cleaned_data.get('payment_date')
                    payer_name = form.cleaned_data.get('payer_name')
                    unit = form.cleaned_data['unit']

                    payer = unit.get_label if unit else payer_name

                    Fund.objects.create(
                        unit=unit,
                        payer_name=payer,
                        user=request.user,
                        bank=bank,
                        house=income.house,
                        content_object=income,
                        amount=income.amount,
                        debtor_amount=income.amount,
                        creditor_amount=0,
                        payment_date=payment_date,
                        transaction_no=reference,
                        payment_gateway='کارت به کارت',
                        payment_description=f'درآمد: {income.category.subject}',
                        is_paid=True,
                    )

                    BankTransactionService.deposit(
                        user=request.user,
                        bank=bank,
                        unit=unit,
                        amount=Decimal(income.amount),
                        description=f'درآمد: {income.category.subject}',
                        content_object=income,
                        payment_date=payment_date,
                        transaction_no=reference,
                        gateway='کارت به کارت',
                        payer_name=payer,
                        house=income.house
                    )

                    income.is_paid = True
                    income.bank = bank
                    income.transaction_reference = reference
                    income.payment_date = payment_date
                    income.unit = unit
                    income.payer_name = payer

                    income.save(update_fields=[
                        'is_paid',
                        'bank',
                        'transaction_reference',
                        'payment_date',
                        'unit',
                        'payer_name'
                    ])

                messages.success(request, 'دریافت با موفقیت انجام شد')
                return redirect('middle_register_income')

            except ValidationError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f'خطا در دریافت: {e}')

    else:
        form = IncomePayForm(user=request.user)

    return render(request, 'middle_income_templates/income_pay.html', {
        'income': income,
        'form': form
    })


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def income_cancel_pay_view(request, income_id):
    income = get_object_or_404(
        Income,
        id=income_id,
        is_paid=True,
        is_active=True,
        user=request.user
    )

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # پیدا کردن Fund مربوطه
                fund = Fund.objects.filter(
                    content_type__model='income',
                    object_id=income.id,
                    user=request.user,
                    is_paid=True
                ).first()

                if not fund:
                    messages.error(request, 'Fund مرتبط با این پرداخت پیدا نشد!')
                    return redirect(request.META.get('HTTP_REFERER'))

                # حذف Fund
                fund.delete()
                bank_fund = BankFund.objects.filter(
                    content_type__model='income',
                    object_id=income.id
                ).first()
                print(f'bank_fund: {bank_fund}')

                if bank_fund:
                    bank = bank_fund.bank

                    # اصلاح موجودی بانک
                    if bank_fund.transaction_type == 'deposit':
                        bank.current_balance -= bank_fund.amount
                    else:
                        bank.current_balance += bank_fund.amount

                    bank.save(update_fields=['current_balance'])

                    bank_fund.delete()

                # بازمحاسبه موجودی صندوق از این Fund به بعد
                Fund.recalc_final_amounts_from(fund)

                # بازگرداندن Expense به حالت پرداخت‌نشده
                income.is_paid = False
                income.bank = None
                income.transaction_reference = None
                income.payment_date = None
                income.payer_name = None
                income.save(update_fields=[
                    'is_paid',
                    'bank',
                    'transaction_reference',
                    'payment_date',
                    'payer_name',
                ])

                messages.success(request, 'دریافت با موفقیت لغو شد و صندوق و بانک اصلاح شد.')
                return redirect(request.META.get('HTTP_REFERER'))

        except Exception as e:
            messages.error(request, f'خطا در لغو دریافت: {e}')
            return redirect(request.META.get('HTTP_REFERER'))

    # اگر GET باشد، فقط برگرد به صفحه قبل
    return redirect(request.META.get('HTTP_REFERER'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_income_edit(request, pk):
    income = get_object_or_404(Income, pk=pk)

    if income.is_paid:
        messages.warning(request, 'این درآمد پرداخت شده است و قابل ویرایش نیست.')
        return redirect('middle_register_income')

    if request.method != 'POST':
        return redirect('middle_register_income')

    form = IncomeForm(request.POST, request.FILES, instance=income, user=request.user)

    if not form.is_valid():
        messages.error(request, 'خطا در ویرایش فرم درآمد. لطفا دوباره تلاش کنید.')
        return redirect('middle_register_income')

    try:
        with transaction.atomic():
            files = request.FILES.getlist('document')
            if len(files) > 2:
                messages.error(request,
                               "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
                return redirect('middle_register_income')
            house = MyHouse.objects.filter(user=request.user, is_active=True).first()
            form.instance.house = house
            income = form.save()

            # 🔹 ذخیره فایل‌های جدید بدون حذف قبلی
            # files = request.FILES.getlist('document')
            for f in files:
                IncomeDocument.objects.create(income=income, document=f)

        messages.success(request, 'درآمد با موفقیت ویرایش شد.')
        return redirect('middle_register_income')

    except Exception as e:
        messages.error(request, 'خطا در ویرایش درآمد.')
        return redirect('middle_register_income')


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_income_delete(request, pk):
    income = get_object_or_404(Income, id=pk)

    if income.is_paid:
        messages.warning(request, 'بدلیل ثبت پرداخت، حذف امکان پذیر نیست!')
        return redirect(reverse('middle_register_income'))

    try:
        with transaction.atomic():
            # حذف Fund مربوطه
            income_ct = ContentType.objects.get_for_model(Income)
            Fund.objects.filter(content_type=income_ct, object_id=income.id).delete()

            # حذف خود Expense
            income.delete()

        messages.success(request, 'درآمد با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, "امکان حذف وجود ندارد!")
    except Exception as e:
        messages.error(request, f"خطا در حذف: {str(e)}")

    return redirect(reverse('middle_register_income'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
@csrf_exempt
def middle_delete_income_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        income_id = request.POST.get('income_id')

        if not image_url or not income_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            income = get_object_or_404(Income, id=income_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = IncomeDocument.objects.filter(income=income, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_income_pdf(request):
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()
    incomes = Income.objects.filter(user=request.user)

    filter_fields = {
        'category': 'category__id',
        'bank': 'bank__id',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
        'is_paid': 'is_paid'
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            incomes = incomes.filter(**filter_expression)

    # فیلتر تاریخ
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__lte=to_date)
    except ValueError:
        incomes = Income.objects.none()

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # رندر قالب HTML
    template = get_template("middle_income_templates/income_pdf.html")
    context = {
        'incomes': incomes,
        'font_path': font_url,
        'house': house,
        'today': timezone.now()
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="incomes.pdf"'

    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_income_excel(request):
    incomes = Income.objects.filter(user=request.user)

    # Filter fields
    filter_fields = {
        'category': 'category__id',
        'bank': 'bank__id',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
        'is_paid': 'is_paid'
    }
    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            incomes = incomes.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__lte=to_date)
    except ValueError:
        expenses = Expense.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "incomes"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست درآمدها")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'موضوع درآمد', 'شرح سند', ' شماره سند', 'مبلغ', 'تاریخ سند', 'توضیحات', 'پرداخت کننده',
               'تاریخ پرداخت']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, income in enumerate(incomes, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        ws.cell(row=row_num, column=2, value=income.category.subject)
        ws.cell(row=row_num, column=3, value=income.description)
        ws.cell(row=row_num, column=4, value=income.doc_number)
        ws.cell(row=row_num, column=5, value=income.amount)
        ws.cell(row=row_num, column=6, value=show_jalali(income.doc_date))
        ws.cell(row=row_num, column=7, value=income.details)
        ws.cell(row=row_num, column=8, value=income.payer_name)
        ws.cell(row=row_num, column=9, value=show_jalali(income.payment_date))

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=incomes.xlsx'
    wb.save(response)
    return response


# ============================ ReceiveMoneyView ==========================
@method_decorator(middle_admin_required, name='dispatch')
class MiddleReceiveMoneyCreateView(CreateView):
    model = ReceiveMoney
    form_class = ReceiveMoneyForm
    template_name = 'MiddleReceiveMoney/add_receive_money.html'
    success_url = reverse_lazy('middle_register_receive')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            files = self.request.FILES.getlist('document')
            if len(files) > 2:
                messages.error(self.request,
                               "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
                return redirect('middle_register_receive')
            # 🔹 مشخص کردن خانه کاربر
            house = MyHouse.objects.filter(user=self.request.user, is_active=True).first()
            form.instance.house = house  # ذخیره خانه در ReceiveMoney

            self.object = form.save(commit=False)
            # self.object.payer_name = self.object.get_payer_display()
            # self.object.is_received_money = True
            self.object.save()
            form.save_m2m()

            # content_type = ContentType.objects.get_for_model(self.object)
            # # payer_name_for_fund = self.object.payer_name
            #
            #
            # # 🔹 ذخیره سند در Fund با خانه
            # Fund.objects.create(
            #     user=self.request.user,
            #     content_type=content_type,
            #     object_id=self.object.id,
            #     bank=self.object.bank,
            #     unit=self.object.unit,
            #     house=house,  # ← اضافه شد
            #     amount=self.object.amount or 0,
            #     debtor_amount=self.object.amount or 0,
            #     creditor_amount=0,
            #     doc_number=self.object.doc_number,
            #     payer_name=self.object.payer_name,
            #     payment_gateway='کارت به کارت',
            #     transaction_no=self.object.transaction_reference,
            #     payment_date=self.object.payment_date,
            #     payment_description=f"حسابهای دریافتنی: {self.object.description[:50]}",
            #     is_paid=True,
            #     is_received_money=True
            # )
            # BankTransactionService.deposit(
            #     user=self.request.user,
            #     bank=self.object.bank,
            #     unit=self.object.unit,
            #     amount=Decimal(self.object.amount),
            #     description=f"حسابهای دریافتنی: {self.object.description[:50]}",
            #     content_object=self.object,
            #     transaction_no=self.object.transaction_reference,
            #     payment_date=self.object.payment_date,
            #     gateway='کارت به کارت',
            #     payer_name=self.object.payer_name,
            #     house=house
            # )

            # files = self.request.FILES.getlist('document')
            for f in files:
                ReceiveDocument.objects.create(receive=self.object, document=f)

            messages.success(self.request, 'سند دریافت با موفقیت ثبت گردید!')
            return super().form_valid(form)

        except Exception as e:
            messages.error(self.request, f'خطا در ثبت! {e}')
            return self.form_invalid(form)

    def get_queryset(self):
        queryset = ReceiveMoney.objects.filter(user=self.request.user).order_by('-created_at')

        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank__id=bank_id)

        # فیلتر بر اساس amount
        amount = self.request.GET.get('amount')
        if amount:
            queryset = queryset.filter(amount__icontains=amount)

        payer_name = self.request.GET.get('payer_name')
        if payer_name:
            queryset = queryset.filter(payer_name__icontains=payer_name)

        description = self.request.GET.get('description')
        if description:
            queryset = queryset.filter(description__icontains=description)

        doc_number = self.request.GET.get('doc_number')
        if doc_number:
            queryset = queryset.filter(doc_number__icontains=doc_number)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس date
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')
        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y-%m-%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(doc_date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y-%m-%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(doc_date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')

        is_paid = self.request.GET.get('is_paid')
        if is_paid == '1':
            queryset = queryset.filter(is_paid=True)
        elif is_paid == '0':
            queryset = queryset.filter(is_paid=False)
        return queryset

    # def get_form_kwargs(self):
    #     kwargs = super().get_form_kwargs()
    #     kwargs['user'] = self.request.user
    #     return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receives = self.get_queryset()
        paginator = Paginator(receives, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_receives'] = ReceiveMoney.objects.filter(user=self.request.user).count()
        context['receives'] = ReceiveMoney.objects.filter(user=self.request.user)
        context['banks'] = Bank.objects.filter(user=self.request.user)
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def receive_pay_view(request, receive_id):
    receive = get_object_or_404(
        ReceiveMoney,
        id=receive_id,
        is_paid=False,
        is_active=True
    )

    if request.method == 'POST':
        form = ReceivePayForm(
            request.POST,
            user=request.user,
            instance=receive
        )

        if form.is_valid():
            try:
                with transaction.atomic():

                    receive = ReceiveMoney.objects.select_for_update().get(
                        id=receive_id,
                        is_paid=False,
                        is_active=True
                    )

                    bank = form.cleaned_data['bank']
                    reference = form.cleaned_data.get('transaction_reference')
                    payment_date = form.cleaned_data.get('payment_date')
                    payer_name = form.cleaned_data.get('payer_name')
                    unit = form.cleaned_data['unit']

                    payer = unit.get_label if unit else payer_name

                    content_type = ContentType.objects.get_for_model(receive)

                    Fund.objects.create(
                        user=request.user,
                        content_type=content_type,
                        object_id=receive.id,
                        bank=bank,
                        unit=unit,
                        house=receive.house,
                        amount=receive.amount or 0,
                        debtor_amount=receive.amount or 0,
                        creditor_amount=0,
                        doc_number=receive.doc_number,
                        payer_name=payer,
                        payment_gateway='کارت به کارت',
                        transaction_no=reference,
                        payment_date=payment_date,
                        payment_description=f"حسابهای دریافتنی: {receive.description[:50]}",
                        is_paid=True,
                        is_received_money=True
                    )

                    BankTransactionService.deposit(
                        user=request.user,
                        bank=bank,
                        unit=unit,
                        amount=Decimal(receive.amount),
                        description=f'اسناد دریافتنی: {receive.description}',
                        content_object=receive,
                        payment_date=payment_date,
                        transaction_no=reference,
                        gateway='کارت به کارت',
                        payer_name=payer,
                        house=receive.house
                    )

                    receive.is_paid = True
                    receive.bank = bank
                    receive.transaction_reference = reference
                    receive.payment_date = payment_date
                    receive.unit = unit
                    receive.payer_name = payer

                    receive.save(update_fields=[
                        'is_paid',
                        'bank',
                        'transaction_reference',
                        'payment_date',
                        'unit',
                        'payer_name'
                    ])

                messages.success(request, 'دریافت با موفقیت انجام شد')
                return redirect('middle_register_receive')

            except ValidationError as e:
                messages.error(request, e.message)

            except Exception as e:
                messages.error(request, f'خطا در دریافت: {e}')

    else:
        form = ReceivePayForm(
            user=request.user,
            instance=receive
        )

    return render(
        request,
        'MiddleReceiveMoney/receive_pay.html',
        {
            'receive': receive,
            'form': form
        }
    )


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def receive_cancel_pay_view(request, receive_id):
    receive = get_object_or_404(
        ReceiveMoney,
        id=receive_id,
        is_paid=True,
        is_active=True,
        user=request.user
    )

    if request.method == 'POST':

        try:
            with transaction.atomic():

                content_type = ContentType.objects.get_for_model(receive)

                fund = Fund.objects.filter(
                    content_type=content_type,
                    object_id=receive.id,
                    user=request.user,
                    is_paid=True
                ).first()

                if not fund:
                    messages.error(request, 'Fund مرتبط با این پرداخت پیدا نشد!')
                    return redirect(request.META.get('HTTP_REFERER'))

                bank_fund = BankFund.objects.filter(
                    content_type=content_type,
                    object_id=receive.id
                ).first()

                # اصلاح موجودی بانک
                if bank_fund:

                    bank = bank_fund.bank

                    if bank_fund.transaction_type == 'deposit':
                        bank.current_balance -= bank_fund.amount

                    elif bank_fund.transaction_type == 'withdraw':
                        bank.current_balance += bank_fund.amount

                    bank.save(update_fields=['current_balance'])

                    bank_fund.delete()

                # حذف Fund
                fund.delete()

                # باز محاسبه صندوق
                Fund.recalc_final_amounts_from(fund)

                # بازگرداندن وضعیت دریافت
                receive.is_paid = False
                receive.bank = None
                receive.transaction_reference = None
                receive.payment_date = None
                receive.payer_name = ''

                receive.save(update_fields=[
                    'is_paid',
                    'bank',
                    'transaction_reference',
                    'payment_date',
                    'payer_name',
                ])

                messages.success(
                    request,
                    'دریافت با موفقیت لغو شد و صندوق و بانک اصلاح شد.'
                )

                return redirect(request.META.get('HTTP_REFERER'))

        except Exception as e:

            messages.error(
                request,
                f'خطا در لغو دریافت: {e}'
            )

            return redirect(request.META.get('HTTP_REFERER'))

    return redirect(request.META.get('HTTP_REFERER'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_receive_edit(request, pk):
    receive = get_object_or_404(
        ReceiveMoney,
        pk=pk,
        user=request.user
    )
    if receive.is_paid:
        messages.warning(request, 'این سند بدلیل ثبت رکورد پرداخت قابل ویرایش نیست')
        return redirect('middle_register_receive')

    if request.method == 'POST':
        form = ReceiveMoneyForm(
            request.POST,
            request.FILES,
            instance=receive,
        )

        if form.is_valid():
            try:
                with transaction.atomic():
                    files = request.FILES.getlist('document')
                    if len(files) > 2:
                        messages.error(request,
                                       "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
                        return redirect('middle_register_receive')
                    receive = form.save(commit=False)
                    # receive.payer_name = receive.get_payer_display()
                    receive.is_received_money = True
                    receive.save()
                    form.save_m2m()

                    # 📎 ذخیره فایل‌ها
                    # files = request.FILES.getlist('document')
                    for f in files:
                        ReceiveDocument.objects.create(
                            receive=receive,
                            document=f
                        )

                messages.success(request, 'سند با موفقیت ویرایش گردید.')
                return redirect('middle_register_receive')

            except Exception as e:
                messages.error(request, f'خطا در ذخیره اطلاعات: {e}')

        else:
            messages.error(request, 'خطا در ویرایش فرم. لطفاً دوباره تلاش کنید.')

    else:
        form = ReceiveMoneyForm(instance=receive)

    return render(
        request,
        'MiddleReceiveMoney/add_receive_money.html',
        {
            'form': form,
            'receive': receive,
            'open_modal': True,  # ← پرچم برای JS که modal را باز کند
        }
    )


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_receive_delete(request, pk):
    receive = get_object_or_404(ReceiveMoney, id=pk)
    if receive.is_paid:
        messages.warning(request, 'بدلیل ثبت پرداخت، حذف امکان پذیر نیست!')
        return redirect(reverse('middle_register_receive'))
    try:
        with transaction.atomic():
            # حذف Fund مربوطه
            receive_ct = ContentType.objects.get_for_model(ReceiveMoney)
            Fund.objects.filter(content_type=receive_ct, object_id=receive.id).delete()

            # حذف خود Expense
            receive.delete()
            messages.success(request, ' سند با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('middle_register_receive'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
@csrf_exempt
def middle_delete_receive_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        receive_id = request.POST.get('receive_id')

        if not image_url or not receive_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            receive = get_object_or_404(ReceiveMoney, id=receive_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = ReceiveDocument.objects.filter(receive=receive, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_receive_pdf(request):
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()
    receives = ReceiveMoney.objects.select_related('bank').filter(user=request.user)

    filter_fields = {
        'bank': 'bank__id',
        'payer_name': 'payer_name__icontains',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            receives = receives.filter(**filter_expression)

    # فیلتر تاریخ
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            receives = receives.filter(doc_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            receives = receives.filter(doc_date__lte=to_date)
    except ValueError:
        receives = ReceiveMoney.objects.none()

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # رندر قالب HTML
    template = get_template("MiddleReceiveMoney/receive_pdf.html")
    context = {
        'receives': receives,
        'font_path': font_url,
        'house': house,
        'today': timezone.now()
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    # response['Content-Disposition'] = f'attachment; filename="receives.pdf"'
    response['Content-Disposition'] = (
        f'attachment; filename="receives_{int(time.time())}.pdf"'
    )

    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def export_receive_excel(request):
    receives = ReceiveMoney.objects.filter(user=request.user)

    filter_fields = {
        'bank': 'bank__id',
        'payer_name': 'payer_name__icontains',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
    }

    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            receives = receives.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            receives = receives.filter(doc_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            receives = receives.filter(doc_date__lte=to_date)
    except ValueError:
        receives = ReceiveMoney.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "receives"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست اسناد دریافتنی")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'شماره جساب', 'دریافت کننده', 'شرح سند', ' شماره سند', 'مبلغ', 'تاریخ سند', 'توضیحات',
               'تاریخ پرداخت', 'شماره پیگیری']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, receive in enumerate(receives, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        bank_account = ""
        if receive.bank and receive.bank.account_no:
            bank_account = f"{receive.bank.bank_name} - {receive.bank.account_no}"
        payer_name = (
            str(receive.unit)
            if receive.unit
            else receive.payer_name
        )
        ws.cell(row=row_num, column=2, value=bank_account)
        ws.cell(row=row_num, column=3, value=payer_name)
        ws.cell(row=row_num, column=4, value=receive.description)
        ws.cell(row=row_num, column=5, value=receive.doc_number)
        ws.cell(row=row_num, column=6, value=receive.amount)
        ws.cell(row=row_num, column=7, value=show_jalali(receive.doc_date))
        ws.cell(row=row_num, column=8, value=receive.details)
        ws.cell(row=row_num, column=9, value=show_jalali(receive.payment_date))
        ws.cell(row=row_num, column=10, value=receive.transaction_reference)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=receive_doc.xlsx'
    wb.save(response)
    return response


# ============================ PaymentMoneyView ==========================
@method_decorator(middle_admin_required, name='dispatch')
class MiddlePaymentMoneyCreateView(CreateView):
    model = PayMoney
    form_class = PayerMoneyForm
    template_name = 'MiddlePayMoney/add_pay_money.html'
    success_url = reverse_lazy('middle_register_pay')

    def form_valid(self, form):
        files = self.request.FILES.getlist('document')
        if len(files) > 2:
            messages.error(self.request,
                           "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
            return redirect('middle_register_pay')
        form.instance.user = self.request.user
        house = MyHouse.objects.filter(user=self.request.user, is_active=True).first()
        form.instance.house = house

        with transaction.atomic():

            self.object = form.save(commit=False)

            # bank = self.object.bank
            # amount = self.object.amount
            #
            # bank_funds = Fund.objects.filter(
            #     user=self.request.user,
            #     bank=bank
            # )
            #
            # total_debit = bank_funds.aggregate(
            #     Sum('debtor_amount')
            # )['debtor_amount__sum'] or 0
            #
            # total_credit = bank_funds.aggregate(
            #     Sum('creditor_amount')
            # )['creditor_amount__sum'] or 0
            #
            # current_final = Decimal(total_debit) - Decimal(total_credit)
            #
            # if current_final < amount:
            #     messages.error(self.request, 'موجودی صندوق کافی نیست')
            #     return self.form_invalid(form)

            # self.object.receiver_name = self.object.get_receiver_display
            # self.object.is_paid_money = True
            self.object.save()
            form.save_m2m()

            # content_type = ContentType.objects.get_for_model(self.object)
            #
            # receiver_name_for_fund = (
            #     self.object.receiver_name
            #     if not self.object.unit else f"{self.object.unit}"
            # )
            #
            # Fund.objects.create(
            #     user=self.request.user,
            #     content_type=content_type,
            #     object_id=self.object.id,
            #     bank=bank,
            #     house=house,
            #     unit=self.object.unit,
            #     amount=self.object.amount,
            #     debtor_amount=0,
            #     receiver_name=receiver_name_for_fund,
            #     creditor_amount=self.object.amount,
            #     payment_gateway='کارت به کارت',
            #     payment_date=self.object.payment_date,
            #     doc_number=self.object.document_number,
            #     payment_description=f"حسابهای پرداختنی: {self.object.description[:50]}",
            #     is_paid=True,
            #     is_paid_money=True
            # )

            # files = self.request.FILES.getlist('document')
            for f in files:
                PayDocument.objects.create(payment=self.object, document=f)

            messages.success(self.request, 'سند پرداخت با موفقیت ثبت گردید!')
            return super().form_valid(form)

        # except Exception as e:
        #     messages.error(self.request, f'خطا در ثبت: {e}')
        #     return self.form_invalid(form)

    def get_queryset(self):
        queryset = PayMoney.objects.filter(user=self.request.user).order_by('-created_at')

        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank__id=bank_id)

        # فیلتر بر اساس amount
        amount = self.request.GET.get('amount')
        if amount:
            queryset = queryset.filter(amount__icontains=amount)

        receiver_name = self.request.GET.get('payer_name')
        if receiver_name:
            queryset = queryset.filter(receiver_name__icontains=receiver_name)

        description = self.request.GET.get('description')
        if description:
            queryset = queryset.filter(description__icontains=description)

        document_number = self.request.GET.get('document_number')
        if document_number:
            queryset = queryset.filter(document_number__icontains=document_number)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس date
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')
        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y-%m-%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(document_date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y-%m-%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(document_date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')

        is_paid = self.request.GET.get('is_paid')
        if is_paid == '1':
            queryset = queryset.filter(is_paid=True)
        elif is_paid == '0':
            queryset = queryset.filter(is_paid=False)
        return queryset

    # def get_form_kwargs(self):
    #     kwargs = super().get_form_kwargs()
    #     kwargs['user'] = self.request.user
    #     return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receives = self.get_queryset()
        paginator = Paginator(receives, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_payments'] = PayMoney.objects.filter(user=self.request.user).count()
        context['payments'] = PayMoney.objects.filter(user=self.request.user)
        context['banks'] = Bank.objects.filter(user=self.request.user)
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_pay_edit(request, pk):
    # گرفتن رکورد پرداخت موجود
    payment = get_object_or_404(PayMoney, pk=pk)

    if payment.is_paid:
        messages.warning(request, 'این سند بدلیل ثبت رکورد پرداخت قابل ویرایش نیست')
        return redirect('middle_register_pay')

    if request.method == 'POST':
        # فرم با instance برای ویرایش
        form = PayerMoneyForm(request.POST, request.FILES, instance=payment)

        if form.is_valid():
            files = request.FILES.getlist('document')
            if len(files) > 2:
                messages.error(request,
                               "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
                return redirect('middle_register_pay')
            payment = form.save(commit=False)
            # payment.receiver_name = payment.get_receiver_display
            # payment.is_paid_money = True
            payment.save()
            form.save_m2m()

            # if payment.unit:
            #     receiver_name_for_fund = str(payment.unit)
            # else:
            #     receiver_name_for_fund = payment.receiver_name
            #
            # content_type = ContentType.objects.get_for_model(PayMoney)
            # fund = Fund.objects.filter(content_type=content_type, object_id=payment.id).first()
            #
            # if fund:
            #     # بروزرسانی رکورد موجود
            #     fund.bank = payment.bank
            #     fund.unit = payment.unit
            #     fund.debtor_amount = 0
            #     fund.amount = payment.amount or 0
            #     fund.creditor_amount = payment.amount or 0
            #     fund.payment_date = payment.document_date
            #     fund.doc_number = payment.document_number
            #     fund.receiver_name = receiver_name_for_fund
            #     fund.payment_gateway = 'کارت به کارت'
            #     fund.payment_description = f"حسابهای پرداختنی: {(payment.description or '')[:50]}"
            #     fund.is_paid_money = True
            #     fund.is_paid = True
            #     fund.save()  # موجودی بانک بروزرسانی می‌شود
            #     Fund.recalc_final_amounts_from(fund)
            #
            #
            # else:
            #     # ایجاد فقط اگر رکورد موجود نبود
            #     Fund.objects.create(
            #         content_type=content_type,
            #         object_id=payment.id,
            #         bank=payment.bank,
            #         unit=payment.unit,
            #         debtor_amount=0,
            #         amount=payment.amount or 0,
            #         creditor_amount=payment.amount or 0,
            #         user=request.user,
            #         receiver_name=receiver_name_for_fund,
            #         payment_date=payment.document_date,
            #         doc_number=payment.document_number,
            #         payment_gateway='کارت به کارت',
            #         payment_description=f"حسابهای پرداختنی: {(payment.description or '')[:50]}",
            #         is_paid=True,
            #         is_paid_money=True
            #     )

            # ثبت فایل‌های پیوست جدید
            # files = request.FILES.getlist('document')
            for f in files:
                PayDocument.objects.create(payment=payment, document=f)

            messages.success(request, 'سند با موفقیت ویرایش گردید.')
            return redirect(reverse('middle_register_pay'))  # Adjust redirect as necessary

        else:
            messages.error(request, 'خطا در ویرایش فرم . لطفا دوباره تلاش کنید.')
            return render(request, 'MiddlePayMoney/add_pay_money.html', {
                'form': form,
                'payment': payment,
                'open_modal': True,
            })
    else:
        form = PayerMoneyForm(instance=payment)
        return render(request, 'MiddlePayMoney/add_pay_money.html',
                      {'form': form,
                       'payment': payment,
                       'open_modal': True,
                       })


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def payment_pay_money_view(request, payment_id):
    payment = get_object_or_404(
        PayMoney,
        id=payment_id,
        is_paid=False,
        is_active=True
    )

    if request.method == 'POST':
        form = PayMoneyForm(
            request.POST,
            user=request.user,
            instance=payment
        )

        if form.is_valid():
            try:
                with transaction.atomic():

                    payment = PayMoney.objects.select_for_update().get(
                        id=payment_id,
                        is_paid=False,
                        is_active=True
                    )

                    bank = form.cleaned_data['bank']
                    reference = form.cleaned_data.get('transaction_reference')
                    payment_date = form.cleaned_data.get('payment_date')
                    receiver_name = form.cleaned_data.get('receiver_name')
                    unit = form.cleaned_data['unit']

                    current_balance = bank.current_balance

                    if current_balance < payment.amount:
                        messages.error(
                            request,
                            f'موجودی بانک کافی نیست. موجودی فعلی: {current_balance:,} تومان'
                        )
                        return redirect(request.META.get('HTTP_REFERER'))

                    receiver = unit.get_label if unit else receiver_name

                    content_type = ContentType.objects.get_for_model(payment)

                    Fund.objects.create(
                        user=request.user,
                        content_type=content_type,
                        object_id=payment.id,
                        bank=bank,
                        unit=unit,
                        house=payment.house,
                        amount=payment.amount or 0,
                        debtor_amount=0,
                        creditor_amount=payment.amount or 0,
                        doc_number=payment.document_number,
                        receiver_name=receiver,
                        payment_gateway='کارت به کارت',
                        transaction_no=reference,
                        payment_date=payment_date,
                        payment_description=f"حسابهای پرداختنی: {payment.description[:50]}",
                        is_paid=True,
                        is_paid_money=True
                    )

                    BankTransactionService.withdraw(
                        user=request.user,
                        bank=bank,
                        unit=unit,
                        amount=Decimal(payment.amount),
                        description=f'اسناد پرداختنی: {payment.description}',
                        content_object=payment,
                        payment_date=payment_date,
                        transaction_no=reference,
                        gateway='کارت به کارت',
                        receiver_name=receiver,
                        house=payment.house
                    )

                    payment.is_paid = True
                    payment.bank = bank
                    payment.transaction_reference = reference
                    payment.payment_date = payment_date
                    payment.unit = unit
                    payment.receiver_name = receiver

                    payment.save(update_fields=[
                        'is_paid',
                        'bank',
                        'transaction_reference',
                        'payment_date',
                        'unit',
                        'receiver_name'
                    ])

                messages.success(request, 'پرداخت با موفقیت انجام شد')
                return redirect('middle_register_pay')

            except ValidationError as e:
                messages.error(request, e.message)

            except Exception as e:
                messages.error(request, f'خطا در پرداخت: {e}')

    else:
        form = PayMoneyForm(
            user=request.user,
            instance=payment
        )

    return render(
        request,
        'MiddlePayMoney/payment_pay_money.html',
        {
            'payment': payment,
            'form': form
        }
    )


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def payment_cancel_pay_money_view(request, payment_id):
    payment = get_object_or_404(
        PayMoney,
        id=payment_id,
        is_paid=True,
        is_active=True,
        user=request.user
    )

    if request.method == 'POST':

        try:
            with transaction.atomic():

                content_type = ContentType.objects.get_for_model(payment)

                fund = Fund.objects.filter(
                    content_type=content_type,
                    object_id=payment.id,
                    user=request.user,
                    is_paid=True
                ).first()

                if not fund:
                    messages.error(request, 'Fund مرتبط با این پرداخت پیدا نشد!')
                    return redirect(request.META.get('HTTP_REFERER'))

                bank_fund = BankFund.objects.filter(
                    content_type=content_type,
                    object_id=payment.id
                ).first()

                # اصلاح موجودی بانک
                if bank_fund:

                    bank = bank_fund.bank

                    if bank_fund.transaction_type == 'deposit':
                        bank.current_balance -= bank_fund.amount

                    elif bank_fund.transaction_type == 'withdraw':
                        bank.current_balance += bank_fund.amount

                    bank.save(update_fields=['current_balance'])

                    bank_fund.delete()

                # حذف Fund
                fund.delete()

                # باز محاسبه صندوق
                Fund.recalc_final_amounts_from(fund)

                # بازگرداندن وضعیت دریافت
                payment.is_paid = False
                payment.bank = None
                payment.transaction_reference = None
                payment.payment_date = None
                payment.receiver_name = ''

                payment.save(update_fields=[
                    'is_paid',
                    'bank',
                    'transaction_reference',
                    'payment_date',
                    'receiver_name',
                ])

                messages.success(
                    request,
                    'پرداخت با موفقیت لغو شد و صندوق و بانک اصلاح شد.'
                )

                return redirect(request.META.get('HTTP_REFERER'))

        except Exception as e:

            messages.error(
                request,
                f'خطا در لغو دریافت: {e}'
            )

            return redirect(request.META.get('HTTP_REFERER'))

    return redirect(request.META.get('HTTP_REFERER'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_pay_delete(request, pk):
    payment = get_object_or_404(PayMoney, id=pk)
    if payment.is_paid:
        messages.warning(request, 'بدلیل ثبت پرداخت، حذف امکان پذیر نیست!')
        return redirect(reverse('middle_register_pay'))
    try:
        with transaction.atomic():
            # حذف Fund مربوطه
            payment_ct = ContentType.objects.get_for_model(PayMoney)
            Fund.objects.filter(content_type=payment_ct, object_id=payment.id).delete()

            # حذف خود Expense
            payment.delete()
        messages.success(request, ' سند با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('middle_register_pay'))


@login_required(login_url=settings.LOGIN_URL_ADMIN)
@csrf_exempt
def middle_delete_pay_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        payment_id = request.POST.get('payment_id')

        if not image_url or not payment_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            payment = get_object_or_404(PayMoney, id=payment_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = PayDocument.objects.filter(payment=payment, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_pay_pdf(request):
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()
    payments = PayMoney.objects.filter(user=request.user)

    filter_fields = {
        'bank': 'bank__id',
        'receiver_name': 'payer_name',
        'amount': 'amount__icontains',
        'document_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
        'is_paid': 'is_paid'
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            payments = payments.filter(**filter_expression)

    # فیلتر تاریخ
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            payments = payments.filter(document_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            payments = payments.filter(document_date__lte=to_date)
    except ValueError:
        payments = PayMoney.objects.none()

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # رندر قالب HTML
    template = get_template("MiddlePayMoney/pay_pdf.html")
    context = {
        'payments': payments,
        'font_path': font_url,
        'house': house,
        'today': timezone.now()
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="payments.pdf"'

    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_pay_excel(request):
    payments = PayMoney.objects.filter(user=request.user)

    filter_fields = {
        'bank': 'bank__id',
        'receiver_name': 'receiver_name__icontains',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
        'is_paid': 'is_paid'
    }

    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            payments = payments.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            payments = payments.filter(document_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            payments = payments.filter(document_date__lte=to_date)
    except ValueError:
        payments = PayMoney.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "payments"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست اسناد پرداختنی")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'شماره جساب', 'دریافت کننده', 'شرح سند', ' شماره سند', 'مبلغ', 'تاریخ سند', 'توضیحات',
               'تاریخ پرداخت', 'شماره پیگیری']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, payment in enumerate(payments, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        bank_account = ""
        if payment.bank and payment.bank.account_no:
            bank_account = f"{payment.bank.bank_name} - {payment.bank.account_no}"

        receiver_name = (
            str(payment.unit)
            if payment.unit
            else payment.receiver_name
        )

        ws.cell(row=row_num, column=2, value=bank_account)
        ws.cell(row=row_num, column=3, value=receiver_name)
        ws.cell(row=row_num, column=4, value=payment.description)
        ws.cell(row=row_num, column=5, value=payment.document_number)
        ws.cell(row=row_num, column=6, value=payment.amount)
        ws.cell(row=row_num, column=7, value=show_jalali(payment.document_date))
        ws.cell(row=row_num, column=8, value=payment.details)
        ws.cell(row=row_num, column=9, value=show_jalali(payment.payment_date))
        ws.cell(row=row_num, column=10, value=payment.transaction_reference)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=payment.xlsx'
    wb.save(response)
    return response


# ============================ PropertyView ==========================
@method_decorator(middle_admin_required, name='dispatch')
class MiddlePropertyCreateView(CreateView):
    model = Property
    template_name = 'middleProperty/manage_property.html'
    form_class = PropertyForm
    success_url = reverse_lazy('middle_register_property')

    def form_valid(self, form):
        form.instance.user = self.request.user
        house = MyHouse.objects.filter(user=self.request.user, is_active=True).first()
        form.instance.house = house

        with transaction.atomic():
            files = self.request.FILES.getlist('document')
            if len(files) > 2:
                messages.error(self.request,
                               "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
                return redirect('middle_register_property')
            self.object = form.save(commit=False)
            self.object.save()
            form.save_m2m()
            #
            # content_type = ContentType.objects.get_for_model(self.object)
            #
            # Fund.objects.create(
            #     user=self.request.user,
            #     content_type=content_type,
            #     object_id=self.object.id,
            #     bank=bank,
            #     house=house,
            #     unit=None,
            #     amount=self.object.property_price,
            #     debtor_amount=0,
            #     receiver_name=self.object.receiver_name,
            #     creditor_amount=self.object.property_price,
            #     payment_gateway='کارت به کارت',
            #     payment_date=self.object.payment_date,
            #     doc_number=self.object.document_number,
            #     payment_description=f"خرید اموال: {self.object.property_name[:50]}",
            #     is_paid=True,
            #
            # )

            # files = self.request.FILES.getlist('document')
            for f in files:
                PropertyDocument.objects.create(property=self.object, document=f)

            messages.success(self.request, 'سند با موفقیت ثبت گردید!')
            return super().form_valid(form)

    def get_queryset(self):
        queryset = Property.objects.filter(user=self.request.user).order_by('-id')

        # فیلتر بر اساس amount
        property_name = self.request.GET.get('property_name')
        if property_name:
            queryset = queryset.filter(property_name__icontains=property_name)

        property_unit = self.request.GET.get('property_unit')
        if property_unit:
            queryset = queryset.filter(property_unit__icontains=property_unit)

        property_location = self.request.GET.get('property_location')
        if property_location:
            queryset = queryset.filter(property_location__icontains=property_location)

        property_code = self.request.GET.get('property_code')
        if property_code:
            queryset = queryset.filter(property_code__icontains=property_code)

        property_price = self.request.GET.get('property_price')
        if property_price:
            queryset = queryset.filter(property_price__icontains=property_price)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس date
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')

        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y-%m-%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(property_purchase_date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y-%m-%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(property_purchase_date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receives = self.get_queryset()
        paginator = Paginator(receives, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_properties'] = Property.objects.filter(user=self.request.user).count()
        context['properties'] = Property.objects.filter(user=self.request.user)
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def property_pay_money_view(request, property_id):
    prop = get_object_or_404(
        Property,
        id=property_id,
        is_paid=False,
        is_active=True
    )

    if request.method == 'POST':
        form = PropertyPayForm(
            request.POST,
            user=request.user,
            instance=prop
        )

        if form.is_valid():
            try:
                with transaction.atomic():

                    prop = Property.objects.select_for_update().get(
                        id=property_id,
                        is_paid=False,
                        is_active=True
                    )

                    bank = form.cleaned_data['bank']
                    reference = form.cleaned_data.get('transaction_reference')
                    payment_date = form.cleaned_data.get('payment_date')
                    receiver_name = form.cleaned_data.get('receiver_name')

                    current_balance = bank.current_balance

                    if current_balance < prop.property_price:
                        messages.error(
                            request,
                            f'موجودی بانک کافی نیست. موجودی فعلی: {current_balance:,} تومان'
                        )
                        return redirect(request.META.get('HTTP_REFERER'))

                    content_type = ContentType.objects.get_for_model(prop)

                    Fund.objects.create(
                        user=request.user,
                        content_type=content_type,
                        object_id=prop.id,
                        bank=bank,
                        house=prop.house,
                        unit=None,
                        amount=prop.property_price,
                        debtor_amount=0,
                        receiver_name=receiver_name,
                        creditor_amount=prop.property_price,
                        payment_gateway='کارت به کارت',
                        payment_date=payment_date,
                        transaction_no=reference,
                        doc_number=prop.document_number,
                        payment_description=f"خرید اموال: {prop.property_name[:50]}",
                        is_paid=True,

                    )

                    BankTransactionService.withdraw(
                        user=request.user,
                        bank=bank,
                        unit=None,
                        amount=Decimal(prop.property_price),
                        description=f"خرید اموال: {prop.property_name[:50]}",
                        content_object=prop,
                        payment_date=payment_date,
                        transaction_no=reference,
                        gateway='کارت به کارت',
                        receiver_name=receiver_name,
                        house=prop.house
                    )

                    prop.is_paid = True
                    prop.bank = bank
                    prop.transaction_reference = reference
                    prop.payment_date = payment_date
                    prop.receiver_name = receiver_name

                    prop.save(update_fields=[
                        'is_paid',
                        'bank',
                        'transaction_reference',
                        'payment_date',

                        'receiver_name'
                    ])

                messages.success(request, 'پرداخت با موفقیت انجام شد')
                return redirect('middle_register_property')

            except ValidationError as e:
                messages.error(request, e.message)

            except Exception as e:
                messages.error(request, f'خطا در پرداخت: {e}')

    else:
        form = PropertyPayForm(
            user=request.user,
            instance=prop
        )

    return render(
        request,
        'middleProperty/property_pay.html',
        {
            'prop': prop,
            'form': form
        }
    )


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def property_cancel_pay_money_view(request, property_id):
    prop = get_object_or_404(
        Property,
        id=property_id,
        is_paid=True,
        is_active=True,
        user=request.user
    )

    if request.method == 'POST':

        try:
            with transaction.atomic():

                content_type = ContentType.objects.get_for_model(prop)

                fund = Fund.objects.filter(
                    content_type=content_type,
                    object_id=prop.id,
                    user=request.user,
                    is_paid=True
                ).first()

                if not fund:
                    messages.error(request, 'Fund مرتبط با این پرداخت پیدا نشد!')
                    return redirect(request.META.get('HTTP_REFERER'))

                bank_fund = BankFund.objects.filter(
                    content_type=content_type,
                    object_id=prop.id
                ).first()

                # اصلاح موجودی بانک
                if bank_fund:

                    bank = bank_fund.bank

                    if bank_fund.transaction_type == 'deposit':
                        bank.current_balance -= bank_fund.amount

                    elif bank_fund.transaction_type == 'withdraw':
                        bank.current_balance += bank_fund.amount

                    bank.save(update_fields=['current_balance'])

                    bank_fund.delete()

                # حذف Fund
                fund.delete()

                # باز محاسبه صندوق
                Fund.recalc_final_amounts_from(fund)

                # بازگرداندن وضعیت دریافت
                prop.is_paid = False
                prop.bank = None
                prop.transaction_reference = None
                prop.payment_date = None
                prop.receiver_name = ''

                prop.save(update_fields=[
                    'is_paid',
                    'bank',
                    'transaction_reference',
                    'payment_date',
                    'receiver_name',
                ])

                messages.success(
                    request,
                    'پرداخت با موفقیت لغو شد و صندوق و بانک اصلاح شد.'
                )

                return redirect(request.META.get('HTTP_REFERER'))

        except Exception as e:

            messages.error(
                request,
                f'خطا در لغو دریافت: {e}'
            )

            return redirect(request.META.get('HTTP_REFERER'))

    return redirect(request.META.get('HTTP_REFERER'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_property_edit(request, pk):
    property_d = get_object_or_404(Property, pk=pk)

    if property_d.is_paid:
        messages.warning(request, 'این سند بدلیل ثبت رکورد پرداخت قابل ویرایش نیست')
        return redirect('middle_register_property')

    if request.method == 'POST':
        # فرم با instance برای ویرایش
        form = PropertyForm(request.POST, request.FILES, instance=property_d)

        if form.is_valid():
            files = request.FILES.getlist('document')
            if len(files) > 2:
                messages.error(request,
                               "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
                return redirect('middle_register_property')
            property_d = form.save(commit=False)
            property_d.save()
            form.save_m2m()

            # ثبت فایل‌های پیوست جدید
            # files = request.FILES.getlist('document')
            for f in files:
                PropertyDocument.objects.create(property=property_d, document=f)

            messages.success(request, 'سند با موفقیت ویرایش گردید.')
            return redirect(reverse('middle_register_property'))  # Adjust redirect as necessary

        else:
            messages.error(request, 'خطا در ویرایش فرم . لطفا دوباره تلاش کنید.')
            return render(request, 'middleProperty/manage_property.html', {
                'form': form,
                'property_d': property_d,
                'open_modal': True,
            })
    else:
        form = PropertyForm(instance=property_d, user=request.user)
        return render(request, 'middleProperty/manage_property.html',
                      {'form': form,
                       'property_d': property_d,
                       'open_modal': True,
                       })

    # if request.method == 'POST':
    #     form = PropertyForm(request.POST, request.FILES, instance=property_d)
    #
    #     if form.is_valid():
    #         property_d = form.save()  # Save the form (updates or creates expense)
    #
    #         # Handle multiple file uploads
    #         files = request.FILES.getlist('document')
    #         if files:
    #             for f in files:
    #                 PropertyDocument.objects.create(property=property_d, document=f)
    #
    #         messages.success(request, 'سند با موفقیت ویرایش شد.')
    #         return redirect('middle_register_property')  # Adjust redirect as necessary
    #
    #     else:
    #         messages.error(request, 'خطا در ویرایش فرم . لطفا دوباره تلاش کنید.')
    #         return redirect('middle_register_property')
    # else:
    #     # If the request is not POST, redirect to the appropriate page
    #     return redirect('middle_register_property')


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_property_delete(request, pk):
    property_d = get_object_or_404(Property, id=pk)
    if property_d.is_paid:
        messages.warning(request, 'بدلیل ثبت پرداخت، حذف امکان پذیر نیست!')
        return redirect(reverse('middle_register_property'))
    try:
        property_d.delete()
        messages.success(request, ' اموال با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('middle_register_property'))


@login_required(login_url=settings.LOGIN_URL_ADMIN)
@csrf_exempt
def middle_delete_property_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        property_id = request.POST.get('property_id')

        print(f'property_id: {property_id}')

        if not image_url or not property_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            property = get_object_or_404(Property, id=property_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = PropertyDocument.objects.filter(property=property, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


# ============================ MaintenanceView ==========================
@method_decorator(middle_admin_required, name='dispatch')
class MiddleMaintenanceCreateView(CreateView):
    model = Maintenance
    template_name = 'middleMaintenance/add_maintenance.html'
    form_class = MaintenanceForm
    success_url = reverse_lazy('middle_register_maintenance')

    def form_valid(self, form):
        form.instance.user = self.request.user
        house = MyHouse.objects.filter(user=self.request.user, is_active=True).first()
        form.instance.house = house

        with transaction.atomic():
            files = self.request.FILES.getlist('document')
            if len(files) > 2:
                messages.error(self.request,
                               "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
                return redirect('middle_register_maintenance')
            self.object = form.save(commit=False)
            self.object.save()

            # bank = self.object.bank
            # maintenance_price = self.object.maintenance_price
            #
            # bank_funds = Fund.objects.filter(
            #     user=self.request.user,
            #     bank=bank
            # )
            #
            # total_debit = bank_funds.aggregate(
            #     Sum('debtor_amount')
            # )['debtor_amount__sum'] or 0
            # print(f"d-{total_debit}")
            #
            # total_credit = bank_funds.aggregate(
            #     Sum('creditor_amount')
            # )['creditor_amount__sum'] or 0

            #
            # current_final = Decimal(total_debit) - Decimal(total_credit)

            #
            # if current_final < maintenance_price:
            #     messages.error(self.request, 'موجودی صندوق کافی نیست')
            #     return self.form_invalid(form)
            #
            # # self.object.receiver_name = self.object.get_receiver_display
            # if self.object.payment_date:
            #     self.object.is_paid = True
            # self.object.save()
            # form.save_m2m()
            #
            # content_type = ContentType.objects.get_for_model(self.object)
            #
            # Fund.objects.create(
            #     user=self.request.user,
            #     content_type=content_type,
            #     object_id=self.object.id,
            #     bank=bank,
            #     house=house,
            #     unit=None,
            #     amount=self.object.maintenance_price,
            #     debtor_amount=0,
            #     receiver_name=self.object.receiver_name,
            #     creditor_amount=self.object.maintenance_price,
            #     payment_gateway='کارت به کارت',
            #     payment_date=self.object.payment_date,
            #     doc_number=self.object.maintenance_document_no,
            #     payment_description=f"تعیمر و نگهداری: {self.object.maintenance_description[:50]}",
            #     is_paid=True,
            #
            # )

            # files = self.request.FILES.getlist('document')
            for f in files:
                MaintenanceDocument.objects.create(maintenance=self.object, document=f)

            messages.success(self.request, 'سند با موفقیت ثبت گردید!')
            return super().form_valid(form)

        # except Exception as e:
        #     messages.error(self.request, f'خطا در ثبت: {e}')
        #     return self.form_invalid(form)

    def get_queryset(self):
        queryset = Maintenance.objects.filter(user=self.request.user).order_by('-id')

        # فیلتر بر اساس amount
        maintenance_description = self.request.GET.get('maintenance_description')
        if maintenance_description:
            queryset = queryset.filter(maintenance_description__icontains=maintenance_description)

        maintenance_price = self.request.GET.get('maintenance_price')
        if maintenance_price:
            queryset = queryset.filter(maintenance_price__icontains=maintenance_price)

        maintenance_status = self.request.GET.get('maintenance_status')
        if maintenance_status:
            queryset = queryset.filter(maintenance_status__icontains=maintenance_status)

        service_company = self.request.GET.get('service_company')
        if service_company:
            queryset = queryset.filter(service_company__icontains=service_company)

        maintenance_start_date = self.request.GET.get('maintenance_start_date')
        if maintenance_start_date and isinstance(maintenance_start_date, str):
            try:
                j_start = jdatetime.date.fromisoformat(maintenance_start_date)
                g_start = j_start.togregorian()
                queryset = queryset.filter(maintenance_start_date__gte=g_start)
            except (ValueError, TypeError) as e:
                print("Invalid date format:", maintenance_start_date, e)

        maintenance_end_date = self.request.GET.get('maintenance_end_date')
        if maintenance_end_date and isinstance(maintenance_end_date, str):
            try:
                j_start = jdatetime.date.fromisoformat(maintenance_end_date)
                g_start = j_start.togregorian()
                queryset = queryset.filter(maintenance_end_date__lte=g_start)
            except (ValueError, TypeError) as e:
                print("Invalid date format:", maintenance_end_date, e)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        return queryset

    # def get_form_kwargs(self):
    #     kwargs = super().get_form_kwargs()
    #     kwargs['user'] = self.request.user
    #     return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        maintenances = self.get_queryset()
        paginator = Paginator(maintenances, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['total_maintenances'] = maintenances.filter(user=self.request.user).count()
        context['maintenances'] = page_obj.object_list
        context['banks'] = Bank.objects.filter(user=self.request.user)
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def middle_maintenance_edit(request, pk):
    maintenance = get_object_or_404(Maintenance, pk=pk)

    if request.method == 'POST':
        # فرم با instance برای ویرایش
        form = MaintenanceForm(request.POST, request.FILES, instance=maintenance)

        if form.is_valid():
            files = request.FILES.getlist('document')
            if len(files) > 2:
                messages.error(request,
                               "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
                return redirect('middle_register_maintenance')
            maintenance = form.save(commit=False)
            maintenance.save()

            # ثبت فایل‌های پیوست جدید
            # files = request.FILES.getlist('document')
            for f in files:
                MaintenanceDocument.objects.create(maintenance=maintenance, document=f)

            messages.success(request, 'سند با موفقیت ویرایش گردید.')
            return redirect(reverse('middle_register_maintenance'))  # Adjust redirect as necessary

        else:
            messages.error(request, 'خطا در ویرایش فرم . لطفا دوباره تلاش کنید.')
            return render(request, 'middleMaintenance/add_maintenance.html', {
                'form': form,
                'maintenance': maintenance,
                'open_modal': True,
            })
    else:
        form = PayerMoneyForm(instance=maintenance)
        return render(request, 'middleMaintenance/add_maintenance.html',
                      {'form': form,
                       'maintenance': maintenance,
                       'open_modal': True,
                       })


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def maintenance_pay_money_view(request, maintenance_id):
    maintenance = get_object_or_404(
        Maintenance,
        id=maintenance_id,
        is_paid=False,
        is_active=True
    )

    if request.method == 'POST':
        form = MaintenancePayForm(
            request.POST,
            user=request.user,
            instance=maintenance
        )

        if form.is_valid():
            try:
                with transaction.atomic():

                    maintenance = Maintenance.objects.select_for_update().get(
                        id=maintenance_id,
                        is_paid=False,
                        is_active=True
                    )

                    bank = form.cleaned_data['bank']
                    reference = form.cleaned_data.get('transaction_reference')
                    payment_date = form.cleaned_data.get('payment_date')
                    receiver_name = form.cleaned_data.get('receiver_name')

                    current_balance = bank.current_balance

                    if current_balance < maintenance.maintenance_price:
                        messages.error(
                            request,
                            f'موجودی بانک کافی نیست. موجودی فعلی: {current_balance:,} تومان'
                        )
                        return redirect(request.META.get('HTTP_REFERER'))

                    content_type = ContentType.objects.get_for_model(maintenance)

                    Fund.objects.create(
                        user=request.user,
                        content_type=content_type,
                        object_id=maintenance.id,
                        bank=bank,
                        house=maintenance.house,
                        unit=None,
                        amount=maintenance.maintenance_price,
                        debtor_amount=0,
                        receiver_name=receiver_name,
                        creditor_amount=maintenance.maintenance_price,
                        payment_gateway='کارت به کارت',
                        payment_date=payment_date,
                        transaction_no=reference,
                        doc_number=maintenance.maintenance_document_no,
                        payment_description=f"هزینه تعمیر و نگهداری: {maintenance.maintenance_description[:50]}",
                        is_paid=True,

                    )

                    BankTransactionService.withdraw(
                        user=request.user,
                        bank=bank,
                        unit=None,
                        amount=Decimal(maintenance.maintenance_price),
                        description=f"هزینه تعمیر و نگهداری: {maintenance.maintenance_description[:50]}",
                        content_object=maintenance,
                        payment_date=payment_date,
                        transaction_no=reference,
                        gateway='کارت به کارت',
                        receiver_name=receiver_name,
                        house=maintenance.house
                    )

                    maintenance.is_paid = True
                    maintenance.bank = bank
                    maintenance.transaction_reference = reference
                    maintenance.payment_date = payment_date
                    maintenance.receiver_name = receiver_name

                    maintenance.save(update_fields=[
                        'is_paid',
                        'bank',
                        'transaction_reference',
                        'payment_date',
                        'receiver_name'
                    ])

                messages.success(request, 'پرداخت با موفقیت انجام شد')
                return redirect('middle_register_maintenance')

            except ValidationError as e:
                messages.error(request, e.message)

            except Exception as e:
                messages.error(request, f'خطا در پرداخت: {e}')

    else:
        form = MaintenancePayForm(
            user=request.user,
            instance=maintenance
        )

    return render(
        request,
        'middleMaintenance/maintenance_pay.html',
        {
            'maintenance': maintenance,
            'form': form
        }
    )


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def maintenance_cancel_pay_money_view(request, maintenance_id):
    maintenance = get_object_or_404(
        Maintenance,
        id=maintenance_id,
        is_paid=True,
        is_active=True,
        user=request.user
    )

    if request.method == 'POST':

        try:
            with transaction.atomic():

                content_type = ContentType.objects.get_for_model(maintenance)

                fund = Fund.objects.filter(
                    content_type=content_type,
                    object_id=maintenance.id,
                    user=request.user,
                    is_paid=True
                ).first()

                if not fund:
                    messages.error(request, 'Fund مرتبط با این پرداخت پیدا نشد!')
                    return redirect(request.META.get('HTTP_REFERER'))

                bank_fund = BankFund.objects.filter(
                    content_type=content_type,
                    object_id=maintenance.id
                ).first()

                # اصلاح موجودی بانک
                if bank_fund:

                    bank = bank_fund.bank

                    if bank_fund.transaction_type == 'deposit':
                        bank.current_balance -= bank_fund.amount

                    elif bank_fund.transaction_type == 'withdraw':
                        bank.current_balance += bank_fund.amount

                    bank.save(update_fields=['current_balance'])

                    bank_fund.delete()

                # حذف Fund
                fund.delete()

                # باز محاسبه صندوق
                Fund.recalc_final_amounts_from(fund)

                # بازگرداندن وضعیت دریافت
                maintenance.is_paid = False
                maintenance.bank = None
                maintenance.transaction_reference = None
                maintenance.payment_date = None
                maintenance.receiver_name = ''

                maintenance.save(update_fields=[
                    'is_paid',
                    'bank',
                    'transaction_reference',
                    'payment_date',
                    'receiver_name',
                ])

                messages.success(
                    request,
                    'پرداخت با موفقیت لغو شد و صندوق و بانک اصلاح شد.'
                )

                return redirect(request.META.get('HTTP_REFERER'))

        except Exception as e:

            messages.error(
                request,
                f'خطا در لغو دریافت: {e}'
            )

            return redirect(request.META.get('HTTP_REFERER'))

    return redirect(request.META.get('HTTP_REFERER'))


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def middle_maintenance_delete(request, pk):
    maintenance = get_object_or_404(Maintenance, id=pk)
    if maintenance.is_paid:
        messages.warning(request, 'این سند بدلیل ثبت رکورد پرداخت قابل حذف نیست')
        return redirect('middle_register_maintenance')

    try:
        with transaction.atomic():
            # حذف Fund مربوطه
            payment_ct = ContentType.objects.get_for_model(Maintenance)
            Fund.objects.filter(content_type=payment_ct, object_id=maintenance.id).delete()

            # حذف خود Expense
            maintenance.delete()
        messages.success(request, ' سند با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('middle_register_maintenance'))


@login_required(login_url=settings.LOGIN_URL_ADMIN)
@csrf_exempt
def middle_delete_maintenance_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        maintenance_id = request.POST.get('maintenance_id')

        print(f'maintenance_id: {maintenance_id}')

        if not image_url or not maintenance_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            maintenance = get_object_or_404(Maintenance, id=maintenance_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = MaintenanceDocument.objects.filter(maintenance=maintenance, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


# ========================Sewage Views ======================================
class SewageCostManage(CreateView):
    model = SewageManage
    template_name = 'middelSewage/sewage_register.html'
    form_class = SewageForm
    success_url = reverse_lazy('middle_register_sewage')

    def form_valid(self, form):
        files = self.request.FILES.getlist('document')
        if len(files) > 2:
            messages.error(self.request, "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
            return redirect('middle_register_sewage')

        form.instance.user = self.request.user
        form.instance.house = MyHouse.objects.filter(user=self.request.user, is_active=True).first()
        self.object = form.save()

        for f in files:
            SewageDocument.objects.create(sewage=self.object, document=f)

        messages.success(self.request, "سند با موفقیت ثبت شد.")
        return redirect('middle_register_sewage')

    def get_queryset(self):
        queryset = SewageManage.objects.filter(user=self.request.user)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sewages = self.get_queryset().order_by('-created_at')
        paginator = Paginator(sewages, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['total_sewage'] = sewages.filter(user=self.request.user).count()
        context['sewages'] = page_obj.object_list
        context['units'] = Unit.objects.filter(
            is_active=True,
            user__manager=self.request.user).count()

        return context


def sewage_cost_edit(request, pk):
    sewage = get_object_or_404(SewageManage, pk=pk)

    if request.method == 'POST':
        # فرم با instance برای ویرایش
        form = SewageForm(request.POST, request.FILES, instance=sewage)

        if form.is_valid():
            files = request.FILES.getlist('document')
            if len(files) > 2:
                messages.error(request,
                               "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
                return redirect('middle_register_sewage')

            sewage = form.save(commit=False)
            sewage.user = request.user  # اختصاص کاربر فعال
            sewage.save()

            for f in files:
                SewageDocument.objects.create(sewage=sewage, document=f)

            # ثبت فایل‌های پیوست جدید
            # files = request.FILES.getlist('document')
            # for f in files:
            #     CivilDocument.objects.create(civil=civil, document=f)

            messages.success(request, 'رکورد با موفقیت ویرایش گردید.')
            return redirect(reverse('middle_register_sewage'))  # Adjust redirect as necessary

        else:
            messages.error(request, 'خطا در ویرایش فرم . لطفا دوباره تلاش کنید.')
            return render(request, 'middelSewage/sewage_register.html', {
                'form': form,
                'sewage': sewage,
                'open_modal': True,
            })
    else:
        form = SewageForm(instance=sewage, user=request.user)
        return render(request, 'middleCharge/civil_charge_manage.html',
                      {'form': form,
                       'sewage': sewage,
                       'open_modal': True,
                       })


@login_required(login_url=settings.LOGIN_URL_ADMIN)
@csrf_exempt
def middle_delete_sewage_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        sewage_id = request.POST.get('sewage_id')

        if not image_url or not sewage_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            sewage = get_object_or_404(SewageManage, id=sewage_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = SewageDocument.objects.filter(sewage=sewage, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


def middle_sewage_delete(request, pk):
    sewage = get_object_or_404(SewageManage, id=pk)
    if sewage.count_sent_units() > 0:
        messages.error(request, 'به علت ارسال هزینه فاضلاب به ساکنین، حذف امکان‌پذیر نیست!')
        return redirect(reverse('middle_register_sewage'))  # ← توقف کامل

    try:
        with transaction.atomic():
            sewage.delete()
        messages.success(request, ' هزینه فاضلاب با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('middle_register_sewage'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_show_sewage_form(request, pk):
    sewage = get_object_or_404(SewageManage, id=pk, user=request.user)
    sewages = SewageManage.objects.filter(user=request.user, id=pk).order_by('id')

    units = Unit.objects.filter(
        is_active=True,
        user__manager=request.user
    ).order_by('unit')

    units_with_details = []
    for unit in units:
        # ✅ بررسی اینکه آیا برای این واحد اقساط مربوط به این civil ارسال شده‌اند
        has_been_sent = SewageInstallment.objects.filter(
            sewage_manage=sewage,
            unit=unit,
            send_notification=True
        ).exists()

        units_with_details.append({
            'unit': unit,
            'sent_status': has_been_sent,
        })

    return render(request, 'middelSewage/sewage_notify.html', {
        'sewage': sewage,
        'units_with_details': units_with_details,
        'sewages': sewages,
    })


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_send_sewage(request, pk):
    sewage = get_object_or_404(SewageManage, id=pk, user=request.user)
    all_units_count = Unit.objects.filter(
        is_active=True,
        user__manager=request.user
    ).count()
    logger = logging.getLogger(__name__)

    logger.info(f"all_units: {all_units_count}")

    if request.method == "POST":
        selected_units = request.POST.getlist('units')

        if not selected_units:
            messages.warning(request, 'هیچ واحدی انتخاب نشده است.')
            return redirect('middle_show_send_sewage_form', pk=pk)

        units_qs = Unit.objects.filter(is_active=True, user__manager=request.user)

        if 'all' in selected_units:
            units_to_notify = units_qs
        else:
            units_to_notify = units_qs.filter(id__in=selected_units)

        if not units_to_notify.exists():
            messages.warning(request, 'هیچ واحد معتبری برای ارسال پیدا نشد.')
            return redirect('middle_show_send_sewage_form', pk=pk)

        already_sent = 0
        created_count = 0

        with transaction.atomic():

            total_amount = sewage.amount
            prepayment_total = sewage.prepayment or 0
            installment_count = sewage.installment_count or 1

            if installment_count <= 0:
                installment_count = 1

            # ✅ تعداد کل واحدهای ساختمان

            # ✅ محاسبه سهم هر واحد
            share_per_unit = total_amount // all_units_count
            print(f'share_per_unit : {share_per_unit}')

            prepayment_per_unit = prepayment_total // all_units_count
            print(f'prepayment_per_unit : {prepayment_per_unit}', flush=True)

            remaining_total = total_amount - prepayment_total
            print(f'remaining_total : {remaining_total}')

            remaining_per_unit = remaining_total // all_units_count
            print(f'remaining_per_unit : {remaining_per_unit}')

            monthly_installment = (
                remaining_per_unit // installment_count
                if installment_count > 0 else remaining_per_unit
            )
            print(f'monthly_installment : {monthly_installment}')

            # واحدهایی که قبلاً ارسال شده‌اند
            sent_unit_ids = set(
                SewageInstallment.objects.filter(
                    sewage_manage=sewage,
                    unit__in=units_to_notify,
                    send_notification=True
                ).values_list('unit_id', flat=True)
            )

            for unit in units_to_notify:

                if unit.id in sent_unit_ids:
                    already_sent += 1
                    continue

                SewageInstallment.objects.filter(
                    sewage_manage=sewage,
                    unit=unit
                ).delete()

                base_date = sewage.first_due_date or timezone.now()

                installments = []
                print(sewage.first_due_date)
                print(type(sewage.first_due_date))

                print("installment_count:", installment_count)
                print("type:", type(installment_count))
                # ✅ 1️⃣ ایجاد رکورد پیش‌پرداخت (فقط یکبار)
                if prepayment_per_unit > 0:
                    installments.append(
                        SewageInstallment(
                            sewage_manage=sewage,
                            unit=unit,
                            installment_number=0,  # پیش پرداخت
                            amount=prepayment_per_unit,
                            due_date=base_date,
                            send_notification=True,
                            send_notification_date=timezone.now(),
                            prepayment_per_unit=prepayment_per_unit,
                            house=sewage.house
                        )
                    )

                # ✅ 2️⃣ ایجاد اقساط معمولی (بدون پیش‌پرداخت)
                for i in range(1, installment_count + 1):
                    due_date = base_date + relativedelta(months=i)

                    installments.append(
                        SewageInstallment(
                            sewage_manage=sewage,
                            unit=unit,
                            installment_number=i,
                            amount=monthly_installment,
                            due_date=due_date,
                            send_notification=True,
                            send_notification_date=timezone.now(),
                            prepayment_per_unit=0,  # ❌ دیگر تکرار نمی‌شود
                            house=sewage.house
                        )
                    )

                SewageInstallment.objects.bulk_create(installments)
                created_count += 1

        # پیام نهایی
        if created_count > 0 and already_sent > 0:
            messages.success(
                request,
                f'{created_count} واحد جدید ارسال شد. {already_sent} واحد قبلاً ارسال شده بودند.'
            )
        elif created_count > 0:
            messages.success(
                request,
                f'برای {created_count} واحد، پیش‌پرداخت و اقساط ایجاد شد.'
            )
        else:
            messages.warning(
                request,
                'اعلان قبلاً برای همه واحدهای انتخابی ارسال شده بود.'
            )

        return redirect('middle_register_sewage')

    # GET
    units = Unit.objects.filter(
        is_active=True,
        user__manager=request.user
    ).order_by('unit')

    units_with_details = [{'unit': unit} for unit in units]

    return render(request, 'middelSewage/sewage_notify.html', {
        'sewage': sewage,
        'units_with_details': units_with_details,
        'all_unit': all_units_count
    })


# ========================Civil Charge Views ======================================
class CivilChargeManage(CreateView):
    model = CivilManage
    template_name = 'middleCharge/civil_charge_manage.html'
    form_class = CivilForm
    success_url = reverse_lazy('civil_charge_manage')

    def form_valid(self, form):
        files = self.request.FILES.getlist('document')
        if len(files) > 2:
            messages.error(self.request, "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
            return redirect('civil_charge_manage')

        form.instance.user = self.request.user
        form.instance.house = MyHouse.objects.filter(user=self.request.user, is_active=True).first()
        self.object = form.save()

        for f in files:
            CivilDocument.objects.create(civil=self.object, document=f)

        messages.success(self.request, "هزینه فاضلاب با موفقیت ثبت شد.")
        return redirect('civil_charge_manage')

    def get_queryset(self):
        queryset = CivilManage.objects.filter(user=self.request.user)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        civiles = self.get_queryset().order_by('-created_at')
        paginator = Paginator(civiles, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['total_civil'] = civiles.filter(user=self.request.user).count()
        context['civils'] = page_obj.object_list
        context['units'] = Unit.objects.filter(
            is_active=True,
            user__manager=self.request.user).count()

        return context


def civil_charge_edit(request, pk):
    civil = get_object_or_404(CivilManage, pk=pk)

    if request.method == 'POST':
        # فرم با instance برای ویرایش
        form = CivilForm(request.POST, request.FILES, instance=civil)

        if form.is_valid():
            files = request.FILES.getlist('document')
            if len(files) > 2:
                messages.error(request,
                               "حداکثر دو فایل مجاز است. در صورت لزوم فایل را بصورت pdf یا zip آپلود کنید")
                return redirect('middle_c')

            civil = form.save(commit=False)
            civil.user = request.user  # اختصاص کاربر فعال
            civil.save()

            for f in files:
                CivilDocument.objects.create(civil=civil, document=f)

            # ثبت فایل‌های پیوست جدید
            # files = request.FILES.getlist('document')
            # for f in files:
            #     CivilDocument.objects.create(civil=civil, document=f)

            messages.success(request, 'رکورد با موفقیت ویرایش گردید.')
            return redirect(reverse('civil_charge_manage'))  # Adjust redirect as necessary

        else:
            messages.error(request, 'خطا در ویرایش فرم . لطفا دوباره تلاش کنید.')
            return render(request, 'middleCharge/civil_charge_manage.html', {
                'form': form,
                'civil': civil,
                'open_modal': True,
            })
    else:
        form = CivilForm(instance=civil)
        return render(request, 'middleCharge/civil_charge_manage.html',
                      {'form': form,
                       'civil': civil,
                       'open_modal': True,
                       })


@login_required(login_url=settings.LOGIN_URL_ADMIN)
@csrf_exempt
def middle_delete_civil_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        civil_id = request.POST.get('civil_id')

        print(f'maintenance_id: {civil_id}')

        if not image_url or not civil_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            civil = get_object_or_404(CivilManage, id=civil_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = CivilDocument.objects.filter(civil=civil, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_civil_delete(request, pk):
    civil = get_object_or_404(CivilManage, id=pk)
    if civil.count_sent_units() > 0:
        messages.error(request, 'به علت ارسال شارژ به ساکنین، حذف امکان‌پذیر نیست!')
        return redirect(reverse('civil_charge_manage'))  # ← توقف کامل

    try:
        with transaction.atomic():
            civil.delete()
        messages.success(request, ' شارژ با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('civil_charge_manage'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_show_civil_form(request, pk):
    civil = get_object_or_404(CivilManage, id=pk, user=request.user)
    civils = CivilManage.objects.filter(user=request.user, id=pk).order_by('id')

    units = Unit.objects.filter(
        is_active=True,
        user__manager=request.user
    ).order_by('unit')

    units_with_details = []
    for unit in units:
        # ✅ بررسی اینکه آیا برای این واحد اقساط مربوط به این civil ارسال شده‌اند
        has_been_sent = CivilInstallment.objects.filter(
            civil_manage=civil,
            unit=unit,
            send_notification=True
        ).exists()

        units_with_details.append({
            'unit': unit,
            'sent_status': has_been_sent,
        })

    return render(request, 'middleCharge/civil_charge_notify.html', {
        'civil': civil,
        'units_with_details': units_with_details,
        'civils': civils,
    })


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_send_civil_charge(request, pk):
    civil = get_object_or_404(CivilManage, id=pk, user=request.user)
    all_units_count = Unit.objects.filter(
        is_active=True,
        user__manager=request.user
    ).count()
    logger = logging.getLogger(__name__)

    logger.info(f"all_units: {all_units_count}")

    if request.method == "POST":
        selected_units = request.POST.getlist('units')

        if not selected_units:
            messages.warning(request, 'هیچ واحدی انتخاب نشده است.')
            return redirect('middle_show_send_civil_form', pk=pk)

        units_qs = Unit.objects.filter(is_active=True, user__manager=request.user)

        if 'all' in selected_units:
            units_to_notify = units_qs
        else:
            units_to_notify = units_qs.filter(id__in=selected_units)

        if not units_to_notify.exists():
            messages.warning(request, 'هیچ واحد معتبری برای ارسال پیدا نشد.')
            return redirect('middle_show_send_civil_form', pk=pk)

        already_sent = 0
        created_count = 0

        with transaction.atomic():

            total_amount = civil.amount
            prepayment_total = civil.prepayment or 0
            installment_count = civil.installment_count or 1

            if installment_count <= 0:
                installment_count = 1

            # ✅ تعداد کل واحدهای ساختمان

            # ✅ محاسبه سهم هر واحد
            share_per_unit = total_amount // all_units_count
            print(f'share_per_unit : {share_per_unit}')

            prepayment_per_unit = prepayment_total // all_units_count
            print(f'prepayment_per_unit : {prepayment_per_unit}', flush=True)

            remaining_total = total_amount - prepayment_total
            print(f'remaining_total : {remaining_total}')

            remaining_per_unit = remaining_total // all_units_count
            print(f'remaining_per_unit : {remaining_per_unit}')

            monthly_installment = (
                remaining_per_unit // installment_count
                if installment_count > 0 else remaining_per_unit
            )
            print(f'monthly_installment : {monthly_installment}')

            # واحدهایی که قبلاً ارسال شده‌اند
            sent_unit_ids = set(
                CivilInstallment.objects.filter(
                    civil_manage=civil,
                    unit__in=units_to_notify,
                    send_notification=True
                ).values_list('unit_id', flat=True)
            )

            for unit in units_to_notify:

                if unit.id in sent_unit_ids:
                    already_sent += 1
                    continue

                CivilInstallment.objects.filter(
                    civil_manage=civil,
                    unit=unit
                ).delete()

                base_date = civil.first_due_date or timezone.now()

                installments = []

                # ✅ 1️⃣ ایجاد رکورد پیش‌پرداخت (فقط یکبار)
                if prepayment_per_unit > 0:
                    installments.append(
                        CivilInstallment(
                            civil_manage=civil,
                            unit=unit,
                            installment_number=0,  # پیش پرداخت
                            amount=prepayment_per_unit,
                            due_date=base_date,
                            send_notification=True,
                            send_notification_date=timezone.now(),
                            prepayment_per_unit=prepayment_per_unit,
                            house=civil.house
                        )
                    )

                # ✅ 2️⃣ ایجاد اقساط معمولی (بدون پیش‌پرداخت)
                for i in range(1, installment_count + 1):
                    due_date = base_date + relativedelta(months=i)

                    installments.append(
                        CivilInstallment(
                            civil_manage=civil,
                            unit=unit,
                            installment_number=i,
                            amount=monthly_installment,
                            due_date=due_date,
                            send_notification=True,
                            send_notification_date=timezone.now(),
                            prepayment_per_unit=0,  # ❌ دیگر تکرار نمی‌شود
                            house=civil.house
                        )
                    )

                CivilInstallment.objects.bulk_create(installments)
                created_count += 1

        # پیام نهایی
        if created_count > 0 and already_sent > 0:
            messages.success(
                request,
                f'{created_count} واحد جدید ارسال شد. {already_sent} واحد قبلاً ارسال شده بودند.'
            )
        elif created_count > 0:
            messages.success(
                request,
                f'برای {created_count} واحد، پیش‌پرداخت و اقساط ایجاد شد.'
            )
        else:
            messages.warning(
                request,
                'اعلان قبلاً برای همه واحدهای انتخابی ارسال شده بود.'
            )

        return redirect('civil_charge_manage')

    # GET
    units = Unit.objects.filter(
        is_active=True,
        user__manager=request.user
    ).order_by('unit')

    units_with_details = [{'unit': unit} for unit in units]

    return render(request, 'middleCharge/civil_charge_notify.html', {
        'civil': civil,
        'units_with_details': units_with_details,
        'all_unit': all_units_count
    })


# ================================================================================================
@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_charge_view(request):
    user = request.user

    if not user.is_middle_admin:
        return HttpResponseForbidden()

    context = {
        "allowed_methods": user.charge_method_codes
    }

    return render(
        request,
        "middleCharge/add_charge.html",
        context
    )


@method_decorator(middle_admin_required, name='dispatch')
class MiddleFixChargeCreateView(CreateView):
    model = FixCharge
    template_name = 'middleCharge/fix_charge_template.html'
    form_class = FixChargeForm
    success_url = reverse_lazy('middle_add_fixed_charge')

    def form_valid(self, form):
        charge_name = form.cleaned_data.get('name') or 'شارژ ثابت'

        # گرفتن کاربران تحت مدیریت
        managed_users = self.request.user.managed_users.all()

        unit_count = UnifiedCharge.objects.filter(
            user=self.request.user,
            unit__is_active=True
        ).values('unit').distinct().count()
        form.instance.unit_count = unit_count
        print(f'unit-{unit_count}')

        units = Unit.objects.filter(
            is_active=True
        ).filter(
            Q(user=self.request.user) | Q(user__in=managed_users)
        ).distinct()
        print(units.count())

        all_units = Unit.objects.filter(is_active=True)
        filtered_units = all_units.filter(
            Q(user=self.request.user) | Q(user__in=managed_users)
        )

        print("All active:", all_units.count())
        print("After filter:", filtered_units.count())

        excluded = all_units.exclude(
            Q(user=self.request.user) | Q(user__in=managed_users)
        )

        print("Excluded:", excluded)

        if not units.exists():
            messages.error(
                self.request,
                'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.'
            )
            return redirect('middle_manage_unit')

        # ذخیره FixCharge
        fix_charge = form.save(commit=False)
        fix_charge.user = self.request.user
        fix_charge.name = charge_name
        fix_charge.save()

        # انتخاب Calculator
        calculator = CALCULATORS.get(fix_charge.charge_type)
        if not calculator:
            messages.error(self.request, 'نوع شارژ پشتیبانی نمی‌شود')
            return redirect(self.success_url)

        unified_charges = []

        for unit in units:
            # محاسبه مبلغ پایه برای هر واحد
            base_amount = calculator.calculate(unit, fix_charge)

            # مطمئن شدن که فیلدهای عددی عدد هستند
            civil_amount = fix_charge.civil or 0
            other_amount = fix_charge.other_cost_amount or 0
            total_monthly_charge = base_amount + civil_amount + other_amount

            unified_charges.append(
                UnifiedCharge(
                    user=self.request.user,
                    unit=unit,
                    bank=None,
                    amount=base_amount,
                    house=unit.myhouse,
                    charge_type=fix_charge.charge_type,
                    base_charge=total_monthly_charge,
                    main_charge=fix_charge,
                    penalty_percent=fix_charge.payment_penalty_amount,
                    civil=civil_amount,
                    other_cost_amount=other_amount,
                    penalty_amount=0,
                    total_charge_month=total_monthly_charge,
                    details=fix_charge.details or '',
                    title=fix_charge.name,
                    send_notification=False,  # ⛔ اعلان ارسال نشده
                    send_notification_date=None,
                    payment_deadline_date=fix_charge.payment_deadline,
                    content_type=ContentType.objects.get_for_model(FixCharge),
                    object_id=fix_charge.id,
                )
            )

        # ایجاد همه UnifiedCharge ها در دیتابیس یکجا
        UnifiedCharge.objects.bulk_create(unified_charges)

        messages.success(self.request, 'شارژ با موفقیت ثبت گردید.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        manager_units = Unit.objects.filter(is_active=True, user=self.request.user)
        managed_units = Unit.objects.filter(is_active=True, user__manager=self.request.user)

        # ترکیب و distinct
        all_units = (manager_units | managed_units).distinct()
        context['unit_count'] = all_units.count()
        context['units'] = all_units

        charges = FixCharge.objects.filter(user=self.request.user).annotate(
            total_units=Count('unified_charges'),  # همه واحدهای مرتبط
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            )
        ).order_by('-created_at')
        context['charges'] = charges
        paginate_by = self.request.GET.get('paginate', '20')

        if paginate_by == '1000':  # نمایش همه
            paginator = Paginator(charges, charges.count() or 20)
        else:
            paginate_by = int(paginate_by)
            paginator = Paginator(charges, paginate_by)

        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['charges'] = page_obj
        context['page_obj'] = page_obj
        context['paginate'] = paginate_by
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_fix_charge_edit(request, pk):
    fix_charge = get_object_or_404(FixCharge, pk=pk, user=request.user)

    # بررسی اینکه برای این شارژ اعلان ارسال شده باشد
    any_notify = UnifiedCharge.objects.filter(
        content_type=ContentType.objects.get_for_model(FixCharge),
        object_id=fix_charge.id,
        send_notification=True
    ).exists()
    if any_notify:
        messages.error(request, 'برای این شارژ اعلان ارسال شده و قابل ویرایش نیست.')
        return redirect('middle_add_fixed_charge')

    if request.method == 'POST':
        form = FixChargeForm(request.POST, request.FILES, instance=fix_charge)
        if form.is_valid():
            with transaction.atomic():
                fix_charge = form.save(commit=False)
                fix_charge.name = fix_charge.name or 'شارژ ثابت'
                fix_charge.save()

                # کاربران و واحدهای تحت مدیریت
                managed_users = request.user.managed_users.all()
                units = Unit.objects.filter(is_active=True, user__in=managed_users)
                if not units.exists():
                    messages.error(request, 'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.')
                    return redirect('middle_manage_unit')

                # انتخاب Calculator
                calculator = CALCULATORS.get(fix_charge.charge_type)
                if not calculator:
                    messages.error(request, 'نوع شارژ پشتیبانی نمی‌شود.')
                    return redirect('middle_add_fixed_charge')

                unified_charges = []

                for unit in units:
                    # محاسبه مبلغ پایه برای هر واحد
                    base_amount = calculator.calculate(unit, fix_charge)
                    civil_amount = fix_charge.civil or 0
                    other_amount = fix_charge.other_cost_amount or 0
                    total_monthly = base_amount + civil_amount + other_amount

                    # آپدیت یا ایجاد UnifiedCharge برای این واحد
                    UnifiedCharge.objects.update_or_create(
                        user=request.user,
                        unit=unit,
                        content_type=ContentType.objects.get_for_model(FixCharge),
                        object_id=fix_charge.id,
                        defaults={
                            'bank': None,
                            'charge_type': fix_charge.charge_type,
                            'amount': base_amount,
                            'main_charge': fix_charge,
                            'base_charge': total_monthly,
                            'penalty_percent': fix_charge.payment_penalty_amount or 0,
                            'civil': civil_amount,
                            'other_cost_amount': other_amount,
                            'penalty_amount': 0,
                            'total_charge_month': total_monthly,
                            'details': fix_charge.details or '',
                            'title': fix_charge.name,
                            'send_notification': False,
                            'send_notification_date': None,
                            'payment_deadline_date': fix_charge.payment_deadline,
                        }
                    )

                messages.success(request, 'شارژ با موفقیت ویرایش شد.')
                return redirect('middle_add_fixed_charge')
        else:
            messages.error(request, 'خطا در ویرایش فرم. لطفا دوباره تلاش کنید.')
    else:
        form = FixChargeForm(instance=fix_charge)

    return render(request, 'middleCharge/fix_charge_template.html', {'form': form, 'charge': fix_charge})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_fix_charge_delete(request, pk):
    charge = get_object_or_404(FixCharge, id=pk, user=request.user)

    content_type = ContentType.objects.get_for_model(FixCharge)

    # بررسی اینکه هیچ رکورد UnifiedCharge با is_paid=True وجود نداشته باشد
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            is_paid=True
    ).exists():
        messages.error(request, "امکان حذف شارژ وجود ندارد چون پرداخت شارژ توسط واحد ثبت شده است.")
        return redirect(reverse('middle_add_fixed_charge'))

    # چک کردن وجود رکوردهایی که send_notification == True هستند
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            send_notification=True
    ).exists():
        messages.error(request, "برای این شارژ اطلاعیه صادر شده است. ابتدا اطلاعیه شارژ را حذف و مجدداً تلاش نمایید!")
        return redirect(reverse('middle_add_fixed_charge'))

    try:
        charge.delete()
        messages.success(request, f'{charge.name} با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, "امکان حذف این شارژ به دلیل وابستگی وجود ندارد!")

    return redirect(reverse('middle_add_fixed_charge'))


def toggle_unit_selection(request, pk):
    session_key = f"charge_{pk}_selected_units"

    selected = set(request.session.get(session_key, []))

    unit_id = request.POST.get("unit_id")

    if not unit_id:
        return JsonResponse({"ok": False})

    unit_id = int(unit_id)

    if unit_id in selected:
        selected.remove(unit_id)
        checked = False
    else:
        selected.add(unit_id)
        checked = True

    # 🔥 مهم‌ترین بخش
    request.session[session_key] = list(selected)
    request.session.modified = True

    return JsonResponse({
        "ok": True,
        "checked": checked,
        "count": len(selected),
        "saved": True
    })


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_fix_charge_notification_view(request, pk):
    charge = get_object_or_404(FixCharge, id=pk, user=request.user)

    house = MyHouse.objects.filter(user=request.user).first()

    units = Unit.objects.filter(
        Q(user=request.user) |
        Q(user__manager=request.user),
        is_active=True,
        myhouse=house
    ).distinct().order_by('unit')

    content_type = ContentType.objects.get_for_model(FixCharge)

    # -----------------------------
    # ساخت UnifiedCharge فقط برای واحدهای جدید
    # -----------------------------
    existing_uc_unit_ids = UnifiedCharge.objects.filter(
        content_type=content_type,
        object_id=charge.id,
        unit__in=units
    ).values_list('unit_id', flat=True)

    new_units = units.exclude(id__in=existing_uc_unit_ids)

    calculator = CALCULATORS.get(charge.charge_type)

    with transaction.atomic():
        for unit in new_units:
            base_amount = calculator.calculate(unit, charge)
            civil_amount = charge.civil or 0
            other_amount = charge.other_cost_amount or 0
            total = base_amount + civil_amount + other_amount

            UnifiedCharge.objects.create(
                user=request.user,
                unit=unit,
                house=unit.myhouse,
                bank=None,
                amount=base_amount,
                charge_type=charge.charge_type,
                base_charge=total,
                main_charge=charge,
                penalty_percent=charge.payment_penalty_amount,
                civil=civil_amount,
                other_cost_amount=other_amount,
                penalty_amount=0,
                total_charge_month=total,
                details=charge.details or '',
                title=charge.name,
                send_notification=False,
                send_notification_date=None,
                payment_deadline_date=charge.payment_deadline,
                content_type=content_type,
                object_id=charge.id,
            )

    # -----------------------------
    # FILTERS
    # -----------------------------
    search_query = request.GET.get('search', '').strip()
    resident_type = request.GET.get('resident_type', '').strip()

    if search_query:
        units = units.filter(
            Q(unit__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(renters__renter_name__icontains=search_query)
        ).distinct()

    if resident_type == 'owner':
        units = units.exclude(renters__renter_is_active=True)

    elif resident_type == 'renter':
        units = units.filter(renters__renter_is_active=True).distinct()

    elif resident_type == 'empty':
        units = units.filter(status_residence='empty')

    # -----------------------------
    # PAGINATION
    # -----------------------------
    per_page = int(request.GET.get('per_page', 30))
    paginator = Paginator(units, per_page)
    page_units = paginator.get_page(request.GET.get('page'))

    # -----------------------------
    # SESSION SELECTION STATE
    # -----------------------------
    session_key = f"charge_{pk}_selected_units"
    selected_units = set(request.session.get(session_key, []))
    # -----------------------------
    # POST: SEND NOTIFICATION / SMS
    # -----------------------------
    if request.method == "POST":

        send_type = request.POST.get("send_type", "notify")
        selected = request.POST.getlist("units")

        if not selected:
            messages.warning(request, 'هیچ واحدی انتخاب نشده')
            return redirect(request.path)

        qs = UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=selected
        ).select_related("unit")

        if not qs.exists():
            messages.info(request, 'اطلاعیه‌ای برای ارسال وجود ندارد')
            return redirect(request.path)

        # ثبت اعلان سیستمی
        qs.update(
            send_notification=True,
            send_notification_date=timezone.now().date()
        )

        # ---------- ارسال SMS ----------
        if send_type == "sms":

            result = SmsService.send_for_unified_charges(
                user=request.user,
                unified_charges=qs,
                meta_callback=lambda total_sms, total_price: qs.update(
                    send_sms=True,
                    send_sms_date=timezone.now().date(),
                    sms_count=total_sms,
                    sms_price=Decimal(settings.SMS_PRICE),
                    sms_total_price=total_price
                )
            )

            if result.success:
                messages.success(
                    request,
                    f'اطلاعیه سیستمی و پیامکی برای {qs.count()} واحد ارسال شد'
                )
            else:
                messages.error(request, result.message)

        else:
            messages.success(
                request,
                f'اطلاعیه سیستمی برای {qs.count()} واحد ثبت شد'
            )

            # پاکسازی انتخاب‌ها بعد از ارسال موفق
            request.session[session_key] = []
            request.session.modified = True
            return redirect(request.path)

    # -----------------------------
    # BUILD CONTEXT DATA
    # -----------------------------
    uc_map = UnifiedCharge.objects.filter(
        content_type=content_type,
        object_id=charge.id,
        unit_id__in=[u.id for u in page_units]
    ).select_related('unit', 'unit__user', 'bank')

    uc_dict = {uc.unit_id: uc for uc in uc_map}

    for i, unit in enumerate(page_units.object_list):
        uc = uc_dict.get(unit.id)
        renter = unit.renters.filter(renter_is_active=True).first()

        current_charge = uc.total_charge_month if uc else 0
        previous_debt = sum(uc.get_previous_debt_by_type().values()) if uc else 0

        page_units.object_list[i] = {
            'unit': unit,
            'renter': renter,
            'is_paid': uc.is_paid if uc else False,
            'is_notified': uc.send_notification if uc else False,
            'send_sms': uc.send_sms if uc else False,
            'sms_date': uc.send_sms_date if uc else None,
            'current_charge': current_charge,
            'previous_debt': previous_debt,
            'total_payable': current_charge + previous_debt,
        }

    return render(request, 'middleCharge/notify_fix_charge_template.html', {
        'charge': charge,
        'page_obj': page_units,
        'selected_units': selected_units,  # مهم برای checkbox sync
    })


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_remove_send_notification_fix(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'فقط درخواست‌های POST مجاز است.'}, status=400)

    charge = get_object_or_404(FixCharge, id=pk, user=request.user)
    selected_units = request.POST.getlist('units[]')

    # -----------------------------
    # SESSION SELECTION STATE
    # -----------------------------
    # session_key = f"charge_{pk}_selected_units"
    # selected_units = set(request.session.get(session_key, []))

    if not selected_units:
        return JsonResponse({'warning': 'هیچ واحدی انتخاب نشده است.'})

    try:
        with transaction.atomic():
            content_type = ContentType.objects.get_for_model(FixCharge)

            # رکوردهایی که باید غیرفعال شوند
            if selected_units == ['all']:
                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    is_paid=False,
                    send_notification=True,
                    send_sms=True
                )
            else:
                try:
                    selected_unit_ids = [int(uid) for uid in selected_units]
                except ValueError:
                    return JsonResponse({'error': 'شناسه واحد نامعتبر است.'}, status=400)

                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    unit_id__in=selected_unit_ids,
                    is_paid=False,
                    send_notification=True,  # فقط رکوردهای فعال

                )

            updated_count = qs.update(
                send_notification=False,
                send_notification_date=None,
                send_sms=False,
                send_sms_date=None
            )

            # اگر هیچ رکوردی با send_notification=True باقی نماند → شارژ را غیرفعال کن
            if not UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    send_notification=True
            ).exists():
                charge.send_notification = False
                charge.save()

        session_key = f"charge_{pk}_selected_units"

        if updated_count:
            request.session.pop(session_key, None)
            request.session.modified = True

            return JsonResponse({
                'success': f'{updated_count} اطلاعیه غیرفعال شد.'
            })

        else:
            return JsonResponse({'info': 'رکوردی برای غیرفعال کردن یافت نشد.'})

    except Exception as e:
        return JsonResponse({'error': f'خطایی هنگام غیرفعال کردن اطلاعیه‌ها رخ داد: {str(e)}'}, status=500)


# ========================================== Area Charge =======================
@method_decorator(middle_admin_required, name='dispatch')
class MiddleAreaChargeCreateView(CreateView):
    model = AreaCharge
    template_name = 'middleCharge/area_charge_template.html'
    form_class = AreaChargeForm
    success_url = reverse_lazy('middle_add_area_charge')

    def form_valid(self, form):
        charge_name = form.cleaned_data.get('name')

        # گرفتن کاربران تحت مدیریت
        managed_users = self.request.user.managed_users.all()

        unit_count = UnifiedCharge.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            unit__is_active=True
        ).values('unit').distinct().count()
        form.instance.unit_count = unit_count

        units = Unit.objects.filter(
            is_active=True
        ).filter(
            Q(user=self.request.user) | Q(user__in=managed_users)
        ).distinct()

        total_area = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('area'))['total'] or 0
        form.instance.total_area = total_area

        if not units.exists():
            messages.error(
                self.request,
                'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.'
            )
            return redirect('middle_manage_unit')

        # ذخیره FixCharge
        area_charge = form.save(commit=False)
        area_charge.user = self.request.user
        area_charge.name = charge_name
        area_charge.save()

        # انتخاب Calculator
        calculator = CALCULATORS.get(area_charge.charge_type)
        if not calculator:
            messages.error(self.request, 'نوع شارژ پشتیبانی نمی‌شود')
            return redirect(self.success_url)

        unified_charges = []

        for unit in units:
            # محاسبه مبلغ پایه برای هر واحد
            base_amount = calculator.calculate(unit, area_charge)

            # مطمئن شدن که فیلدهای عددی عدد هستند
            civil_amount = area_charge.civil or 0
            other_amount = area_charge.other_cost_amount or 0
            total_monthly_charge = base_amount + civil_amount + other_amount

            unified_charges.append(
                UnifiedCharge(
                    user=self.request.user,
                    unit=unit,
                    bank=None,
                    house=unit.myhouse,
                    main_charge=area_charge,
                    charge_type=area_charge.charge_type,
                    amount=base_amount,
                    base_charge=total_monthly_charge,
                    penalty_percent=area_charge.payment_penalty_amount,
                    civil=civil_amount,
                    other_cost_amount=other_amount,
                    penalty_amount=0,
                    total_charge_month=total_monthly_charge,
                    details=area_charge.details or '',
                    title=area_charge.name,
                    send_notification=False,  # ⛔ اعلان ارسال نشده
                    send_notification_date=None,
                    payment_deadline_date=area_charge.payment_deadline,
                    content_type=ContentType.objects.get_for_model(AreaCharge),
                    object_id=area_charge.id,
                )
            )

        # ایجاد همه UnifiedCharge ها در دیتابیس یکجا
        UnifiedCharge.objects.bulk_create(unified_charges)

        messages.success(self.request, 'شارژ با موفقیت ثبت گردید.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        managed_users = self.request.user.managed_users.all()
        unit_count = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).count()

        context['unit_count'] = unit_count
        print(unit_count)

        total_area = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('area'))['total'] or 0
        context['total_area'] = total_area

        charges = AreaCharge.objects.filter(user=self.request.user).annotate(
            total_units=Count('unified_charges'),  # همه واحدهای مرتبط
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            )
        ).order_by('-created_at')
        context['charges'] = charges
        paginate_by = self.request.GET.get('paginate', '20')

        if paginate_by == '1000':  # نمایش همه
            paginator = Paginator(charges, charges.count() or 20)
        else:
            paginate_by = int(paginate_by)
            paginator = Paginator(charges, paginate_by)

        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['charges'] = page_obj
        context['page_obj'] = page_obj
        context['paginate'] = paginate_by

        context.update({
            'unit_count': unit_count,
            'total_area': total_area,
            # 'total_people': total_people,
            'charges': charges,
        })
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_area_charge_edit(request, pk):
    charge = get_object_or_404(AreaCharge, pk=pk, user=request.user)

    # بررسی اینکه برای این شارژ اعلان ارسال شده باشد
    any_notify = UnifiedCharge.objects.filter(
        content_type=ContentType.objects.get_for_model(AreaCharge),
        object_id=charge.id,
        send_notification=True
    ).exists()
    if any_notify:
        messages.error(request, 'برای این شارژ اعلان ارسال شده و قابل ویرایش نیست.')
        return redirect('middle_add_area_charge')

    if request.method == 'POST':
        form = AreaChargeForm(request.POST, request.FILES, instance=charge)
        if form.is_valid():
            with transaction.atomic():
                charge = form.save(commit=False)
                charge.name = charge.name or 'شارژ ثابت'
                charge.save()

                # کاربران و واحدهای تحت مدیریت
                managed_users = request.user.managed_users.all()
                units = Unit.objects.filter(
                    Q(user=request.user) | Q(user__in=managed_users),
                    is_active=True)
                if not units.exists():
                    messages.error(request, 'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.')
                    return redirect('middle_manage_unit')

                # انتخاب Calculator
                calculator = CALCULATORS.get(charge.charge_type)
                if not calculator:
                    messages.error(request, 'نوع شارژ پشتیبانی نمی‌شود.')
                    return redirect('middle_add_fixed_charge')

                unified_charges = []

                for unit in units:
                    # محاسبه مبلغ پایه برای هر واحد
                    base_amount = calculator.calculate(unit, charge)
                    civil_amount = charge.civil or 0
                    other_amount = charge.other_cost_amount or 0
                    total_monthly = base_amount + civil_amount + other_amount

                    # آپدیت یا ایجاد UnifiedCharge برای این واحد
                    UnifiedCharge.objects.update_or_create(
                        user=request.user,
                        unit=unit,
                        content_type=ContentType.objects.get_for_model(AreaCharge),
                        object_id=charge.id,
                        defaults={
                            'bank': None,
                            'charge_type': charge.charge_type,
                            'fix_amount': 0,
                            'amount': base_amount,
                            'main_charge': charge,
                            'base_charge': total_monthly,
                            'charge_by_person_amount': 0,
                            'charge_by_area_amount': charge.area_amount,
                            'fix_person_variable_amount': 0,
                            'fix_area_variable_amount': 0,
                            'penalty_percent': charge.payment_penalty_amount or 0,
                            'civil': civil_amount,
                            'other_cost_amount': other_amount,
                            'penalty_amount': 0,
                            'total_charge_month': total_monthly,
                            'details': charge.details or '',
                            'title': charge.name,
                            'send_notification': False,
                            'send_notification_date': None,
                            'payment_deadline_date': charge.payment_deadline,
                        }
                    )

                messages.success(request, 'شارژ با موفقیت ویرایش شد.')
                return redirect('middle_add_area_charge')
        else:
            messages.error(request, 'خطا در ویرایش فرم. لطفا دوباره تلاش کنید.')
    else:
        form = FixChargeForm(instance=charge)

    return render(request, 'middleCharge/area_charge_template.html', {'form': form, 'charge': charge})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_area_charge_delete(request, pk):
    charge = get_object_or_404(AreaCharge, id=pk, user=request.user)

    content_type = ContentType.objects.get_for_model(AreaCharge)

    # بررسی اینکه هیچ رکورد UnifiedCharge با is_paid=True وجود نداشته باشد
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            is_paid=True
    ).exists():
        messages.error(request, "امکان حذف شارژ وجود ندارد چون پرداخت شارژ توسط واحد ثبت شده است.")
        return redirect(reverse('middle_add_area_charge'))

    # چک کردن وجود رکوردهایی که send_notification == True هستند
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            send_notification=True
    ).exists():
        messages.error(request, "برای این شارژ اطلاعیه صادر شده است. ابتدا اطلاعیه شارژ را حذف و مجدداً تلاش نمایید!")
        return redirect(reverse('middle_add_area_charge'))

    try:
        charge.delete()
        messages.success(request, f'{charge.name} با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, "امکان حذف این شارژ به دلیل وابستگی وجود ندارد!")

    return redirect(reverse('middle_add_area_charge'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_area_charge_notification_view(request, pk):
    charge = get_object_or_404(AreaCharge, id=pk, user=request.user)
    content_type = ContentType.objects.get_for_model(AreaCharge)

    # ------------------ واحدها ------------------
    units = Unit.objects.filter(
        is_active=True
    ).filter(
        Q(user=request.user) | Q(user__manager=request.user)
    ).distinct().order_by('unit')

    # ------------------ ساخت UnifiedCharge ------------------
    existing_ids = UnifiedCharge.objects.filter(
        content_type=content_type,
        object_id=charge.id
    ).values_list('unit_id', flat=True)

    new_units = units.exclude(id__in=existing_ids)

    calculator = CALCULATORS.get(charge.charge_type)

    with transaction.atomic():
        for unit in new_units:
            base = calculator.calculate(unit, charge)
            civil = charge.civil or 0
            other = charge.other_cost_amount or 0
            total = base + civil + other

            UnifiedCharge.objects.create(
                user=request.user,
                unit=unit,
                amount=base,
                house=unit.myhouse,
                base_charge=total,
                total_charge_month=total,
                title=charge.name,
                main_charge=charge,
                charge_type=charge.charge_type,
                penalty_percent=charge.payment_penalty_amount,
                civil=civil,
                other_cost_amount=other,
                payment_deadline_date=charge.payment_deadline,
                content_type=content_type,
                object_id=charge.id,
            )

    # ------------------ جستجو ------------------
    search_query = request.GET.get('search', '').strip()
    resident_type = request.GET.get('resident_type', '').strip()

    # -----------------------------
    # Search
    # -----------------------------
    if search_query:
        units = units.filter(
            Q(unit__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(renters__renter_name__icontains=search_query)
        ).distinct()

    # -----------------------------
    # Filter Resident Type
    # -----------------------------
    if resident_type == 'owner':

        # فقط مالک
        units = units.exclude(
            renters__renter_is_active=True
        )

    elif resident_type == 'renter':

        # فقط مستاجر فعال
        units = units.filter(
            renters__renter_is_active=True
        ).distinct()

    elif resident_type == 'empty':

        # واحد خالی
        units = units.filter(
            status_residence='empty'
        )

    # ------------------ pagination ------------------
    try:
        per_page = int(request.GET.get('per_page', 30))
    except ValueError:
        per_page = 30

    paginator = Paginator(units, per_page)
    page_units = paginator.get_page(request.GET.get('page'))

    # -----------------------------
    # SESSION SELECTION STATE
    # -----------------------------
    session_key = f"charge_{pk}_selected_units"
    selected_units = set(request.session.get(session_key, []))

    # ------------------ POST ------------------
    if request.method == "POST":

        send_type = request.POST.get("send_type", "notify")
        selected = request.POST.getlist("units")

        if not selected:
            messages.warning(request, 'هیچ واحدی انتخاب نشده')
            return redirect(request.path)

        qs = UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=selected
        ).select_related("unit")

        if not qs.exists():
            messages.info(request, 'اطلاعیه‌ای برای ارسال وجود ندارد')
            return redirect(request.path)

        # ثبت اعلان سیستمی
        qs.update(
            send_notification=True,
            send_notification_date=timezone.now().date()
        )

        # ---------- ارسال SMS ----------
        if send_type == "sms":

            result = SmsService.send_for_unified_charges(
                user=request.user,
                unified_charges=qs,
                meta_callback=lambda total_sms, total_price: qs.update(
                    send_sms=True,
                    send_sms_date=timezone.now().date(),
                    sms_count=total_sms,
                    sms_price=Decimal(settings.SMS_PRICE),
                    sms_total_price=total_price
                )
            )

            if result.success:
                messages.success(
                    request,
                    f'اطلاعیه سیستمی و پیامکی برای {qs.count()} واحد ارسال شد'
                )
            else:
                messages.error(request, result.message)

        else:
            messages.success(
                request,
                f'اطلاعیه سیستمی برای {qs.count()} واحد ثبت شد'
            )

            # پاکسازی انتخاب‌ها بعد از ارسال موفق
            request.session[session_key] = []
            request.session.modified = True
            return redirect(request.path)

    # ------------------ آماده‌سازی template ------------------
    unit_ids = [unit.id for unit in page_units.object_list]

    uc_map = {
        uc.unit_id: uc
        for uc in UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=unit_ids
        ).select_related('unit', 'unit__user')
    }

    # حالا uc_map همان uc_dict است
    for i, unit in enumerate(page_units):
        uc = uc_map.get(unit.id)
        renter = unit.renters.filter(renter_is_active=True).first()

        current_charge = uc.total_charge_month if uc else 0

        # جمع عددی بدهی‌های معوقه
        previous_debt_total = sum(uc.get_previous_debt_by_type().values()) if uc else 0

        total_payable = current_charge + previous_debt_total

        page_units.object_list[i] = {
            'unit': unit,
            'renter': renter,
            'is_paid': uc.is_paid if uc else False,
            'is_notified': uc.send_notification if uc else False,
            'send_sms': uc.send_sms if uc else False,
            'sms_date': uc.send_sms_date if uc else None,
            'current_charge': current_charge,
            'previous_debt': previous_debt_total,  # عددی
            'total_payable': total_payable,
        }

    return render(request, 'middleCharge/notify_area_charge_template.html', {
        "charge": charge,
        "page_obj": page_units,
        "selected_units": selected_units

        # "paginator": paginator,
    })


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_remove_send_notification_area(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'فقط درخواست‌های POST مجاز است.'}, status=400)

    charge = get_object_or_404(AreaCharge, id=pk, user=request.user)
    selected_units = request.POST.getlist('units[]')

    if not selected_units:
        return JsonResponse({'warning': 'هیچ واحدی انتخاب نشده است.'})

    try:
        with transaction.atomic():
            content_type = ContentType.objects.get_for_model(AreaCharge)

            # رکوردهایی که باید غیرفعال شوند
            if selected_units == ['all']:
                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )
            else:
                try:
                    selected_unit_ids = [int(uid) for uid in selected_units]
                except ValueError:
                    return JsonResponse({'error': 'شناسه واحد نامعتبر است.'}, status=400)

                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    unit_id__in=selected_unit_ids,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )

            updated_count = qs.update(
                send_notification=False,
                send_notification_date=None
            )

            # اگر هیچ رکوردی با send_notification=True باقی نماند → شارژ را غیرفعال کن
            if not UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    send_notification=True
            ).exists():
                charge.send_notification = False
                charge.save()

        session_key = f"charge_{pk}_selected_units"

        if updated_count:
            request.session.pop(session_key, None)
            request.session.modified = True

            return JsonResponse({
                'success': f'{updated_count} اطلاعیه غیرفعال شد.'
            })

        else:
            return JsonResponse({'info': 'رکوردی برای غیرفعال کردن یافت نشد.'})

    except Exception as e:
        return JsonResponse({'error': f'خطایی هنگام غیرفعال کردن اطلاعیه‌ها رخ داد: {str(e)}'}, status=500)


# ======================= Person Charge ===============
@method_decorator(middle_admin_required, name='dispatch')
class MiddlePersonChargeCreateView(CreateView):
    model = PersonCharge
    template_name = 'middleCharge/person_charge_template.html'
    form_class = PersonChargeForm
    success_url = reverse_lazy('middle_add_person_charge')

    def form_valid(self, form):
        charge_name = form.cleaned_data.get('name')

        # گرفتن کاربران تحت مدیریت
        managed_users = self.request.user.managed_users.all()

        unit_count = UnifiedCharge.objects.filter(
            user=self.request.user,
            unit__is_active=True
        ).values('unit').distinct().count()
        form.instance.unit_count = unit_count

        units = Unit.objects.filter(
            is_active=True
        ).filter(
            Q(user=self.request.user) | Q(user__in=managed_users)
        ).distinct()

        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('people_count'))['total'] or 0
        form.instance.total_people = total_people

        if not units.exists():
            messages.error(
                self.request,
                'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.'
            )
            return redirect('middle_manage_unit')

        # ذخیره FixCharge
        person_charge = form.save(commit=False)
        person_charge.user = self.request.user
        person_charge.name = charge_name
        person_charge.save()

        # انتخاب Calculator
        calculator = CALCULATORS.get(person_charge.charge_type)
        if not calculator:
            messages.error(self.request, 'نوع شارژ پشتیبانی نمی‌شود')
            return redirect(self.success_url)

        unified_charges = []

        for unit in units:
            # محاسبه مبلغ پایه برای هر واحد
            base_amount = calculator.calculate(unit, person_charge)

            # مطمئن شدن که فیلدهای عددی عدد هستند
            civil_amount = person_charge.civil or 0
            other_amount = person_charge.other_cost_amount or 0
            total_monthly_charge = base_amount + civil_amount + other_amount

            unified_charges.append(
                UnifiedCharge(
                    user=self.request.user,
                    unit=unit,
                    bank=None,
                    house=unit.myhouse,
                    charge_type=person_charge.charge_type,
                    amount=base_amount,
                    base_charge=total_monthly_charge,
                    main_charge=person_charge,
                    penalty_percent=person_charge.payment_penalty_amount,
                    civil=civil_amount,
                    other_cost_amount=other_amount,
                    penalty_amount=0,
                    total_charge_month=total_monthly_charge,
                    details=person_charge.details or '',
                    title=person_charge.name,
                    send_notification=False,  # ⛔ اعلان ارسال نشده
                    send_notification_date=None,
                    payment_deadline_date=person_charge.payment_deadline,
                    content_type=ContentType.objects.get_for_model(PersonCharge),
                    object_id=person_charge.id,
                )
            )

        # ایجاد همه UnifiedCharge ها در دیتابیس یکجا
        UnifiedCharge.objects.bulk_create(unified_charges)

        messages.success(self.request, 'شارژ با موفقیت ثبت گردید.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        managed_users = self.request.user.managed_users.all()
        unit_count = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).count()
        context['unit_count'] = unit_count
        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(
            total=Sum('people_count'))['total'] or 0
        context['total_people'] = total_people

        charges = PersonCharge.objects.filter(user=self.request.user).annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        context['charges'] = charges
        paginate_by = self.request.GET.get('paginate', '20')

        if paginate_by == '1000':  # نمایش همه
            paginator = Paginator(charges, charges.count() or 20)
        else:
            paginate_by = int(paginate_by)
            paginator = Paginator(charges, paginate_by)

        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['charges'] = page_obj
        context['page_obj'] = page_obj
        context['paginate'] = paginate_by

        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_person_charge_edit(request, pk):
    charge = get_object_or_404(PersonCharge, pk=pk, user=request.user)

    # بررسی اینکه برای این شارژ اعلان ارسال شده باشد
    any_notify = UnifiedCharge.objects.filter(
        content_type=ContentType.objects.get_for_model(PersonCharge),
        object_id=charge.id,
        send_notification=True
    ).exists()
    if any_notify:
        messages.error(request, 'برای این شارژ اعلان ارسال شده و قابل ویرایش نیست.')
        return redirect('middle_add_person_charge')

    if request.method == 'POST':
        form = PersonChargeForm(request.POST, request.FILES, instance=charge)
        if form.is_valid():
            with transaction.atomic():
                charge = form.save(commit=False)
                charge.name = charge.name or 'شارژ ثابت'
                charge.save()

                # کاربران و واحدهای تحت مدیریت
                managed_users = request.user.managed_users.all()
                units = Unit.objects.filter(is_active=True, user__in=managed_users)
                if not units.exists():
                    messages.error(request, 'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.')
                    return redirect('middle_manage_unit')

                # انتخاب Calculator
                calculator = CALCULATORS.get(charge.charge_type)
                if not calculator:
                    messages.error(request, 'نوع شارژ پشتیبانی نمی‌شود.')
                    return redirect('middle_add_person_charge')

                unified_charges = []

                for unit in units:
                    # محاسبه مبلغ پایه برای هر واحد
                    base_amount = calculator.calculate(unit, charge)
                    civil_amount = charge.civil or 0
                    other_amount = charge.other_cost_amount or 0
                    total_monthly = base_amount + civil_amount + other_amount

                    # آپدیت یا ایجاد UnifiedCharge برای این واحد
                    UnifiedCharge.objects.update_or_create(
                        user=request.user,
                        unit=unit,
                        content_type=ContentType.objects.get_for_model(PersonCharge),
                        object_id=charge.id,
                        defaults={
                            'bank': None,
                            'charge_type': charge.charge_type,
                            'amount': base_amount,
                            'base_charge': total_monthly,
                            'main-charge': charge,
                            'penalty_percent': charge.payment_penalty_amount or 0,
                            'civil': civil_amount,
                            'other_cost_amount': other_amount,
                            'penalty_amount': 0,
                            'total_charge_month': total_monthly,
                            'details': charge.details or '',
                            'title': charge.name,
                            'send_notification': False,
                            'send_notification_date': None,
                            'payment_deadline_date': charge.payment_deadline,
                        }
                    )

                messages.success(request, 'شارژ با موفقیت ویرایش شد.')
                return redirect('middle_add_person_charge')
        else:
            messages.error(request, 'خطا در ویرایش فرم. لطفا دوباره تلاش کنید.')
    else:
        form = FixChargeForm(instance=charge)

    return render(request, 'middleCharge/person_charge_template.html', {'form': form, 'charge': charge})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_person_charge_delete(request, pk):
    charge = get_object_or_404(PersonCharge, id=pk, user=request.user)

    content_type = ContentType.objects.get_for_model(PersonCharge)

    # بررسی اینکه هیچ رکورد UnifiedCharge با is_paid=True وجود نداشته باشد
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            is_paid=True
    ).exists():
        messages.error(request, "امکان حذف شارژ وجود ندارد چون پرداخت شارژ توسط واحد ثبت شده است.")
        return redirect(reverse('middle_add_person_charge'))

    # چک کردن وجود رکوردهایی که send_notification == True هستند
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            send_notification=True
    ).exists():
        messages.error(request, "برای این شارژ اطلاعیه صادر شده است. ابتدا اطلاعیه شارژ را حذف و مجدداً تلاش نمایید!")
        return redirect(reverse('middle_add_person_charge'))

    try:
        charge.delete()
        messages.success(request, f'{charge.name} با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, "امکان حذف این شارژ به دلیل وابستگی وجود ندارد!")

    return redirect(reverse('middle_add_person_charge'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_person_charge_notification_view(request, pk):
    charge = get_object_or_404(PersonCharge, id=pk, user=request.user)
    content_type = ContentType.objects.get_for_model(PersonCharge)

    # ------------------ واحدها ------------------
    units = Unit.objects.filter(
        is_active=True
    ).filter(
        Q(user=request.user) | Q(user__manager=request.user)
    ).distinct().order_by('unit')

    # ------------------ ساخت UnifiedCharge ------------------
    existing_ids = UnifiedCharge.objects.filter(
        content_type=content_type,
        object_id=charge.id
    ).values_list('unit_id', flat=True)

    new_units = units.exclude(id__in=existing_ids)

    calculator = CALCULATORS.get(charge.charge_type)

    with transaction.atomic():
        for unit in new_units:
            base = calculator.calculate(unit, charge)
            civil = charge.civil or 0
            other = charge.other_cost_amount or 0
            total = base + civil + other

            UnifiedCharge.objects.create(
                user=request.user,
                unit=unit,
                amount=base,
                base_charge=total,
                house=unit.myhouse,
                total_charge_month=total,
                title=charge.name,
                main_charge=charge,
                charge_type=charge.charge_type,
                penalty_percent=charge.payment_penalty_amount,
                civil=civil,
                other_cost_amount=other,
                payment_deadline_date=charge.payment_deadline,
                content_type=content_type,
                object_id=charge.id,
            )

    # ------------------ جستجو ------------------
    search_query = request.GET.get('search', '').strip()
    resident_type = request.GET.get('resident_type', '').strip()

    # -----------------------------
    # Search
    # -----------------------------
    if search_query:
        units = units.filter(
            Q(unit__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(renters__renter_name__icontains=search_query)
        ).distinct()

    # -----------------------------
    # Filter Resident Type
    # -----------------------------
    if resident_type == 'owner':

        # فقط مالک
        units = units.exclude(
            renters__renter_is_active=True
        )

    elif resident_type == 'renter':

        # فقط مستاجر فعال
        units = units.filter(
            renters__renter_is_active=True
        ).distinct()

    elif resident_type == 'empty':

        # واحد خالی
        units = units.filter(
            status_residence='empty'
        )

    # ------------------ Pagination ------------------
    try:
        per_page = int(request.GET.get('per_page', 30))
    except ValueError:
        per_page = 30

    paginator = Paginator(units, per_page)
    page_units = paginator.get_page(request.GET.get('page'))

    # -----------------------------
    # SESSION SELECTION STATE
    # -----------------------------
    session_key = f"charge_{pk}_selected_units"
    selected_units = set(request.session.get(session_key, []))

    # ------------------ POST: ارسال اطلاعیه یا پیامک ------------------
    if request.method == "POST":
        send_type = request.POST.get("send_type", "notify")
        selected = request.POST.getlist("units")

        if not selected:
            messages.warning(request, "هیچ واحدی انتخاب نشده")
            return redirect(request.path)

        qs = UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=selected
        ).select_related("unit")

        if not qs.exists():
            messages.info(request, "اطلاعیه‌ای برای ارسال وجود ندارد")
            return redirect(request.path)

        # ثبت اطلاعیه سیستمی
        qs.update(
            send_notification=True,
            send_notification_date=timezone.now().date()
        )

        # ---------- ارسال پیامک ----------
        if send_type == "sms":
            result = SmsService.send_for_unified_charges(
                user=request.user,
                unified_charges=qs,
                meta_callback=lambda total_sms, total_price: qs.update(
                    send_sms=True,
                    send_sms_date=timezone.now().date(),
                    sms_count=total_sms,
                    sms_price=Decimal(settings.SMS_PRICE),
                    sms_total_price=total_price
                )
            )

            if result.success:
                messages.success(
                    request,
                    f"اطلاعیه سیستمی و پیامکی برای {qs.count()} واحد ارسال شد"
                )
            else:
                messages.error(request, result.message)

        else:
            messages.success(
                request,
                f"اطلاعیه سیستمی برای {qs.count()} واحد ثبت شد"
            )

        # پاکسازی انتخاب‌ها بعد از ارسال موفق
        request.session[session_key] = []
        request.session.modified = True
        return redirect(request.path)

    # ------------------ آماده‌سازی template ------------------
    unit_ids = [unit.id for unit in page_units.object_list]

    uc_map = {
        uc.unit_id: uc
        for uc in UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=unit_ids
        ).select_related('unit', 'unit__user')
    }

    # حالا uc_map همان uc_dict است
    for i, unit in enumerate(page_units):
        uc = uc_map.get(unit.id)
        renter = unit.renters.filter(renter_is_active=True).first()

        current_charge = uc.total_charge_month if uc else 0

        # جمع عددی بدهی‌های معوقه
        previous_debt_total = sum(uc.get_previous_debt_by_type().values()) if uc else 0

        total_payable = current_charge + previous_debt_total

        page_units.object_list[i] = {
            'unit': unit,
            'renter': renter,
            'is_paid': uc.is_paid if uc else False,
            'is_notified': uc.send_notification if uc else False,
            'send_sms': uc.send_sms if uc else False,
            'sms_date': uc.send_sms_date if uc else None,
            'current_charge': current_charge,
            'previous_debt': previous_debt_total,  # عددی
            'total_payable': total_payable,
        }

    return render(request, "middleCharge/notify_person_charge_template.html", {
        "charge": charge,
        "page_obj": page_units,
        "selected_units": selected_units
        # "paginator": paginator,
    })


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_remove_send_notification_person(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'فقط درخواست‌های POST مجاز است.'}, status=400)

    charge = get_object_or_404(PersonCharge, id=pk, user=request.user)
    selected_units = request.POST.getlist('units[]')

    if not selected_units:
        return JsonResponse({'warning': 'هیچ واحدی انتخاب نشده است.'})

    try:
        with transaction.atomic():
            content_type = ContentType.objects.get_for_model(PersonCharge)

            # رکوردهایی که باید غیرفعال شوند
            if selected_units == ['all']:
                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )
            else:
                try:
                    selected_unit_ids = [int(uid) for uid in selected_units]
                except ValueError:
                    return JsonResponse({'error': 'شناسه واحد نامعتبر است.'}, status=400)

                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    unit_id__in=selected_unit_ids,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )

            updated_count = qs.update(
                send_notification=False,
                send_notification_date=None
            )

            # اگر هیچ رکوردی با send_notification=True باقی نماند → شارژ را غیرفعال کن
            if not UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    send_notification=True
            ).exists():
                charge.send_notification = False
                charge.save()

        session_key = f"charge_{pk}_selected_units"

        if updated_count:
            request.session.pop(session_key, None)
            request.session.modified = True

            return JsonResponse({
                'success': f'{updated_count} اطلاعیه غیرفعال شد.'
            })
        else:
            return JsonResponse({'info': 'رکوردی برای غیرفعال کردن یافت نشد.'})

    except Exception as e:
        return JsonResponse({'error': f'خطایی هنگام غیرفعال کردن اطلاعیه‌ها رخ داد: {str(e)}'}, status=500)


# ==================== Fix Area Charge    =============================
@method_decorator(middle_admin_required, name='dispatch')
class MiddleFixAreaChargeCreateView(CreateView):
    model = FixAreaCharge
    template_name = 'middleCharge/fix_area_charge_template.html'
    form_class = FixAreaChargeForm
    success_url = reverse_lazy('middle_add_fix_area_charge')

    def form_valid(self, form):
        charge_name = form.cleaned_data.get('name')

        # گرفتن کاربران تحت مدیریت
        managed_users = self.request.user.managed_users.all()

        unit_count = UnifiedCharge.objects.filter(
            user=self.request.user,
            unit__is_active=True
        ).values('unit').distinct().count()
        form.instance.unit_count = unit_count

        units = Unit.objects.filter(
            is_active=True
        ).filter(
            Q(user=self.request.user) | Q(user__in=managed_users)
        ).distinct()
        total_area = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('area'))['total'] or 0
        form.instance.total_area = total_area

        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('people_count'))['total'] or 0
        form.instance.total_people = total_people

        if not units.exists():
            messages.error(
                self.request,
                'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.'
            )
            return redirect('middle_manage_unit')

        # ذخیره FixCharge
        fix_area_charge = form.save(commit=False)
        fix_area_charge.user = self.request.user
        fix_area_charge.name = charge_name
        fix_area_charge.save()

        # انتخاب Calculator
        calculator = CALCULATORS.get(fix_area_charge.charge_type)
        if not calculator:
            messages.error(self.request, 'نوع شارژ پشتیبانی نمی‌شود')
            return redirect(self.success_url)

        unified_charges = []

        for unit in units:
            # محاسبه مبلغ پایه برای هر واحد
            base_amount = calculator.calculate(unit, fix_area_charge)

            # مطمئن شدن که فیلدهای عددی عدد هستند
            civil_amount = fix_area_charge.civil or 0
            other_amount = fix_area_charge.other_cost_amount or 0
            total_monthly_charge = base_amount + civil_amount + other_amount

            unified_charges.append(
                UnifiedCharge(
                    user=self.request.user,
                    unit=unit,
                    bank=None,
                    house=unit.myhouse,
                    charge_type=fix_area_charge.charge_type,
                    main_charge=fix_area_charge,
                    amount=base_amount,
                    base_charge=total_monthly_charge,
                    penalty_percent=fix_area_charge.payment_penalty_amount,
                    civil=civil_amount,
                    other_cost_amount=other_amount,
                    penalty_amount=0,
                    total_charge_month=total_monthly_charge,
                    details=fix_area_charge.details or '',
                    title=fix_area_charge.name,
                    send_notification=False,  # ⛔ اعلان ارسال نشده
                    send_notification_date=None,
                    payment_deadline_date=fix_area_charge.payment_deadline,
                    content_type=ContentType.objects.get_for_model(FixAreaCharge),
                    object_id=fix_area_charge.id,
                )
            )

        # ایجاد همه UnifiedCharge ها در دیتابیس یکجا
        UnifiedCharge.objects.bulk_create(unified_charges)

        messages.success(self.request, 'شارژ با موفقیت ثبت گردید.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        managed_users = self.request.user.managed_users.all()
        unit_count = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).count()
        context['unit_count'] = unit_count
        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(
            total=Sum('people_count'))['total'] or 0
        context['total_people'] = total_people
        context['total_area'] = \
            Unit.objects.filter(Q(user=self.request.user) | Q(user__in=managed_users), is_active=True).aggregate(
                total=Sum('area'))[
                'total'] or 0

        charges = FixAreaCharge.objects.filter(user=self.request.user).annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        context['charges'] = charges
        paginate_by = self.request.GET.get('paginate', '20')

        if paginate_by == '1000':  # نمایش همه
            paginator = Paginator(charges, charges.count() or 20)
        else:
            paginate_by = int(paginate_by)
            paginator = Paginator(charges, paginate_by)

        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['charges'] = page_obj
        context['page_obj'] = page_obj
        context['paginate'] = paginate_by
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_fix_area_charge_edit(request, pk):
    charge = get_object_or_404(FixAreaCharge, pk=pk, user=request.user)

    # بررسی اینکه برای این شارژ اعلان ارسال شده باشد
    any_notify = UnifiedCharge.objects.filter(
        content_type=ContentType.objects.get_for_model(FixAreaCharge),
        object_id=charge.id,
        send_notification=True
    ).exists()
    if any_notify:
        messages.error(request, 'برای این شارژ اعلان ارسال شده و قابل ویرایش نیست.')
        return redirect('middle_add_fix_area_charge')

    if request.method == 'POST':
        form = FixAreaChargeForm(request.POST, request.FILES, instance=charge)
        if form.is_valid():
            with transaction.atomic():
                charge = form.save(commit=False)
                charge.name = charge.name or 'شارژ ثابت'
                charge.save()

                # کاربران و واحدهای تحت مدیریت
                managed_users = request.user.managed_users.all()
                units = Unit.objects.filter(is_active=True, user__in=managed_users)
                if not units.exists():
                    messages.error(request, 'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.')
                    return redirect('middle_manage_unit')

                # انتخاب Calculator
                calculator = CALCULATORS.get(charge.charge_type)
                if not calculator:
                    messages.error(request, 'نوع شارژ پشتیبانی نمی‌شود.')
                    return redirect('middle_add_fix_area_charge')

                unified_charges = []

                for unit in units:
                    # محاسبه مبلغ پایه برای هر واحد
                    base_amount = calculator.calculate(unit, charge)
                    civil_amount = charge.civil or 0
                    other_amount = charge.other_cost_amount or 0
                    total_monthly = base_amount + civil_amount + other_amount

                    # آپدیت یا ایجاد UnifiedCharge برای این واحد
                    UnifiedCharge.objects.update_or_create(
                        user=request.user,
                        unit=unit,
                        house=unit.myhouse,
                        content_type=ContentType.objects.get_for_model(FixAreaCharge),
                        object_id=charge.id,
                        defaults={
                            'bank': None,
                            'charge_type': charge.charge_type,
                            'main_charge': charge,
                            'amount': base_amount,
                            'base_charge': total_monthly,
                            'penalty_percent': charge.payment_penalty_amount or 0,
                            'civil': civil_amount,
                            'other_cost_amount': other_amount,
                            'penalty_amount': 0,
                            'total_charge_month': total_monthly,
                            'details': charge.details or '',
                            'title': charge.name,
                            'send_notification': False,
                            'send_notification_date': None,
                            'payment_deadline_date': charge.payment_deadline,
                        }
                    )

                messages.success(request, 'شارژ با موفقیت ویرایش شد.')
                return redirect('middle_add_fix_area_charge')
        else:
            messages.error(request, 'خطا در ویرایش فرم. لطفا دوباره تلاش کنید.')
    else:
        form = FixChargeForm(instance=charge)

    return render(request, 'middleCharge/fix_area_charge_template.html', {'form': form, 'charge': charge})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_fix_area_charge_delete(request, pk):
    charge = get_object_or_404(FixAreaCharge, id=pk, user=request.user)

    content_type = ContentType.objects.get_for_model(FixAreaCharge)

    # بررسی اینکه هیچ رکورد UnifiedCharge با is_paid=True وجود نداشته باشد
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            is_paid=True
    ).exists():
        messages.error(request, "امکان حذف شارژ وجود ندارد چون پرداخت شارژ توسط واحد ثبت شده است.")
        return redirect(reverse('middle_add_fix_area_charge'))

    # چک کردن وجود رکوردهایی که send_notification == True هستند
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            send_notification=True
    ).exists():
        messages.error(request, "برای این شارژ اطلاعیه صادر شده است. ابتدا اطلاعیه شارژ را حذف و مجدداً تلاش نمایید!")
        return redirect(reverse('middle_add_fix_area_charge'))

    try:
        charge.delete()
        messages.success(request, f'{charge.name} با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, "امکان حذف این شارژ به دلیل وابستگی وجود ندارد!")

    return redirect(reverse('middle_add_fix_area_charge'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_show_fix_area_charge_notification_form(request, pk):
    charge = get_object_or_404(FixAreaCharge, id=pk, user=request.user)
    content_type = ContentType.objects.get_for_model(FixAreaCharge)

    # ------------------ واحدها ------------------
    units = Unit.objects.filter(
        is_active=True
    ).filter(
        Q(user=request.user) | Q(user__manager=request.user)
    ).distinct().order_by('unit')

    # ------------------ ساخت UnifiedCharge ------------------
    existing_ids = UnifiedCharge.objects.filter(
        content_type=content_type,
        object_id=charge.id
    ).values_list('unit_id', flat=True)

    new_units = units.exclude(id__in=existing_ids)

    calculator = CALCULATORS.get(charge.charge_type)

    with transaction.atomic():
        for unit in new_units:
            base = calculator.calculate(unit, charge)
            civil = charge.civil or 0
            other = charge.other_cost_amount or 0
            total = base + civil + other

            UnifiedCharge.objects.create(
                user=request.user,
                unit=unit,
                amount=base,
                house=unit.myhouse,
                base_charge=total,
                total_charge_month=total,
                title=charge.name,
                main_charge=charge,
                charge_type=charge.charge_type,
                penalty_percent=charge.payment_penalty_amount,
                civil=civil,
                other_cost_amount=other,
                payment_deadline_date=charge.payment_deadline,
                content_type=content_type,
                object_id=charge.id,
            )

    # جستجو
    search_query = request.GET.get('search', '').strip()
    resident_type = request.GET.get('resident_type', '').strip()

    # -----------------------------
    # Search
    # -----------------------------
    if search_query:
        units = units.filter(
            Q(unit__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(renters__renter_name__icontains=search_query)
        ).distinct()

    # -----------------------------
    # Filter Resident Type
    # -----------------------------
    if resident_type == 'owner':

        # فقط مالک
        units = units.exclude(
            renters__renter_is_active=True
        )

    elif resident_type == 'renter':

        # فقط مستاجر فعال
        units = units.filter(
            renters__renter_is_active=True
        ).distinct()

    elif resident_type == 'empty':

        # واحد خالی
        units = units.filter(
            status_residence='empty'
        )

    # ------------------ Pagination ------------------
    try:
        per_page = int(request.GET.get('per_page', 30))
    except ValueError:
        per_page = 30

    paginator = Paginator(units, per_page)
    page_units = paginator.get_page(request.GET.get('page'))

    # -----------------------------
    # SESSION SELECTION STATE
    # -----------------------------
    session_key = f"charge_{pk}_selected_units"
    selected_units = set(request.session.get(session_key, []))

    # ------------------ POST: ارسال اطلاعیه یا پیامک ------------------
    if request.method == "POST":
        send_type = request.POST.get("send_type", "notify")
        selected = request.POST.getlist("units")

        if not selected:
            messages.warning(request, "هیچ واحدی انتخاب نشده")
            return redirect(request.path)

        qs = UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=selected
        ).select_related("unit")

        if not qs.exists():
            messages.info(request, "اطلاعیه‌ای برای ارسال وجود ندارد")
            return redirect(request.path)

        # ثبت اطلاعیه سیستمی
        qs.update(
            send_notification=True,
            send_notification_date=timezone.now().date()
        )

        # ---------- ارسال پیامک ----------
        if send_type == "sms":
            result = SmsService.send_for_unified_charges(
                user=request.user,
                unified_charges=qs,
                meta_callback=lambda total_sms, total_price: qs.update(
                    send_sms=True,
                    send_sms_date=timezone.now().date(),
                    sms_count=total_sms,
                    sms_price=Decimal(settings.SMS_PRICE),
                    sms_total_price=total_price
                )
            )

            if result.success:
                messages.success(
                    request,
                    f"اطلاعیه سیستمی و پیامکی برای {qs.count()} واحد ارسال شد"
                )
            else:
                messages.error(request, result.message)

        else:
            messages.success(
                request,
                f"اطلاعیه سیستمی برای {qs.count()} واحد ثبت شد"
            )

            # پاکسازی انتخاب‌ها بعد از ارسال موفق
        request.session[session_key] = []
        request.session.modified = True
        return redirect(request.path)

        # ------------------ آماده‌سازی template ------------------
    unit_ids = [unit.id for unit in page_units.object_list]

    uc_map = {
        uc.unit_id: uc
        for uc in UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=unit_ids
        ).select_related('unit', 'unit__user')
    }

    # حالا uc_map همان uc_dict است
    for i, unit in enumerate(page_units):
        uc = uc_map.get(unit.id)
        renter = unit.renters.filter(renter_is_active=True).first()

        current_charge = uc.total_charge_month if uc else 0

        # جمع عددی بدهی‌های معوقه
        previous_debt_total = sum(uc.get_previous_debt_by_type().values()) if uc else 0

        total_payable = current_charge + previous_debt_total

        page_units.object_list[i] = {
            'unit': unit,
            'renter': renter,
            'is_paid': uc.is_paid if uc else False,
            'is_notified': uc.send_notification if uc else False,
            'send_sms': uc.send_sms if uc else False,
            'sms_date': uc.send_sms_date if uc else None,
            'current_charge': current_charge,
            'previous_debt': previous_debt_total,  # عددی
            'total_payable': total_payable,
        }

    context = {
        'charge': charge,
        'page_obj': page_units,
        "selected_units": selected_units
        # 'paginator': paginator,
    }

    return render(request, 'middleCharge/notify_area_fix_charge_template.html', context)


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_remove_send_notification_fix_area(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'فقط درخواست‌های POST مجاز است.'}, status=400)

    charge = get_object_or_404(FixAreaCharge, id=pk, user=request.user)
    selected_units = request.POST.getlist('units[]')

    if not selected_units:
        return JsonResponse({'warning': 'هیچ واحدی انتخاب نشده است.'})

    try:
        with transaction.atomic():
            content_type = ContentType.objects.get_for_model(FixAreaCharge)

            # رکوردهایی که باید غیرفعال شوند
            if selected_units == ['all']:
                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )
            else:
                try:
                    selected_unit_ids = [int(uid) for uid in selected_units]
                except ValueError:
                    return JsonResponse({'error': 'شناسه واحد نامعتبر است.'}, status=400)

                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    unit_id__in=selected_unit_ids,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )

            updated_count = qs.update(
                send_notification=False,
                send_notification_date=None
            )

            # اگر هیچ رکوردی با send_notification=True باقی نماند → شارژ را غیرفعال کن
            if not UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    send_notification=True
            ).exists():
                charge.send_notification = False
                charge.save()

        session_key = f"charge_{pk}_selected_units"

        if updated_count:
            request.session.pop(session_key, None)
            request.session.modified = True

            return JsonResponse({
                'success': f'{updated_count} اطلاعیه غیرفعال شد.'
            })
        else:
            return JsonResponse({'info': 'رکوردی برای غیرفعال کردن یافت نشد.'})

    except Exception as e:
        return JsonResponse({'error': f'خطایی هنگام غیرفعال کردن اطلاعیه‌ها رخ داد: {str(e)}'}, status=500)


# ======================= Fix Person Charge  ==========================
@method_decorator(middle_admin_required, name='dispatch')
class MiddleFixPersonChargeCreateView(CreateView):
    model = FixPersonCharge
    template_name = 'middleCharge/fix_person_charge_template.html'
    form_class = FixPersonChargeForm
    success_url = reverse_lazy('middle_add_fix_person_charge')

    def form_valid(self, form):
        charge_name = form.cleaned_data.get('name')

        # گرفتن کاربران تحت مدیریت
        managed_users = self.request.user.managed_users.all()

        unit_count = UnifiedCharge.objects.filter(
            user=self.request.user,
            unit__is_active=True
        ).values('unit').distinct().count()
        form.instance.unit_count = unit_count

        units = Unit.objects.filter(
            is_active=True
        ).filter(
            Q(user=self.request.user) | Q(user__in=managed_users)
        ).distinct()
        total_area = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('area'))['total'] or 0
        form.instance.total_area = total_area

        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('people_count'))['total'] or 0
        form.instance.total_people = total_people

        if not units.exists():
            messages.error(
                self.request,
                'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.'
            )
            return redirect('middle_manage_unit')

        # ذخیره FixCharge
        fix_person_charge = form.save(commit=False)
        fix_person_charge.user = self.request.user
        fix_person_charge.name = charge_name
        fix_person_charge.save()

        # انتخاب Calculator
        calculator = CALCULATORS.get(fix_person_charge.charge_type)
        if not calculator:
            messages.error(self.request, 'نوع شارژ پشتیبانی نمی‌شود')
            return redirect(self.success_url)

        unified_charges = []

        for unit in units:
            # محاسبه مبلغ پایه برای هر واحد
            base_amount = calculator.calculate(unit, fix_person_charge)
            # مطمئن شدن که فیلدهای عددی عدد هستند
            civil_amount = fix_person_charge.civil or 0
            other_amount = fix_person_charge.other_cost_amount or 0
            total_monthly_charge = base_amount + civil_amount + other_amount

            unified_charges.append(
                UnifiedCharge(
                    user=self.request.user,
                    unit=unit,
                    bank=None,
                    house=unit.myhouse,
                    charge_type=fix_person_charge.charge_type,
                    main_charge=fix_person_charge,
                    amount=base_amount,
                    base_charge=total_monthly_charge,
                    penalty_percent=fix_person_charge.payment_penalty_amount,
                    civil=civil_amount,
                    other_cost_amount=other_amount,
                    penalty_amount=0,
                    total_charge_month=total_monthly_charge,
                    details=fix_person_charge.details or '',
                    title=fix_person_charge.name,
                    send_notification=False,  # ⛔ اعلان ارسال نشده
                    send_notification_date=None,
                    payment_deadline_date=fix_person_charge.payment_deadline,
                    content_type=ContentType.objects.get_for_model(FixPersonCharge),
                    object_id=fix_person_charge.id,
                )
            )

        # ایجاد همه UnifiedCharge ها در دیتابیس یکجا
        UnifiedCharge.objects.bulk_create(unified_charges)

        messages.success(self.request, 'شارژ با موفقیت ثبت گردید.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        managed_users = self.request.user.managed_users.all()
        unit_count = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).count()
        context['unit_count'] = unit_count
        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(
            total=Sum('people_count'))['total'] or 0
        context['total_people'] = total_people
        context['total_area'] = \
            Unit.objects.filter(Q(user=self.request.user) | Q(user__in=managed_users), is_active=True).aggregate(
                total=Sum('area'))[
                'total'] or 0

        charges = FixPersonCharge.objects.filter(user=self.request.user).annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        context['charges'] = charges
        paginate_by = self.request.GET.get('paginate', '20')

        if paginate_by == '1000':  # نمایش همه
            paginator = Paginator(charges, charges.count() or 20)
        else:
            paginate_by = int(paginate_by)
            paginator = Paginator(charges, paginate_by)

        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['charges'] = page_obj
        context['page_obj'] = page_obj
        context['paginate'] = paginate_by
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_fix_person_charge_edit(request, pk):
    charge = get_object_or_404(FixPersonCharge, pk=pk, user=request.user)

    # بررسی اینکه برای این شارژ اعلان ارسال شده باشد
    any_notify = UnifiedCharge.objects.filter(
        content_type=ContentType.objects.get_for_model(FixPersonCharge),
        object_id=charge.id,
        send_notification=True
    ).exists()
    if any_notify:
        messages.error(request, 'برای این شارژ اعلان ارسال شده و قابل ویرایش نیست.')
        return redirect('middle_add_fix_person_charge')

    if request.method == 'POST':
        form = FixPersonChargeForm(request.POST, request.FILES, instance=charge)
        if form.is_valid():
            with transaction.atomic():
                charge = form.save(commit=False)
                charge.name = charge.name or 'شارژ ثابت'
                charge.save()

                # کاربران و واحدهای تحت مدیریت
                managed_users = request.user.managed_users.all()
                units = Unit.objects.filter(is_active=True, user__in=managed_users)
                if not units.exists():
                    messages.error(request, 'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.')
                    return redirect('middle_manage_unit')

                # انتخاب Calculator
                calculator = CALCULATORS.get(charge.charge_type)
                if not calculator:
                    messages.error(request, 'نوع شارژ پشتیبانی نمی‌شود.')
                    return redirect('middle_add_fix_person_charge')

                unified_charges = []

                for unit in units:
                    # محاسبه مبلغ پایه برای هر واحد
                    base_amount = calculator.calculate(unit, charge)
                    civil_amount = charge.civil or 0
                    other_amount = charge.other_cost_amount or 0
                    total_monthly = base_amount + civil_amount + other_amount

                    # آپدیت یا ایجاد UnifiedCharge برای این واحد
                    UnifiedCharge.objects.update_or_create(
                        user=request.user,
                        unit=unit,
                        house=unit.myhouse,
                        content_type=ContentType.objects.get_for_model(FixPersonCharge),
                        object_id=charge.id,
                        defaults={
                            'bank': None,
                            'charge_type': charge.charge_type,
                            'main_charge': charge,
                            'amount': base_amount,
                            'base_charge': base_amount,
                            'penalty_percent': charge.payment_penalty_amount or 0,
                            'civil': civil_amount,
                            'other_cost_amount': other_amount,
                            'penalty_amount': 0,
                            'total_charge_month': total_monthly,
                            'details': charge.details or '',
                            'title': charge.name,
                            'send_notification': False,
                            'send_notification_date': None,
                            'payment_deadline_date': charge.payment_deadline,
                        }
                    )

                messages.success(request, 'شارژ با موفقیت ویرایش شد.')
                return redirect('middle_add_fix_person_charge')
        else:
            messages.error(request, 'خطا در ویرایش فرم. لطفا دوباره تلاش کنید.')
    else:
        form = FixChargeForm(instance=charge)

    return render(request, 'middleCharge/fix_person_charge_template.html', {'form': form, 'charge': charge})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_fix_person_charge_delete(request, pk):
    charge = get_object_or_404(FixPersonCharge, id=pk, user=request.user)
    content_type = ContentType.objects.get_for_model(FixPersonCharge)

    # بررسی اینکه هیچ رکورد UnifiedCharge با is_paid=True وجود نداشته باشد
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            is_paid=True
    ).exists():
        messages.error(request, "امکان حذف شارژ وجود ندارد چون پرداخت شارژ توسط واحد ثبت شده است.")
        return redirect(reverse('middle_add_fix_person_charge'))

    # چک کردن وجود رکوردهایی که send_notification == True هستند
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            send_notification=True
    ).exists():
        messages.error(request, "برای این شارژ اطلاعیه صادر شده است. ابتدا اطلاعیه شارژ را حذف و مجدداً تلاش نمایید!")
        return redirect(reverse('middle_add_fix_person_charge'))

    try:
        charge.delete()
        messages.success(request, f'{charge.name} با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, "امکان حذف این شارژ به دلیل وابستگی وجود ندارد!")

    return redirect(reverse('middle_add_fix_person_charge'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_show_fix_person_charge_notification_form(request, pk):
    charge = get_object_or_404(FixPersonCharge, id=pk, user=request.user)
    content_type = ContentType.objects.get_for_model(FixPersonCharge)

    # ------------------ واحدها ------------------
    units = Unit.objects.filter(
        is_active=True
    ).filter(
        Q(user=request.user) | Q(user__manager=request.user)
    ).distinct().order_by('unit')

    # ------------------ ساخت UnifiedCharge ------------------
    existing_ids = UnifiedCharge.objects.filter(
        content_type=content_type,
        object_id=charge.id
    ).values_list('unit_id', flat=True)

    new_units = units.exclude(id__in=existing_ids)

    calculator = CALCULATORS.get(charge.charge_type)

    with transaction.atomic():
        for unit in new_units:
            base = calculator.calculate(unit, charge)
            civil = charge.civil or 0
            other = charge.other_cost_amount or 0
            total = base + civil + other

            UnifiedCharge.objects.create(
                user=request.user,
                unit=unit,
                amount=base,
                base_charge=total,
                house=unit.myhouse,
                total_charge_month=total,
                title=charge.name,
                main_charge=charge,
                charge_type=charge.charge_type,
                penalty_percent=charge.payment_penalty_amount,
                civil=civil,
                other_cost_amount=other,
                payment_deadline_date=charge.payment_deadline,
                content_type=content_type,
                object_id=charge.id,
            )

    # جستجو
    search_query = request.GET.get('search', '').strip()
    resident_type = request.GET.get('resident_type', '').strip()

    # -----------------------------
    # Search
    # -----------------------------
    if search_query:
        units = units.filter(
            Q(unit__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(renters__renter_name__icontains=search_query)
        ).distinct()

    # -----------------------------
    # Filter Resident Type
    # -----------------------------
    if resident_type == 'owner':

        # فقط مالک
        units = units.exclude(
            renters__renter_is_active=True
        )

    elif resident_type == 'renter':

        # فقط مستاجر فعال
        units = units.filter(
            renters__renter_is_active=True
        ).distinct()

    elif resident_type == 'empty':

        # واحد خالی
        units = units.filter(
            status_residence='empty'
        )

        # ------------------ Pagination ------------------
    try:
        per_page = int(request.GET.get('per_page', 30))
    except ValueError:
        per_page = 30

    paginator = Paginator(units, per_page)
    page_units = paginator.get_page(request.GET.get('page'))
    # -----------------------------
    # SESSION SELECTION STATE
    # -----------------------------
    session_key = f"charge_{pk}_selected_units"
    selected_units = set(request.session.get(session_key, []))

    # ------------------ POST: ارسال اطلاعیه یا پیامک ------------------
    if request.method == "POST":
        send_type = request.POST.get("send_type", "notify")
        selected = request.POST.getlist("units")

        if not selected:
            messages.warning(request, "هیچ واحدی انتخاب نشده")
            return redirect(request.path)

        qs = UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=selected
        ).select_related("unit")

        if not qs.exists():
            messages.info(request, "اطلاعیه‌ای برای ارسال وجود ندارد")
            return redirect(request.path)

        # ثبت اطلاعیه سیستمی
        qs.update(
            send_notification=True,
            send_notification_date=timezone.now().date()
        )

        # ---------- ارسال پیامک ----------
        if send_type == "sms":
            result = SmsService.send_for_unified_charges(
                user=request.user,
                unified_charges=qs,
                meta_callback=lambda total_sms, total_price: qs.update(
                    send_sms=True,
                    send_sms_date=timezone.now().date(),
                    sms_count=total_sms,
                    sms_price=Decimal(settings.SMS_PRICE),
                    sms_total_price=total_price
                )
            )

            if result.success:
                messages.success(
                    request,
                    f"اطلاعیه سیستمی و پیامکی برای {qs.count()} واحد ارسال شد"
                )
            else:
                messages.error(request, result.message)

        else:
            messages.success(
                request,
                f"اطلاعیه سیستمی برای {qs.count()} واحد ثبت شد"
            )

            # پاکسازی انتخاب‌ها بعد از ارسال موفق
        request.session[session_key] = []
        request.session.modified = True
        return redirect(request.path)

        # ------------------ آماده‌سازی template ------------------
    unit_ids = [unit.id for unit in page_units.object_list]

    uc_map = {
        uc.unit_id: uc
        for uc in UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=unit_ids
        ).select_related('unit', 'unit__user')
    }

    # حالا uc_map همان uc_dict است
    for i, unit in enumerate(page_units):
        uc = uc_map.get(unit.id)
        renter = unit.renters.filter(renter_is_active=True).first()

        current_charge = uc.total_charge_month if uc else 0

        # جمع عددی بدهی‌های معوقه
        previous_debt_total = sum(uc.get_previous_debt_by_type().values()) if uc else 0

        total_payable = current_charge + previous_debt_total

        page_units.object_list[i] = {
            'unit': unit,
            'renter': renter,
            'is_paid': uc.is_paid if uc else False,
            'is_notified': uc.send_notification if uc else False,
            'send_sms': uc.send_sms if uc else False,
            'sms_date': uc.send_sms_date if uc else None,
            'current_charge': current_charge,
            'previous_debt': previous_debt_total,  # عددی
            'total_payable': total_payable,
        }

    context = {
        'charge': charge,
        'page_obj': page_units,  # حالا فقط واحدهای دارای UnifiedCharge هستند
        "selected_units": selected_units
        # 'paginator': paginator,
    }
    return render(request, 'middleCharge/notify_person_fix_charge_template.html', context)


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_remove_send_notification_fix_person(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'فقط درخواست‌های POST مجاز است.'}, status=400)

    charge = get_object_or_404(FixPersonCharge, id=pk, user=request.user)
    selected_units = request.POST.getlist('units[]')

    if not selected_units:
        return JsonResponse({'warning': 'هیچ واحدی انتخاب نشده است.'})

    try:
        with transaction.atomic():
            content_type = ContentType.objects.get_for_model(FixPersonCharge)

            # رکوردهایی که باید غیرفعال شوند
            if selected_units == ['all']:
                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )
            else:
                try:
                    selected_unit_ids = [int(uid) for uid in selected_units]
                except ValueError:
                    return JsonResponse({'error': 'شناسه واحد نامعتبر است.'}, status=400)

                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    unit_id__in=selected_unit_ids,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )

            updated_count = qs.update(
                send_notification=False,
                send_notification_date=None
            )

            # اگر هیچ رکوردی با send_notification=True باقی نماند → شارژ را غیرفعال کن
            if not UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    send_notification=True
            ).exists():
                charge.send_notification = False
                charge.save()

        session_key = f"charge_{pk}_selected_units"

        if updated_count:
            request.session.pop(session_key, None)
            request.session.modified = True

            return JsonResponse({
                'success': f'{updated_count} اطلاعیه غیرفعال شد.'
            })
        else:
            return JsonResponse({'info': 'رکوردی برای غیرفعال کردن یافت نشد.'})

    except Exception as e:
        return JsonResponse({'error': f'خطایی هنگام غیرفعال کردن اطلاعیه‌ها رخ داد: {str(e)}'}, status=500)


# ============================== Person Area Charge ============================
@method_decorator(middle_admin_required, name='dispatch')
class MiddlePersonAreaChargeCreateView(CreateView):
    model = ChargeByPersonArea
    template_name = 'middleCharge/person_area_charge_template.html'
    form_class = PersonAreaChargeForm
    success_url = reverse_lazy('middle_add_person_area_charge')

    def form_valid(self, form):
        charge_name = form.cleaned_data.get('name')

        # گرفتن کاربران تحت مدیریت
        managed_users = self.request.user.managed_users.all()

        unit_count = UnifiedCharge.objects.filter(
            user=self.request.user,
            unit__is_active=True
        ).values('unit').distinct().count()
        form.instance.unit_count = unit_count

        units = Unit.objects.filter(
            is_active=True
        ).filter(
            Q(user=self.request.user) | Q(user__in=managed_users)
        ).distinct()
        total_area = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('area'))['total'] or 0
        form.instance.total_area = total_area

        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('people_count'))['total'] or 0
        form.instance.total_people = total_people

        if not units.exists():
            messages.error(
                self.request,
                'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.'
            )
            return redirect('middle_manage_unit')

        # ذخیره FixCharge
        person_area_charge = form.save(commit=False)
        person_area_charge.user = self.request.user
        person_area_charge.name = charge_name
        person_area_charge.save()

        # انتخاب Calculator
        calculator = CALCULATORS.get(person_area_charge.charge_type)
        if not calculator:
            messages.error(self.request, 'نوع شارژ پشتیبانی نمی‌شود')
            return redirect(self.success_url)

        unified_charges = []

        for unit in units:
            # محاسبه مبلغ پایه برای هر واحد
            base_amount = calculator.calculate(unit, person_area_charge)
            # مطمئن شدن که فیلدهای عددی عدد هستند
            civil_amount = person_area_charge.civil or 0
            other_amount = person_area_charge.other_cost_amount or 0
            total_monthly_charge = base_amount + civil_amount + other_amount

            unified_charges.append(
                UnifiedCharge(
                    user=self.request.user,
                    unit=unit,
                    bank=None,
                    house=unit.myhouse,
                    charge_type=person_area_charge.charge_type,
                    main_charge=person_area_charge,
                    amount=base_amount,
                    base_charge=total_monthly_charge,
                    penalty_percent=person_area_charge.payment_penalty_amount,
                    civil=civil_amount,
                    other_cost_amount=other_amount,
                    penalty_amount=0,
                    total_charge_month=total_monthly_charge,
                    details=person_area_charge.details or '',
                    title=person_area_charge.name,
                    send_notification=False,  # ⛔ اعلان ارسال نشده
                    send_notification_date=None,
                    payment_deadline_date=person_area_charge.payment_deadline,
                    content_type=ContentType.objects.get_for_model(ChargeByPersonArea),
                    object_id=person_area_charge.id,
                )
            )

        # ایجاد همه UnifiedCharge ها در دیتابیس یکجا
        UnifiedCharge.objects.bulk_create(unified_charges)

        messages.success(self.request, 'شارژ با موفقیت ثبت گردید.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        managed_users = self.request.user.managed_users.all()
        unit_count = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).count()
        context['unit_count'] = unit_count
        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(
            total=Sum('people_count'))['total'] or 0
        context['total_people'] = total_people
        context['total_area'] = \
            Unit.objects.filter(Q(user=self.request.user) | Q(user__in=managed_users), is_active=True).aggregate(
                total=Sum('area'))[
                'total'] or 0

        charges = ChargeByPersonArea.objects.filter(user=self.request.user).annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        context['charges'] = charges
        paginate_by = self.request.GET.get('paginate', '20')

        if paginate_by == '1000':  # نمایش همه
            paginator = Paginator(charges, charges.count() or 20)
        else:
            paginate_by = int(paginate_by)
            paginator = Paginator(charges, paginate_by)

        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['charges'] = page_obj
        context['page_obj'] = page_obj
        context['paginate'] = paginate_by
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_person_area_charge_edit(request, pk):
    charge = get_object_or_404(ChargeByPersonArea, pk=pk, user=request.user)

    # بررسی اینکه برای این شارژ اعلان ارسال شده باشد
    any_notify = UnifiedCharge.objects.filter(
        content_type=ContentType.objects.get_for_model(ChargeByPersonArea),
        object_id=charge.id,
        send_notification=True
    ).exists()
    if any_notify:
        messages.error(request, 'برای این شارژ اعلان ارسال شده و قابل ویرایش نیست.')
        return redirect('middle_add_person_area_charge')

    if request.method == 'POST':
        form = PersonAreaChargeForm(request.POST, request.FILES, instance=charge)
        if form.is_valid():
            with transaction.atomic():
                charge = form.save(commit=False)
                charge.name = charge.name or 'شارژ ثابت'
                charge.save()

                # کاربران و واحدهای تحت مدیریت
                managed_users = request.user.managed_users.all()
                units = Unit.objects.filter(is_active=True, user__in=managed_users)
                if not units.exists():
                    messages.error(request, 'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.')
                    return redirect('middle_manage_unit')

                # انتخاب Calculator
                calculator = CALCULATORS.get(charge.charge_type)
                if not calculator:
                    messages.error(request, 'نوع شارژ پشتیبانی نمی‌شود.')
                    return redirect('middle_add_fix_person_charge')

                unified_charges = []

                for unit in units:
                    # محاسبه مبلغ پایه برای هر واحد
                    base_amount = calculator.calculate(unit, charge)
                    civil_amount = charge.civil or 0
                    other_amount = charge.other_cost_amount or 0
                    total_monthly = base_amount + civil_amount + other_amount

                    # آپدیت یا ایجاد UnifiedCharge برای این واحد
                    UnifiedCharge.objects.update_or_create(
                        user=request.user,
                        house=unit.myhouse,
                        unit=unit,
                        content_type=ContentType.objects.get_for_model(ChargeByPersonArea),
                        object_id=charge.id,
                        defaults={
                            'bank': None,
                            'charge_type': charge.charge_type,
                            'main_charge': charge,
                            'amount': base_amount,
                            'base_charge': total_monthly,
                            'penalty_percent': charge.payment_penalty_amount or 0,
                            'civil': civil_amount,
                            'other_cost_amount': other_amount,
                            'penalty_amount': 0,
                            'total_charge_month': total_monthly,
                            'details': charge.details or '',
                            'title': charge.name,
                            'send_notification': False,
                            'send_notification_date': None,
                            'payment_deadline_date': charge.payment_deadline,
                        }
                    )

                messages.success(request, 'شارژ با موفقیت ویرایش شد.')
                return redirect('middle_add_person_area_charge')
        else:
            messages.error(request, 'خطا در ویرایش فرم. لطفا دوباره تلاش کنید.')
    else:
        form = FixChargeForm(instance=charge)

    return render(request, 'middleCharge/person_area_charge_template.html', {'form': form, 'charge': charge})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_person_area_charge_delete(request, pk):
    charge = get_object_or_404(ChargeByPersonArea, id=pk, user=request.user)
    content_type = ContentType.objects.get_for_model(ChargeByPersonArea)

    # بررسی اینکه هیچ رکورد UnifiedCharge با is_paid=True وجود نداشته باشد
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            is_paid=True
    ).exists():
        messages.error(request, "امکان حذف شارژ وجود ندارد چون پرداخت شارژ توسط واحد ثبت شده است.")
        return redirect(reverse('middle_add_person_area_charge'))

    # چک کردن وجود رکوردهایی که send_notification == True هستند
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            send_notification=True
    ).exists():
        messages.error(request, "برای این شارژ اطلاعیه صادر شده است. ابتدا اطلاعیه شارژ را حذف و مجدداً تلاش نمایید!")
        return redirect(reverse('middle_add_person_area_charge'))

    try:
        charge.delete()
        messages.success(request, f'{charge.name} با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, "امکان حذف این شارژ به دلیل وابستگی وجود ندارد!")

    return redirect(reverse('middle_add_person_area_charge'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_show_person_area_charge_notification_form(request, pk):
    charge = get_object_or_404(ChargeByPersonArea, id=pk, user=request.user)
    content_type = ContentType.objects.get_for_model(ChargeByPersonArea)

    # ------------------ واحدها ------------------
    units = Unit.objects.filter(
        is_active=True
    ).filter(
        Q(user=request.user) | Q(user__manager=request.user)
    ).distinct().order_by('unit')

    # ------------------ ساخت UnifiedCharge ------------------
    existing_ids = UnifiedCharge.objects.filter(
        content_type=content_type,
        object_id=charge.id
    ).values_list('unit_id', flat=True)

    new_units = units.exclude(id__in=existing_ids)

    calculator = CALCULATORS.get(charge.charge_type)

    with transaction.atomic():
        for unit in new_units:
            base = calculator.calculate(unit, charge)
            civil = charge.civil or 0
            other = charge.other_cost_amount or 0
            total = base + civil + other

            UnifiedCharge.objects.create(
                user=request.user,
                unit=unit,
                amount=base,
                house=unit.myhouse,
                base_charge=total,
                total_charge_month=total,
                title=charge.name,
                main_charge=charge,
                charge_type=charge.charge_type,
                penalty_percent=charge.payment_penalty_amount,
                civil=civil,
                other_cost_amount=other,
                payment_deadline_date=charge.payment_deadline,
                content_type=content_type,
                object_id=charge.id,
            )

    # جستجو
    search_query = request.GET.get('search', '').strip()
    resident_type = request.GET.get('resident_type', '').strip()

    # -----------------------------
    # Search
    # -----------------------------
    if search_query:
        units = units.filter(
            Q(unit__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(renters__renter_name__icontains=search_query)
        ).distinct()

    # -----------------------------
    # Filter Resident Type
    # -----------------------------
    if resident_type == 'owner':

        # فقط مالک
        units = units.exclude(
            renters__renter_is_active=True
        )

    elif resident_type == 'renter':

        # فقط مستاجر فعال
        units = units.filter(
            renters__renter_is_active=True
        ).distinct()

    elif resident_type == 'empty':

        # واحد خالی
        units = units.filter(
            status_residence='empty'
        )

    # ------------------ Pagination ------------------
    try:
        per_page = int(request.GET.get('per_page', 30))
    except ValueError:
        per_page = 30

    paginator = Paginator(units, per_page)
    page_units = paginator.get_page(request.GET.get('page'))

    # -----------------------------
    # SESSION SELECTION STATE
    # -----------------------------
    session_key = f"charge_{pk}_selected_units"
    selected_units = set(request.session.get(session_key, []))

    # ------------------ POST: ارسال اطلاعیه یا پیامک ------------------
    if request.method == "POST":
        send_type = request.POST.get("send_type", "notify")
        selected = request.POST.getlist("units")

        if not selected:
            messages.warning(request, "هیچ واحدی انتخاب نشده")
            return redirect(request.path)

        qs = UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=selected
        ).select_related("unit")

        if not qs.exists():
            messages.info(request, "اطلاعیه‌ای برای ارسال وجود ندارد")
            return redirect(request.path)

        # ثبت اطلاعیه سیستمی
        qs.update(
            send_notification=True,
            send_notification_date=timezone.now().date()
        )

        # ---------- ارسال پیامک ----------
        if send_type == "sms":
            result = SmsService.send_for_unified_charges(
                user=request.user,
                unified_charges=qs,
                meta_callback=lambda total_sms, total_price: qs.update(
                    send_sms=True,
                    send_sms_date=timezone.now().date(),
                    sms_count=total_sms,
                    sms_price=Decimal(settings.SMS_PRICE),
                    sms_total_price=total_price
                )
            )

            if result.success:
                messages.success(
                    request,
                    f"اطلاعیه سیستمی و پیامکی برای {qs.count()} واحد ارسال شد"
                )
            else:
                messages.error(request, result.message)

        else:
            messages.success(
                request,
                f"اطلاعیه سیستمی برای {qs.count()} واحد ثبت شد"
            )

            # پاکسازی انتخاب‌ها بعد از ارسال موفق
        request.session[session_key] = []
        request.session.modified = True
        return redirect(request.path)

        # ------------------ آماده‌سازی template ------------------
    unit_ids = [unit.id for unit in page_units.object_list]

    uc_map = {
        uc.unit_id: uc
        for uc in UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=unit_ids
        ).select_related('unit', 'unit__user')
    }

    # حالا uc_map همان uc_dict است
    for i, unit in enumerate(page_units):
        uc = uc_map.get(unit.id)
        renter = unit.renters.filter(renter_is_active=True).first()

        current_charge = uc.total_charge_month if uc else 0

        # جمع عددی بدهی‌های معوقه
        previous_debt_total = sum(uc.get_previous_debt_by_type().values()) if uc else 0

        total_payable = current_charge + previous_debt_total

        page_units.object_list[i] = {
            'unit': unit,
            'renter': renter,
            'is_paid': uc.is_paid if uc else False,
            'is_notified': uc.send_notification if uc else False,
            'send_sms': uc.send_sms if uc else False,
            'sms_date': uc.send_sms_date if uc else None,
            'current_charge': current_charge,
            'previous_debt': previous_debt_total,  # عددی
            'total_payable': total_payable,
        }
    context = {
        'charge': charge,
        'page_obj': page_units,  # حالا فقط واحدهای دارای UnifiedCharge هستند
        "selected_units": selected_units
        # 'paginator': paginator,
    }
    return render(request, 'middleCharge/notify_person_area_charge_template.html', context)


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_remove_send_notification_person_area(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'فقط درخواست‌های POST مجاز است.'}, status=400)

    charge = get_object_or_404(ChargeByPersonArea, id=pk, user=request.user)
    selected_units = request.POST.getlist('units[]')

    if not selected_units:
        return JsonResponse({'warning': 'هیچ واحدی انتخاب نشده است.'})

    try:
        with transaction.atomic():
            content_type = ContentType.objects.get_for_model(ChargeByPersonArea)

            # رکوردهایی که باید غیرفعال شوند
            if selected_units == ['all']:
                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )
            else:
                try:
                    selected_unit_ids = [int(uid) for uid in selected_units]
                except ValueError:
                    return JsonResponse({'error': 'شناسه واحد نامعتبر است.'}, status=400)

                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    unit_id__in=selected_unit_ids,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )

            updated_count = qs.update(
                send_notification=False,
                send_notification_date=None
            )

            # اگر هیچ رکوردی با send_notification=True باقی نماند → شارژ را غیرفعال کن
            if not UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    send_notification=True
            ).exists():
                charge.send_notification = False
                charge.save()

        session_key = f"charge_{pk}_selected_units"

        if updated_count:
            request.session.pop(session_key, None)
            request.session.modified = True

            return JsonResponse({
                'success': f'{updated_count} اطلاعیه غیرفعال شد.'
            })
        else:
            return JsonResponse({'info': 'رکوردی برای غیرفعال کردن یافت نشد.'})

    except Exception as e:
        return JsonResponse({'error': f'خطایی هنگام غیرفعال کردن اطلاعیه‌ها رخ داد: {str(e)}'}, status=500)


# ==========================Fix Person Area Charge ================================
@method_decorator(middle_admin_required, name='dispatch')
class MiddlePersonAreaFixChargeCreateView(CreateView):
    model = ChargeByFixPersonArea
    template_name = 'middleCharge/person_area_fix_charge_template.html'
    form_class = PersonAreaFixChargeForm
    success_url = reverse_lazy('middle_add_person_area_fix_charge')

    def form_valid(self, form):
        charge_name = form.cleaned_data.get('name')

        # گرفتن کاربران تحت مدیریت
        managed_users = self.request.user.managed_users.all()

        unit_count = UnifiedCharge.objects.filter(
            user=self.request.user,
            unit__is_active=True
        ).values('unit').distinct().count()
        form.instance.unit_count = unit_count

        units = Unit.objects.filter(
            is_active=True
        ).filter(
            Q(user=self.request.user) | Q(user__in=managed_users)
        ).distinct()
        total_area = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('area'))['total'] or 0
        form.instance.total_area = total_area

        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('people_count'))['total'] or 0
        form.instance.total_people = total_people

        if not units.exists():
            messages.error(
                self.request,
                'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.'
            )
            return redirect('middle_manage_unit')

        # ذخیره FixCharge
        fix_person_area_charge = form.save(commit=False)
        fix_person_area_charge.user = self.request.user
        fix_person_area_charge.name = charge_name
        fix_person_area_charge.save()

        # انتخاب Calculator
        calculator = CALCULATORS.get(fix_person_area_charge.charge_type)
        if not calculator:
            messages.error(self.request, 'نوع شارژ پشتیبانی نمی‌شود')
            return redirect(self.success_url)

        unified_charges = []

        for unit in units:
            # محاسبه مبلغ پایه برای هر واحد
            base_amount = calculator.calculate(unit, fix_person_area_charge)
            # مطمئن شدن که فیلدهای عددی عدد هستند
            civil_amount = fix_person_area_charge.civil or 0
            other_amount = fix_person_area_charge.other_cost_amount or 0
            total_monthly_charge = base_amount + civil_amount + other_amount

            unified_charges.append(
                UnifiedCharge(
                    user=self.request.user,
                    unit=unit,
                    bank=None,
                    house=unit.myhouse,
                    charge_type=fix_person_area_charge.charge_type,
                    main_charge=fix_person_area_charge,
                    amount=base_amount,
                    base_charge=total_monthly_charge,
                    penalty_percent=fix_person_area_charge.payment_penalty_amount,
                    civil=civil_amount,
                    other_cost_amount=other_amount,
                    penalty_amount=0,
                    total_charge_month=total_monthly_charge,
                    details=fix_person_area_charge.details or '',
                    title=fix_person_area_charge.name,
                    send_notification=False,  # ⛔ اعلان ارسال نشده
                    send_notification_date=None,
                    payment_deadline_date=fix_person_area_charge.payment_deadline,
                    content_type=ContentType.objects.get_for_model(ChargeByFixPersonArea),
                    object_id=fix_person_area_charge.id,
                )
            )

        # ایجاد همه UnifiedCharge ها در دیتابیس یکجا
        UnifiedCharge.objects.bulk_create(unified_charges)

        messages.success(self.request, 'شارژ با موفقیت ثبت گردید.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        managed_users = self.request.user.managed_users.all()
        unit_count = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).count()
        context['unit_count'] = unit_count
        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(
            total=Sum('people_count'))['total'] or 0
        context['total_people'] = total_people
        context['total_area'] = \
            Unit.objects.filter(Q(user=self.request.user) | Q(user__in=managed_users), is_active=True).aggregate(
                total=Sum('area'))[
                'total'] or 0

        charges = ChargeByFixPersonArea.objects.filter(user=self.request.user).annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        context['charges'] = charges
        paginate_by = self.request.GET.get('paginate', '20')

        if paginate_by == '1000':  # نمایش همه
            paginator = Paginator(charges, charges.count() or 20)
        else:
            paginate_by = int(paginate_by)
            paginator = Paginator(charges, paginate_by)

        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['charges'] = page_obj
        context['page_obj'] = page_obj
        context['paginate'] = paginate_by
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_person_area_fix_charge_edit(request, pk):
    charge = get_object_or_404(ChargeByFixPersonArea, pk=pk, user=request.user)

    # بررسی اینکه برای این شارژ اعلان ارسال شده باشد
    any_notify = UnifiedCharge.objects.filter(
        content_type=ContentType.objects.get_for_model(ChargeByFixPersonArea),
        object_id=charge.id,
        send_notification=True
    ).exists()
    if any_notify:
        messages.error(request, 'برای این شارژ اعلان ارسال شده و قابل ویرایش نیست.')
        return redirect('middle_add_person_area_fix_charge')

    if request.method == 'POST':
        form = PersonAreaFixChargeForm(request.POST, request.FILES, instance=charge)
        if form.is_valid():
            with transaction.atomic():
                charge = form.save(commit=False)
                charge.name = charge.name or 'شارژ ثابت'
                charge.save()

                # کاربران و واحدهای تحت مدیریت
                managed_users = request.user.managed_users.all()
                units = Unit.objects.filter(is_active=True, user__in=managed_users)
                if not units.exists():
                    messages.error(request, 'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.')
                    return redirect('middle_manage_unit')

                # انتخاب Calculator
                calculator = CALCULATORS.get(charge.charge_type)
                if not calculator:
                    messages.error(request, 'نوع شارژ پشتیبانی نمی‌شود.')
                    return redirect('middle_add_person_area_fix_charge')

                unified_charges = []

                for unit in units:
                    # محاسبه مبلغ پایه برای هر واحد
                    base_amount = calculator.calculate(unit, charge)
                    civil_amount = charge.civil or 0
                    other_amount = charge.other_cost_amount or 0
                    total_monthly = base_amount + civil_amount + other_amount

                    # آپدیت یا ایجاد UnifiedCharge برای این واحد
                    UnifiedCharge.objects.update_or_create(
                        user=request.user,
                        unit=unit,
                        house=unit.myhouse,
                        content_type=ContentType.objects.get_for_model(ChargeByFixPersonArea),
                        object_id=charge.id,
                        defaults={
                            'bank': None,
                            'charge_type': charge.charge_type,
                            'main_charge': charge,
                            'amount': base_amount,
                            'base_charge': total_monthly,
                            'penalty_percent': charge.payment_penalty_amount or 0,
                            'civil': civil_amount,
                            'other_cost_amount': other_amount,
                            'penalty_amount': 0,
                            'total_charge_month': total_monthly,
                            'details': charge.details or '',
                            'title': charge.name,
                            'send_notification': False,
                            'send_notification_date': None,
                            'payment_deadline_date': charge.payment_deadline,
                        }
                    )

                messages.success(request, 'شارژ با موفقیت ویرایش شد.')
                return redirect('middle_add_person_area_fix_charge')
        else:
            messages.error(request, 'خطا در ویرایش فرم. لطفا دوباره تلاش کنید.')
    else:
        form = FixChargeForm(instance=charge)
        return render(request, 'middleCharge/person_area_fix_charge_template.html',
                      {'form': form, 'charge': charge})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_person_area_fix_delete(request, pk):
    charge = get_object_or_404(ChargeByFixPersonArea, id=pk, user=request.user)
    content_type = ContentType.objects.get_for_model(ChargeByFixPersonArea)

    # بررسی اینکه هیچ رکورد UnifiedCharge با is_paid=True وجود نداشته باشد
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            is_paid=True
    ).exists():
        messages.error(request, "امکان حذف شارژ وجود ندارد چون پرداخت شارژ توسط واحد ثبت شده است.")
        return redirect(reverse('middle_add_person_area_fix_charge'))

    # چک کردن وجود رکوردهایی که send_notification == True هستند
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            send_notification=True
    ).exists():
        messages.error(request, "برای این شارژ اطلاعیه صادر شده است. ابتدا اطلاعیه شارژ را حذف و مجدداً تلاش نمایید!")
        return redirect(reverse('middle_add_person_area_fix_charge'))

    try:
        charge.delete()
        messages.success(request, f'{charge.name} با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, "امکان حذف این شارژ به دلیل وابستگی وجود ندارد!")

    return redirect(reverse('middle_add_person_area_fix_charge'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_show_fix_person_area_charge_notification_form(request, pk):
    charge = get_object_or_404(ChargeByFixPersonArea, id=pk, user=request.user)
    content_type = ContentType.objects.get_for_model(ChargeByFixPersonArea)

    # ------------------ واحدها ------------------
    units = Unit.objects.filter(
        is_active=True
    ).filter(
        Q(user=request.user) | Q(user__manager=request.user)
    ).distinct().order_by('unit')

    # ------------------ ساخت UnifiedCharge ------------------
    existing_ids = UnifiedCharge.objects.filter(
        content_type=content_type,
        object_id=charge.id
    ).values_list('unit_id', flat=True)

    new_units = units.exclude(id__in=existing_ids)

    calculator = CALCULATORS.get(charge.charge_type)

    with transaction.atomic():
        for unit in new_units:
            base = calculator.calculate(unit, charge)
            civil = charge.civil or 0
            other = charge.other_cost_amount or 0
            total = base + civil + other

            UnifiedCharge.objects.create(
                user=request.user,
                unit=unit,
                amount=base,
                house=unit.myhouse,
                base_charge=total,
                total_charge_month=total,
                title=charge.name,
                main_charge=charge,
                charge_type=charge.charge_type,
                penalty_percent=charge.payment_penalty_amount,
                civil=civil,
                other_cost_amount=other,
                payment_deadline_date=charge.payment_deadline,
                content_type=content_type,
                object_id=charge.id,
            )

    # جستجو
    search_query = request.GET.get('search', '').strip()
    resident_type = request.GET.get('resident_type', '').strip()

    # -----------------------------
    # Search
    # -----------------------------
    if search_query:
        units = units.filter(
            Q(unit__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(renters__renter_name__icontains=search_query)
        ).distinct()

    # -----------------------------
    # Filter Resident Type
    # -----------------------------
    if resident_type == 'owner':

        # فقط مالک
        units = units.exclude(
            renters__renter_is_active=True
        )

    elif resident_type == 'renter':

        # فقط مستاجر فعال
        units = units.filter(
            renters__renter_is_active=True
        ).distinct()

    elif resident_type == 'empty':

        # واحد خالی
        units = units.filter(
            status_residence='empty'
        )

    # ------------------ Pagination ------------------
    try:
        per_page = int(request.GET.get('per_page', 30))
    except ValueError:
        per_page = 30

    paginator = Paginator(units, per_page)
    page_units = paginator.get_page(request.GET.get('page'))

    # -----------------------------
    # SESSION SELECTION STATE
    # -----------------------------
    session_key = f"charge_{pk}_selected_units"
    selected_units = set(request.session.get(session_key, []))

    # ------------------ POST: ارسال اطلاعیه یا پیامک ------------------
    if request.method == "POST":
        send_type = request.POST.get("send_type", "notify")
        selected = request.POST.getlist("units")

        if not selected:
            messages.warning(request, "هیچ واحدی انتخاب نشده")
            return redirect(request.path)

        qs = UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=selected
        ).select_related("unit")

        if not qs.exists():
            messages.info(request, "اطلاعیه‌ای برای ارسال وجود ندارد")
            return redirect(request.path)

        # ثبت اطلاعیه سیستمی
        qs.update(
            send_notification=True,
            send_notification_date=timezone.now().date()
        )

        # ---------- ارسال پیامک ----------
        if send_type == "sms":
            result = SmsService.send_for_unified_charges(
                user=request.user,
                unified_charges=qs,
                meta_callback=lambda total_sms, total_price: qs.update(
                    send_sms=True,
                    send_sms_date=timezone.now().date(),
                    sms_count=total_sms,
                    sms_price=Decimal(settings.SMS_PRICE),
                    sms_total_price=total_price
                )
            )

            if result.success:
                messages.success(
                    request,
                    f"اطلاعیه سیستمی و پیامکی برای {qs.count()} واحد ارسال شد"
                )
            else:
                messages.error(request, result.message)

        else:
            messages.success(
                request,
                f"اطلاعیه سیستمی برای {qs.count()} واحد ثبت شد"
            )

            # پاکسازی انتخاب‌ها بعد از ارسال موفق
        request.session[session_key] = []
        request.session.modified = True
        return redirect(request.path)

        # ------------------ آماده‌سازی template ------------------
    unit_ids = [unit.id for unit in page_units.object_list]

    uc_map = {
        uc.unit_id: uc
        for uc in UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=unit_ids
        ).select_related('unit', 'unit__user')
    }

    # حالا uc_map همان uc_dict است
    for i, unit in enumerate(page_units):
        uc = uc_map.get(unit.id)
        renter = unit.renters.filter(renter_is_active=True).first()

        current_charge = uc.total_charge_month if uc else 0

        # جمع عددی بدهی‌های معوقه
        previous_debt_total = sum(uc.get_previous_debt_by_type().values()) if uc else 0

        total_payable = current_charge + previous_debt_total

        page_units.object_list[i] = {
            'unit': unit,
            'renter': renter,
            'is_paid': uc.is_paid if uc else False,
            'is_notified': uc.send_notification if uc else False,
            'send_sms': uc.send_sms if uc else False,
            'sms_date': uc.send_sms_date if uc else None,
            'current_charge': current_charge,
            'previous_debt': previous_debt_total,  # عددی
            'total_payable': total_payable,
        }

    context = {
        'charge': charge,
        'page_obj': page_units,  # حالا فقط واحدهای دارای UnifiedCharge هستند
        "selected_units": selected_units
        # 'paginator': paginator,
    }
    return render(request, 'middleCharge/notify_fix_person_area_charge_template.html', context)


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_remove_send_notification_fix_person_area(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'فقط درخواست‌های POST مجاز است.'}, status=400)

    charge = get_object_or_404(ChargeByFixPersonArea, id=pk, user=request.user)
    selected_units = request.POST.getlist('units[]')

    if not selected_units:
        return JsonResponse({'warning': 'هیچ واحدی انتخاب نشده است.'})

    try:
        with transaction.atomic():
            content_type = ContentType.objects.get_for_model(ChargeByFixPersonArea)

            # رکوردهایی که باید غیرفعال شوند
            if selected_units == ['all']:
                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )
            else:
                try:
                    selected_unit_ids = [int(uid) for uid in selected_units]
                except ValueError:
                    return JsonResponse({'error': 'شناسه واحد نامعتبر است.'}, status=400)

                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    unit_id__in=selected_unit_ids,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )

            updated_count = qs.update(
                send_notification=False,
                send_notification_date=None
            )

            # اگر هیچ رکوردی با send_notification=True باقی نماند → شارژ را غیرفعال کن
            if not UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    send_notification=True
            ).exists():
                charge.send_notification = False
                charge.save()

        session_key = f"charge_{pk}_selected_units"

        if updated_count:
            request.session.pop(session_key, None)
            request.session.modified = True

            return JsonResponse({
                'success': f'{updated_count} اطلاعیه غیرفعال شد.'
            })
        else:
            return JsonResponse({'info': 'رکوردی برای غیرفعال کردن یافت نشد.'})

    except Exception as e:
        return JsonResponse({'error': f'خطایی هنگام غیرفعال کردن اطلاعیه‌ها رخ داد: {str(e)}'}, status=500)


# =========================ّFix Variable Charge =================================
@method_decorator(middle_admin_required, name='dispatch')
class MiddleVariableFixChargeCreateView(CreateView):
    model = ChargeFixVariable
    template_name = 'middleCharge/variable_fix_charge_template.html'
    form_class = VariableFixChargeForm
    success_url = reverse_lazy('middle_add_variable_fix_charge')

    def form_valid(self, form):
        charge_name = form.cleaned_data.get('name')

        # گرفتن کاربران تحت مدیریت
        managed_users = self.request.user.managed_users.all()

        unit_count = UnifiedCharge.objects.filter(
            user=self.request.user,
            unit__is_active=True
        ).values('unit').distinct().count()
        form.instance.unit_count = unit_count

        units = Unit.objects.filter(
            is_active=True
        ).filter(
            Q(user=self.request.user) | Q(user__in=managed_users)
        ).distinct()
        total_area = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('area'))['total'] or 0
        form.instance.total_area = total_area

        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(total=Sum('people_count'))['total'] or 0
        form.instance.total_people = total_people

        if not units.exists():
            messages.error(
                self.request,
                'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.'
            )
            return redirect('middle_manage_unit')

        # ذخیره FixCharge
        fix_variable = form.save(commit=False)
        fix_variable.user = self.request.user
        fix_variable.name = charge_name
        fix_variable.save()

        # انتخاب Calculator
        calculator = CALCULATORS.get(fix_variable.charge_type)
        if not calculator:
            messages.error(self.request, 'نوع شارژ پشتیبانی نمی‌شود')
            return redirect(self.success_url)

        unified_charges = []

        for unit in units:
            # محاسبه مبلغ پایه برای هر واحد
            base_amount = calculator.calculate(unit, fix_variable)
            # مطمئن شدن که فیلدهای عددی عدد هستند
            civil_amount = fix_variable.civil or 0
            other_amount = fix_variable.other_cost_amount or 0
            parking_count = getattr(unit, 'parking_counts', 0) or 0

            parking_price = fix_variable.extra_parking_amount or 0
            parking_total = parking_count * parking_price

            total_monthly_charge = base_amount + civil_amount + other_amount + parking_total

            unified_charges.append(
                UnifiedCharge(
                    user=self.request.user,
                    unit=unit,
                    bank=None,
                    house=unit.myhouse,
                    charge_type=fix_variable.charge_type,
                    extra_parking_price=parking_total,
                    main_charge=fix_variable,
                    amount=base_amount,
                    base_charge=total_monthly_charge,
                    penalty_percent=fix_variable.payment_penalty_amount,
                    civil=civil_amount,
                    other_cost_amount=other_amount,
                    penalty_amount=0,
                    total_charge_month=total_monthly_charge,
                    details=fix_variable.details or '',
                    title=fix_variable.name,
                    send_notification=False,  # ⛔ اعلان ارسال نشده
                    send_notification_date=None,
                    payment_deadline_date=fix_variable.payment_deadline,
                    content_type=ContentType.objects.get_for_model(ChargeFixVariable),
                    object_id=fix_variable.id,
                )
            )

        # ایجاد همه UnifiedCharge ها در دیتابیس یکجا
        UnifiedCharge.objects.bulk_create(unified_charges)

        messages.success(self.request, 'شارژ با موفقیت ثبت گردید.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        managed_users = self.request.user.managed_users.all()
        unit_count = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).count()
        context['unit_count'] = unit_count
        total_people = Unit.objects.filter(
            Q(user=self.request.user) | Q(user__in=managed_users),
            is_active=True,
        ).aggregate(
            total=Sum('people_count'))['total'] or 0
        context['total_people'] = total_people
        context['total_area'] = \
            Unit.objects.filter(Q(user=self.request.user) | Q(user__in=managed_users), is_active=True).aggregate(
                total=Sum('area'))[
                'total'] or 0

        charges = ChargeFixVariable.objects.filter(user=self.request.user).annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        context['charges'] = charges
        paginate_by = self.request.GET.get('paginate', '20')

        if paginate_by == '1000':  # نمایش همه
            paginator = Paginator(charges, charges.count() or 20)
        else:
            paginate_by = int(paginate_by)
            paginator = Paginator(charges, paginate_by)

        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['charges'] = page_obj
        context['page_obj'] = page_obj
        context['paginate'] = paginate_by
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_variable_fix_charge_edit(request, pk):
    charge = get_object_or_404(ChargeFixVariable, pk=pk, user=request.user)

    # بررسی اینکه برای این شارژ اعلان ارسال شده باشد
    any_notify = UnifiedCharge.objects.filter(
        content_type=ContentType.objects.get_for_model(ChargeFixVariable),
        object_id=charge.id,
        send_notification=True
    ).exists()
    if any_notify:
        messages.error(request, 'برای این شارژ اعلان ارسال شده و قابل ویرایش نیست.')
        return redirect('middle_add_variable_fix_charge')

    if request.method == 'POST':
        form = VariableFixChargeForm(request.POST, request.FILES, instance=charge)
        if form.is_valid():
            with transaction.atomic():
                charge = form.save(commit=False)
                charge.name = charge.name or 'شارژ ثابت'
                charge.save()

                # کاربران و واحدهای تحت مدیریت
                managed_users = request.user.managed_users.all()
                units = Unit.objects.filter(is_active=True, user__in=managed_users)
                if not units.exists():
                    messages.error(request, 'هیچ واحد فعالی یافت نشد. لطفا ابتدا واحدها را ثبت کنید.')
                    return redirect('middle_manage_unit')

                # انتخاب Calculator
                calculator = CALCULATORS.get(charge.charge_type)
                if not calculator:
                    messages.error(request, 'نوع شارژ پشتیبانی نمی‌شود.')
                    return redirect('middle_add_variable_fix_charge')

                unified_charges = []

                for unit in units:
                    # محاسبه مبلغ پایه برای هر واحد
                    base_amount = calculator.calculate(unit, charge)
                    civil_amount = charge.civil or 0
                    other_amount = charge.other_cost_amount or 0
                    parking_count = getattr(unit, 'parking_counts', 0) or 0
                    parking_price = charge.extra_parking_amount or 0
                    parking_total = parking_count * parking_price

                    total_monthly = base_amount + civil_amount + other_amount + parking_total

                    # آپدیت یا ایجاد UnifiedCharge برای این واحد
                    UnifiedCharge.objects.update_or_create(
                        user=request.user,
                        unit=unit,
                        house=unit.myhouse,
                        content_type=ContentType.objects.get_for_model(ChargeFixVariable),
                        object_id=charge.id,
                        defaults={
                            'bank': None,
                            'charge_type': charge.charge_type,
                            'main_charge': charge,
                            'amount': base_amount,
                            'base_charge': total_monthly,
                            'extra_parking_price': parking_total,
                            'penalty_percent': charge.payment_penalty_amount or 0,
                            'civil': civil_amount,
                            'other_cost_amount': other_amount,
                            'penalty_amount': 0,
                            'total_charge_month': total_monthly,
                            'details': charge.details or '',
                            'title': charge.name,
                            'send_notification': False,
                            'send_notification_date': None,
                            'payment_deadline_date': charge.payment_deadline,
                        }
                    )

                messages.success(request, 'شارژ با موفقیت ویرایش شد.')
                return redirect('middle_add_variable_fix_charge')
        else:
            messages.error(request, 'خطا در ویرایش فرم. لطفا دوباره تلاش کنید.')
    else:
        form = FixChargeForm(instance=charge)
        return render(request, 'middleCharge/variable_fix_charge_template.html',
                      {'form': form, 'charge': charge})


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_variable_fix_charge_delete(request, pk):
    charge = get_object_or_404(ChargeFixVariable, id=pk, user=request.user)
    content_type = ContentType.objects.get_for_model(ChargeFixVariable)

    # بررسی اینکه هیچ رکورد UnifiedCharge با is_paid=True وجود نداشته باشد
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            is_paid=True
    ).exists():
        messages.error(request, "امکان حذف شارژ وجود ندارد چون پرداخت شارژ توسط واحد ثبت شده است.")
        return redirect(reverse('middle_add_variable_fix_charge'))

    # چک کردن وجود رکوردهایی که send_notification == True هستند
    if UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            send_notification=True
    ).exists():
        messages.error(request, "برای این شارژ اطلاعیه صادر شده است. ابتدا اطلاعیه شارژ را حذف و مجدداً تلاش نمایید!")
        return redirect(reverse('middle_add_variable_fix_charge'))

    try:
        charge.delete()
        messages.success(request, f'{charge.name} با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, "امکان حذف این شارژ به دلیل وابستگی وجود ندارد!")

    return redirect(reverse('middle_add_variable_fix_charge'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_show_fix_variable_notification_form(request, pk):
    charge = get_object_or_404(ChargeFixVariable, id=pk, user=request.user)
    content_type = ContentType.objects.get_for_model(ChargeFixVariable)

    # ------------------ واحدها ------------------
    units = Unit.objects.filter(
        is_active=True
    ).filter(
        Q(user=request.user) | Q(user__manager=request.user)
    ).distinct().order_by('unit')

    # ------------------ ساخت UnifiedCharge ------------------
    existing_ids = UnifiedCharge.objects.filter(
        content_type=content_type,
        object_id=charge.id
    ).values_list('unit_id', flat=True)

    new_units = units.exclude(id__in=existing_ids)

    calculator = CALCULATORS.get(charge.charge_type)

    with transaction.atomic():
        for unit in new_units:
            base = calculator.calculate(unit, charge)
            civil = charge.civil or 0
            other = charge.other_cost_amount or 0
            total = base + civil + other

            UnifiedCharge.objects.create(
                user=request.user,
                unit=unit,
                amount=base,
                house=unit.myhouse,
                base_charge=total,
                total_charge_month=total,
                title=charge.name,
                main_charge=charge,
                charge_type=charge.charge_type,
                penalty_percent=charge.payment_penalty_amount,
                civil=civil,
                other_cost_amount=other,
                payment_deadline_date=charge.payment_deadline,
                content_type=content_type,
                object_id=charge.id,
            )

    # جستجو
    search_query = request.GET.get('search', '').strip()
    resident_type = request.GET.get('resident_type', '').strip()

    # -----------------------------
    # Search
    # -----------------------------
    if search_query:
        units = units.filter(
            Q(unit__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(renters__renter_name__icontains=search_query)
        ).distinct()

    # -----------------------------
    # Filter Resident Type
    # -----------------------------
    if resident_type == 'owner':

        # فقط مالک
        units = units.exclude(
            renters__renter_is_active=True
        )

    elif resident_type == 'renter':

        # فقط مستاجر فعال
        units = units.filter(
            renters__renter_is_active=True
        ).distinct()

    elif resident_type == 'empty':

        # واحد خالی
        units = units.filter(
            status_residence='empty'
        )

    # ------------------ Pagination ------------------
    try:
        per_page = int(request.GET.get('per_page', 30))
    except ValueError:
        per_page = 30

    paginator = Paginator(units, per_page)
    page_units = paginator.get_page(request.GET.get('page'))

    # -----------------------------
    # SESSION SELECTION STATE
    # -----------------------------
    session_key = f"charge_{pk}_selected_units"
    selected_units = set(request.session.get(session_key, []))

    # ------------------ POST: ارسال اطلاعیه یا پیامک ------------------
    if request.method == "POST":
        send_type = request.POST.get("send_type", "notify")
        selected = request.POST.getlist("units")

        if not selected:
            messages.warning(request, "هیچ واحدی انتخاب نشده")
            return redirect(request.path)

        qs = UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=selected
        ).select_related("unit")

        if not qs.exists():
            messages.info(request, "اطلاعیه‌ای برای ارسال وجود ندارد")
            return redirect(request.path)

        # ثبت اطلاعیه سیستمی
        qs.update(
            send_notification=True,
            send_notification_date=timezone.now().date()
        )

        # ---------- ارسال پیامک ----------
        if send_type == "sms":
            result = SmsService.send_for_unified_charges(
                user=request.user,
                unified_charges=qs,
                meta_callback=lambda total_sms, total_price: qs.update(
                    send_sms=True,
                    send_sms_date=timezone.now().date(),
                    sms_count=total_sms,
                    sms_price=Decimal(settings.SMS_PRICE),
                    sms_total_price=total_price
                )
            )

            if result.success:
                messages.success(
                    request,
                    f"اطلاعیه سیستمی و پیامکی برای {qs.count()} واحد ارسال شد"
                )
            else:
                messages.error(request, result.message)

        else:
            messages.success(
                request,
                f"اطلاعیه سیستمی برای {qs.count()} واحد ثبت شد"
            )

            # پاکسازی انتخاب‌ها بعد از ارسال موفق
        request.session[session_key] = []
        request.session.modified = True
        return redirect(request.path)

        # ------------------ آماده‌سازی template ------------------
    unit_ids = [unit.id for unit in page_units.object_list]

    uc_map = {
        uc.unit_id: uc
        for uc in UnifiedCharge.objects.filter(
            content_type=content_type,
            object_id=charge.id,
            unit_id__in=unit_ids
        ).select_related('unit', 'unit__user')
    }

    # حالا uc_map همان uc_dict است
    for i, unit in enumerate(page_units):
        uc = uc_map.get(unit.id)
        renter = unit.renters.filter(renter_is_active=True).first()

        current_charge = uc.total_charge_month if uc else 0

        # جمع عددی بدهی‌های معوقه
        previous_debt = sum(uc.get_previous_debt_by_type().values()) if uc else 0
        print(f'p-{previous_debt}')

        total_payable = current_charge + previous_debt

        page_units.object_list[i] = {
            'unit': unit,
            'renter': renter,
            'is_paid': uc.is_paid if uc else False,
            'is_notified': uc.send_notification if uc else False,
            'send_sms': uc.send_sms if uc else False,
            'sms_date': uc.send_sms_date if uc else None,
            'current_charge': current_charge,
            'previous_debt': previous_debt,  # عددی
            'total_payable': total_payable,
        }

    context = {
        'charge': charge,
        'page_obj': page_units,  # حالا فقط واحدهای دارای UnifiedCharge هستند
        "selected_units": selected_units
        # 'paginator': paginator,
    }
    return render(request, 'middleCharge/notify_fix_variable_charge_template.html', context)


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_remove_send_notification_fix_variable(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'فقط درخواست‌های POST مجاز است.'}, status=400)

    charge = get_object_or_404(ChargeFixVariable, id=pk, user=request.user)
    selected_units = request.POST.getlist('units[]')

    if not selected_units:
        return JsonResponse({'warning': 'هیچ واحدی انتخاب نشده است.'})

    try:
        with transaction.atomic():
            content_type = ContentType.objects.get_for_model(ChargeFixVariable)

            # رکوردهایی که باید غیرفعال شوند
            if selected_units == ['all']:
                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )
            else:
                try:
                    selected_unit_ids = [int(uid) for uid in selected_units]
                except ValueError:
                    return JsonResponse({'error': 'شناسه واحد نامعتبر است.'}, status=400)

                qs = UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    unit_id__in=selected_unit_ids,
                    is_paid=False,
                    send_notification=True  # فقط رکوردهای فعال
                )

            updated_count = qs.update(
                send_notification=False,
                send_notification_date=None
            )

            # اگر هیچ رکوردی با send_notification=True باقی نماند → شارژ را غیرفعال کن
            if not UnifiedCharge.objects.filter(
                    content_type=content_type,
                    object_id=charge.id,
                    send_notification=True
            ).exists():
                charge.send_notification = False
                charge.save()

        session_key = f"charge_{pk}_selected_units"

        if updated_count:
            request.session.pop(session_key, None)
            request.session.modified = True

            return JsonResponse({
                'success': f'{updated_count} اطلاعیه غیرفعال شد.'
            })
        else:
            return JsonResponse({'info': 'رکوردی برای غیرفعال کردن یافت نشد.'})

    except Exception as e:
        return JsonResponse({'error': f'خطایی هنگام غیرفعال کردن اطلاعیه‌ها رخ داد: {str(e)}'}, status=500)


# ================================================================================================

# class MiddleExpenseCharge(CreateView):
#     model = ChargeByExpense
#     form_class = ExpenseChargeForm
#     template_name = "middleCharge/variable_fix_charge_template.html"
#     success_url = reverse_lazy('middle_add_expense_charge')
#
#     def get_initial(self):
#         # گرفتن دسته‌بندی‌ها
#         try:
#             power_category = ExpenseCategory.objects.get(title='برق', user=self.request.user)
#         except ExpenseCategory.DoesNotExist:
#             power_category = None
#
#         try:
#             water_category = ExpenseCategory.objects.get(title='آب', user=self.request.user)
#         except ExpenseCategory.DoesNotExist:
#             water_category = None
#
#         try:
#             gas_category = ExpenseCategory.objects.get(title='گاز', user=self.request.user)
#         except ExpenseCategory.DoesNotExist:
#             gas_category = None
#
#         # جمع هزینه‌ها
#         total_power = Expense.objects.filter(category=power_category).aggregate(total=Sum('amount'))[
#                           'total'] or 0 if power_category else 0
#         total_water = Expense.objects.filter(category=water_category).aggregate(total=Sum('amount'))[
#                           'total'] or 0 if water_category else 0
#         total_gas = Expense.objects.filter(category=gas_category).aggregate(total=Sum('amount'))[
#                         'total'] or 0 if gas_category else 0
#
#         return {
#             'unit_power_amount': total_power,
#             'unit_water_amount': total_water,
#             'unit_gas_amount': total_gas,
#         }
#
#     def form_valid(self, form):
#         charge = form.save(commit=False)
#         charge.user = self.request.user
#         charge.save()
#
#         units = Unit.objects.filter(is_active=True, user__manager=self.request.user)
#         if not units.exists():
#             messages.error(self.request, "هیچ واحد فعالی پیدا نشد.")
#             return redirect('middle_manage_unit')
#
#         unified_charges = []
#
#         for unit in units:
#             total_amount = (
#                 (form.cleaned_data.get('unit_power_amount') or 0) +
#                 (form.cleaned_data.get('unit_water_amount') or 0) +
#                 (form.cleaned_data.get('unit_gas_amount') or 0) +
#                 (form.cleaned_data.get('civil') or 0) +
#                 (form.cleaned_data.get('other_cost_amount') or 0)
#             )
#
#             unified_charges.append(
#                 UnifiedCharge(
#                     user=self.request.user,
#                     unit=unit,
#                     main_charge=charge,
#                     charge_type=ChargeByExpense.charge_type,
#                     amount=total_amount,
#                     base_charge=total_amount,
#                     civil=form.cleaned_data.get('civil') or 0,
#                     other_cost_amount=form.cleaned_data.get('other_cost_amount') or 0,
#                     penalty_percent=form.cleaned_data.get('payment_penalty_amount') or 0,
#                     penalty_amount=0,
#                     total_charge_month=total_amount,
#                     extra_parking_price=0,
#                     details=form.cleaned_data.get('details') or '',
#                     title=form.cleaned_data.get('name'),
#                     send_notification=False,
#                     send_notification_date=None,
#                     payment_deadline_date=form.cleaned_data.get('payment_deadline'),
#                     content_type=ContentType.objects.get_for_model(ChargeByExpense),
#                     object_id=charge.id
#                 )
#             )
#
#         UnifiedCharge.objects.bulk_create(unified_charges)
#         messages.success(self.request, "شارژ هزینه‌ها با موفقیت ثبت شد.")
#         return super().form_valid(form)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#
#         # جمع کل هر شارژ
#         charges = ChargeByExpense.objects.annotate(
#             notified_count=Count('unified_charges', filter=Q(unified_charges__send_notification=True)),
#             total_units=Count('unified_charges')
#         ).order_by('-created_at')
#
#         for charge in charges:
#             charge.total_power = charge.unit_power_amount or 0
#             charge.total_water = charge.unit_water_amount or 0
#             charge.total_gas = charge.unit_gas_amount or 0
#             charge.total_civil = charge.civil or 0
#             charge.total_other = charge.other_cost_amount or 0
#
#         context['charges'] = charges
#         context['unit_count'] = Unit.objects.filter(is_active=True, user__manager=self.request.user).count()
#         return context


# ==============================================================================================

def get_all_base_charges(user):
    all_charges = chain(
        FixCharge.objects.filter(user=user),
        AreaCharge.objects.filter(user=user),
        PersonCharge.objects.filter(user=user),
        FixPersonCharge.objects.filter(user=user),
        FixAreaCharge.objects.filter(user=user),
        ChargeByPersonArea.objects.filter(user=user),
        ChargeByFixPersonArea.objects.filter(user=user),
        ChargeFixVariable.objects.filter(user=user),
    )

    return sorted(
        all_charges,
        key=lambda x: x.created_at,
        reverse=True
    )


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def base_charge_list(request):
    query = request.GET.get('q', '').strip()
    paginate = int(request.GET.get('paginate', 20))

    charges = get_all_base_charges(request.user)

    # سرچ
    if query:
        query_lower = query.lower()

        charges = [
            c for c in charges
            if query_lower in (c.name or '').lower()
        ]

    # -----------------------------
    # آماده‌سازی object_id ها
    # -----------------------------
    grouped_ids = defaultdict(list)

    for charge in charges:
        content_type = ContentType.objects.get_for_model(charge.__class__)
        grouped_ids[content_type.id].append(charge.id)

    # -----------------------------
    # گرفتن آمار همه charge ها یکجا
    # -----------------------------
    stats_map = {}

    for content_type_id, object_ids in grouped_ids.items():

        unified_stats = (
            UnifiedCharge.objects
            .filter(
                content_type_id=content_type_id,
                object_id__in=object_ids,
                unit__isnull=False
            )
            .values('object_id')
            .annotate(
                notified_count=Count(
                    'unit',
                    filter=Q(
                        send_notification=True,
                        send_notification_date__isnull=False
                    ),
                    distinct=True
                ),

                total_count=Count(
                    'unit',
                    distinct=True
                ),

                paid_count=Count(
                    'unit',
                    filter=Q(is_paid=True),
                    distinct=True
                )
            )
        )

        for item in unified_stats:
            stats_map[(content_type_id, item['object_id'])] = item

    # -----------------------------
    # ساخت داده نهایی
    # -----------------------------
    charges_data = []

    for charge in charges:
        content_type = ContentType.objects.get_for_model(charge.__class__)

        stats = stats_map.get(
            (content_type.id, charge.id),
            {}
        )

        data = charge.to_dict()

        data['notified_count'] = stats.get('notified_count', 0)
        data['total_count'] = stats.get('total_count', 0)
        data['paid_count'] = stats.get('paid_count', 0)

        charges_data.append(data)

    # pagination
    paginator = Paginator(charges_data, paginate)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        'middleCharge/middle_charges_list.html',
        {
            'charges': page_obj,
            'query': query,
            'paginate': paginate,
            'page_obj': page_obj,
        }
    )


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def middle_base_charges_pdf(request):
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(
            residents=request.user
        ).order_by('-created_at').first()

    managed_users = request.user.managed_users.all()
    unit_count = Unit.objects.filter(
        is_active=True,
        user__in=managed_users
    ).count()

    # 🔹 همان منبع ویوی لیست
    charges = get_all_base_charges(request.user)

    # 🔍 جستجو
    query = request.GET.get('q', '').strip()
    if query:
        charges = [
            c for c in charges
            if (
                    query.lower() in (c.name or '').lower() or
                    query.lower() in (getattr(c, 'details', '') or '').lower()
            )
        ]

    charges_data = []

    for charge in charges:
        data = charge.to_dict()

        # ✅ محاسبه notified_count (دقیقاً مثل base_charge_list)
        data['notified_count'] = (
            charge.unified_charges
            .filter(
                send_notification=True,
                send_notification_date__isnull=False,
                unit__isnull=False
            )
            .values_list('unit_id', flat=True)
            .distinct()
            .count()
        )

        charges_data.append(data)

    # 🧾 HTML برای PDF
    html_string = render_to_string(
        'middleCharge/middle_charges_list_pdf.html',
        {
            'charges': charges_data,
            'query': query,
            'today': datetime.now(),
            'house': house,
            'unit_count': unit_count,
            'font_url': request.build_absolute_uri('/static/fonts/Vazir.ttf')
        }
    )

    # 🎨 فونت و CSS
    font_url = request.build_absolute_uri(static('fonts/Vazir.ttf'))
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        @font-face {{
            font-family: 'Vazir';
            src: url('{font_url}');
        }}
        body {{
            font-family: 'Vazir', sans-serif;
        }}
    """)

    html = HTML(string=html_string)
    pdf_file = html.write_pdf(stylesheets=[css])

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="charge_main.pdf"'
    return response


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_base_charges_excel(request):
    managed_users = request.user.managed_users.all()
    unit_count = Unit.objects.filter(
        is_active=True,
        user__in=managed_users
    ).count()

    # 🔹 منبع یکسان با HTML و PDF
    charges = get_all_base_charges()

    # 🔍 جستجو
    query = request.GET.get('q', '').strip()
    if query:
        charges = [
            c for c in charges
            if (
                    query.lower() in (c.name or '').lower()
                    or query.lower() in (getattr(c, 'details', '') or '').lower()
            )
        ]

    # -------------------------
    # Create Excel workbook
    # -------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Main Charges"
    ws.sheet_view.rightToLeft = True

    # Title
    title_cell = ws.cell(row=1, column=1, value="لیست شارژهای اصلی ساختمان")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    # Headers
    headers = ['#', 'عنوان', 'تاریخ ثبت', 'جریمه دیرکرد(%)', 'مهلت پرداخت', 'توضیحات', 'اعلام شارژ']
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Data
    row = 3
    for index, charge in enumerate(charges, start=1):
        notified_count = (
            charge.unified_charges
            .filter(
                send_notification=True,
                send_notification_date__isnull=False,
                unit__isnull=False
            )
            .values_list('unit_id', flat=True)
            .distinct()
            .count()
        )

        ws.cell(row=row, column=1, value=index)
        ws.cell(row=row, column=2, value=charge.name)
        ws.cell(row=row, column=3, value=show_jalali(charge.created_at))
        ws.cell(row=row, column=4, value=getattr(charge, 'payment_penalty_amount', None))
        ws.cell(row=row, column=5, value=show_jalali(getattr(charge, 'payment_deadline', None)))
        ws.cell(row=row, column=6, value=getattr(charge, 'details', ''))
        ws.cell(
            row=row,
            column=7,
            value=f"{notified_count} از {unit_count} واحد"
        )
        row += 1

    # -------------------------
    # Response
    # -------------------------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=main_charges.xlsx'
    wb.save(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def charge_units_list(request, app_label, model_name, charge_id):
    model = apps.get_model(app_label, model_name)
    charge = get_object_or_404(model, id=charge_id)

    # 🔥 بررسی نوع مدل و گرفتن unified charges
    if hasattr(charge, 'unified_charges'):
        # مدل اصلی شارژ → گرفتن همه UnifiedCharge ها
        unified_qs = charge.unified_charges.all()
    elif model_name.lower() == 'unifiedcharge':
        unified_qs = model.objects.filter(
            content_type=charge.content_type,
            object_id=charge.object_id
        )
    else:
        # هر حالت غیرمنتظره
        unified_qs = model.objects.none()

    # 🔥 آپدیت جریمه همه UnifiedCharge ها
    for uc in unified_qs:
        if hasattr(uc, 'update_penalty'):
            uc.update_penalty(save=True)

    # -------------------------
    # 🔍 جستجو
    # -------------------------
    query = request.GET.get('q', '').strip()

    unified_charges = unified_qs.filter(
        send_notification_date__isnull=False
    ).select_related('unit', 'unit__user')

    if query:
        search_q = (
                Q(unit__unit__icontains=query) |
                Q(unit__user__full_name__icontains=query)
        )

        # اگر عدد بود → جستجو روی مقادیر عددی شارژ
        if query.isdigit():
            search_q |= (
                    Q(penalty_amount=query) |
                    Q(total_charge_month=query) |
                    Q(base_charge=query)
            )

        unified_charges = unified_charges.filter(search_q)

    unified_charges = unified_qs.filter(
        send_notification_date__isnull=False
    ).select_related('unit', 'unit__user').order_by('unit__unit')

    # -------------------------
    # 📄 pagination
    # -------------------------
    paginate = int(request.GET.get('paginate', 20))
    paginator = Paginator(unified_charges, paginate)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # واحدها (مطابق صفحه‌بندی)
    units = [uc.unit for uc in page_obj if uc.unit]

    return render(
        request,
        'middleCharge/middle_charges_detail.html',
        {
            'charge': charge,
            'units': units,
            'unified_charges': page_obj,
            'query': query,
            'paginate': paginate,
            'page_obj': page_obj,
            'app_label': app_label,
            'model_name': model_name,
        }
    )


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def charge_units_list_pdf(request, app_label, model_name, charge_id):
    model = apps.get_model(app_label, model_name)
    charge = get_object_or_404(model, id=charge_id)
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    unified_qs = charge.unified_charges.all()

    # 🔍 جستجو
    query = request.GET.get('q', '').strip()

    unified_charges = unified_qs.filter(
        send_notification_date__isnull=False
    ).select_related('unit', 'unit__user')

    if query:
        search_q = (
                Q(unit__unit__icontains=query) |
                Q(unit__user__full_name__icontains=query)
        )

        try:
            value = Decimal(query)
            search_q |= (
                    Q(penalty_amount=value) |
                    Q(total_charge_month=value) |
                    Q(base_charge=value)
            )
        except:
            pass

        unified_charges = unified_charges.filter(search_q)

    unified_charges = unified_charges.order_by('-created_at')

    html_string = render_to_string(
        'middleCharge/middle_charges_detail_pdf.html',
        {
            'charge': charge,
            'unified_charges': unified_charges,
            'query': query,
            'today': datetime.now(),
            'house': house,
            'font_url': request.build_absolute_uri('/static/fonts/Vazir.ttf')

        }
    )

    font_url = request.build_absolute_uri(static('fonts/Vazir.ttf'))
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        @font-face {{
            font-family: 'Vazir';
            src: url('{font_url}');
        }}
        body {{
            font-family: 'Vazir', sans-serif;
        }}
    """)

    pdf = HTML(string=html_string).write_pdf(stylesheets=[css])

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="charge_units.pdf"'
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def charge_units_list_excel(request, app_label, model_name, charge_id):
    model = apps.get_model(app_label, model_name)
    charge = get_object_or_404(model, id=charge_id)

    unified_qs = charge.unified_charges.all()

    # 🔍 جستجو
    query = request.GET.get('q', '').strip()

    unified_charges = unified_qs.filter(
        send_notification_date__isnull=False
    ).select_related('unit', 'unit__user')

    if query:
        search_q = (
                Q(unit__unit__icontains=query) |
                Q(unit__user__full_name__icontains=query)
        )

        try:
            value = Decimal(query)
            search_q |= (
                    Q(penalty_amount=value) |
                    Q(total_charge_month=value) |
                    Q(base_charge=value)
            )
        except:
            pass

        unified_charges = unified_charges.filter(search_q)

    unified_charges = unified_charges.order_by('-created_at')

    # Excel
    # -------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Charge Units"
    ws.sheet_view.rightToLeft = True

    # عنوان اصلی
    title_cell = ws.cell(row=1, column=1, value="لیست تراکنش های من")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)  # 9 ستون

    # هدرها
    headers = [
        '#', 'واحد', 'مالک / مستاجر', 'مبلغ پایه', 'جریمه',
        'مبلغ نهایی', 'تاریخ اعلام', 'مهلت پرداخت', 'وضعیت پرداخت'
    ]
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # داده‌ها
    row = 3  # داده‌ها از ردیف بعد از هدر شروع می‌شوند
    for index, uc in enumerate(unified_charges, start=1):
        ws.cell(row=row, column=1, value=index)
        ws.cell(row=row, column=2, value=uc.title)
        ws.cell(row=row, column=3, value=uc.unit.get_label())
        ws.cell(row=row, column=4, value=uc.base_charge)
        ws.cell(row=row, column=5, value=uc.penalty_amount)
        ws.cell(row=row, column=6, value=uc.total_charge_month)
        ws.cell(row=row, column=7, value=show_jalali(uc.send_notification_date))
        ws.cell(row=row, column=8, value=show_jalali(uc.payment_deadline_date))
        ws.cell(row=row, column=9, value="پرداخت شده" if uc.is_paid else "پرداخت نشده")
        row += 1

    # پاسخ Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=middle_charge_units.xlsx'
    wb.save(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def charge_units_pdf(request, charge_id):
    charge = get_object_or_404(UnifiedCharge, id=charge_id)
    units = Unit.objects.filter(unified_charges=charge, is_active=True).order_by('unit')
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()
    bank = Bank.get_default(request.user, house)
    html_string = render_to_string('middleCharge/single_charge_pdf.html', {
        'charge': charge,
        'units': units,
        'house': house,
        'bank': bank,
        'font_url': request.build_absolute_uri('/static/fonts/Vazir.ttf')
    })
    css = CSS(string=f"""
        @page {{ size: A5 portrait; margin: 0.8cm; }}
        body {{
            font-family: 'Vazir', sans-serif;
        }}
        @font-face {{
            font-family: 'Vazira', sans-serif;

        }}
    """)

    pdf_file = io.BytesIO()
    HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[css])
    pdf_file.seek(0)
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment;filename=charge_unit:{charge.unit.unit}.pdf'
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def all_invoices_pdf(request, app_label, model_name, charge_id):
    house = None
    if request.user.is_authenticated:
        house = MyHouse.objects.filter(residents=request.user).order_by('-created_at').first()

    model = apps.get_model(app_label, model_name)
    charge = get_object_or_404(model, id=charge_id)

    charges = (
        charge.unified_charges
        .filter(send_notification_date__isnull=False)
        .select_related('unit', 'unit__user')
        .order_by('unit__unit')
    )
    bank = Bank.get_default(request.user, house)

    html_string = render_to_string(
        'middleCharge/all_invoices_pdf.html',
        {
            'charge': charge,
            'charges': charges,
            'today': datetime.now(),
            'house': house,
            'bank': bank,
            'font_url': request.build_absolute_uri('/static/fonts/Vazir.ttf')
        }
    )

    font_base = request.build_absolute_uri('/static/fonts/')
    css = CSS(string=f"""
        @page {{
            size: A5 portrait;
            margin: 1cm;
        }}
        
    @font-face {{
        font-family: 'Vazir';
        src: url('{font_base}Vazir-Regular.ttf') format('truetype');
        font-weight: 400;
    }}

    @font-face {{
        font-family: 'Vazir';
        src: url('{font_base}Vazir-Bold.ttf') format('truetype');
        font-weight: 700;
    }}


    body {{
        font-family: 'Vazir';
        direction: rtl;
    }}

    h1, h2, h3 {{
        font-weight: 700;
    }}

    .page-break {{
        page-break-after: always;
    }}
""")

    pdf = HTML(string=html_string).write_pdf(stylesheets=[css])

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="all_invoices.pdf"'
    return response


# =================================================================================================
def add_sms_credit(request):
    current_credit = (
            SmsCredit.objects
            .filter(user=request.user, is_paid=True)
            .aggregate(total=Sum('amount'))['total']
            or Decimal('0')
    )
    context = {
        'current_credit': current_credit
    }
    return render(request, 'middle_admin/middle_credit_sms.html', context)


@method_decorator(middle_admin_required, name='dispatch')
class MiddleSmsManagementView(CreateView):
    model = SmsManagement
    template_name = 'middle_admin/middle_register_sms.html'
    form_class = SmsForm
    success_url = reverse_lazy('middle_register_sms')

    def form_valid(self, form):
        sms = form.save(commit=False)
        sms.user = self.request.user

        try:
            sms.house = MyHouse.objects.filter(user=self.request.user).first()
            sms.save()
            self.object = sms
            messages.success(self.request, 'پیامک موفقیت ثبت گردید')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_sms'] = SmsManagement.objects.filter(user=self.request.user, send_notification=False).order_by(
            '-created_at')
        context['units'] = Unit.objects.all()

        return context


@method_decorator(middle_admin_required, name='dispatch')
class MiddleSmsUpdateView(UpdateView):
    model = SmsManagement
    template_name = 'middle_admin/middle_register_sms.html'
    form_class = SmsForm
    success_url = reverse_lazy('middle_register_sms')

    def form_valid(self, form):
        edit_instance = form.instance
        self.object = form.save(commit=False)
        messages.success(self.request, 'پیامک با موفقیت ویرایش گردید!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_sms'] = SmsManagement.objects.filter(
            is_active=True,
            user=self.request.user,
            send_notification=False
        ).order_by('-created_at')
        return context


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_sms_delete(request, pk):
    sms = get_object_or_404(SmsManagement, id=pk)
    print(sms.id)

    try:
        sms.delete()
        messages.success(request, 'پیامک با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('middle_register_sms'))


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_show_send_sms_form(request, pk):
    sms = get_object_or_404(SmsManagement, id=pk, user=request.user)
    house = MyHouse.objects.filter(user=request.user)
    units = (Unit.objects.filter(is_active=True, user__manager=request.user, myhouse__in=house)
             .prefetch_related('renters').order_by('unit'))

    units_with_details = []
    for unit in units:
        active_renter = unit.renters.filter(renter_is_active=True).first()
        units_with_details.append({
            'unit': unit,
            'active_renter': active_renter
        })
    all_sms = SmsManagement.objects.filter(user=request.user, send_notification=False, pk=pk).order_by(
        '-created_at')

    return render(request, 'middle_admin/middle_send_sms.html', {
        'sms': sms,
        'units_with_details': units_with_details,
        'all_sms': all_sms
    })


@login_required(login_url=settings.LOGIN_URL_MIDDLE_ADMIN)
def middle_send_sms(request, pk):
    sms = get_object_or_404(SmsManagement, id=pk, user=request.user)

    # ❌ اگر قبلاً ارسال شده
    if sms.send_notification:
        messages.warning(request, 'این پیامک قبلاً ارسال شده است.')
        return redirect('middle_sms_management')

    if request.method == "POST":

        selected_units = request.POST.getlist('units')
        if not selected_units:
            messages.warning(request, 'هیچ واحدی انتخاب نشده است.')
            return redirect('middle_register_sms')

        units_qs = Unit.objects.filter(
            is_active=True,
            user__manager=request.user
        )

        if 'all' in selected_units:
            units_to_notify = units_qs
        else:
            units_to_notify = units_qs.filter(id__in=selected_units)

        if not units_to_notify.exists():
            messages.warning(request, 'هیچ واحد معتبری برای ارسال پیامک پیدا نشد.')
            return redirect('middle_register_sms')

        # 1️⃣ محاسبه تعداد پیامک
        unit_count = units_to_notify.count()
        sms_per_message = sms.sms_count
        total_sms_needed = unit_count * sms_per_message

        # 2️⃣ محاسبه مبلغ
        sms_price = Decimal(str(settings.SMS_PRICE))
        total_price = total_sms_needed * sms_price

        # 3️⃣ محاسبه شارژ موجود
        total_credit = SmsCredit.objects.filter(
            user=request.user,
            is_paid=True
        ).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')

        if total_credit < total_price:
            messages.error(
                request,
                f'شارژ پیامکی کافی نیست. مبلغ مورد نیاز: {total_price:,} تومان'
            )
            return redirect('add_sms_credit')

        # 4️⃣ عملیات اتمیک
        notified_units = []

        with transaction.atomic():

            # 🔻 کسر شارژ (FIFO)
            remaining_price = total_price
            credits = SmsCredit.objects.filter(
                user=request.user,
                is_paid=True,
                amount__gt=0
            ).order_by('created_at')

            for credit in credits:
                if remaining_price <= 0:
                    break

                if credit.amount >= remaining_price:
                    credit.amount -= remaining_price
                    remaining_price = 0
                else:
                    remaining_price -= credit.amount
                    credit.amount = 0

                credit.save()

            # 5️⃣ ارسال پیامک
            for unit in units_to_notify:
                if unit.user and unit.user.mobile:
                    helper.send_sms_to_user(
                        mobile=unit.user.mobile,
                        message=sms.message,
                        full_name=unit.user.full_name,
                    )
                    notified_units.append(unit)

            # 6️⃣ ثبت نهایی
            if notified_units:
                # 6️⃣ ثبت نهایی اطلاعات پیامک
                sms.total_units_sent = unit_count
                sms.sms_per_message = sms_per_message
                sms.total_sms_sent = total_sms_needed
                sms.total_price = total_price

                sms.notified_units.set(notified_units)
                sms.send_notification = True
                sms.send_notification_date = timezone.now().date()
                sms.save()

        messages.success(
            request,
            f'پیامک با موفقیت برای {len(notified_units)} واحد ارسال شد.'
        )
        return redirect('middle_sms_management')

    # GET
    units_with_details = Unit.objects.filter(is_active=True)
    return render(request, 'middle_admin/middle_send_sms.html', {
        'sms': sms,
        'units_with_details': units_with_details,
    })


@method_decorator(middle_admin_required, name='dispatch')
class MiddleSmsListView(ListView):
    model = SmsManagement
    template_name = 'middle_admin/middle_sms_management.html'
    context_object_name = 'all_sms'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')
        queryset = SmsManagement.objects.filter(
            user=self.request.user,
            is_active=True,
            send_notification=True,
        )
        if query:
            queryset = queryset.filter(
                Q(subject__icontains=query) |
                Q(message__icontains=query)
            )
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        context['sms_list'] = SmsManagement.objects.filter(user=self.request.user).annotate(
            unit_number=F('notified_units__unit'),
            user_full_name=F('notified_units__user__full_name')
        )
        return context


# =============================================================================================
@login_required(login_url=settings.LOGIN_URL_ADMIN)
def waive_penalty_bulk(request):
    try:
        ids = request.POST.getlist('charge_ids[]')
        if not ids:
            return JsonResponse({'success': False, 'error': 'هیچ موردی انتخاب نشده'}, status=400)

        charges = UnifiedCharge.objects.filter(id__in=ids, is_paid=False)

        if not charges.exists():
            return JsonResponse({'success': False, 'error': 'شارژی برای حذف جریمه یافت نشد'}, status=400)

        titles = []
        with transaction.atomic():
            for charge in charges:
                result = charge.waive_penalty(request.user)
                if result:
                    titles.append(result['title'])

        first_charge = charges.first()
        # app_label = first_charge._meta.app_label
        # model_name = first_charge._meta.model_name

        redirect_url = reverse(
            'charge_units_list',  # همونی که داری
            args=[
                first_charge._meta.app_label,
                first_charge._meta.model_name,
                first_charge.id  # ❗ حتماً باید باشد
            ]
        )

        return JsonResponse({
            'success': True,
            'titles': titles,
            'redirect_url': redirect_url,
            'message': 'جریمه با موفقیت حذف شد'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def restore_penalty_bulk(request):
    try:
        ids = request.POST.getlist('charge_ids[]')
        if not ids:
            return JsonResponse({'success': False, 'error': 'هیچ موردی انتخاب نشده'}, status=400)

        charges = UnifiedCharge.objects.filter(id__in=ids, is_paid=False)

        if not charges.exists():
            return JsonResponse({'success': False, 'error': 'شارژی برای بازگردانی جریمه یافت نشد'}, status=400)

        titles = []
        with transaction.atomic():
            for charge in charges:
                result = charge.restore_penalty()
                if result:
                    titles.append(result['title'])

        first_charge = charges.first()
        redirect_url = reverse(
            'charge_units_list',  # همونی که داری
            args=[
                first_charge._meta.app_label,
                first_charge._meta.model_name,
                first_charge.id  # ❗ حتماً باید باشد
            ]
        )

        return JsonResponse({
            'success': True,
            'titles': titles,
            'redirect_url': redirect_url,
            'message': 'جریمه با موفقیت بازگردانده شد'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
