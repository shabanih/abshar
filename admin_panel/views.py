import io
import json
import os
from datetime import timedelta
from decimal import Decimal

import sweetify
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.contenttypes.models import ContentType
from django.db.models.functions import Cast
from django.utils import timezone
import jdatetime
import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import ProtectedError, Q, Sum, Count, Prefetch, CharField, F
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import get_template
from django.urls import reverse_lazy, reverse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, UpdateView, DetailView, ListView
from openpyxl.styles import PatternFill, Font, Alignment
from pypdf import PdfWriter
from weasyprint import HTML, CSS

from absharProject.settings import LOGIN_URL_ADMIN
from admin_panel import helper
from admin_panel.forms import announcementForm, UnitForm, ExpenseForm, ExpenseCategoryForm, \
    IncomeForm, IncomeCategoryForm, BankForm, ReceiveMoneyForm, PayerMoneyForm, PropertyForm, \
    MaintenanceForm, FixChargeForm, PersonAreaChargeForm, AreaChargeForm, PersonChargeForm, FixAreaChargeForm, \
    FixPersonChargeForm, PersonAreaFixChargeForm, VariableFixChargeForm, UserRegistrationForm, SmsForm, MyHouseForm, \
    ChargeCategoryForm, AdminSmsForm, SubscriptionPlanForm
from admin_panel.models import Announcement, Expense, ExpenseCategory, ExpenseDocument, Income, IncomeDocument, \
    IncomeCategory, ReceiveMoney, ReceiveDocument, PayMoney, PayDocument, Property, PropertyDocument, Maintenance, \
    MaintenanceDocument, ChargeByPersonArea, \
    ChargeByFixPersonArea, FixCharge, AreaCharge, PersonCharge, \
    FixPersonCharge, FixAreaCharge, ChargeFixVariable, \
    SmsManagement, Fund, UnifiedCharge, AdminSmsManagement, SmsCredit, ImpersonationLog, SubscriptionPlan, Subscription
from notifications.models import AdminTicket
from polls.templatetags.poll_extras import jalali_to_gregorian
from user_app.models import Unit, Bank, Renter, User, MyHouse, ChargeMethod, CalendarNote, UnitResidenceHistory


def admin_required(view_func):
    return user_passes_test(lambda u: u.is_superuser, login_url=settings.LOGIN_URL_ADMIN)(view_func)


@method_decorator(admin_required, name='dispatch')
class AddSubscriptionPlan(CreateView):
    model = SubscriptionPlan
    form_class = SubscriptionPlanForm
    template_name = 'admin_panel/add_plan.html'
    success_url = reverse_lazy('subscription_plan')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['plans'] = SubscriptionPlan.objects.all().order_by('-created_at')
        return context


@method_decorator(admin_required, name='dispatch')
class SubscriptionPlanUpdate(UpdateView):
    model = SubscriptionPlan
    form_class = SubscriptionPlanForm
    template_name = 'admin_panel/add_plan.html'
    success_url = reverse_lazy('subscription_plan')

    def form_valid(self, form):
        old_plan = self.get_object()
        new_price = form.cleaned_data['price_per_unit']

        # اگر قیمت تغییر نکرده، فقط برگرد
        if old_plan.price_per_unit == new_price:
            return redirect(self.success_url)

        # غیرفعال کردن رکورد قبلی
        old_plan.is_active = False
        old_plan.end_date = timezone.now()
        old_plan.save()

        # ساخت رکورد جدید با قیمت جدید
        SubscriptionPlan.objects.create(
            duration=old_plan.duration,
            price_per_unit=new_price,
            is_active=True
        )

        return redirect(self.success_url)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['plans'] = SubscriptionPlan.objects.all().order_by('-created_at')
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def subscription_plan_delete(request, pk):
    plan = get_object_or_404(SubscriptionPlan, pk=pk)
    plan.delete()
    messages.success(request, 'اشتراک با موفقیت حذف گردید')
    return redirect('subscription_plan')


@method_decorator(admin_required, name='dispatch')
class UserManagementListView(ListView):
    model = User
    template_name = 'admin_panel/user_management.html'
    context_object_name = 'users'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        qs = (
            User.objects.all()
        )

        if query:
            qs = qs.filter(
                Q(full_name__icontains=query) |
                Q(mobile__icontains=query) |
                Q(username__icontains=query)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '1')
        context.update(
            User.objects.aggregate(
                admin_count=Count('id', filter=Q(is_superuser=True)),
                middle_count=Count('id', filter=Q(is_middle_admin=True, is_superuser=False)),
                unit_count=Count('id', filter=Q(is_unit=True, is_superuser=False, is_middle_admin=False)),
            )
        )
        return context


def dashboard_redirect(request):
    """
    بعد از login یا impersonation، کاربر به پنل مناسب هدایت شود
    """
    user = request.user

    if request.session.get('impersonator_id'):
        # اگر سوپریوزر در حالت impersonate است، به پنل کاربر target برو
        if user.is_middle_admin:
            return redirect('middle_admin_dashboard')
        elif user.is_unit:
            return redirect('user_panel')
        else:
            return redirect('user_panel')  # کاربران بدون نقش خاص

    # کاربر معمولی
    if user.is_superuser:
        return redirect('/admin-panel/')  # پنل سوپریوزر
    elif user.is_middle_admin:
        return redirect('middle_admin_dashboard')
    elif user.is_unit:
        return redirect('user_panel')
    else:
        return redirect('user_panel')


def impersonate_user(request, user_id):
    target = get_object_or_404(User, id=user_id)

    if not request.user.is_superuser:
        return redirect("/")

    impersonator_id = request.user.id  # ذخیره موقت

    log = ImpersonationLog.objects.create(
        admin=request.user,
        target_user=target,
        ip_address=request.META.get("REMOTE_ADDR"),
    )

    # backend لازم
    target.backend = 'django.contrib.auth.backends.ModelBackend'

    login(request, target)  # اول login

    # بعدش session رو پر کن
    request.session["impersonator_id"] = impersonator_id
    request.session["impersonation_log"] = log.id
    request.session["is_impersonating"] = True
    request.session.save()

    print("AFTER LOGIN SESSION:", request.session.items())

    if target.is_middle_admin:
        return redirect('middle_admin_dashboard')
    else:
        return redirect('user_panel')


def stop_impersonation(request):
    admin_id = request.session.get("impersonator_id")

    if admin_id:
        admin_user = User.objects.get(id=admin_id)
        admin_user.backend = 'django.contrib.auth.backends.ModelBackend'
        login(request, admin_user)

        # پاک کردن session
        request.session.pop("impersonator_id", None)
        request.session.pop("impersonation_log", None)

    return redirect('/admin-panel/')


@method_decorator(admin_required, name='dispatch')
class MiddleAdminCreateView(CreateView):
    model = User
    template_name = 'admin_panel/add_middleAdmin.html'
    form_class = UserRegistrationForm
    success_url = reverse_lazy('create_middle_admin')

    def form_valid(self, form):
        # ایجاد کاربر
        self.object = form.save(commit=False)
        raw_password = form.cleaned_data.get('password')
        if raw_password:
            self.object.set_password(raw_password)
        self.object.is_middle_admin = True
        self.object.manager = self.request.user
        self.object.save()

        # ثبت charge_methods
        charge_methods = form.cleaned_data.get('charge_methods')
        if charge_methods:
            self.object.charge_methods.set(charge_methods)


        # اگر Trial فعال است
        if form.cleaned_data.get('is_trial') == '1':
            Subscription.objects.create(
                user=self.object,
                units_count=5,
                is_trial=True,
                start_date=timezone.now(),
                end_date=timezone.now() + timedelta(days=1)
            )
            messages.success(
                self.request,
                'مدیر ساختمان با موفقیت ثبت شد!'
            )
        else:
            messages.success(
                self.request,
                'مدیر ساختمان با موفقیت ثبت شد!!'
            )

        return redirect(self.success_url)


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['middleAdmins'] = User.objects.filter(
            is_middle_admin=True
        ).order_by('-created_time')
        context['users'] = User.objects.filter(is_active=True).order_by('created_time')

        if self.request.user.charge_methods.exists():
            context['allowed_methods'] = list(
                self.request.user.charge_methods.values_list('id', flat=True)
            )
        else:
            context['allowed_methods'] = []

        return context



@method_decorator(admin_required, name='dispatch')
class MiddleAdminUpdateView(UpdateView):
    model = User
    template_name = 'admin_panel/add_middleAdmin.html'
    form_class = UserRegistrationForm
    success_url = reverse_lazy('create_middle_admin')

    def form_valid(self, form):
        obj = form.save(commit=False)

        # گرفتن رمز جدید
        raw_password = form.cleaned_data.get('password')
        if raw_password:
            obj.set_password(raw_password)
        else:
            old_user = User.objects.get(pk=obj.pk)
            obj.password = old_user.password

        obj.manager = self.request.user
        obj.save()

        # ست کردن روش‌های شارژ
        charge_methods = form.cleaned_data.get('charge_methods')
        if charge_methods is not None:
            obj.charge_methods.set(charge_methods)

        subscription, created = Subscription.objects.get_or_create(
            user=obj,
            defaults={
                'units_count': 1,
                'is_trial': True,
                'start_date': timezone.now(),
                'end_date': timezone.now() + timedelta(days=35)
            }
        )
        if not created:
            # اگر قبلا وجود داشت، فقط مقادیر فرم را آپدیت کن، تاریخ Trial را تغییر نده
            subscription.units_count = 1
            # فقط اگر خواستی اشتراک پولی یا روش‌های دیگر اضافه شود می‌توانی اینجا تغییر دهی
            subscription.save()

        messages.success(self.request, 'اطلاعات مدیر ساختمان با موفقیت ویرایش گردید!')
        return super().form_valid(form)

    def form_invalid(self, form):
        # اگر فرم نامعتبر بود (مثلاً موبایل تکراری)، همان قالب را با خطا نمایش بده
        middle_admins = User.objects.filter(is_middle_admin=True).order_by('-created_time')
        for middle in middle_admins:
            # ایجاد attribute موقت برای قالب
            middle.charge_method_ids_list = list(middle.charge_methods.values_list('id', flat=True))

        context = {
            'form': form,
            'middleAdmins': middle_admins,
            'users': User.objects.filter(is_active=True).order_by('-created_time'),
        }
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        middle_admins = User.objects.filter(is_middle_admin=True).order_by('-created_time')
        for middle in middle_admins:
            # ایجاد attribute موقت برای قالب
            middle.charge_method_ids_list = list(middle.charge_methods.values_list('id', flat=True))
        context['middleAdmins'] = middle_admins
        context['users'] = User.objects.filter(is_active=True).order_by('-created_time')
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def middleAdmin_delete(request, pk):
    middleAdmin = get_object_or_404(User, id=pk)
    print(middleAdmin.id)

    try:
        middleAdmin.delete()
        messages.success(request, 'مدیر ساختمان با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('create_middle_admin'))


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def admin_dashboard(request):
    announcements = Announcement.objects.filter(is_active=True).order_by('-created_at')[:4]
    tickets = AdminTicket.objects.filter(user__manager=request.user).order_by('-created_at')[:5]
    middle_count = User.objects.filter(is_middle_admin=True).count()
    tickets_count = AdminTicket.objects.all().count()
    context = {
        'announcements': announcements,
        'tickets': tickets,
        'middle_count': middle_count,
        'tickets_count': tickets_count,
    }
    return render(request, 'shared/home_template.html', context)


def admin_login_view(request):
    if request.method == 'POST':
        mobile = request.POST.get('mobile')
        password = request.POST.get('password1')

        user = authenticate(request, mobile=mobile, password=password)
        if user is not None:
            if user.is_superuser:
                login(request, user)
                sweetify.success(request, f"{user.username} عزیز، با موفقیت وارد بخش ادمین شدید!")
                return redirect(reverse('admin_dashboard'))
            else:
                logout(request)  # Log out any non-superuser who authenticated successfully
                messages.error(request, 'شما مجوز دسترسی به بخش ادمین را ندارید!')
                return redirect(reverse('login_admin'))
        else:
            messages.error(request, 'نام کاربری و یا رمز عبور اشتباه است!')
            return redirect(reverse('login_admin'))

    return render(request, 'shared/login.html')


def logout_admin(request):
    logout(request)
    return redirect('login_admin')


def site_header_component(request):
    context = {
        'user': request.user,
    }
    return render(request, 'shared/notification_template.html', context)


# ====================== Announcement =================================
@method_decorator(admin_required, name='dispatch')
class AnnouncementView(ListView):
    model = Announcement
    template_name = 'admin_panel/announcement.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        qs = (
            MyHouse.objects
            .filter(announcement__is_active=True)
            .annotate(
                total_announcements=Count(
                    'announcement',
                    filter=Q(announcement__is_active=True)
                )
            )
            .distinct()
        )

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '1')
        return context


@method_decorator(admin_required, name='dispatch')
class ManagerAnnouncementsDetailView(ListView):
    model = Announcement
    template_name = "admin_panel/detail_announcement.html"
    context_object_name = "announcements"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = Announcement.objects.filter(
            house_id=house_id,
            is_active=True
        )

        if query:
            qs = qs.filter(
                Q(title__icontains=query)
            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '1')
        # اضافه کردن اطلاعات خانه
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def announcement_delete(request, pk):
    # فقط ادمین اجازه دارد
    if not request.user.is_superuser:
        messages.error(request, "شما دسترسی ندارید!")
        return redirect('house_announcements')

    announce = get_object_or_404(Announcement, id=pk)

    try:
        announce.delete()
        messages.success(request, 'اطلاعیه با موفقیت حذف شد!')
    except ProtectedError:
        messages.error(request, 'امکان حذف این اطلاعیه وجود ندارد!')

    # برگشت به لیست اطلاعیه‌های همان خانه
    return redirect(reverse('house_announcements', kwargs={'house_id': announce.house.id}))


# ========================== My House Views ========================
@method_decorator(admin_required, name='dispatch')
class AddMyHouseView(ListView):
    model = MyHouse
    template_name = 'admin_panel/add_my_house.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        queryset = MyHouse.objects.all()

        # فیلتر جستجو
        if query:
            queryset = queryset.filter(
                Q(user__full_name__icontains=query) |
                Q(name__icontains=query) |
                Q(user_type__icontains=query) |
                Q(address__icontains=query) |
                Q(city__icontains=query)
            ).distinct()

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


# =============================== banks =========================
@method_decorator(admin_required, name='dispatch')
class AddBankView(ListView):
    model = MyHouse
    template_name = 'admin_panel/add_my_bank.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        qs = (
            MyHouse.objects
            .annotate(
                total_banks=Count(
                    'banks',
                    filter=Q(banks__is_active=True)
                )
            )
            .filter(total_banks__gt=0)
        )

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@method_decorator(admin_required, name='dispatch')
class HouseBanksDetailView(ListView):
    model = Bank
    template_name = "admin_panel/middle_bank_list.html"
    context_object_name = "banks"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = Bank.objects.filter(
            house_id=house_id,
            is_active=True
        )

        if query:
            qs = qs.filter(
                Q(bank_name__icontains=query) |
                Q(account_holder_name__icontains=query)
            )

        return qs.order_by('-create_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        context['house'] = get_object_or_404(MyHouse, id=self.kwargs['house_id'])
        return context


# =========================== unit Views ================================

@method_decorator(admin_required, name='dispatch')
class UnitRegisterView(ListView):
    model = MyHouse
    template_name = 'unit_templates/unit_management.html'
    context_object_name = 'houses'

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        # حالا خانه‌ها را با تعداد واحد فعال
        qs = MyHouse.objects.annotate(
            total_units=Count('units', filter=Q(units__is_active=True))
        ).filter(total_units__gt=0)

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@method_decorator(admin_required, name='dispatch')
class UnitHouseDetailView(ListView):
    model = Unit
    template_name = 'unit_templates/unit_list.html'
    context_object_name = 'object_list'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None
        return int(paginate) if paginate and paginate.isdigit() else 20

    def get_queryset(self):
        house_id = self.kwargs['house_id']

        # پایه: همه واحدهای خانه و فعال
        qs = Unit.objects.filter(myhouse_id=house_id, is_active=True).prefetch_related('renters').order_by('unit')

        params = self.request.GET
        filters = Q()

        # فیلترهای اختیاری فقط در صورتی اعمال شود که مقدار معتبر داشته باشند
        if params.get('unit') and params['unit'].isdigit():
            filters &= Q(unit=int(params['unit']))
        if params.get('owner_name'):
            filters &= Q(owner_name__icontains=params['owner_name'])
        if params.get('owner_mobile'):
            filters &= Q(owner_mobile__icontains=params['owner_mobile'])
        if params.get('area') and params['area'].isdigit():
            filters &= Q(area=int(params['area']))
        if params.get('bedrooms_count') and params['bedrooms_count'].isdigit():
            filters &= Q(bedrooms_count=int(params['bedrooms_count']))
        if params.get('renter_name'):
            filters &= Q(renters__renter_name__icontains=params['renter_name'])
        if params.get('renter_mobile'):
            filters &= Q(renters__renter_mobile__icontains=params['renter_mobile'])
        if params.get('people_count') and params['people_count'].isdigit():
            filters &= Q(owner_people_count=int(params['people_count']))
        if params.get('status_residence'):
            filters &= Q(status_residence__icontains=params['status_residence'])
        if params.get('is_renter'):
            filters &= Q(is_renter__icontains=params['is_renter'])

        # اعمال فیلتر
        qs = qs.filter(filters).distinct()

        # آماده کردن مستاجر فعال برای هر واحد
        for unit in qs:
            unit.active_renters = unit.renters.filter(renter_is_active=True)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        house_id = self.kwargs.get('house_id')
        if house_id:
            context['house'] = MyHouse.objects.get(id=house_id)
            context['total_units'] = Unit.objects.filter(myhouse_id=house_id, is_active=True).count()
        context['query_params'] = self.request.GET.urlencode()
        return context


@method_decorator(admin_required, name='dispatch')
class UnitInfoView(DetailView):
    model = Unit
    template_name = 'unit_templates/unit_info.html'
    context_object_name = 'unit'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unit = self.object
        context['renters'] = unit.renters.order_by('-renter_is_active', '-start_date')

        # اضافه کردن خانه مربوطه
        context['house'] = unit.myhouse  # اینجا می‌توانید house.id هم استفاده کنید
        return context


def to_jalali(date_obj):
    if not date_obj:
        return ''
    jalali_date = jdatetime.date.fromgregorian(date=date_obj)
    return jalali_date.strftime('%Y/%m/%d')


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_units_excel(request, house_id):
    # پایه: همان واحدهایی که در لیست دیده می‌شوند
    qs = Unit.objects.filter(myhouse_id=house_id, is_active=True).prefetch_related('renters').order_by('unit')

    params = request.GET
    filters = Q()

    if params.get('unit') and params['unit'].isdigit():
        filters &= Q(unit=int(params['unit']))
    if params.get('owner_name'):
        filters &= Q(owner_name__icontains=params['owner_name'])
    if params.get('owner_mobile'):
        filters &= Q(owner_mobile__icontains=params['owner_mobile'])
    if params.get('area') and params['area'].isdigit():
        filters &= Q(area=int(params['area']))
    if params.get('bedrooms_count') and params['bedrooms_count'].isdigit():
        filters &= Q(bedrooms_count=int(params['bedrooms_count']))
    if params.get('renter_name'):
        filters &= Q(renters__renter_name__icontains=params['renter_name'])
    if params.get('renter_mobile'):
        filters &= Q(renters__renter_mobile__icontains=params['renter_mobile'])
    if params.get('people_count') and params['people_count'].isdigit():
        filters &= Q(owner_people_count=int(params['people_count']))
    if params.get('status_residence'):
        filters &= Q(status_residence__icontains=params['status_residence'])

    qs = qs.filter(filters).distinct()

    # آماده کردن مستاجرهای فعال
    for unit in qs:
        unit.active_renters = unit.renters.filter(renter_is_active=True)

    # ساخت فایل اکسل
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "units"
    ws.sheet_view.rightToLeft = True

    # عنوان
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=22)
    title_cell = ws.cell(row=1, column=1, value="لیست واحدهای ساختمان")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # سرستون‌ها
    headers = [
        'واحد', 'طبقه', 'متراژ', 'تعداد خواب', 'شماره تلفن',
        'وضعیت سکونت', 'نام مالک', 'تلفن مالک', 'تعداد نفرات مالک',
        'نام مستاجر', 'تلفن مستاجر', 'کد ملی مستاجر',
        'تاریخ اجاره', 'تاریخ پایان', 'شماره قرارداد'
    ]

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

    # داده‌ها
    for row_num, unit in enumerate(qs, start=3):
        ws.cell(row=row_num, column=1, value=unit.unit)
        ws.cell(row=row_num, column=2, value=unit.floor_number)
        ws.cell(row=row_num, column=3, value=unit.area)
        ws.cell(row=row_num, column=4, value=unit.bedrooms_count)
        ws.cell(row=row_num, column=5, value=unit.unit_phone)
        ws.cell(row=row_num, column=6, value=unit.status_residence)
        ws.cell(row=row_num, column=7, value=unit.owner_name)
        ws.cell(row=row_num, column=8, value=unit.owner_mobile)
        ws.cell(row=row_num, column=9, value=unit.owner_people_count)

        # فقط اولین مستاجر فعال (می‌توان همه مستاجرها را هم اضافه کرد)
        renter = unit.active_renters.first()
        if renter:
            ws.cell(row=row_num, column=10, value=renter.renter_name)
            ws.cell(row=row_num, column=11, value=renter.renter_mobile)
            ws.cell(row=row_num, column=12, value=renter.renter_national_code)
            ws.cell(row=row_num, column=13, value=to_jalali(renter.start_date) if renter.start_date else '')
            ws.cell(row=row_num, column=14, value=to_jalali(renter.end_date) if renter.end_date else '')
            ws.cell(row=row_num, column=15, value=renter.contract_number)

    # ارسال فایل
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=units_house_{house_id}.xlsx'
    wb.save(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_units_pdf(request, house_id):
    # پایه: همه واحدهای خانه با prefetch مستاجرهای فعال
    qs = Unit.objects.filter(myhouse_id=house_id, is_active=True).prefetch_related(
        Prefetch(
            'renters',
            queryset=Renter.objects.filter(renter_is_active=True),
            to_attr='active_renters'
        )
    ).order_by('unit')

    params = request.GET
    filters = Q()

    # فیلترهای اختیاری (مثل اکسل)
    if params.get('unit') and params['unit'].isdigit():
        filters &= Q(unit=int(params['unit']))
    if params.get('owner_name'):
        filters &= Q(owner_name__icontains=params['owner_name'])
    if params.get('owner_mobile'):
        filters &= Q(owner_mobile__icontains=params['owner_mobile'])
    if params.get('area') and params['area'].isdigit():
        filters &= Q(area=int(params['area']))
    if params.get('bedrooms_count') and params['bedrooms_count'].isdigit():
        filters &= Q(bedrooms_count=int(params['bedrooms_count']))
    if params.get('people_count') and params['people_count'].isdigit():
        filters &= Q(owner_people_count=int(params['people_count']))
    if params.get('status_residence'):
        filters &= Q(status_residence__icontains=params['status_residence'])

    qs = qs.filter(filters).distinct()

    # آماده کردن template
    template = get_template("unit_templates/unit_pdf.html")
    context = {
        'units': qs,
        'house_id': house_id,
    }
    html_content = template.render(context)

    # تنظیم فونت و راست‌چین
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
        @page {{ size: A4 landscape; margin: 1cm; }}
        body {{
            font-family: 'BYekan', sans-serif;
        }}
        @font-face {{
            font-family: 'BYekan';
            src: url('{font_url}');
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 4px;
            text-align: center;
        }}
        th {{
            background-color: #FFD700;
        }}
    """)

    # ساخت PDF
    pdf_file = io.BytesIO()
    HTML(string=html_content, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[css])
    pdf_file.seek(0)

    response = HttpResponse(pdf_file.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="units_house_{house_id}.pdf"'
    return response


# ================================= Expense Views ==============================
@method_decorator(admin_required, name='dispatch')
class ExpenseCategoryView(CreateView):
    model = ExpenseCategory
    template_name = 'expense_templates/add_category_expense.html'
    form_class = ExpenseCategoryForm
    success_url = reverse_lazy('add_category_expense')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            self.object = form.save()
            messages.success(self.request, 'موضوع هزینه با موفقیت ثبت گردید!')
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'خطا در ثبت !')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = ExpenseCategory.objects.all()
        return context


@method_decorator(admin_required, name='dispatch')
class ExpenseCategoryUpdate(UpdateView):
    model = ExpenseCategory
    template_name = 'expense_templates/add_category_expense.html'
    form_class = ExpenseCategoryForm
    success_url = reverse_lazy('add_category_expense')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            edit_instance = form.instance
            self.object = form.save()
            messages.success(self.request, f' موضوع هزینه با موفقیت ویرایش گردید!')
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'خطا در ثبت !')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = ExpenseCategory.objects.all()
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def expense_category_delete(request, pk):
    category = get_object_or_404(ExpenseCategory, id=pk)
    try:
        category.delete()
        messages.success(request, 'موضوع هزینه با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('add_category_expense'))


@method_decorator(admin_required, name='dispatch')
class ExpenseView(CreateView):
    model = Expense
    template_name = 'expense_templates/expense_register.html'
    form_class = ExpenseForm
    success_url = reverse_lazy('add_expense')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            self.object = form.save()
            content_type = ContentType.objects.get_for_model(self.object)

            Fund.objects.create(
                content_type=content_type,
                object_id=self.object.id,
                debtor_amount=0,
                creditor_amount=self.object.amount or 0,
                payment_date=self.object.date,
                payment_description=f"هزینه: {self.object.description[:50]}",
            )
            files = self.request.FILES.getlist('document')

            # ذخیره فایل‌ها در مدل ExpenseDocument
            for f in files:
                ExpenseDocument.objects.create(expense=self.object, document=f)
            messages.success(self.request, 'هزینه با موفقیت ثبت گردید')
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'خطا در ثبت هزینه!')
            return self.form_invalid(form)

    def get_queryset(self):
        queryset = Expense.objects.all()

        # فیلتر بر اساس category__title
        category_id = self.request.GET.get('category')
        if category_id:
            queryset = queryset.filter(category__id=category_id)

        # فیلتر بر اساس amount
        amount = self.request.GET.get('amount')
        if amount:
            queryset = queryset.filter(amount__icontains=amount)

        description = self.request.GET.get('description')
        if description:
            queryset = queryset.filter(description__icontains=description)

        doc_no = self.request.GET.get('doc_no')
        if doc_no:
            queryset = queryset.filter(doc_no__icontains=doc_no)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس date
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')

        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        expenses = self.get_queryset()  # از get_queryset برای دریافت داده‌های فیلتر شده استفاده می‌کنیم
        paginator = Paginator(expenses, 10)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_expense'] = Expense.objects.count()
        context['categories'] = ExpenseCategory.objects.all()

        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def expense_edit(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES, instance=expense)
        if form.is_valid():
            expense = form.save()  # Save the form (updates or creates expense)
            # Handle multiple file uploads
            files = request.FILES.getlist('document')
            if files:
                for f in files:
                    ExpenseDocument.objects.create(expense=expense, document=f)
            messages.success(request, 'هزینه با موفقیت ویرایش شد.')
            return redirect('add_expense')  # Adjust redirect as necessary
        else:
            messages.error(request, 'خطا در ویرایش فرم هزینه. لطفا دوباره تلاش کنید.')
            return redirect('add_expense')
    else:
        # If the request is not POST, redirect to the appropriate page
        return redirect('add_expense')


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, id=pk)
    try:
        expense.delete()
        messages.success(request, ' هزینه با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('add_expense'))


@csrf_exempt
def delete_expense_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        expense_id = request.POST.get('expense_id')

        if not image_url or not expense_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            expense = get_object_or_404(Expense, id=expense_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = ExpenseDocument.objects.filter(expense=expense, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_expense_pdf(request):
    expenses = Expense.objects.all()

    filter_fields = {
        'category': 'category__id',
        'amount': 'amount__icontains',
        'doc_no': 'doc_no__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
    }

    # Apply filters based on GET parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            expenses = expenses.filter(**filter_expression)

    # Handle date filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__lte=to_date)
    except ValueError:
        expenses = Expense.objects.none()

    # Log the filtered expenses for debugging
    print(expenses)

    # Font setup
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # Render HTML template
    template = get_template("expense_templates/expense_pdf.html")
    context = {
        'expenses': expenses,
        'font_path': font_url,
    }
    html = template.render(context)

    # Generate PDF
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # Generate the final PDF response
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="expenses.pdf"'
    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_expense_excel(request):
    expenses = Expense.objects.all()

    # Filter fields
    filter_fields = {
        'category': 'category__id',
        'amount': 'amount__icontains',
        'doc_no': 'doc_no__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
    }

    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            expenses = expenses.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            expenses = expenses.filter(date__lte=to_date)
    except ValueError:
        expenses = Expense.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "expenses"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست هزینه‌ها")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'موضوع هزینه', 'شرح سند', ' شماره سند', 'مبلغ', 'تاریخ سند', 'توضیحات']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, expense in enumerate(expenses, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        ws.cell(row=row_num, column=2, value=expense.category.title)
        ws.cell(row=row_num, column=3, value=expense.description)
        ws.cell(row=row_num, column=4, value=expense.doc_no)
        ws.cell(row=row_num, column=5, value=expense.amount)
        jalali_date = jdatetime.date.fromgregorian(date=expense.date).strftime('%Y/%m/%d')
        ws.cell(row=row_num, column=6, value=jalali_date)
        ws.cell(row=row_num, column=7, value=expense.details)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=expenses.xlsx'
    wb.save(response)
    return response


# =========================== Income Views =========================
@method_decorator(admin_required, name='dispatch')
class IncomeCategoryView(CreateView):
    model = IncomeCategory
    template_name = 'income_templates/add_category_income.html'
    form_class = IncomeCategoryForm
    success_url = reverse_lazy('add_category_income')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            self.object = form.save()
            messages.success(self.request, 'موضوع درآمد با موفقیت ثبت گردید!')
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'خطا در ثبت !')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['income_categories'] = IncomeCategory.objects.all()
        return context


@method_decorator(admin_required, name='dispatch')
class IncomeCategoryUpdate(UpdateView):
    model = IncomeCategory
    template_name = 'income_templates/add_category_income.html'
    form_class = IncomeCategoryForm
    success_url = reverse_lazy('add_category_income')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            edit_instance = form.instance
            self.object = form.save()
            messages.success(self.request, f' موضوع درآمد با موفقیت ویرایش گردید!')
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'خطا در ثبت !')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['income_categories'] = IncomeCategory.objects.all()
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def income_category_delete(request, pk):
    income_category = get_object_or_404(IncomeCategory, id=pk)
    try:
        income_category.delete()
        messages.success(request, 'موضوع درآمد با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('add_category_income'))


@method_decorator(admin_required, name='dispatch')
class IncomeView(CreateView):
    model = Income
    template_name = 'income_templates/income_register.html'
    form_class = IncomeForm
    success_url = reverse_lazy('add_income')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            self.object = form.save()
            content_type = ContentType.objects.get_for_model(self.object)

            Fund.objects.create(
                content_type=content_type,
                object_id=self.object.id,
                debtor_amount=self.object.amount or 0,
                creditor_amount=0,
                payment_date=self.object.doc_date,
                payment_description=f"درآمد: {self.object.description[:50]}",
            )
            files = self.request.FILES.getlist('document')

            # ذخیره فایل‌ها در مدل ExpenseDocument
            for f in files:
                IncomeDocument.objects.create(income=self.object, document=f)
            messages.success(self.request, 'درآمد با موفقیت ثبت گردید')
            return super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, 'خطا در ثبت درآمد!')
            return self.form_invalid(form)

    def get_queryset(self):
        queryset = Income.objects.all()

        # فیلتر بر اساس category__title
        category_id = self.request.GET.get('category')
        if category_id:
            queryset = queryset.filter(category__id=category_id)

        # فیلتر بر اساس amount
        amount = self.request.GET.get('amount')
        if amount:
            queryset = queryset.filter(amount__icontains=amount)

        description = self.request.GET.get('description')
        if description:
            queryset = queryset.filter(description__icontains=description)

        doc_number = self.request.GET.get('doc_number')
        if doc_number:
            queryset = queryset.filter(doc_number__icontains=doc_number)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس date
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')

        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(doc_date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(doc_date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        incomes = self.get_queryset()  # از get_queryset برای دریافت داده‌های فیلتر شده استفاده می‌کنیم
        paginator = Paginator(incomes, 10)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_incomes'] = Income.objects.count()
        context['categories'] = IncomeCategory.objects.all()

        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def income_edit(request, pk):
    income = get_object_or_404(Income, pk=pk)

    if request.method == 'POST':
        form = IncomeForm(request.POST, request.FILES, instance=income)

        if form.is_valid():
            income = form.save()  # Save the form (updates or creates expense)

            # Handle multiple file uploads
            files = request.FILES.getlist('document')
            if files:
                for f in files:
                    IncomeDocument.objects.create(income=income, document=f)

            messages.success(request, 'درآمد با موفقیت ویرایش شد.')
            return redirect('add_income')  # Adjust redirect as necessary

        else:
            messages.error(request, 'خطا در ویرایش فرم درآمد. لطفا دوباره تلاش کنید.')
            return redirect('add_income')
    else:
        # If the request is not POST, redirect to the appropriate page
        return redirect('add_income')


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def income_delete(request, pk):
    income = get_object_or_404(Income, id=pk)
    try:
        income.delete()
        messages.success(request, ' درآمد با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('add_income'))


@csrf_exempt
def delete_income_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        income_id = request.POST.get('income_id')

        if not image_url or not income_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            income = get_object_or_404(Income, id=income_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = IncomeDocument.objects.filter(income=income, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_income_pdf(request):
    incomes = Income.objects.all()

    filter_fields = {
        'category': 'category__id',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            incomes = incomes.filter(**filter_expression)

    # فیلتر تاریخ
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__lte=to_date)
    except ValueError:
        incomes = Income.objects.none()

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # رندر قالب HTML
    template = get_template("income_templates/income_pdf.html")
    context = {
        'incomes': incomes,
        'font_path': font_url,
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="incomes.pdf"'

    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_income_excel(request):
    incomes = Income.objects.all()

    # Filter fields
    filter_fields = {
        'category': 'category__id',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
    }

    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            incomes = incomes.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            incomes = incomes.filter(doc_date__lte=to_date)
    except ValueError:
        expenses = Expense.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "incomes"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست درآمدها")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'موضوع درآمد', 'شرح سند', ' شماره سند', 'مبلغ', 'تاریخ سند', 'توضیحات']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, income in enumerate(incomes, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        ws.cell(row=row_num, column=2, value=income.category.subject)
        ws.cell(row=row_num, column=3, value=income.description)
        ws.cell(row=row_num, column=4, value=income.doc_number)
        ws.cell(row=row_num, column=5, value=income.amount)
        jalali_date = jdatetime.date.fromgregorian(date=income.doc_date).strftime('%Y/%m/%d')
        ws.cell(row=row_num, column=6, value=jalali_date)
        ws.cell(row=row_num, column=7, value=income.details)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=incomes.xlsx'
    wb.save(response)
    return response


# ============================ ReceiveMoneyView ==========================
@method_decorator(admin_required, name='dispatch')
class ReceiveMoneyCreateView(CreateView):
    model = ReceiveMoney
    form_class = ReceiveMoneyForm
    template_name = 'receiveMoney/add_receive_money.html'
    success_url = reverse_lazy('add_receive')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            self.object = form.save()
            files = self.request.FILES.getlist('document')

            # ذخیره فایل‌ها در مدل ExpenseDocument
            for f in files:
                ReceiveDocument.objects.create(receive=self.object, document=f)
            messages.success(self.request, 'سند دریافت با موفقیت ثبت گردید!')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_queryset(self):
        queryset = ReceiveMoney.objects.all()

        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank__id=bank_id)

        # فیلتر بر اساس amount
        amount = self.request.GET.get('amount')
        if amount:
            queryset = queryset.filter(amount__icontains=amount)

        payer_name = self.request.GET.get('payer_name')
        if payer_name:
            queryset = queryset.filter(payer_name__icontains=payer_name)

        description = self.request.GET.get('description')
        if description:
            queryset = queryset.filter(description__icontains=description)

        doc_number = self.request.GET.get('doc_number')
        if doc_number:
            queryset = queryset.filter(doc_number__icontains=doc_number)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس date
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')

        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(doc_date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(doc_date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')
        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receives = self.get_queryset()
        paginator = Paginator(receives, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_receives'] = ReceiveMoney.objects.count()
        context['receives'] = ReceiveMoney.objects.all()
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def receive_edit(request, pk):
    receive = get_object_or_404(ReceiveMoney, pk=pk)

    if request.method == 'POST':
        form = ReceiveMoneyForm(request.POST, request.FILES, instance=receive, user=request.user)

        if form.is_valid():
            receive = form.save()  # Save the form (updates or creates expense)

            # Handle multiple file uploads
            files = request.FILES.getlist('document')
            if files:
                for f in files:
                    ReceiveDocument.objects.create(receive=receive, document=f)

            messages.success(request, 'سند با موفقیت ویرایش شد.')
            return redirect('add_receive')  # Adjust redirect as necessary


        else:

            messages.error(request, 'خطا در ویرایش فرم درآمد. لطفا دوباره تلاش کنید.')
            return render(request, 'receiveMoney/add_receive_money.html', {'form': form, 'receive': receive})

    else:

        form = ReceiveMoneyForm(instance=receive, user=request.user)
        return render(request, 'receiveMoney/add_receive_money.html', {'form': form, 'receive': receive})


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def receive_delete(request, pk):
    receive = get_object_or_404(ReceiveMoney, id=pk)
    try:
        receive.delete()
        messages.success(request, ' سند با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('add_receive'))


@login_required(login_url=settings.LOGIN_URL_ADMIN)
@csrf_exempt
def delete_receive_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        receive_id = request.POST.get('receive_id')

        if not image_url or not receive_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            receive = get_object_or_404(ReceiveMoney, id=receive_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = ReceiveDocument.objects.filter(receive=receive, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_receive_pdf(request):
    receives = ReceiveMoney.objects.all()

    filter_fields = {
        'bank': 'bank__id',
        'payer_name': 'payer_name',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            receives = receives.filter(**filter_expression)

    # فیلتر تاریخ
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            receives = receives.filter(doc_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            receives = receives.filter(doc_date__lte=to_date)
    except ValueError:
        receives = ReceiveMoney.objects.none()

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # رندر قالب HTML
    template = get_template("receiveMoney/receive_pdf.html")
    context = {
        'receives': receives,
        'font_path': font_url,
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="receives.pdf"'

    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_receive_excel(request):
    receives = ReceiveMoney.objects.all()

    filter_fields = {
        'bank': 'bank__id',
        'payer_name': 'payer_name__icontains',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
    }

    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            receives = receives.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            receives = receives.filter(doc_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            receives = receives.filter(doc_date__lte=to_date)
    except ValueError:
        receives = ReceiveMoney.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "receives"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست اسناد دریافتنی")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'شماره جساب', 'دریافت کننده', 'شرح سند', ' شماره سند', 'مبلغ', 'تاریخ سند', 'توضیحات']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, receive in enumerate(receives, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        bank_account = ""
        if receive.bank and receive.bank.account_number:
            bank_account = str(receive.bank.account_number)

        ws.cell(row=row_num, column=2, value=bank_account)
        ws.cell(row=row_num, column=3, value=receive.payer_name)
        ws.cell(row=row_num, column=4, value=receive.description)
        ws.cell(row=row_num, column=5, value=receive.doc_number)
        ws.cell(row=row_num, column=6, value=receive.amount)
        jalali_date = jdatetime.date.fromgregorian(date=receive.doc_date).strftime('%Y/%m/%d')
        ws.cell(row=row_num, column=7, value=jalali_date)
        ws.cell(row=row_num, column=8, value=receive.details)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=receive.xlsx'
    wb.save(response)
    return response


# ============================ PaymentMoneyView ==========================
@method_decorator(admin_required, name='dispatch')
class PaymentMoneyCreateView(CreateView):
    model = PayMoney
    form_class = PayerMoneyForm
    template_name = 'payMoney/add_pay_money.html'
    success_url = reverse_lazy('add_pay')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            self.object = form.save()
            files = self.request.FILES.getlist('document')

            # ذخیره فایل‌ها در مدل ExpenseDocument
            for f in files:
                PayDocument.objects.create(payment=self.object, document=f)
            messages.success(self.request, 'سند پرداخت با موفقیت ثبت گردید!')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_queryset(self):
        queryset = PayMoney.objects.all()

        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank__id=bank_id)

        # فیلتر بر اساس amount
        amount = self.request.GET.get('amount')
        if amount:
            queryset = queryset.filter(amount__icontains=amount)

        receiver_name = self.request.GET.get('payer_name')
        if receiver_name:
            queryset = queryset.filter(receiver_name__icontains=receiver_name)

        description = self.request.GET.get('description')
        if description:
            queryset = queryset.filter(description__icontains=description)

        doc_number = self.request.GET.get('doc_number')
        if doc_number:
            queryset = queryset.filter(doc_number__icontains=doc_number)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس date
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')

        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(document_date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(document_date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')
        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receives = self.get_queryset()
        paginator = Paginator(receives, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_payments'] = PayMoney.objects.count()
        context['payments'] = PayMoney.objects.all()
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def pay_edit(request, pk):
    payment = get_object_or_404(PayMoney, pk=pk)

    if request.method == 'POST':
        form = PayerMoneyForm(request.POST, request.FILES, instance=payment)

        if form.is_valid():
            payment = form.save()  # Save the form (updates or creates expense)

            # Handle multiple file uploads
            files = request.FILES.getlist('document')
            if files:
                for f in files:
                    PayDocument.objects.create(payment=payment, document=f)

            messages.success(request, 'سند با موفقیت ویرایش شد.')
            return redirect('add_pay')  # Adjust redirect as necessary

        else:
            messages.error(request, 'خطا در ویرایش فرم درآمد. لطفا دوباره تلاش کنید.')
            return redirect('add_pay')
    else:
        # If the request is not POST, redirect to the appropriate page
        return redirect('add_pay')


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def pay_delete(request, pk):
    payment = get_object_or_404(PayMoney, id=pk)
    try:
        payment.delete()
        messages.success(request, ' سند با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('add_pay'))


@login_required(login_url=settings.LOGIN_URL_ADMIN)
@csrf_exempt
def delete_pay_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        payment_id = request.POST.get('payment_id')

        if not image_url or not payment_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            payment = get_object_or_404(PayMoney, id=payment_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = PayDocument.objects.filter(payment=payment, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_pay_pdf(request):
    payments = PayMoney.objects.all()

    filter_fields = {
        'bank': 'bank__id',
        'receiver_name': 'payer_name',
        'amount': 'amount__icontains',
        'document_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            payments = payments.filter(**filter_expression)

    # فیلتر تاریخ
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            payments = payments.filter(document_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            payments = payments.filter(document_date__lte=to_date)
    except ValueError:
        payments = PayMoney.objects.none()

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # رندر قالب HTML
    template = get_template("payMoney/pay_pdf.html")
    context = {
        'payments': payments,
        'font_path': font_url,
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="payments.pdf"'

    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_pay_excel(request):
    payments = PayMoney.objects.all()

    filter_fields = {
        'bank': 'bank__id',
        'receiver_name': 'receiver_name__icontains',
        'amount': 'amount__icontains',
        'doc_number': 'doc_number__icontains',
        'description': 'description__icontains',
        'details': 'details__icontains',
    }

    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            payments = payments.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            payments = payments.filter(document_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            payments = payments.filter(document_date__lte=to_date)
    except ValueError:
        payments = PayMoney.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "payments"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست اسناد پرداختنی")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'شماره جساب', 'دریافت کننده', 'شرح سند', ' شماره سند', 'مبلغ', 'تاریخ سند', 'توضیحات']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, payment in enumerate(payments, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        bank_account = ""
        if payment.bank and payment.bank.account_number:
            bank_account = str(payment.bank.account_number)

        ws.cell(row=row_num, column=2, value=bank_account)
        ws.cell(row=row_num, column=3, value=payment.receiver_name)
        ws.cell(row=row_num, column=4, value=payment.description)
        ws.cell(row=row_num, column=5, value=payment.document_number)
        ws.cell(row=row_num, column=6, value=payment.amount)
        jalali_date = jdatetime.date.fromgregorian(date=payment.document_date).strftime('%Y/%m/%d')
        ws.cell(row=row_num, column=7, value=jalali_date)
        ws.cell(row=row_num, column=8, value=payment.details)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=payment.xlsx'
    wb.save(response)
    return response


# ============================ PropertyView ==========================
@method_decorator(admin_required, name='dispatch')
class PropertyCreateView(CreateView):
    model = Property
    template_name = 'property/manage_property.html'
    form_class = PropertyForm
    success_url = reverse_lazy('add_property')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            self.object = form.save()
            files = self.request.FILES.getlist('document')

            for f in files:
                PropertyDocument.objects.create(property=self.object, document=f)
            messages.success(self.request, 'اموال با موفقیت ثبت گردید!')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_queryset(self):
        queryset = Property.objects.all()

        # فیلتر بر اساس amount
        property_name = self.request.GET.get('property_name')
        if property_name:
            queryset = queryset.filter(property_name__icontains=property_name)

        property_unit = self.request.GET.get('property_unit')
        if property_unit:
            queryset = queryset.filter(property_unit__icontains=property_unit)

        property_location = self.request.GET.get('property_location')
        if property_location:
            queryset = queryset.filter(property_location__icontains=property_location)

        property_code = self.request.GET.get('property_code')
        if property_code:
            queryset = queryset.filter(property_code__icontains=property_code)

        property_price = self.request.GET.get('property_price')
        if property_price:
            queryset = queryset.filter(property_price__icontains=property_price)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        # فیلتر بر اساس date
        from_date_str = self.request.GET.get('from_date')
        to_date_str = self.request.GET.get('to_date')

        try:
            if from_date_str:
                jalali_from = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d')
                gregorian_from = jalali_from.togregorian().date()
                queryset = queryset.filter(property_purchase_date__gte=gregorian_from)

            if to_date_str:
                jalali_to = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d')
                gregorian_to = jalali_to.togregorian().date()
                queryset = queryset.filter(property_purchase_date__lte=gregorian_to)
        except ValueError:
            messages.warning(self.request, 'فرمت تاریخ وارد شده صحیح نیست.')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receives = self.get_queryset()
        paginator = Paginator(receives, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['total_properties'] = Property.objects.count()
        context['properties'] = Property.objects.all()
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def property_edit(request, pk):
    property_d = get_object_or_404(Property, pk=pk)

    if request.method == 'POST':
        form = PropertyForm(request.POST, request.FILES, instance=property_d)

        if form.is_valid():
            property_d = form.save()  # Save the form (updates or creates expense)

            # Handle multiple file uploads
            files = request.FILES.getlist('document')
            if files:
                for f in files:
                    PropertyDocument.objects.create(property_d=property_d, document=f)

            messages.success(request, 'اموال با موفقیت ویرایش شد.')
            return redirect('add_property')  # Adjust redirect as necessary

        else:
            messages.error(request, 'خطا در ویرایش فرم درآمد. لطفا دوباره تلاش کنید.')
            return redirect('add_property')
    else:
        # If the request is not POST, redirect to the appropriate page
        return redirect('add_property')


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def property_delete(request, pk):
    property_d = get_object_or_404(Property, id=pk)
    try:
        property_d.delete()
        messages.success(request, ' اموال با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('add_property'))


@login_required(login_url=settings.LOGIN_URL_ADMIN)
@csrf_exempt
def delete_property_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        property_id = request.POST.get('property_id')

        print(f'property_id: {property_id}')

        if not image_url or not property_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            property = get_object_or_404(Property, id=property_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = PropertyDocument.objects.filter(property=property, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_property_pdf(request):
    properties = Property.objects.all()

    filter_fields = {
        'property_name': 'property_name__icontains',
        'property_unit': 'property_unit__icontains',
        'property_location': 'property_location__icontains',
        'property_code': 'property_code__icontains',
        'property_price': 'property_price__icontains',
        'details': 'details__icontains',

    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            properties = properties.filter(**filter_expression)

    # فیلتر تاریخ
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            properties = properties.filter(property_purchase_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            properties = properties.filter(property_purchase_date__lte=to_date)
    except ValueError:
        properties = Property.objects.none()

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # رندر قالب HTML
    template = get_template("property/property_pdf.html")
    context = {
        'properties': properties,
        'font_path': font_url,
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="properties.pdf"'

    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_property_excel(request):
    properties = Property.objects.all()

    filter_fields = {
        'property_name': 'property_name__icontains',
        'property_unit': 'property_unit__icontains',
        'property_location': 'property_location__icontains',
        'property_code': 'property_code__icontains',
        'property_price': 'property_price__icontains',
        'details': 'details__icontains',

    }

    # Apply filters based on query parameters
    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            filter_expression = {lookup: value}
            properties = properties.filter(**filter_expression)

    # Date range filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    try:
        if from_date_str:
            from_date = jdatetime.datetime.strptime(from_date_str, '%Y/%m/%d').togregorian().date()
            properties = properties.filter(property_purchase_date__gte=from_date)
        if to_date_str:
            to_date = jdatetime.datetime.strptime(to_date_str, '%Y/%m/%d').togregorian().date()
            properties = properties.filter(property_purchase_date__lte=to_date)
    except ValueError:
        properties = Property.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "properties"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست اموال ساختمان")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'نام اموال', 'واحد', ' شماره اموال', ' موقعیت ', 'ارزش', 'تاریخ خرید', 'توضیحات']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, property in enumerate(properties, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        ws.cell(row=row_num, column=2, value=property.property_name)
        ws.cell(row=row_num, column=3, value=property.property_unit)
        ws.cell(row=row_num, column=4, value=property.property_code)
        ws.cell(row=row_num, column=5, value=property.property_location)
        ws.cell(row=row_num, column=6, value=property.property_price)
        jalali_date = jdatetime.date.fromgregorian(date=property.property_purchase_date).strftime('%Y/%m/%d')
        ws.cell(row=row_num, column=7, value=jalali_date)
        ws.cell(row=row_num, column=8, value=property.details)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=properties.xlsx'
    wb.save(response)
    return response


# ============================ MaintenanceView ==========================
@method_decorator(admin_required, name='dispatch')
class MaintenanceCreateView(CreateView):
    model = Maintenance
    template_name = 'maintenance/add_maintenance.html'
    form_class = MaintenanceForm
    success_url = reverse_lazy('add_maintenance')

    def form_valid(self, form):
        form.instance.user = self.request.user
        try:
            self.object = form.save()
            files = self.request.FILES.getlist('document')

            for f in files:
                MaintenanceDocument.objects.create(maintenance=self.object, document=f)
            messages.success(self.request, 'سند با موفقیت ثبت گردید!')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_queryset(self):
        queryset = Maintenance.objects.all()

        # فیلتر بر اساس amount
        maintenance_description = self.request.GET.get('maintenance_description')
        if maintenance_description:
            queryset = queryset.filter(maintenance_description__icontains=maintenance_description)

        maintenance_price = self.request.GET.get('maintenance_price')
        if maintenance_price:
            queryset = queryset.filter(maintenance_price__icontains=maintenance_price)

        maintenance_status = self.request.GET.get('maintenance_status')
        if maintenance_status:
            queryset = queryset.filter(maintenance_status__icontains=maintenance_status)

        service_company = self.request.GET.get('service_company')
        if service_company:
            queryset = queryset.filter(service_company__icontains=service_company)

        maintenance_start_date = self.request.GET.get('maintenance_start_date')
        if maintenance_start_date and isinstance(maintenance_start_date, str):
            try:
                j_start = jdatetime.date.fromisoformat(maintenance_start_date)
                g_start = j_start.togregorian()
                queryset = queryset.filter(maintenance_start_date__gte=g_start)
            except (ValueError, TypeError) as e:
                print("Invalid date format:", maintenance_start_date, e)

        maintenance_end_date = self.request.GET.get('maintenance_end_date')
        if maintenance_end_date and isinstance(maintenance_end_date, str):
            try:
                j_start = jdatetime.date.fromisoformat(maintenance_end_date)
                g_start = j_start.togregorian()
                queryset = queryset.filter(maintenance_end_date__lte=g_start)
            except (ValueError, TypeError) as e:
                print("Invalid date format:", maintenance_end_date, e)

        details = self.request.GET.get('details')
        if details:
            queryset = queryset.filter(details__icontains=details)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        maintenances = self.get_queryset()
        paginator = Paginator(maintenances, 50)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['total_maintenances'] = maintenances.count()
        context['maintenances'] = page_obj.object_list
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def maintenance_edit(request, pk):
    maintenance = get_object_or_404(Maintenance, pk=pk)

    if request.method == 'POST':
        form = MaintenanceForm(request.POST, request.FILES, instance=maintenance)

        if form.is_valid():
            maintenance = form.save()  # Save the form (updates or creates expense)

            # Handle multiple file uploads
            files = request.FILES.getlist('document')
            if files:
                for f in files:
                    MaintenanceDocument.objects.create(maintenance=maintenance, document=f)

            messages.success(request, 'سند با موفقیت ویرایش شد.')
            return redirect('add_maintenance')  # Adjust redirect as necessary

        else:
            messages.error(request, 'خطا در ویرایش فرم درآمد. لطفا دوباره تلاش کنید.')
            return redirect('add_maintenance')
    else:
        # If the request is not POST, redirect to the appropriate page
        return redirect('add_maintenance')


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def maintenance_delete(request, pk):
    maintenance = get_object_or_404(Maintenance, id=pk)
    try:
        maintenance.delete()
        messages.success(request, ' سند با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('add_maintenance'))


@login_required(login_url=settings.LOGIN_URL_ADMIN)
@csrf_exempt
def delete_maintenance_document(request):
    if request.method == 'POST':
        image_url = request.POST.get('url')
        maintenance_id = request.POST.get('maintenance_id')

        print(f'maintenance_id: {maintenance_id}')

        if not image_url or not maintenance_id:
            return JsonResponse({'status': 'error', 'message': 'URL یا ID هزینه مشخص نیست'})

        try:
            maintenance = get_object_or_404(Maintenance, id=maintenance_id)

            relative_path = image_url.replace(settings.MEDIA_URL, '')  # دقیق کردن مسیر
            doc = MaintenanceDocument.objects.filter(maintenance=maintenance, document=relative_path).first()

            if doc:
                # Delete the file from filesystem
                if doc.document:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.document.name)
                    if os.path.exists(file_path):
                        os.remove(file_path)

                doc.delete()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'تصویر مرتبط پیدا نشد'})

        except Expense.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'هزینه یافت نشد'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'خطا در حذف تصویر: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'درخواست معتبر نیست'})


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def parse_jalali_to_gregorian(date_str):
    try:
        return jdatetime.date.fromisoformat(date_str.strip()).togregorian()
    except Exception:
        return None


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_maintenance_pdf(request):
    maintenances = Maintenance.objects.all()

    filter_fields = {
        'maintenance_description': 'maintenance_description__icontains',
        'maintenance_start_date': 'maintenance_start_date__gte',
        'maintenance_end_date': 'maintenance_end_date__lte',
        'maintenance_price': 'maintenance_price__icontains',
        'maintenance_status': 'maintenance_status__icontains',
        'service_company': 'service_company__icontains',
        'maintenance_document_no': 'maintenance_document_no__icontains',
        'details': 'details__icontains',
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            if field in ['maintenance_start_date', 'maintenance_end_date']:
                gregorian_date = parse_jalali_to_gregorian(value)
                if gregorian_date:
                    maintenances = maintenances.filter(**{lookup: gregorian_date})
            else:
                maintenances = maintenances.filter(**{lookup: value.strip()})

    # مسیر فونت
    font_url = request.build_absolute_uri('/static/fonts/BYekan.ttf')
    css = CSS(string=f"""
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{
                font-family: 'BYekan', sans-serif;
            }}
            @font-face {{
                font-family: 'BYekan';
                src: url('{font_url}');
            }}
        """)

    # رندر قالب HTML
    template = get_template("maintenance/maintenance_pdf.html")
    context = {
        'maintenances': maintenances,
        'font_path': font_url,
    }

    html = template.render(context)
    page_pdf = io.BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(page_pdf, stylesheets=[css])

    page_pdf.seek(0)

    # تولید پاسخ PDF
    pdf_merger = PdfWriter()
    pdf_merger.append(page_pdf)
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="maintenances.pdf"'

    pdf_merger.write(response)
    return response


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def export_maintenance_excel(request):
    maintenances = Maintenance.objects.all()

    filter_fields = {
        'maintenance_description': 'maintenance_description__icontains',
        'maintenance_start_date': 'maintenance_start_date__gte',
        'maintenance_end_date': 'maintenance_end_date__lte',
        'maintenance_price': 'maintenance_price__icontains',
        'maintenance_status': 'maintenance_status__icontains',
        'service_company': 'service_company__icontains',
        'maintenance_document_no': 'maintenance_document_no__icontains',
        'details': 'details__icontains',
    }

    for field, lookup in filter_fields.items():
        value = request.GET.get(field)
        if value:
            if field in ['maintenance_start_date', 'maintenance_end_date']:
                gregorian_date = parse_jalali_to_gregorian(value)
                if gregorian_date:
                    maintenances = maintenances.filter(**{lookup: gregorian_date})
            else:
                maintenances = maintenances.filter(**{lookup: value.strip()})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "maintenances"
    ws.sheet_view.rightToLeft = True

    # ✅ Add title
    title_cell = ws.cell(row=1, column=1, value="لیست هزینه های تعمیرات و نگهداری")
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    # ✅ Style setup
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    header_font = Font(bold=True, color="000000")  # Black bold text

    headers = ['#', 'شرح کار', 'تاریخ شروع', ' تاریخ پایان', ' اجرت/دستمزد ', 'شرکت خدماتی', 'شماره فاکتور',
               'توضیحات', 'آخرین وضعیت']

    # ✅ Write header (row 2)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    # ✅ Write data (start from row 3)
    for row_num, maintenance in enumerate(maintenances, start=3):
        ws.cell(row=row_num, column=1, value=row_num - 2)  # index starts from 1
        ws.cell(row=row_num, column=2, value=maintenance.maintenance_description)
        jalali_date = jdatetime.date.fromgregorian(date=maintenance.maintenance_start_date).strftime('%Y/%m/%d')
        ws.cell(row=row_num, column=3, value=jalali_date)
        jalali_date = jdatetime.date.fromgregorian(date=maintenance.maintenance_end_date).strftime('%Y/%m/%d')
        ws.cell(row=row_num, column=4, value=jalali_date)
        ws.cell(row=row_num, column=5, value=maintenance.maintenance_price)
        ws.cell(row=row_num, column=6, value=maintenance.service_company)
        ws.cell(row=row_num, column=7, value=maintenance.maintenance_document_no)
        ws.cell(row=row_num, column=8, value=maintenance.details)
        ws.cell(row=row_num, column=9, value=maintenance.maintenance_status)

    # ✅ Return file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=maintenances.xlsx'
    wb.save(response)
    return response


# ======================== Charge Views ======================================
@method_decorator(admin_required, name='dispatch')
class ChargeCategoryCreateView(CreateView):
    model = ChargeMethod
    template_name = 'charge/add_category_charge.html'
    form_class = ChargeCategoryForm
    success_url = reverse_lazy('add_charge_category')

    def form_valid(self, form):
        messages.success(self.request, 'روش شارژ با موفقیت ثبت گردید!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['charges'] = ChargeMethod.objects.all().order_by('code')
        return context


@method_decorator(admin_required, name='dispatch')
class ChargeCategoryUpdateView(UpdateView):
    model = ChargeMethod
    template_name = 'charge/add_category_charge.html'
    form_class = ChargeCategoryForm
    success_url = reverse_lazy('add_charge_category')

    def form_valid(self, form):
        edit_instance = form.instance
        self.object = form.save(commit=False)
        messages.success(self.request, 'روش شارژ با موفقیت ویرایش گردید!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['charges'] = ChargeMethod.objects.all()
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def charge_category_delete(request, pk):
    charge = get_object_or_404(ChargeMethod, id=pk)

    try:
        charge.delete()
        messages.success(request, 'روش شارژ با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('add_charge_category'))


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def charge_view(request):
    return render(request, 'charge/add_charge.html')


# ==================== Fix Charge ================================
@method_decorator(admin_required, name='dispatch')
class FixChargeCreateView(CreateView):
    model = FixCharge
    template_name = 'charge/fix_charge_template.html'
    form_class = FixChargeForm
    success_url = reverse_lazy('add_fixed_charge')
    paginate_by = 50

    def form_valid(self, form):
        form.instance.name = form.cleaned_data.get('name') or 'شارژ ثابت'
        units = Unit.objects.filter(is_active=True)

        if not units.exists():
            messages.warning(self.request, 'هیچ واحد فعالی یافت نشد.')
            return self.form_invalid(form)

        fix_charge = form.save(commit=False)
        fix_charge.user = self.request.user
        fix_charge.unit_count = units.count()
        fix_charge.civil = fix_charge.civil or 0
        fix_charge.payment_penalty_amount = fix_charge.payment_penalty_amount or 0
        fix_charge.save()

        messages.success(self.request, 'شارژ با موفقیت ثبت گردید.')
        return redirect(self.success_url)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unit_count = Unit.objects.filter(is_active=True).count()
        context['unit_count'] = unit_count

        charges = FixCharge.objects.annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')

        paginator = Paginator(charges, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['charges'] = page_obj.object_list
        return context


# ======================   Area Charge  =========================
@method_decorator(admin_required, name='dispatch')
class AreaChargeCreateView(CreateView):
    model = AreaCharge
    template_name = 'charge/area_charge_template.html'
    form_class = AreaChargeForm
    success_url = reverse_lazy('add_area_charge')
    paginate_by = 50

    def form_valid(self, form):
        charge_name = form.cleaned_data.get('name') or 'شارژ متراژ'
        area_charge = form.save(commit=False)
        area_charge.name = charge_name
        area_charge.user = self.request.user

        if area_charge.civil is None:
            area_charge.civil = 0

        try:
            area_charge.save()
            messages.success(self.request, 'محاسبه شارژ با موفقیت ثبت گردید')
            return redirect(self.success_url)
        except Exception as e:
            messages.error(self.request, f'خطا در ثبت! ({e})')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['charges'] = AreaCharge.objects.all()
        unit_count = Unit.objects.filter(is_active=True).count()
        context['unit_count'] = unit_count
        total_area = Unit.objects.filter(is_active=True).aggregate(total=Sum('area'))[
                         'total'] or 0
        context['total_area'] = total_area

        total_people = Unit.objects.filter(is_active=True, user=self.request.user.id).aggregate(
            total=Sum('people_count'))['total'] or 0
        context['total_people'] = total_people

        charges = AreaCharge.objects.annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        paginator = Paginator(charges, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['charges'] = page_obj.object_list
        return context


# ========================  person Charge  =======================
@method_decorator(admin_required, name='dispatch')
class PersonChargeCreateView(CreateView):
    model = PersonCharge
    template_name = 'charge/person_charge_template.html'
    form_class = PersonChargeForm
    success_url = reverse_lazy('add_person_charge')
    paginate_by = 50

    def form_valid(self, form):
        person_charge = form.save(commit=False)

        charge_name = form.cleaned_data.get('name') or 0
        person_charge.name = charge_name
        if person_charge.civil is None:
            person_charge.civil = 0

        total_people_count = Unit.objects.filter(is_active=True).aggregate(total=Sum('people_count'))['total'] or 0
        print(f"Total people count calculated: {total_people_count}")  # Debug line
        person_charge.total_people = total_people_count

        unit_count = Unit.objects.filter(is_active=True).count()
        person_charge.unit_count = unit_count

        try:
            person_charge.save()
            self.object = person_charge
            messages.success(self.request, 'محاسبه شارژ با موفقیت ثبت گردید')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['charges'] = PersonCharge.objects.all()
        unit_count = Unit.objects.filter(is_active=True).count()
        context['unit_count'] = unit_count
        total_area = Unit.objects.filter(is_active=True).aggregate(total=Sum('area'))[
                         'total'] or 0
        context['total_area'] = total_area
        total_people = Unit.objects.filter(is_active=True).aggregate(
            total=Sum('people_count'))['total'] or 0
        context['total_people'] = total_people

        charges = PersonCharge.objects.annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        paginator = Paginator(charges, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['charges'] = page_obj.object_list
        return context


# ======================= Fix Person Charge  ==========================
@method_decorator(admin_required, name='dispatch')
class FixPersonChargeCreateView(CreateView):
    model = FixPersonCharge
    template_name = 'charge/fix_person_charge_template.html'
    form_class = FixPersonChargeForm
    success_url = reverse_lazy('add_fix_person_charge')
    paginate_by = 50

    def form_valid(self, form):
        person_charge = form.save(commit=False)
        charge_name = form.cleaned_data.get('name') or 0
        person_charge.user = self.request.user
        person_charge.name = charge_name
        if person_charge.civil is None:
            person_charge.civil = 0

        total_people_count = Unit.objects.filter(is_active=True).aggregate(total=Sum('people_count'))['total'] or 0
        print(f"Total people count calculated: {total_people_count}")  # Debug line
        person_charge.total_people = total_people_count

        try:
            person_charge.save()
            self.object = person_charge
            messages.success(self.request, 'محاسبه شارژ با موفقیت ثبت گردید')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unit_count'] = Unit.objects.filter(is_active=True).count()
        context['total_area'] = Unit.objects.filter(is_active=True).aggregate(total=Sum('area'))['total'] or 0
        context['total_people'] = Unit.objects.filter(is_active=True).aggregate(total=Sum('people_count'))['total'] or 0

        charges = FixPersonCharge.objects.annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        paginator = Paginator(charges, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['charges'] = page_obj.object_list
        return context


# ==================== Fix Area Charge    =============================
@method_decorator(admin_required, name='dispatch')
class FixAreaChargeCreateView(CreateView):
    model = FixAreaCharge
    template_name = 'charge/fix_area_charge_template.html'
    form_class = FixAreaChargeForm
    success_url = reverse_lazy('add_fix_area_charge')
    paginate_by = 50

    def form_valid(self, form):
        fix_area_charge = form.save(commit=False)

        charge_name = form.cleaned_data.get('name') or 0
        fix_area_charge.name = charge_name
        if fix_area_charge.civil is None:
            fix_area_charge.civil = 0

        total_area = Unit.objects.filter(is_active=True).aggregate(total=Sum('area'))['total'] or 0
        print(f"Total people count calculated: {total_area}")  # Debug line
        fix_area_charge.total_area = total_area

        try:
            fix_area_charge.save()
            self.object = fix_area_charge
            messages.success(self.request, 'محاسبه شارژ با موفقیت ثبت گردید')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unit_count'] = Unit.objects.filter(is_active=True).count()
        context['total_area'] = Unit.objects.filter(is_active=True).aggregate(total=Sum('area'))['total'] or 0
        context['total_people'] = Unit.objects.filter(is_active=True).aggregate(total=Sum('people_count'))['total'] or 0

        charges = FixAreaCharge.objects.annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        paginator = Paginator(charges, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['charges'] = page_obj.object_list
        return context


# ================================= Person Area Charge =========================
@method_decorator(admin_required, name='dispatch')
class PersonAreaChargeCreateView(CreateView):
    model = ChargeByPersonArea
    template_name = 'charge/person_area_charge_template.html'
    form_class = PersonAreaChargeForm
    success_url = reverse_lazy('add_person_area_charge')
    paginate_by = 50

    def form_valid(self, form):
        person_area_charge = form.save(commit=False)
        charge_name = form.cleaned_data.get('name') or 0
        person_area_charge.name = charge_name
        person_area_charge.user = self.request.user
        if person_area_charge.civil is None:
            person_area_charge.civil = 0

        total_area = Unit.objects.filter(is_active=True).aggregate(total=Sum('area'))['total'] or 0
        print(f"Total people count calculated: {total_area}")  # Debug line
        person_area_charge.total_area = total_area

        total_people = Unit.objects.filter(is_active=True).aggregate(total=Sum('people_count'))['total'] or 0
        person_area_charge.total_people = total_people

        try:
            person_area_charge.save()
            self.object = person_area_charge
            messages.success(self.request, 'محاسبه شارژ با موفقیت ثبت گردید')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unit_count'] = Unit.objects.filter(is_active=True).count()
        context['total_area'] = Unit.objects.filter(is_active=True).aggregate(total=Sum('area'))['total'] or 0
        context['total_people'] = Unit.objects.filter(is_active=True).aggregate(total=Sum('people_count'))['total'] or 0

        charges = ChargeByPersonArea.objects.annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        paginator = Paginator(charges, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['charges'] = page_obj.object_list
        return context


# ==========================Fix Person Area Charge ================================
@method_decorator(admin_required, name='dispatch')
class PersonAreaFixChargeCreateView(CreateView):
    model = ChargeByFixPersonArea
    template_name = 'charge/person_area_fix_charge_template.html'
    form_class = PersonAreaFixChargeForm
    success_url = reverse_lazy('add_person_area_fix_charge')
    paginate_by = 50

    def form_valid(self, form):
        fix_person_area_charge = form.save(commit=False)
        charge_name = form.cleaned_data.get('name') or 0
        fix_person_area_charge.name = charge_name
        fix_person_area_charge.user = self.request.user
        if fix_person_area_charge.civil is None:
            fix_person_area_charge.civil = 0

        total_area = Unit.objects.filter(is_active=True).aggregate(total=Sum('area'))['total'] or 0
        print(f"Total people count calculated: {total_area}")  # Debug line
        fix_person_area_charge.total_area = total_area

        total_people = Unit.objects.filter(is_active=True).aggregate(total=Sum('people_count'))['total'] or 0
        fix_person_area_charge.total_people = total_people

        try:
            fix_person_area_charge.save()
            self.object = fix_person_area_charge
            messages.success(self.request, 'محاسبه شارژ با موفقیت ثبت گردید')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unit_count'] = Unit.objects.filter(is_active=True).count()
        context['total_area'] = Unit.objects.filter(is_active=True).aggregate(total=Sum('area'))['total'] or 0
        context['total_people'] = Unit.objects.filter(is_active=True).aggregate(total=Sum('people_count'))['total'] or 0

        charges = ChargeByFixPersonArea.objects.annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        paginator = Paginator(charges, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['charges'] = page_obj.object_list
        return context


# =========================ّFix Variable Charge =================================
@method_decorator(admin_required, name='dispatch')
class VariableFixChargeCreateView(CreateView):
    model = ChargeFixVariable
    template_name = 'charge/variable_fix_charge_template.html'
    form_class = VariableFixChargeForm
    success_url = reverse_lazy('add_variable_fix_charge')
    paginate_by = 50

    def form_valid(self, form):
        fix_variable_charge = form.save(commit=False)
        fix_variable_charge.user = self.request.user

        if fix_variable_charge.civil is None:
            fix_variable_charge.civil = 0

        if fix_variable_charge.extra_parking_amount is None:
            fix_variable_charge.extra_parking_amount = 0

        if fix_variable_charge.other_cost_amount is None:
            fix_variable_charge.other_cost_amount = 0

        total_area = Unit.objects.filter(is_active=True).aggregate(total=Sum('area'))['total'] or 0
        print(f"Total people count calculated: {total_area}")  # Debug line
        fix_variable_charge.total_area = total_area

        total_people = Unit.objects.filter(is_active=True).aggregate(total=Sum('people_count'))['total'] or 0
        fix_variable_charge.total_people = total_people

        try:
            fix_variable_charge.save()
            self.object = fix_variable_charge
            messages.success(self.request, 'محاسبه شارژ با موفقیت ثبت گردید')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unit_count'] = Unit.objects.filter(is_active=True).count()
        context['total_area'] = Unit.objects.filter(is_active=True).aggregate(total=Sum('area'))['total'] or 0
        context['total_people'] = Unit.objects.filter(is_active=True).aggregate(total=Sum('people_count'))['total'] or 0

        charges = ChargeFixVariable.objects.annotate(
            notified_count=Count(
                'unified_charges',
                filter=Q(unified_charges__send_notification=True)
            ),
            total_units=Count('unified_charges')
        ).order_by('-created_at')
        paginator = Paginator(charges, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['charges'] = page_obj.object_list
        return context


# =========================================================================================
@method_decorator(admin_required, name='dispatch')
class AdminSmsManagementView(CreateView):
    model = AdminSmsManagement
    template_name = 'admin_panel/register_sms.html'
    form_class = AdminSmsForm
    success_url = reverse_lazy('admin_register_sms')

    def form_valid(self, form):
        sms = form.save(commit=False)
        sms.user = self.request.user

        try:
            sms.save()
            self.object = sms
            messages.success(self.request, 'پیامک موفقیت ثبت گردید')
            return super().form_valid(form)
        except:
            messages.error(self.request, 'خطا در ثبت!')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_sms'] = AdminSmsManagement.objects.filter(user=self.request.user,
                                                               send_notification=False).order_by(
            '-created_at')
        context['units'] = Unit.objects.all()

        return context


@method_decorator(admin_required, name='dispatch')
class AdminSmsUpdateView(UpdateView):
    model = AdminSmsManagement
    template_name = 'admin_panel/register_sms.html'
    form_class = AdminSmsForm
    success_url = reverse_lazy('admin_register_sms')

    def form_valid(self, form):
        edit_instance = form.instance
        self.object = form.save(commit=False)
        messages.success(self.request, 'پیامک با موفقیت ویرایش گردید!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_sms'] = AdminSmsManagement.objects.filter(
            is_active=True,
            user=self.request.user,
            send_notification=False
        ).order_by('-created_at')
        return context


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def admin_sms_delete(request, pk):
    sms = get_object_or_404(AdminSmsManagement, pk=pk, user=request.user)
    try:
        sms.delete()
        messages.success(request, 'پیامک با موفقیت حذف گردید!')
    except ProtectedError:
        messages.error(request, " امکان حذف وجود ندارد! ")
    return redirect(reverse('admin_register_sms'))


@login_required(login_url=settings.LOGIN_URL_ADMIN)
def admin_show_send_sms_form(request, pk):
    sms = get_object_or_404(AdminSmsManagement, id=pk, user=request.user)

    # همه مدیران ساختمان فعال
    managers = User.objects.filter(is_middle_admin=True, is_active=True)

    return render(request, 'admin_panel/send_sms.html', {
        'sms': sms,
        'managers': managers,
    })


def admin_send_sms(request, pk):
    sms = get_object_or_404(AdminSmsManagement, id=pk, user=request.user)

    # ❌ اگر قبلاً ارسال شده
    if sms.send_notification:
        messages.warning(request, 'این پیامک قبلاً ارسال شده است.')
        return redirect('admin_sms_management')

    if request.method == "POST":
        # دریافت انتخاب‌ها
        selected_units = request.POST.getlist('units')
        selected_managers = request.POST.getlist('managers')  # فرض می‌کنیم checkbox با name="managers"

        if not selected_units and not selected_managers:
            messages.warning(request, 'هیچ  مدیری انتخاب نشده است.')
            return redirect('admin_sms_management')

        # QuerySet واحدها
        units_qs = Unit.objects.filter(
            is_active=True,
            user__manager=request.user
        )

        if 'all' in selected_units:
            units_to_notify = units_qs
        else:
            units_to_notify = units_qs.filter(id__in=selected_units)

        # QuerySet مدیران
        if 'all' in selected_managers:
            managers_to_notify = User.objects.filter(is_middle_admin=True, is_active=True)
        else:
            managers_to_notify = User.objects.filter(id__in=selected_managers, is_middle_admin=True, is_active=True)

        # جمع همه گیرندگان

        notified_users = []

        # ارسال به مدیران
        for manager in managers_to_notify:
            if manager.mobile:
                helper.send_sms_to_middle(
                    mobile=manager.mobile,
                    message=sms.message,
                    full_name=manager.full_name,
                )
                notified_users.append(manager)
        # ❌ اگر هیچ کاربری پیدا نشد
        if not notified_users:
            messages.warning(request, 'هیچ گیرنده معتبری برای ارسال پیامک پیدا نشد.')
            return redirect('admin_sms_management')

        # محاسبه تعداد پیامک و هزینه
        total_recipients_count = len(notified_users)
        sms_per_message = sms.sms_count
        total_sms_needed = total_recipients_count * sms_per_message
        sms_price = Decimal(str(settings.SMS_PRICE))
        total_price = total_sms_needed * sms_price

        # ذخیره اطلاعات پیامک
        sms.sms_per_message = sms_per_message
        sms.total_sms_sent = total_sms_needed
        sms.total_price = total_price

        # ثبت واحدها و مدیران در ManyToMany
        sms.notified_users.set(notified_users)  # اضافه کن اگر فیلد notified_users داری
        sms.send_notification = True
        sms.send_notification_date = timezone.now().date()
        sms.save()

        messages.success(
            request,
            f'پیامک با موفقیت برای {total_recipients_count} گیرنده ارسال شد.'
        )
        return redirect('admin_sms_management')

    # GET → نمایش فرم انتخاب واحدها و مدیران
    units_with_details = []
    units = Unit.objects.filter(is_active=True, user__manager=request.user).prefetch_related('renters').order_by('unit')
    for unit in units:
        active_renter = unit.renters.filter(renter_is_active=True).first()
        units_with_details.append({
            'unit': unit,
            'active_renter': active_renter
        })

    managers = User.objects.filter(is_middle_admin=True, is_active=True)

    return render(request, 'admin_panel/send_sms.html', {
        'sms': sms,
        'units_with_details': units_with_details,
        'managers': managers,
    })


@method_decorator(admin_required, name='dispatch')
class AdminSmsListView(ListView):
    model = AdminSmsManagement
    template_name = 'admin_panel/sms_management.html'
    context_object_name = 'all_sms'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')
        queryset = AdminSmsManagement.objects.filter(
            user=self.request.user,
            is_active=True,
            send_notification=True,
        )
        if query:
            queryset = queryset.filter(
                Q(subject__icontains=query) |
                Q(message__icontains=query)
            )
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        context['sms_list'] = AdminSmsManagement.objects.filter(user=self.request.user).order_by('-created_at')

        return context


@method_decorator(admin_required, name='dispatch')
class ApprovedSms(ListView):
    model = MyHouse
    template_name = 'admin_panel/approved_sms.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        qs = (
            MyHouse.objects
            .filter(house_sms__is_active=True)
            .annotate(
                total_sms=Count(
                    'house_sms',
                    filter=Q(house_sms__is_approved=False)
                )
            )
            .distinct()
        )

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@method_decorator(admin_required, name='dispatch')
class ApprovedSmsDetailView(ListView):
    model = SmsManagement
    template_name = "admin_panel/sms_approved_list.html"
    context_object_name = "all_sms"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = SmsManagement.objects.filter(
            house_id=house_id,

        )

        if query:
            qs = qs.filter(
                Q(title__icontains=query) |
                Q(unit__owner_name__icontains=query) |
                Q(unit__renters__renter_name__icontains=query) |
                Q(total_charge_month_str__icontains=query) |
                Q(base_charge_str__icontains=query) |
                Q(details__icontains=query)
            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # اضافه کردن اطلاعات خانه
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        return context


def approve_sms(request, pk):
    sms = get_object_or_404(SmsManagement, pk=pk)

    sms.is_approved = True
    sms.approved_at = timezone.now()
    sms.save()

    messages.success(request, "پیامک تایید شد.")
    return redirect("sms_approved_list", house_id=sms.house.id)


def disapprove_sms(request, pk):
    sms = get_object_or_404(SmsManagement, pk=pk)

    if sms.send_notification:
        messages.success(request, " عدم تایید پیامک بدلیل ارسال آن توسط مدیر ساختمان امکانپذیر نیست")

    sms.is_approved = False
    sms.approved_at = None
    sms.save()

    messages.success(request, "پیامک تایید نگردید.")
    return redirect("sms_approved_list", house_id=sms.house.id)


@method_decorator(admin_required, name='dispatch')
class middleSmsManagementReport(ListView):
    model = MyHouse
    template_name = 'admin_panel/middle_sms_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        qs = (
            MyHouse.objects
            .filter(house_sms__is_active=True)
            .annotate(
                total_credit_sms=Count(
                    'house_sms',
                    filter=Q(house_sms__is_active=True)
                )
            )
            .distinct()
        )

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@method_decorator(admin_required, name='dispatch')
class middleReportSmsDetailView(ListView):
    model = SmsManagement
    template_name = "admin_panel/middle_sms_report_list.html"
    context_object_name = "all_sms"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = SmsManagement.objects.filter(
            house_id=house_id,

        )

        if query:
            qs = qs.filter(
                Q(subject__icontains=query)

            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        context['sms_list'] = SmsManagement.objects.filter(user=self.request.user).annotate(
            unit_number=F('notified_units__unit'),
            user_full_name=F('notified_units__user__full_name')
        )
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        return context


@method_decorator(admin_required, name='dispatch')
class CreditSmsManagement(ListView):
    model = MyHouse
    template_name = 'admin_panel/credit_sms_management.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        qs = (
            MyHouse.objects
            .filter(sms_credit__is_paid=True)
            .annotate(
                total_credit_sms=Count(
                    'sms_credit',
                    filter=Q(sms_credit__is_paid=True)
                )
            )
            .distinct()
        )

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@method_decorator(admin_required, name='dispatch')
class CreditSmsDetailView(ListView):
    model = SmsCredit
    template_name = "admin_panel/credit_sms_list.html"
    context_object_name = "all_credit"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = SmsCredit.objects.filter(
            house_id=house_id,

        )

        if query:
            qs = qs.filter(
                Q(amount__icontains=query)
            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # اضافه کردن اطلاعات خانه
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        return context


@method_decorator(admin_required, name='dispatch')
class AdminTicketReport(ListView):
    model = MyHouse
    template_name = 'admin_panel/admin_ticket_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        qs = (
            MyHouse.objects
            .filter(admin_tickets__is_sent=True)
            .annotate(
                total_tickets=Count(
                    'admin_tickets',
                    filter=Q(admin_tickets__is_sent=True)
                )
            )
            .distinct()
        )

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@method_decorator(admin_required, name='dispatch')
class AdminTicketDetailView(ListView):
    model = AdminTicket
    template_name = "admin_panel/admin_tickets_report_list.html"
    context_object_name = "admin_tickets"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = AdminTicket.objects.filter(
            house_id=house_id
        )

        if query:
            qs = qs.filter(
                Q(subject__icontains=query)
            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')

        context['house'] = MyHouse.objects.filter(
            id=self.kwargs['house_id']
        ).first()

        return context


# ==================================================================
@login_required
def get_notes(request, year, month):
    notes = CalendarNote.objects.filter(user=request.user, year=year, month=month)
    data = {note.day: note.note for note in notes}
    return JsonResponse(data)


@login_required
@csrf_exempt
def save_note(request):
    if request.method == "POST":
        data = json.loads(request.body)
        year = data.get('year')
        month = data.get('month')
        day = data.get('day')
        note_text = data.get('note', '').strip()
        note, created = CalendarNote.objects.get_or_create(
            user=request.user, year=year, month=month, day=day
        )
        if note_text:
            note.note = note_text
            note.save()
        else:
            note.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)


@csrf_exempt
@login_required
def delete_note(request):
    if request.method == "POST":
        payload = json.loads(request.body)
        year = payload.get("year")
        month = payload.get("month")
        day = payload.get("day")
        CalendarNote.objects.filter(user=request.user, year=year, month=month, day=day).delete()
        return JsonResponse({"status": "ok"})
    return JsonResponse({"status": "error"}, status=400)


# ========================== Fund Report =============================================
@method_decorator(admin_required, name='dispatch')
class AdminFundReport(ListView):
    model = MyHouse
    template_name = 'report/Fund_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        qs = (
            MyHouse.objects
            .filter(house_funds__id__isnull=False)
            .annotate(
                total_fund=Count('house_funds')
            )
            .distinct()
        )

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@method_decorator(admin_required, name='dispatch')
class AdminFundReportDetailView(ListView):
    model = Fund
    template_name = "report/admin_fund_detail.html"
    context_object_name = "funds"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = Fund.objects.filter(
            house_id=house_id,
        )

        if query:
            qs = qs.filter(
                Q(payment_description__icontains=query) |
                Q(payer_name__icontains=query) |
                Q(receiver_name__icontains=query) |
                Q(creditor_amount__icontains=query) |
                Q(debtor_amount__icontains=query)
            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # اضافه کردن اطلاعات خانه
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        return context


# ================================== Bank Report ===================================
@method_decorator(admin_required, name='dispatch')
class AdminBanksReport(ListView):
    model = MyHouse
    template_name = 'report/bank_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        qs = (
            MyHouse.objects
            .filter(banks__is_active=True)
            .annotate(
                total_banks=Count(
                    'banks',
                    filter=Q(banks__is_active=True)
                )
            )
            .distinct()
        )

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@method_decorator(admin_required, name='dispatch')
class AdminBanksListReportView(ListView):
    model = Bank
    template_name = "report/admin_bank_list.html"
    context_object_name = "banks"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = Bank.objects.filter(
            house_id=house_id,
        )

        if query:
            qs = qs.filter(
                Q(bank_name__icontains=query) |
                Q(account_no__icontains=query) |
                Q(account_holder_name__icontains=query) |
                Q(sheba_number__icontains=query) |
                Q(cart_number__icontains=query)

            )

        return qs.order_by('-create_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        context['house'] = MyHouse.objects.filter(
            id=self.kwargs['house_id']
        ).first()

        banks = context['banks'].prefetch_related(
            Prefetch(
                'fund_set',
                queryset=Fund.objects.order_by('doc_number')
            )
        )

        bank_transactions = {}

        for bank in banks:
            running_total = Decimal('0')
            transactions = []

            for f in bank.fund_set.all():
                running_total += (
                        (f.debtor_amount or Decimal('0')) -
                        (f.creditor_amount or Decimal('0'))
                )

                transactions.append({
                    'date': f.payment_date,
                    'description': f.payment_description,
                    'debit': f.debtor_amount,
                    'credit': f.creditor_amount,
                    'balance': running_total
                })

            bank_transactions[bank.id] = {
                'transactions': transactions,
                'balance': running_total
            }

        context['bank_transactions'] = bank_transactions
        return context


@login_required
def admin_bank_detail_view(request, bank_id):
    bank = get_object_or_404(
        Bank.objects.prefetch_related(
            Prefetch(
                'fund_set',
                queryset=Fund.objects.order_by('doc_number')
            )
        ),
        id=bank_id
    )

    funds = bank.fund_set.all()

    # -------- paginate --------
    page_number = request.GET.get('page', 1)
    paginator = Paginator(funds, 20)  # 20 تراکنش در هر صفحه
    page_obj = paginator.get_page(page_number)
    # --------------------------

    running_total = Decimal('0')
    transactions = []

    # محاسبه running balance فقط برای آیتم‌های همین صفحه
    # (اگر running کل تاریخچه میخوای پایین‌تر توضیح دادم)
    for f in page_obj:
        running_total += (
                (f.debtor_amount or Decimal('0')) -
                (f.creditor_amount or Decimal('0'))
        )

        transactions.append({
            'date': f.payment_date,
            'description': f.payment_description,
            'debit': f.debtor_amount,
            'credit': f.creditor_amount,
            'balance': running_total
        })

    context = {
        'bank': bank,
        'house': bank.house,
        'transactions': transactions,
        'page_obj': page_obj,
        'balance': running_total,
    }

    return render(request, 'report/admin_bank_detail.html', context)


# ============================ unit Fund report ====================================

@method_decorator(admin_required, name='dispatch')
class AdminUnitFundReport(ListView):
    model = MyHouse
    template_name = 'report/unit_fund_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        # حالا خانه‌ها را با تعداد واحد فعال
        qs = MyHouse.objects.annotate(
            total_units=Count('units', filter=Q(units__is_active=True))
        ).filter(total_units__gt=0)

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # context['house'] = MyHouse.objects.filter(
        #     id=self.kwargs['house_id']
        # ).first()
        return context


@method_decorator(admin_required, name='dispatch')
class AdminUnitsListReportView(ListView):
    model = Unit
    template_name = 'report/admin_unit_list.html'
    context_object_name = 'object_list'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None
        return int(paginate) if paginate and paginate.isdigit() else 20

    def get_queryset(self):
        house_id = self.kwargs['house_id']

        # پایه: همه واحدهای خانه و فعال + تعداد تراکنش‌ها
        qs = (
            Unit.objects
            .filter(myhouse_id=house_id, is_active=True)
            .annotate(
                total_transaction=Count(
                    'funds',  # همان related_name واقعی
                    filter=Q(funds__isnull=False)
                )
            )

            .prefetch_related(
                Prefetch(
                    'renters',
                    queryset=Renter.objects.filter(renter_is_active=True),
                    to_attr='active_renters'
                )
            )
            .order_by('unit')
        )

        params = self.request.GET
        filters = Q()

        if params.get('unit') and params['unit'].isdigit():
            filters &= Q(unit=int(params['unit']))
        if params.get('owner_name'):
            filters &= Q(owner_name__icontains=params['owner_name'])
        if params.get('owner_mobile'):
            filters &= Q(owner_mobile__icontains=params['owner_mobile'])
        if params.get('area') and params['area'].isdigit():
            filters &= Q(area=int(params['area']))
        if params.get('bedrooms_count') and params['bedrooms_count'].isdigit():
            filters &= Q(bedrooms_count=int(params['bedrooms_count']))
        if params.get('renter_name'):
            filters &= Q(renters__renter_name__icontains=params['renter_name'])
        if params.get('renter_mobile'):
            filters &= Q(renters__renter_mobile__icontains=params['renter_mobile'])
        if params.get('people_count') and params['people_count'].isdigit():
            filters &= Q(owner_people_count=int(params['people_count']))
        if params.get('status_residence'):
            filters &= Q(status_residence__icontains=params['status_residence'])
        if params.get('is_renter'):
            filters &= Q(is_renter__icontains=params['is_renter'])

        qs = qs.filter(filters).distinct()

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        house = MyHouse.objects.filter(
            id=self.kwargs['house_id']
        ).first()
        context['house'] = house
        context['total_units'] = Unit.objects.filter(myhouse_id=house, is_active=True).count()
        context['query_params'] = self.request.GET.urlencode()
        return context


@login_required
def admin_Unit_Fund_detail(request, unit_id):
    # گرفتن واحد برای اطلاعات اضافی در template
    unit = get_object_or_404(Unit, id=unit_id, is_active=True)

    # فقط فاندهای همین واحد
    funds = (
        Fund.objects
        .filter(unit=unit)
        .select_related('bank', 'content_type')
        .order_by('-payment_date')
    )

    # -------- paginate --------
    page_number = request.GET.get('page', 1)
    paginator = Paginator(funds, 20)  # 20 تراکنش در هر صفحه
    page_obj = paginator.get_page(page_number)
    # --------------------------

    # محاسبه running balance فقط برای همین صفحه
    running_total = Decimal('0')
    transactions = []
    for f in page_obj:
        running_total += (f.debtor_amount or Decimal('0')) - (f.creditor_amount or Decimal('0'))
        transactions.append({
            'payment_date': f.payment_date,
            'payment_description': f.payment_description,
            'debtor_amount': f.debtor_amount,
            'creditor_amount': f.creditor_amount,
            'payment_gateway': f.payment_gateway,
            'transaction_no': f.transaction_no,
            'balance': running_total,
            'unit': unit.get_label()
        })

    context = {
        'unit': unit,
        'house': unit.myhouse,  # ← اضافه شد
        'transactions': transactions,
        'page_obj': page_obj,
        'balance': running_total,
    }

    return render(request, 'report/admin_unit_fund_detail.html', context)


# ======================   issued Charge   ====================================

@method_decorator(admin_required, name='dispatch')
class ChargeIssued(ListView):
    model = MyHouse
    template_name = 'report/charge_issued.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        qs = (
            MyHouse.objects
            .filter(unified_charges__send_notification=True)
            .annotate(
                total_charges=Count(
                    'unified_charges',
                    filter=Q(unified_charges__send_notification=True)
                )
            )
            .distinct()
        )

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        return context


@method_decorator(admin_required, name='dispatch')
class ChargeIssuedDetailView(ListView):
    model = UnifiedCharge
    template_name = "report/detail_charges.html"
    context_object_name = "unified_charges"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = UnifiedCharge.objects.filter(
            house_id=house_id,
            send_notification=True
        )

        if query:
            qs = qs.annotate(
                total_charge_month_str=Cast('total_charge_month', CharField()),
                base_charge_str=Cast('base_charge', CharField())
            ).filter(
                Q(title__icontains=query) |
                Q(unit__owner_name__icontains=query) |
                Q(unit__renters__renter_name__icontains=query) |
                Q(total_charge_month_str__icontains=query) |
                Q(base_charge_str__icontains=query) |
                Q(details__icontains=query)
            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # اضافه کردن اطلاعات خانه
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        for charge in context['unified_charges']:
            charge.update_penalty()
        return context


# ==================================== Debtor unit report ========================================
@method_decorator(admin_required, name='dispatch')
class AdminDebtorReport(ListView):
    model = MyHouse
    template_name = 'report/unit_debtor_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')
        qs = MyHouse.objects.annotate(
            total_debtors=Count(
                'units__unified_charges',
                filter=Q(units__unified_charges__is_paid=False)
            )
        ).filter(total_debtors__gt=0)

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query = self.request.GET.get('q', '').strip()
        context['query'] = query
        context['paginate'] = self.request.GET.get('paginate', '20')

        houses_data = []

        for house in context['houses']:
            total_debt_house = UnifiedCharge.objects.filter(
                is_paid=False,
                unit__myhouse=house
            ).aggregate(total=Sum('total_charge_month'))['total'] or 0

            debt_units_count = UnifiedCharge.objects.filter(
                is_paid=False,
                unit__myhouse=house
            ).values('unit').distinct().count()

            houses_data.append({
                'house': house,
                'total_debt_house': total_debt_house,
                'debt_units_count': debt_units_count
            })

        context['houses_data'] = houses_data
        return context


@method_decorator(admin_required, name='dispatch')
class AdminDebtorUnitDetailView(DetailView):
    model = MyHouse
    template_name = 'report/admin_debtor_detail.html'
    context_object_name = 'house'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        house = self.object
        query = self.request.GET.get('q', '').strip()

        units = house.units.filter(is_active=True)

        unit_debts = []
        for unit in units:
            charges = unit.unified_charges.filter(is_paid=False)
            if query:
                charges = charges.filter(
                    Q(unit__unit__icontains=query) |
                    Q(title__icontains=query) |
                    Q(unit__owner_name__icontains=query) |
                    Q(unit__renters__renter_name__icontains=query)
                ).distinct()

            total_debt_unit = sum(c.total_charge_month or 0 for c in charges)
            if total_debt_unit > 0:
                renter = unit.get_active_renter()
                label = f"واحد {unit.unit} - {renter.renter_name}" if renter else f"واحد {unit.unit} - {unit.owner_name}"
                unit_debts.append({
                    'id': unit.id,
                    'label': label,
                    'total_debt': total_debt_unit,
                    'charges': charges
                })

        context['unit_debts'] = unit_debts
        context['total_debt_house'] = sum(u['total_debt'] for u in unit_debts)
        context['query'] = query
        return context


# ==================================== History Unit report ========================================
@method_decorator(admin_required, name='dispatch')
class AdminUnitHistoryReport(ListView):
    model = MyHouse
    template_name = 'report/unit_history_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        # حالا خانه‌ها را با تعداد واحد فعال
        qs = MyHouse.objects.annotate(
            total_units=Count('units', filter=Q(units__is_active=True))
        ).filter(total_units__gt=0)

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # context['house'] = MyHouse.objects.filter(
        #     id=self.kwargs['house_id']
        # ).first()
        return context


class AdminUnitsHistoryListView(ListView):
    model = Unit
    template_name = 'report/admin_unit_history_list.html'
    context_object_name = 'object_list'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None
        return int(paginate) if paginate and paginate.isdigit() else 20

    def get_queryset(self):
        house_id = self.kwargs['house_id']

        # پایه: همه واحدهای خانه و فعال + تعداد تراکنش‌ها
        qs = (
            Unit.objects
            .filter(myhouse_id=house_id, is_active=True)

            .prefetch_related(
                Prefetch(
                    'renters',
                    queryset=Renter.objects.filter(renter_is_active=True),
                    to_attr='active_renters'
                )
            )
            .order_by('unit')
        )

        params = self.request.GET
        filters = Q()

        if params.get('unit') and params['unit'].isdigit():
            filters &= Q(unit=int(params['unit']))
        if params.get('owner_name'):
            filters &= Q(owner_name__icontains=params['owner_name'])
        if params.get('owner_mobile'):
            filters &= Q(owner_mobile__icontains=params['owner_mobile'])
        if params.get('area') and params['area'].isdigit():
            filters &= Q(area=int(params['area']))
        if params.get('bedrooms_count') and params['bedrooms_count'].isdigit():
            filters &= Q(bedrooms_count=int(params['bedrooms_count']))
        if params.get('renter_name'):
            filters &= Q(renters__renter_name__icontains=params['renter_name'])
        if params.get('renter_mobile'):
            filters &= Q(renters__renter_mobile__icontains=params['renter_mobile'])
        if params.get('people_count') and params['people_count'].isdigit():
            filters &= Q(owner_people_count=int(params['people_count']))
        if params.get('status_residence'):
            filters &= Q(status_residence__icontains=params['status_residence'])
        if params.get('is_renter'):
            filters &= Q(is_renter__icontains=params['is_renter'])

        qs = qs.filter(filters).distinct()

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        house = MyHouse.objects.filter(
            id=self.kwargs['house_id']
        ).first()
        context['house'] = house
        context['total_units'] = Unit.objects.filter(myhouse_id=house, is_active=True).count()
        context['query_params'] = self.request.GET.urlencode()
        return context


@login_required
def admin_Unit_history_detail(request, unit_id):
    # گرفتن واحد برای اطلاعات اضافی در template
    unit = get_object_or_404(Unit, id=unit_id, is_active=True)

    # فقط فاندهای همین واحد
    histories = (
        UnitResidenceHistory.objects
        .filter(unit=unit)
        .order_by('-created_at')
    )

    # -------- paginate --------
    page_number = request.GET.get('page', 1)
    paginator = Paginator(histories, 20)  # 20 تراکنش در هر صفحه
    page_obj = paginator.get_page(page_number)
    # --------------------------

    context = {
        'unit': unit,
        'house': unit.myhouse,  # ← اضافه شد
        'page_obj': page_obj,
        'unit_detail': unit.get_label()
    }

    return render(request, 'report/unit_history_detail.html', context)


# ==================================== Expense Report ========================================
@method_decorator(admin_required, name='dispatch')
class AdminExpenseReport(ListView):
    model = MyHouse
    template_name = 'report/expenses_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        # حالا خانه‌ها را با تعداد واحد فعال
        qs = MyHouse.objects.annotate(
            total_expenses=Count('house_expenses', filter=Q(house_expenses__is_active=True))
        ).filter(total_expenses__gt=0)

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # context['house'] = MyHouse.objects.filter(
        #     id=self.kwargs['house_id']
        # ).first()
        return context


@method_decorator(admin_required, name='dispatch')
class AdminExpensesDetailView(ListView):
    model = Expense
    template_name = "report/admin_expense_detail.html"
    context_object_name = "expenses"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = Expense.objects.filter(
            house_id=house_id,
        )

        if query:
            qs = qs.filter(
                Q(category__id__icontains=query) |
                Q(amount__icontains=query) |
                Q(description__icontains=query) |
                Q(doc_no__icontains=query) |
                Q(details__icontains=query) |
                Q(receiver_name__icontains=query)
            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # اضافه کردن اطلاعات خانه
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        return context


# ==================================== Income Report ========================================
@method_decorator(admin_required, name='dispatch')
class AdminIncomeReport(ListView):
    model = MyHouse
    template_name = 'report/incomes_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        # حالا خانه‌ها را با تعداد واحد فعال
        qs = MyHouse.objects.annotate(
            total_incomes=Count('house_incomes', filter=Q(house_incomes__is_active=True))
        ).filter(total_incomes__gt=0)

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # context['house'] = MyHouse.objects.filter(
        #     id=self.kwargs['house_id']
        # ).first()
        return context


@method_decorator(admin_required, name='dispatch')
class AdminIncomesDetailView(ListView):
    model = Income
    template_name = "report/admin_income_detail.html"
    context_object_name = "incomes"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = Income.objects.filter(
            house_id=house_id,
        )

        if query:
            qs = qs.filter(
                Q(category__id__icontains=query) |
                Q(unit__id__icontains=query) |
                Q(amount__icontains=query) |
                Q(description__icontains=query) |
                Q(doc_number__icontains=query) |
                Q(details__icontains=query) |
                Q(payer_name__icontains=query)
            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # اضافه کردن اطلاعات خانه
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        return context


# ==================================== Receive Report ========================================

@method_decorator(admin_required, name='dispatch')
class AdminReceiveReport(ListView):
    model = MyHouse
    template_name = 'report/receive_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        # حالا خانه‌ها را با تعداد واحد فعال
        qs = MyHouse.objects.annotate(
            total_receives=Count('house_receives', filter=Q(house_receives__is_active=True))
        ).filter(total_receives__gt=0)

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # context['house'] = MyHouse.objects.filter(
        #     id=self.kwargs['house_id']
        # ).first()
        return context


@method_decorator(admin_required, name='dispatch')
class AdminReceiveDetailView(ListView):
    model = ReceiveMoney
    template_name = "report/receive_report_detail.html"
    context_object_name = "receives"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = ReceiveMoney.objects.filter(
            house_id=house_id,
        )

        if query:
            qs = qs.filter(
                Q(property_name__icontains=query) |
                Q(property_unit__icontains=query) |
                Q(property_location__icontains=query) |
                Q(property_code__icontains=query) |
                Q(property_price__icontains=query) |
                Q(details__icontains=query)

            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # اضافه کردن اطلاعات خانه
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        return context


# ==================================== Pay Report ========================================

@method_decorator(admin_required, name='dispatch')
class AdminPayReport(ListView):
    model = MyHouse
    template_name = 'report/pay_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        # حالا خانه‌ها را با تعداد واحد فعال
        qs = MyHouse.objects.annotate(
            total_payments=Count('house_payments', filter=Q(house_payments__is_active=True))
        ).filter(total_payments__gt=0)

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # context['house'] = MyHouse.objects.filter(
        #     id=self.kwargs['house_id']
        # ).first()
        return context


@method_decorator(admin_required, name='dispatch')
class AdminPayDetailView(ListView):
    model = PayMoney
    template_name = "report/pay_report_detail.html"
    context_object_name = "payments"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = PayMoney.objects.filter(
            house_id=house_id,
        )

        if query:
            qs = qs.filter(
                Q(property_name__icontains=query) |
                Q(property_unit__icontains=query) |
                Q(property_location__icontains=query) |
                Q(property_code__icontains=query) |
                Q(property_price__icontains=query) |
                Q(details__icontains=query)

            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # اضافه کردن اطلاعات خانه
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        return context


# ==================================== Property Report ========================================
@method_decorator(admin_required, name='dispatch')
class AdminPropertyReport(ListView):
    model = MyHouse
    template_name = 'report/property_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        # حالا خانه‌ها را با تعداد واحد فعال
        qs = MyHouse.objects.annotate(
            total_property=Count('house_property', filter=Q(house_property__is_active=True))
        ).filter(total_property__gt=0)

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # context['house'] = MyHouse.objects.filter(
        #     id=self.kwargs['house_id']
        # ).first()
        return context


@method_decorator(admin_required, name='dispatch')
class AdminPropertyDetailView(ListView):
    model = Property
    template_name = "report/admin_property_detail.html"
    context_object_name = "all_property"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = Property.objects.filter(
            house_id=house_id,
        )

        if query:
            qs = qs.filter(
                Q(property_name__icontains=query) |
                Q(property_unit__icontains=query) |
                Q(property_location__icontains=query) |
                Q(property_code__icontains=query) |
                Q(property_price__icontains=query) |
                Q(details__icontains=query)

            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # اضافه کردن اطلاعات خانه
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        return context


# ==================================== Maintenance Report ========================================
@method_decorator(admin_required, name='dispatch')
class AdminMaintenanceReport(ListView):
    model = MyHouse
    template_name = 'report/maintenance_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        # حالا خانه‌ها را با تعداد واحد فعال
        qs = MyHouse.objects.annotate(
            total_maintenances=Count('house_maintenance', filter=Q(house_maintenance__is_active=True))
        ).filter(total_maintenances__gt=0)

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # context['house'] = MyHouse.objects.filter(
        #     id=self.kwargs['house_id']
        # ).first()
        return context


@method_decorator(admin_required, name='dispatch')
class AdminMaintenanceDetailView(ListView):
    model = Maintenance
    template_name = "report/admin_maintenance_detail.html"

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        house_id = self.kwargs['house_id']
        query = self.request.GET.get('q', '')

        qs = Maintenance.objects.filter(
            house_id=house_id,
        )

        if query:
            qs = qs.filter(
                Q(property_name__icontains=query) |
                Q(property_unit__icontains=query) |
                Q(property_location__icontains=query) |
                Q(property_code__icontains=query) |
                Q(property_price__icontains=query) |
                Q(details__icontains=query)

            )

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # اضافه کردن اطلاعات خانه
        context['house'] = MyHouse.objects.filter(id=self.kwargs['house_id']).first()
        return context


# ==================================== Billan Report ========================================

@method_decorator(admin_required, name='dispatch')
class AdminBillanReport(ListView):
    model = MyHouse
    template_name = 'report/billan_report.html'
    context_object_name = 'houses'

    def get_paginate_by(self, queryset):
        paginate = self.request.GET.get('paginate')
        if paginate == '1000':
            return None  # نمایش همه آیتم‌ها
        return int(paginate or 20)

    def get_queryset(self):
        query = self.request.GET.get('q', '')

        qs = MyHouse.objects.all()

        if query:
            qs = qs.filter(
                Q(name__icontains=query) |
                Q(user__full_name__icontains=query)
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['paginate'] = self.request.GET.get('paginate', '20')
        # context['house'] = MyHouse.objects.filter(
        #     id=self.kwargs['house_id']
        # ).first()
        return context


def admin_house_balance(request, house_id):
    house = get_object_or_404(MyHouse, id=house_id)
    bank_id = request.GET.get('bank')

    start_date_j = request.GET.get('start_date')
    end_date_j = request.GET.get('end_date')

    # تبدیل تاریخ شمسی به میلادی
    start_date = jalali_to_gregorian(start_date_j) if start_date_j else None
    end_date = jalali_to_gregorian(end_date_j) if end_date_j else None

    # ---------- فیلترها ----------

    # برای پرداخت‌ها و دریافت‌ها
    payment_filter = {}
    if start_date:
        payment_filter['payment_date__gte'] = start_date
    if end_date:
        payment_filter['payment_date__lte'] = end_date
    if bank_id:
        payment_filter['bank_id'] = bank_id

    # برای درآمدهای دریافت نشده (doc_date)
    doc_income_filter = {}
    if start_date:
        doc_income_filter['doc_date__gte'] = start_date
    if end_date:
        doc_income_filter['doc_date__lte'] = end_date

    # برای هزینه‌های دریافت نشده (date)
    doc_expense_filter = {}
    if start_date:
        doc_expense_filter['date__gte'] = start_date
    if end_date:
        doc_expense_filter['date__lte'] = end_date

    # ---------- کوئری‌ها ----------

    total_incomes_exclude_unpaid = Income.objects.filter(
        is_paid=False,
        house=house,
        **doc_income_filter
    ).aggregate(Sum('amount'))[
        'amount__sum']

    total_expenses_exclude_unpaid = Expense.objects.filter(
        is_paid=False,
        house=house,
        **doc_expense_filter
    ).aggregate(Sum('amount'))[
        'amount__sum']

    total_incomes = Income.objects.filter(
        house=house,
        is_paid=True,
        **payment_filter,
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_expenses = Expense.objects.filter(
        house=house,
        is_paid=True,
        **payment_filter,
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_pay_money = PayMoney.objects.filter(
        house=house,
        is_paid=True,
        **payment_filter,
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_receive_money = ReceiveMoney.objects.filter(
        house=house,
        is_paid=True,
        **payment_filter,
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_assets = total_incomes + total_receive_money
    total_debts = total_pay_money + total_expenses
    total_amount_assets_debts = total_assets - total_debts

    funds = Fund.objects.filter(
        house=house,
        **payment_filter,

    )

    totals = funds.aggregate(
        total_income=Sum('debtor_amount'),
        total_expense=Sum('creditor_amount'),
    )

    balance = (totals['total_income'] or 0) - (totals['total_expense'] or 0)

    total_charge_unpaid = UnifiedCharge.objects.filter(
        house=house,
        is_paid=False
    ).aggregate(total=Sum('total_charge_month'))['total'] or 0

    context = {

        'house': house,
        'total_incomes': total_incomes,
        'total_expenses': total_expenses,
        'total_pay_money': total_pay_money,
        'total_receive_money': total_receive_money,
        'total_assets': total_assets,
        'total_debts': total_debts,
        'balance': balance,
        'total_charge_unpaid': total_charge_unpaid,
        'total_amount_assets_debts': total_amount_assets_debts,
        # 'start_date': start_date,
        # 'end_date': end_date,
        'start_date': start_date_j,  # همینو به input میدیم
        'end_date': end_date_j,
        'bank_id': bank_id,
        'banks': Bank.objects.filter(house=house),
        'total_incomes_exclude_unpaid': total_incomes_exclude_unpaid,
        'total_expenses_exclude_unpaid': total_expenses_exclude_unpaid,
        'today': timezone.now()
    }

    return render(request, 'report/house_balance_detail.html', context)
